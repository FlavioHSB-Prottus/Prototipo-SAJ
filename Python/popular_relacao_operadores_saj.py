# -*- coding: utf-8 -*-
"""
Popula a tabela `relacao_contrato_operador` a partir do Excel
`Banco/Operadores Iniciais.xlsx` (base SAJ antiga).

Execute este script **antes** da importacao TXT / antes de
`distribuir_funcionarios_cobranca.py`. O script de distribuicao apenas
**consulta** a tabela ja populada.

Cabecalho esperado na planilha: GRUPO, COTA, OPERADOR; opcionalmente
CONTRATO SISTEMA.

Grupo e cota sao normalizados (6 + 4 digitos, ex.: 0190050002).

Ambiente (opcional): DB_HOST, DB_USER, DB_PASSWORD, DB_NAME
                       OPERADORES_SAJ_XLSX (caminho absoluto ou relativo ao projeto)

Uso (na raiz do repositorio):
    python Python/popular_relacao_operadores_saj.py
"""
import os
import sys

import pymysql

try:
    import openpyxl
except ImportError:
    openpyxl = None

_ROOT_DIR = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
_DEFAULT_OPERADORES_XLSX = os.path.join(_ROOT_DIR, "Banco", "Operadores Iniciais.xlsx")


def connect_db():
    return pymysql.connect(
        host=os.environ.get("DB_HOST", "localhost"),
        user=os.environ.get("DB_USER", "root"),
        password=os.environ.get("DB_PASSWORD", "root"),
        database=os.environ.get("DB_NAME", "consorcio_gm"),
        cursorclass=pymysql.cursors.DictCursor,
        charset="utf8mb4",
    )


def normalizar_grupo_cota_int(grupo, cota):
    """Grupo 6 digitos + cota 4 digitos (ex.: 19005 + 2 -> 0190050002)."""
    try:
        g = int(float(str(grupo).strip()))
        ct = int(float(str(cota).strip()))
    except (TypeError, ValueError):
        return None, None, None
    g_s = f"{g:06d}"
    c_s = f"{ct:04d}"
    return g_s, c_s, g_s + c_s


def distribuicao_saj_antigo(cursor, caminho_xlsx=None):
    """Le o Excel e repovoa `relacao_contrato_operador`.

    Retorna numero de linhas inseridas, 0 se vazio, ou -1 em erro/indisponivel.
    """
    if openpyxl is None:
        print("ERRO: instale openpyxl: pip install openpyxl")
        return -1

    path = caminho_xlsx or os.environ.get("OPERADORES_SAJ_XLSX", "").strip() or _DEFAULT_OPERADORES_XLSX
    path = os.path.abspath(path)
    if not os.path.isfile(path):
        print(f"ERRO: arquivo nao encontrado: {path}")
        return -1

    try:
        cursor.execute(
            """
            SELECT COUNT(*) AS n FROM information_schema.tables
            WHERE table_schema = DATABASE() AND table_name = 'relacao_contrato_operador'
            """
        )
        if int((cursor.fetchone() or {}).get("n") or 0) == 0:
            print(
                "ERRO: tabela relacao_contrato_operador inexistente. "
                "Execute Banco/criar_banco.py primeiro."
            )
            return -1
    except Exception as exc:
        print(f"ERRO ao verificar tabela: {exc}")
        return -1

    rows_out = []
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        header_map = {}
        header_row_idx = None
        for ri, row in enumerate(ws.iter_rows(min_row=1, max_row=50, values_only=True), start=1):
            cells = [
                (i, str(x).strip().upper() if x is not None and str(x).strip() != "" else "")
                for i, x in enumerate(row)
            ]
            labels = [c[1] for c in cells]
            if "GRUPO" in labels and "COTA" in labels:
                op_lab = None
                for lab in ("OPERADOR", "OPERADOR ", "OPERADOR SAJ"):
                    if lab.strip() in labels:
                        op_lab = lab.strip()
                        break
                if op_lab is None:
                    for idx, lab in enumerate(labels):
                        if lab.startswith("OPER"):
                            op_lab = lab
                            break
                if op_lab is None:
                    continue
                header_row_idx = ri
                for idx, lab in enumerate(labels):
                    if lab in ("GRUPO", "COTA"):
                        header_map[lab] = idx
                    elif lab.replace(".", "") in ("OPERADOR",):
                        header_map["OPERADOR"] = idx
                    elif "CONTRATO" in lab and "SISTEMA" in lab:
                        header_map["CONTRATO"] = idx
                if "OPERADOR" not in header_map:
                    header_map["OPERADOR"] = labels.index(op_lab) if op_lab in labels else None
                break
        wb.close()
    except Exception as exc:
        print(f"ERRO ao ler Excel ({path}): {exc}")
        return -1

    if not header_row_idx or "GRUPO" not in header_map or "COTA" not in header_map:
        print(f"ERRO: cabecalho GRUPO/COTA/OPERADOR nao localizado em {path}")
        return -1
    op_col = header_map.get("OPERADOR")
    if op_col is None:
        print(f"ERRO: coluna OPERADOR nao localizada em {path}")
        return -1
    ct_col = header_map["CONTRATO"] if "CONTRATO" in header_map else None

    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        ig = header_map["GRUPO"]
        ic = header_map["COTA"]
        for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
            if not row:
                continue
            gv = row[ig] if ig < len(row) else None
            cv = row[ic] if ic < len(row) else None
            ov = row[op_col] if op_col < len(row) else None
            if gv is None or cv is None:
                continue
            if str(gv).strip() == "" and str(cv).strip() == "":
                continue
            op_txt = (str(ov).strip() if ov is not None else "") or ""
            if not op_txt or op_txt.upper() in ("NONE", "NAN"):
                continue
            g_s, c_s, gc = normalizar_grupo_cota_int(gv, cv)
            if not gc:
                continue
            contrato_txt = None
            if ct_col is not None and ct_col < len(row) and row[ct_col] is not None:
                contrato_txt = str(row[ct_col]).strip()
            rows_out.append((contrato_txt, g_s, c_s, gc, op_txt))
        wb.close()
    except Exception as exc:
        print(f"ERRO ao processar linhas do Excel: {exc}")
        return -1

    if not rows_out:
        print(f"AVISO: nenhuma linha valida no arquivo: {path}")
        return 0

    por_gc = {}
    for r in rows_out:
        por_gc[r[3]] = r
    batch = list(por_gc.values())

    cursor.execute("DELETE FROM relacao_contrato_operador")
    cursor.executemany(
        """
        INSERT INTO relacao_contrato_operador
            (numero_contrato_saj, grupo, cota, grupo_cota, nome_operador)
        VALUES (%s, %s, %s, %s, %s)
        """,
        batch,
    )
    n = len(batch)
    print(f"OK: {n} registro(s) em relacao_contrato_operador <- {path}")
    return n


def main():
    try:
        conn = connect_db()
        cursor = conn.cursor()
    except Exception as e:
        print(f"ERRO: conexao com o banco: {e}")
        sys.exit(1)

    n = distribuicao_saj_antigo(cursor)
    if n < 0:
        conn.rollback()
        cursor.close()
        conn.close()
        sys.exit(1)
    conn.commit()
    cursor.close()
    conn.close()
    sys.exit(0 if n >= 0 else 1)


if __name__ == "__main__":
    main()
