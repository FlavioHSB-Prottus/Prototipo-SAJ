import calendar
import datetime
import decimal
import functools
import html as html_module
import io
import json
import mimetypes
import secrets
import os
import re
import shutil
import subprocess
import sys
import tempfile
import threading
import time
import uuid
import importlib.util

import pymysql
import requests
from pymysql.err import IntegrityError, OperationalError
from flask import Flask, Response, render_template, request, redirect, url_for, jsonify, send_file, session, flash, abort, stream_with_context
from werkzeug.utils import secure_filename

app = Flask(__name__)
app.secret_key = os.environ.get('FLASK_SECRET_KEY', 'dev-consorcio-gm-altere-em-producao')
# JSON com caracteres Unicode (acentos, cedilha) sem escape \\uXXXX no fio
app.config['JSON_AS_ASCII'] = False

PROJECT_DIR = os.path.dirname(os.path.abspath(__file__))
try:
    from dotenv import load_dotenv
    load_dotenv(os.path.join(PROJECT_DIR, '.env'))
except ImportError:
    pass
PYTHON_DIR = os.path.join(PROJECT_DIR, 'Python')
PYTHON_EXE = sys.executable


def _load_google_workspace_smtp():
    """Carrega ``Python/google_workspace_smtp.py`` (SMTP Google na Importacao)."""
    path = os.path.join(PYTHON_DIR, 'google_workspace_smtp.py')
    spec = importlib.util.spec_from_file_location('google_workspace_smtp', path)
    if spec is None or spec.loader is None:
        raise ImportError(f'Nao foi possivel carregar {path}')
    mod = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(mod)
    return mod


_smtp_google_workspace = _load_google_workspace_smtp()

_serasa_conv_txt_module = None
_serasa_pefin_erros_module = None


def _get_serasa_pefin_erros():
    """Carrega ``Python/serasa_pefin_erros.py`` (tabela PEFIN + parse de linhas de retorno)."""
    global _serasa_pefin_erros_module
    if _serasa_pefin_erros_module is None:
        path = os.path.join(PYTHON_DIR, 'serasa_pefin_erros.py')
        spec = importlib.util.spec_from_file_location('serasa_pefin_erros', path)
        if spec is None or spec.loader is None:
            raise ImportError(f'Nao foi possivel carregar {path}')
        mod = importlib.util.module_from_spec(spec)
        spec.loader.exec_module(mod)
        _serasa_pefin_erros_module = mod
    return _serasa_pefin_erros_module


def _get_serasa_conv_txt():
    """Carrega ``Python/serasa_conv_txt.py`` sob demanda (layout SERASA-CONVEM 600 chars)."""
    global _serasa_conv_txt_module
    if _serasa_conv_txt_module is None:
        path = os.path.join(PYTHON_DIR, 'serasa_conv_txt.py')
        spec = importlib.util.spec_from_file_location('serasa_conv_txt', path)
        if spec is None or spec.loader is None:
            raise ImportError(f'Nao foi possivel carregar {path}')
        mod = importlib.util.module_from_spec(spec)
        spec.loader.exec_module(mod)
        _serasa_conv_txt_module = mod
    return _serasa_conv_txt_module

_SUBPROCESS_ENV = {**os.environ, 'PYTHONUNBUFFERED': '1'}
_POPEN_EXTRA = {}
if sys.platform == 'win32':
    _POPEN_EXTRA['creationflags'] = subprocess.CREATE_NO_WINDOW

def _db_port():
    """Porta MySQL a partir de DB_PORT ou MYSQL_PORT (compatível com scripts em Python/)."""
    raw = os.environ.get('DB_PORT', os.environ.get('MYSQL_PORT', '3306'))
    try:
        return int(raw)
    except (TypeError, ValueError):
        return 3306


# Preferir variáveis de ambiente (mesmos nomes dos scripts Python/: DB_*).
# MYSQL_* é alias opcional. Sem env, mantém o comportamento anterior (localhost/root).
DB_CONFIG = {
    'host': os.environ.get('DB_HOST', os.environ.get('MYSQL_HOST', 'localhost')),
    'port': _db_port(),
    'user': os.environ.get('DB_USER', os.environ.get('MYSQL_USER', 'root')),
    'password': os.environ.get('DB_PASSWORD', os.environ.get('MYSQL_PASSWORD', 'root')),
    'database': os.environ.get('DB_NAME', os.environ.get('MYSQL_DATABASE', 'consorcio_gm')),
    'charset': 'utf8mb4',
}

# Inatividade: tempo maximo (segundos) sem renovar `session['_idle_last']` antes de encerrar a sessao.
_SESSION_IDLE_MAX_SEC = int(os.environ.get('SESSION_IDLE_MAX_SEC', '3600'))


def _request_skips_idle_touch():
    """Pedidos automaticos que nao devem renovar a marca de actividade da sessao."""
    p = request.path or ''
    if p.startswith('/api/importacao/background/'):
        return True
    if p.startswith('/api/cobranca/ligacao-bloco/'):
        return True
    if p == '/api/notificacoes':
        return True
    return False


def _session_idle_expired_response():
    """Limpa a sessao e devolve resposta adequada (API vs pagina)."""
    session.clear()
    if request.path.startswith('/api/'):
        return jsonify(
            {'error': 'Sessão encerrada por inatividade.', 'sessao_expirada': True}
        ), 401
    flash('Sua sessão expirou por inatividade. Faça login novamente.', 'message')
    return redirect(url_for('login'))


def _enforce_and_touch_session_idle():
    """Se autenticado: expira sessao apos inactividade; actualiza `_idle_last` quando aplicavel."""
    if not session.get('funcionario_id'):
        return None
    now = time.time()
    raw = session.get('_idle_last')
    if raw is not None:
        try:
            last = float(raw)
        except (TypeError, ValueError):
            last = None
        if last is not None and (now - last) > _SESSION_IDLE_MAX_SEC:
            return _session_idle_expired_response()
    if not _request_skips_idle_touch():
        session['_idle_last'] = now
        session.modified = True
    return None


# Status em BD: contrato.status, parcela.status, ocorrencia.status; performance.ocorrencia_status.
# ocorrencia.status (contrato no GM): quitacao por pagamento usa 'pago total' ou 'pago parcial'
# (tracker); parcela/contrato continuam cobranca|pago|indenizado.
_STATUS_BD_COBRANCA = 'cobranca'
_STATUS_BD_PAGO = 'pago'
_STATUS_OCORRENCIA_PAGO_TOTAL = 'pago total'
_LEGACY_STATUS_CONTRATO_MAP = {'aberto': _STATUS_BD_COBRANCA, 'fechado': _STATUS_BD_PAGO}
_STATUS_CONTRATO_VALIDOS = frozenset((_STATUS_BD_COBRANCA, _STATUS_BD_PAGO, 'indenizado'))


def _normalize_contrato_status_arg(val):
    """Aceita valores actuais ou URLs antigas (?status=aberto|fechado)."""
    s = (val or '').strip().lower()
    return _LEGACY_STATUS_CONTRATO_MAP.get(s, s)


def _funcionario_esta_ativo(val):
    """Interpreta a coluna `ativo` (bit(1)) retornada pelo PyMySQL."""
    if val is None:
        return False
    if isinstance(val, (bytes, bytearray)):
        return len(val) > 0 and val != b'\x00'
    if isinstance(val, int):
        return val != 0
    return bool(val)


def _nivel_normalizado(val):
    """Normaliza `funcionario.nivel_acesso` para gestor | administrador | cobranca.

    Aceita legacy 'Cobranca' sem cedilha e 'Cobrança'. Qualquer valor desconhecido
    cai em 'cobranca' (perfil mais restrito).
    """
    s = (val or '').strip().lower()
    if s in ('cobranca', 'cobrança'):
        return 'cobranca'
    if s == 'gestor':
        return 'gestor'
    if s == 'administrador':
        return 'administrador'
    return 'cobranca'


def _nivel_acesso_valor_db_amigavel(raw):
    """Valor gravado no MySQL: Gestor, Administrador ou Cobrança."""
    n = _nivel_normalizado(str(raw or '').strip())
    return {'gestor': 'Gestor', 'administrador': 'Administrador', 'cobranca': 'Cobrança'}.get(n, 'Cobrança')


def _session_pode_gerir_operadores():
    """Página Operadores e APIs /api/admin/funcionario: Gestor ou Administrador."""
    return _nivel_normalizado(session.get('funcionario_nivel_acesso')) in ('gestor', 'administrador')


def _admin_json_forbidden():
    if not _session_pode_gerir_operadores():
        return jsonify({'error': 'Acesso restrito a gestores ou administradores.'}), 403
    return None


_EMAIL_ENVIO_PERFIL_COBRANCA_MSG = 'voce nao tem permissao'


def _cobranca_json_sem_permissao_envio_email():
    """Perfil Cobrança não envia e-mail (MessageCenter / registro_email)."""
    if _nivel_normalizado(session.get('funcionario_nivel_acesso')) == 'cobranca':
        return jsonify({'error': _EMAIL_ENVIO_PERFIL_COBRANCA_MSG}), 403
    return None


_COBRANCA_API_PREFIXES_OK = (
    '/api/busca',
    '/api/contrato/',
    '/api/pessoa/',
    '/api/tramitacao',
    '/api/pasta-virtual',
    '/api/relatorios',
    '/api/agenda',
    '/api/protocolos',
    '/api/protocolo',
    '/api/solicitacoes',
    '/api/solicitacao/moderacao',
    '/api/solicitacao',
    '/api/mensagens',
    '/api/mensagem',
    '/api/cobranca',
    '/api/funcionarios',
    '/api/funcionario/perfil',
    '/api/avisos',
    '/api/notificacoes',
    '/api/automacao/',
    '/api/negativacao/',
    '/api/enviar-sms',
    '/api/enviar-email-html',
    '/api/enviar-whatsapp',
)


def _cobranca_api_ok(path):
    return any(path.startswith(p) for p in _COBRANCA_API_PREFIXES_OK)


def _cobranca_html_ok(path):
    if path in ('/', '/home'):
        return True
    if path.startswith('/minha-foto'):
        return True
    prefixes = (
        '/cobranca',
        '/negativacao',
        '/busca',
        '/pasta-virtual',
        '/relatorios',
        '/agenda',
        '/protocolo',
        '/solicitacao',
        '/mensagem',
    )
    return any(path.startswith(p) for p in prefixes)


@app.before_request
def _session_idle_and_touch():
    """Expira sessao por inactividade e renova `_idle_last` antes dos outros before_request."""
    if request.endpoint is None:
        return None
    if request.endpoint in ('login', 'static', 'recuperar_senha', 'logout'):
        return None
    if not session.get('funcionario_id'):
        return None
    return _enforce_and_touch_session_idle()


@app.before_request
def _enforce_nivel_acesso_modulos():
    """Gestor: tudo. Administrador: tudo exceto Performance JB. Cobrança: módulos limitados."""
    if request.endpoint is None:
        return None
    if request.endpoint in ('login', 'static', 'recuperar_senha', 'logout'):
        return None
    if not session.get('funcionario_id'):
        return None
    path = request.path
    if path.startswith('/static'):
        return None

    n = _nivel_normalizado(session.get('funcionario_nivel_acesso'))

    if n == 'gestor':
        return None

    if n == 'cobranca':
        if path.startswith('/api/'):
            if not _cobranca_api_ok(path):
                return None #jsonify({'error': 'Acesso negado.'}), 403
            return None
        if not _cobranca_html_ok(path):
            flash('Seu perfil não tem acesso a esta área.', 'error')
            return redirect(url_for('home'))
        return None

    if path.startswith('/performance') or path.startswith('/api/performance'):
        if path.startswith('/api/'):
            return jsonify({'error': 'Acesso negado ao módulo Performance JB.'}), 403
        flash('Seu perfil não tem acesso ao módulo Performance JB.', 'error')
        return redirect(url_for('home'))

    return None


def _session_funcionario_empresa():
    return (session.get('funcionario_empresa') or 'GM').strip()


def _session_empresa_ativa():
    return (session.get('empresa_ativa') or 'GM').strip()


def _session_funcionario_login():
    return (session.get('funcionario_login') or '').strip().lower()


def _pode_selecionar_empresa_bradesco():
    """Apenas gestor.jb (multi-empresa) e bradesco.demo podem usar o contexto Bradesco."""
    login = _session_funcionario_login()
    return login in ('gestor.jb', 'bradesco.demo')


def _contexto_empresa_bradesco_ativo():
    return _session_empresa_ativa().strip().lower() == 'bradesco'


def _empresas_dropdown_sessao():
    """Duplas (id, titulo) para o seletor do header."""
    fe = _session_funcionario_empresa().lower()
    login = _session_funcionario_login()
    if fe == 'todas':
        out = [('GM', 'Consórcio GM')]
        if login == 'gestor.jb':
            out.append(('Bradesco', 'Bradesco'))
        return out
    if fe == 'bradesco':
        if not _pode_selecionar_empresa_bradesco():
            return [('GM', 'Consórcio GM')]
        return [('Bradesco', 'Bradesco'), ('GM', 'GM')]
    return [('GM', 'Consórcio GM')]


def _bradesco_contexto_gm_restrito():
    """Usuario vinculado a Bradesco com empresa ativa GM: só Busca + Cadastro."""
    return (
        _session_funcionario_empresa().lower() == 'bradesco'
        and _session_empresa_ativa().upper() == 'GM'
    )


def _path_ok_bradesco_em_gm(path):
    """Rotas permitidas quando Bradesco seleciona contexto GM."""
    if path in ('/', '/home'):
        return True
    if path.startswith('/static') or path.startswith('/minha-foto'):
        return True
    if path.startswith('/busca') or path.startswith('/cadastro'):
        return True
    if not path.startswith('/api/'):
        return False
    prefixos = (
        '/api/busca',
        '/api/consorciados',
        '/api/avalistas',
        '/api/contrato/',
        '/api/pessoa/',
        '/api/discar',
        '/api/enviar-sms',
        '/api/enviar-email-html',
        '/api/enviar-whatsapp',
        '/api/tramitacao',
        '/api/funcionario/perfil',
        '/api/avisos',
        '/api/notificacoes',
        '/api/upload',
    )
    return any(path.startswith(p) for p in prefixos)


@app.before_request
def _corrige_sessao_bradesco_nao_permitida():
    """Quem não é gestor.jb nem bradesco.demo não mantém empresa_ativa Bradesco (ex.: sessão antiga)."""
    if request.endpoint is None:
        return None
    if request.endpoint in ('login', 'static', 'recuperar_senha', 'logout'):
        return None
    if not session.get('funcionario_id'):
        return None
    if _contexto_empresa_bradesco_ativo() and not _pode_selecionar_empresa_bradesco():
        session['empresa_ativa'] = 'GM'
        session.modified = True
    return None


@app.before_request
def _enforce_empresa_escopo_bradesco_gm():
    """Bradesco + contexto GM ativo: apenas Busca e Consorciados/Avalistas (+ APIs auxiliares)."""
    if request.endpoint is None:
        return None
    if request.endpoint in ('login', 'static', 'recuperar_senha', 'logout'):
        return None
    if not session.get('funcionario_id'):
        return None
    path = request.path
    if path.startswith('/static'):
        return None
    # Troca de empresa no servidor
    if path.startswith('/api/sessao/empresa'):
        return None
    if path.startswith('/api/sessao/atividade'):
        return None
    if not _bradesco_contexto_gm_restrito():
        return None
    if _path_ok_bradesco_em_gm(path):
        return None
    if path.startswith('/api/'):
        return jsonify({'error': 'Acesso negado neste contexto de empresa (GM).'}), 403
    flash('No contexto GM você só acessa Busca e Consorciados e Avalistas.', 'error')
    return redirect(url_for('home'))


@app.before_request
def _enforce_bradesco_contexto_sem_modulos():
    """Empresa ativa Bradesco: demonstração sem módulos de negócio (só Home + perfil + troca de empresa)."""
    if request.endpoint is None:
        return None
    if request.endpoint in ('login', 'static', 'recuperar_senha', 'logout'):
        return None
    if not session.get('funcionario_id'):
        return None
    path = request.path
    if path.startswith('/static'):
        return None
    if not _contexto_empresa_bradesco_ativo():
        return None

    if path.startswith('/api/sessao/empresa'):
        return None
    if path.startswith('/api/sessao/atividade'):
        return None
    if path.startswith('/api/funcionario/perfil'):
        return None
    if path.startswith('/api/avisos'):
        return None
    if path.startswith('/minha-foto'):
        return None

    if path in ('/', '/home'):
        return None

    if path.startswith('/api/'):
        return jsonify({'error': 'Módulos indisponíveis no contexto Bradesco (demonstração).'}), 403

    flash(
        'No contexto Bradesco não há módulos implementados. Selecione Consórcio GM no topo '
        'para usar o sistema ou permaneça na Home.',
        'error',
    )
    return redirect(url_for('home'))


@app.context_processor
def inject_perfil_ui():
    na = session.get('funcionario_nivel_acesso')
    n = _nivel_normalizado(na)
    labels = {'gestor': 'Gestor', 'administrador': 'Administrador', 'cobranca': 'Cobrança'}
    label = labels.get(n, (na or '').strip() or 'Usuário')
    pode_op = n in ('gestor', 'administrador')
    br_gm = bool(session.get('funcionario_id')) and _bradesco_contexto_gm_restrito()
    br_demo = bool(session.get('funcionario_id')) and _contexto_empresa_bradesco_ativo()
    ctx = {
        'perfil_header_label': label,
        'is_admin': pode_op,
        'nav_importacao': n in ('gestor', 'administrador'),
        'nav_dashboard': n in ('gestor', 'administrador'),
        'nav_performance': n == 'gestor',
        'nav_cadastro': n in ('gestor', 'administrador'),
        'nav_operadores': pode_op,
        'nav_restrict_bradesco_gm': br_gm,
        'nav_bradesco_demonstracao': br_demo,
        'empresa_ativa_header': _session_empresa_ativa(),
        'empresas_dropdown': _empresas_dropdown_sessao(),
    }
    if br_gm:
        ctx['nav_importacao'] = False
        ctx['nav_dashboard'] = False
        ctx['nav_performance'] = False
        ctx['nav_operadores'] = False
        ctx['nav_cadastro'] = True
    return ctx


FOTO_MAX_BYTES = 5 * 1024 * 1024


def _guess_image_mimetype(data):
    """Retorna MIME se os bytes parecem JPEG, PNG, GIF ou WebP; caso contrário None."""
    if not data or len(data) < 12:
        return None
    if data[:3] == b'\xff\xd8\xff':
        return 'image/jpeg'
    if data[:8] == b'\x89PNG\r\n\x1a\n':
        return 'image/png'
    if len(data) >= 6 and data[:6] in (b'GIF87a', b'GIF89a'):
        return 'image/gif'
    if data[:4] == b'RIFF' and data[8:12] == b'WEBP':
        return 'image/webp'
    return None


def _mimetype_for_stored_blob(data):
    return _guess_image_mimetype(data) or 'application/octet-stream'


def _avatar_placeholder_path():
    return os.path.join(app.root_path, 'static', 'imagens', 'avatar-placeholder.png')


def _send_avatar_placeholder():
    path = _avatar_placeholder_path()
    if not os.path.isfile(path):
        abort(404)
    resp = send_file(path, mimetype='image/png')
    resp.headers['Cache-Control'] = 'private, no-store'
    resp.headers['Vary'] = 'Cookie'
    return resp


@app.route('/', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        login_val = (request.form.get('login') or '').strip()
        senha_val = request.form.get('senha') or ''
        if not login_val:
            flash('Informe o login.', 'error')
            return render_template('login.html')
        conn = None
        row = None
        try:
            conn = _get_db()
            with conn.cursor() as cursor:
                try:
                    cursor.execute(
                        'SELECT id, nome, senha, ativo, nivel_acesso, empresa '
                        'FROM funcionario WHERE login = %s LIMIT 1',
                        (login_val,),
                    )
                    row = cursor.fetchone()
                except OperationalError as op_err:
                    # 1054 = coluna inexistente (banco ainda sem migration da coluna `empresa`)
                    if op_err.args and op_err.args[0] == 1054:
                        app.logger.warning(
                            "login: coluna 'empresa' ausente em funcionario; "
                            "rodar ALTER/seed ou criar_banco. Consultando sem 'empresa'."
                        )
                        cursor.execute(
                            'SELECT id, nome, senha, ativo, nivel_acesso '
                            'FROM funcionario WHERE login = %s LIMIT 1',
                            (login_val,),
                        )
                        row = cursor.fetchone()
                    else:
                        raise
        except Exception:
            app.logger.exception('login: falha ao consultar funcionario')
            flash('Não foi possível validar o login. Tente novamente.', 'error')
            return render_template('login.html')
        finally:
            if conn is not None:
                conn.close()

        if not row or not _funcionario_esta_ativo(row.get('ativo')):
            flash('Login ou senha incorretos.', 'error')
            return render_template('login.html')

        db_senha = row.get('senha')
        if db_senha is None:
            db_senha = ''
        else:
            db_senha = str(db_senha)
        if not secrets.compare_digest(str(senha_val), db_senha):
            flash('Login ou senha incorretos.', 'error')
            return render_template('login.html')

        session.clear()
        session['funcionario_id'] = int(row['id'])
        session['funcionario_login'] = login_val.strip().lower()
        session['funcionario_nome'] = row.get('nome') or ''
        session['funcionario_nivel_acesso'] = row.get('nivel_acesso') or ''
        fe = (row.get('empresa') or 'GM')
        if not isinstance(fe, str):
            fe = 'GM'
        fe = fe.strip() or 'GM'
        session['funcionario_empresa'] = fe
        fl = fe.lower()
        if fl == 'todas':
            session['empresa_ativa'] = 'GM'
        elif fl == 'bradesco':
            session['empresa_ativa'] = 'Bradesco'
        else:
            session['empresa_ativa'] = fe
        session['_idle_last'] = time.time()
        session.modified = True
        return redirect(url_for('home'))

    if session.get('funcionario_id'):
        return redirect(url_for('home'))
    return render_template('login.html')


@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login'))


@app.route('/api/sessao/empresa', methods=['GET', 'POST'])
def api_sessao_empresa():
    """Define empresa ativa no contexto (GM / Bradesco / ...)."""
    if not session.get('funcionario_id'):
        return jsonify({'error': 'Não autenticado.'}), 401
    if request.method == 'GET':
        opcoes = _empresas_dropdown_sessao()
        return jsonify({
            'funcionario_empresa': _session_funcionario_empresa(),
            'empresa_ativa': _session_empresa_ativa(),
            'opcoes': [{'id': o[0], 'titulo': o[1]} for o in opcoes],
        })
    data = request.get_json(silent=True) or {}
    nova = (data.get('empresa') or '').strip()
    permitidas = [x[0] for x in _empresas_dropdown_sessao()]
    if nova not in permitidas:
        return jsonify({'error': 'Empresa não permitida para este usuário.'}), 400
    if nova.strip().lower() == 'bradesco' and not _pode_selecionar_empresa_bradesco():
        return jsonify({'error': 'Empresa Bradesco não disponível para este usuário.'}), 403
    session['empresa_ativa'] = nova
    session.modified = True
    return jsonify({'ok': True, 'empresa_ativa': nova})


@app.route('/api/sessao/atividade', methods=['POST'])
def api_sessao_atividade():
    """Marca uso humano no servidor (renova `_idle_last` via before_request)."""
    if not session.get('funcionario_id'):
        return jsonify({'error': 'Não autenticado.'}), 401
    return jsonify({'ok': True})


@app.route('/minha-foto', methods=['GET', 'POST'])
def minha_foto():
    fid = session.get('funcionario_id')
    if not fid:
        if request.method == 'POST':
            return jsonify({'error': 'Não autenticado'}), 401
        abort(404)

    if request.method == 'GET':
        conn = None
        try:
            conn = _get_db()
            with conn.cursor() as cursor:
                cursor.execute(
                    'SELECT foto FROM funcionario WHERE id = %s',
                    (int(fid),),
                )
                row = cursor.fetchone()
        finally:
            if conn is not None:
                conn.close()

        if not row:
            abort(404)
        foto = row.get('foto')
        if foto is None:
            return _send_avatar_placeholder()
        if isinstance(foto, memoryview):
            foto = foto.tobytes()
        elif not isinstance(foto, (bytes, bytearray)):
            foto = bytes(foto) if foto else b''
        if not foto:
            return _send_avatar_placeholder()
        mt = _mimetype_for_stored_blob(foto)
        return Response(
            foto,
            mimetype=mt,
            headers={
                'Cache-Control': 'private, no-store',
                'Vary': 'Cookie',
            },
        )

    f = request.files.get('foto')
    if not f or not getattr(f, 'filename', None):
        return jsonify({'error': 'Envie um ficheiro de imagem.'}), 400
    raw = f.read()
    if len(raw) > FOTO_MAX_BYTES:
        return jsonify({'error': 'Imagem demasiado grande (máx. 5 MB).'}), 400
    if not _guess_image_mimetype(raw):
        return jsonify({'error': 'Formato não suportado. Use JPEG, PNG, GIF ou WebP.'}), 400

    conn = None
    try:
        conn = _get_db()
        with conn.cursor() as cursor:
            cursor.execute(
                'UPDATE funcionario SET foto = %s WHERE id = %s',
                (raw, int(fid)),
            )
        conn.commit()
    except Exception:
        app.logger.exception('minha_foto POST')
        if conn is not None:
            conn.rollback()
        return jsonify({'error': 'Não foi possível guardar a foto.'}), 500
    finally:
        if conn is not None:
            conn.close()
    return jsonify({'ok': True})


@app.before_request
def _require_login():
    """Exige sessão de funcionário para todas as rotas, exceto login, logout, estáticos e recuperação de senha."""
    if request.endpoint is None:
        return None
    public = {'login', 'static', 'recuperar_senha', 'logout'}
    if request.endpoint in public:
        return None
    if not session.get('funcionario_id'):
        if request.path.startswith('/api/'):
            return jsonify({'error': 'Não autenticado'}), 401
        return redirect(url_for('login'))
    return None

@app.route('/home', methods=['GET', 'POST'])
def home():
    return render_template('home.html')

@app.route('/dashboard')
def dashboard():
    return render_template('dashboard.html')

@app.route('/busca')
def busca():
    return render_template('busca.html')

@app.route('/relatorios')
def relatorios():
    pode_email_massa = _session_pode_gerir_operadores()
    return render_template('relatorios.html', relatorios_email_massa=pode_email_massa)

@app.route('/performance')
def performance():
    return render_template('performance.html')

@app.route('/agenda')
def agenda():
    return render_template('agenda.html')

@app.route('/cadastro')
def cadastro():
    return render_template('cadastro.html')

@app.route('/importacao')
def importacao():
    return render_template('importacao.html')

@app.route('/cobranca')
def cobranca():
    # Perfil Cobrança: painel já abre filtrando pelo operador logado.
    # Gestor / Administrador: mantém "Todos os Operadores" (valor vazio).
    n = _nivel_normalizado(session.get('funcionario_nivel_acesso'))
    cobranca_page_config = {}
    if n == 'cobranca':
        fid = session.get('funcionario_id')
        try:
            fid = int(fid) if fid is not None else None
        except (TypeError, ValueError):
            fid = None
        if fid is not None:
            nome = (session.get('funcionario_nome') or '').strip() or ('#' + str(fid))
            cobranca_page_config = {
                'defaultOperadorId': fid,
                'defaultOperadorNome': nome,
            }
    return render_template('cobranca.html', cobranca_page_config=cobranca_page_config)


@app.route('/negativacao')
def negativacao():
    n = _nivel_normalizado(session.get('funcionario_nivel_acesso'))
    negativacao_page_config = {}
    if n == 'cobranca':
        fid = session.get('funcionario_id')
        try:
            fid = int(fid) if fid is not None else None
        except (TypeError, ValueError):
            fid = None
        if fid is not None:
            nome = (session.get('funcionario_nome') or '').strip() or ('#' + str(fid))
            negativacao_page_config = {
                'defaultFuncionarioCobrancaId': fid,
                'defaultFuncionarioCobrancaNome': nome,
                'perfilCobranca': True,
            }
    return render_template(
        'negativacao.html',
        negativacao_page_config=negativacao_page_config,
    )


@app.route('/operadores')
def operadores():
    if not _session_pode_gerir_operadores():
        flash('Acesso à página Operadores é restrito a gestores ou administradores.', 'error')
        return redirect(url_for('home'))
    return render_template('operadores.html')

@app.route('/protocolo')
def protocolo():
    return render_template('protocolo.html')

@app.route('/solicitacao')
def solicitacao():
    pode_revisar = False
    eh_cobranca = False
    if session.get('funcionario_id'):
        n = _nivel_normalizado(session.get('funcionario_nivel_acesso'))
        pode_revisar = n in ('gestor', 'administrador')
        eh_cobranca = n == 'cobranca'
    return render_template(
        'solicitacao.html',
        solicitacao_mod_cfg={'pode_revisar': pode_revisar, 'eh_cobranca': eh_cobranca},
    )

@app.route('/mensagem')
def mensagem():
    return render_template('mensagem.html')

@app.route('/pasta-virtual')
def pasta_virtual():
    return render_template('pasta_virtual.html')

@app.route('/recuperar_senha')
def recuperar_senha():
    return render_template('recuperar_senha.html')


# ---------------------------------------------------------------------------
# API: Pasta Virtual
# ---------------------------------------------------------------------------

_PASTA_VIRTUAL_TABLE = 'pasta_virtual'
_PV_BINARY_TYPES = {'tinyblob', 'blob', 'mediumblob', 'longblob', 'binary', 'varbinary'}


def _sql_ident(name):
    return '`' + str(name).replace('`', '``') + '`'


def _pv_bool(val):
    if isinstance(val, (bytes, bytearray)):
        return bool(val) and val not in (b'\x00', b'0')
    if isinstance(val, (int, float)):
        return int(val) != 0
    return bool(val)


def _pv_columns(cursor):
    cursor.execute(
        """
        SELECT COLUMN_NAME, DATA_TYPE, COLUMN_KEY, IS_NULLABLE, COLUMN_DEFAULT, EXTRA
        FROM INFORMATION_SCHEMA.COLUMNS
        WHERE TABLE_SCHEMA = %s AND TABLE_NAME = %s
        ORDER BY ORDINAL_POSITION
        """,
        (DB_CONFIG['database'], _PASTA_VIRTUAL_TABLE),
    )
    return cursor.fetchall()


def _pv_detect_meta(columns):
    names = [c['COLUMN_NAME'] for c in columns]
    lower_map = {n.lower(): n for n in names}

    pk_col = None
    for c in columns:
        if str(c.get('COLUMN_KEY') or '').upper() == 'PRI':
            pk_col = c['COLUMN_NAME']
            break
    if not pk_col:
        pk_col = lower_map.get('id') or (names[0] if names else None)

    contrato_col = lower_map.get('id_contrato')
    blob_col = None
    for c in columns:
        if (c.get('DATA_TYPE') or '').lower() in _PV_BINARY_TYPES:
            blob_col = c['COLUMN_NAME']
            break

    nome_col = None
    for cand in ('nome_arquivo', 'arquivo_nome', 'filename', 'file_name', 'nome', 'titulo'):
        if cand in lower_map:
            nome_col = lower_map[cand]
            break

    mime_col = None
    for cand in ('mimetype', 'mime_type', 'tipo_mime', 'content_type'):
        if cand in lower_map:
            mime_col = lower_map[cand]
            break

    order_col = None
    for cand in ('updated_at', 'created_at', 'data_arquivo', 'data', 'id'):
        if cand in lower_map:
            order_col = lower_map[cand]
            break
    if not order_col:
        order_col = pk_col

    funcionario_col = lower_map.get('id_funcionario')

    desc_col = None
    for cand in ('descricao', 'observacao', 'comentario', 'texto', 'notas'):
        if cand in lower_map:
            desc_col = lower_map[cand]
            break

    return {
        'pk_col': pk_col,
        'contrato_col': contrato_col,
        'funcionario_col': funcionario_col,
        'desc_col': desc_col,
        'blob_col': blob_col,
        'nome_col': nome_col,
        'mime_col': mime_col,
        'order_col': order_col,
    }


def _pv_required_insert_cols(columns, meta):
    required = []
    for c in columns:
        name = c['COLUMN_NAME']
        if name == meta.get('pk_col'):
            continue
        extra = str(c.get('EXTRA') or '').lower()
        if 'auto_increment' in extra:
            continue
        if (c.get('IS_NULLABLE') or '').upper() == 'YES':
            continue
        if c.get('COLUMN_DEFAULT') is not None:
            continue
        required.append(name)
    return required


@app.route('/api/pasta-virtual')
def api_pasta_virtual_list():
    conn = _get_db()
    cursor = conn.cursor()
    try:
        cols = _pv_columns(cursor)
        if not cols:
            return jsonify({'error': 'Tabela `pasta_virtual` não encontrada ou sem colunas.'}), 404
        meta = _pv_detect_meta(cols)
        if not meta.get('pk_col'):
            return jsonify({'error': 'Tabela `pasta_virtual` sem chave identificadora.'}), 500

        select_cols = []
        for c in cols:
            col_name = c['COLUMN_NAME']
            dt = (c.get('DATA_TYPE') or '').lower()
            if dt in _PV_BINARY_TYPES:
                continue
            select_cols.append(col_name)

        for req_col in (meta['pk_col'], meta['contrato_col'], meta['nome_col'], meta['mime_col']):
            if req_col and req_col not in select_cols:
                select_cols.append(req_col)

        select_expr = [f"pv.{_sql_ident(c)} AS {_sql_ident(c)}" for c in select_cols]
        if meta['blob_col']:
            select_expr.append(
                f"(pv.{_sql_ident(meta['blob_col'])} IS NOT NULL "
                f"AND OCTET_LENGTH(pv.{_sql_ident(meta['blob_col'])}) > 0) AS has_arquivo"
            )
        else:
            select_expr.append("0 AS has_arquivo")

        if meta['contrato_col']:
            select_expr.extend([
                "c.grupo AS contrato_grupo",
                "c.cota AS contrato_cota",
                "c.numero_contrato AS contrato_numero",
                "c.status AS contrato_status",
                "p.nome_completo AS contrato_nome_devedor",
            ])

        if meta.get('funcionario_col'):
            select_expr.append("f.nome AS pasta_virtual_funcionario_nome")

        sql = (
            "SELECT " + ", ".join(select_expr) +
            f" FROM {_sql_ident(_PASTA_VIRTUAL_TABLE)} pv "
        )
        if meta['contrato_col']:
            sql += (
                f"LEFT JOIN contrato c ON c.id = pv.{_sql_ident(meta['contrato_col'])} "
                "LEFT JOIN pessoa p ON p.id = c.id_pessoa "
            )
        if meta.get('funcionario_col'):
            sql += f"LEFT JOIN funcionario f ON f.id = pv.{_sql_ident(meta['funcionario_col'])} "

        where_clauses = []
        params = []

        contrato_arg = (request.args.get('contrato') or '').strip()
        if contrato_arg and meta.get('contrato_col'):
            if '/' in contrato_arg:
                parts = contrato_arg.split('/', 1)
                grupo_f = parts[0].strip()
                cota_f = (parts[1] if len(parts) > 1 else '').strip()
                if grupo_f:
                    where_clauses.append("c.grupo = %s")
                    params.append(grupo_f)
                    if cota_f:
                        where_clauses.append("c.cota = %s")
                        params.append(cota_f)
            else:
                where_clauses.append("c.grupo = %s")
                params.append(contrato_arg)

        id_func_arg = (request.args.get('id_funcionario') or '').strip()
        if id_func_arg and meta.get('funcionario_col'):
            try:
                id_func_int = int(id_func_arg)
            except (TypeError, ValueError):
                id_func_int = None
            if id_func_int is not None:
                where_clauses.append(f"pv.{_sql_ident(meta['funcionario_col'])} = %s")
                params.append(id_func_int)

        if where_clauses:
            sql += " WHERE " + " AND ".join(where_clauses)
        sql += f" ORDER BY pv.{_sql_ident(meta['order_col'])} DESC"

        if params:
            cursor.execute(sql, tuple(params))
        else:
            cursor.execute(sql)
        rows = _clean_rows(cursor.fetchall())

        items = []
        for r in rows:
            id_contrato = r.get(meta['contrato_col']) if meta['contrato_col'] else None
            g = r.get('contrato_grupo')
            cota = r.get('contrato_cota')
            contrato_label = None
            if g is not None and cota is not None:
                contrato_label = f"{g}/{cota}"
            elif id_contrato is not None:
                contrato_label = f"Contrato #{id_contrato}"
            _pv_campos_excl = {
                'has_arquivo', 'contrato_grupo', 'contrato_cota', 'contrato_numero', 'contrato_status',
                'contrato_nome_devedor', 'pasta_virtual_funcionario_nome',
            }
            if meta.get('desc_col'):
                _pv_campos_excl.add(meta['desc_col'])

            item = {
                'id': r.get(meta['pk_col']),
                'id_contrato': id_contrato,
                'contrato_label': contrato_label,
                'contrato_numero': r.get('contrato_numero'),
                'contrato_status': r.get('contrato_status'),
                'contrato_nome_devedor': r.get('contrato_nome_devedor'),
                'arquivo_nome': r.get(meta['nome_col']) if meta['nome_col'] else None,
                'mime_type': r.get(meta['mime_col']) if meta['mime_col'] else None,
                'funcionario_nome': r.get('pasta_virtual_funcionario_nome')
                if meta.get('funcionario_col') else None,
                'descricao': r.get(meta['desc_col']) if meta.get('desc_col') else None,
                'has_arquivo': _pv_bool(r.get('has_arquivo')),
                'campos': {k: v for k, v in r.items() if k not in _pv_campos_excl},
            }
            items.append(item)

        return jsonify({'items': items})
    finally:
        try:
            cursor.close()
        except Exception:
            pass
        try:
            conn.close()
        except Exception:
            pass


@app.route('/api/pasta-virtual/meta')
def api_pasta_virtual_meta():
    conn = _get_db()
    cursor = conn.cursor()
    try:
        cols = _pv_columns(cursor)
        if not cols:
            return jsonify({'error': 'Tabela `pasta_virtual` não encontrada.'}), 404
        meta = _pv_detect_meta(cols)
        required = _pv_required_insert_cols(cols, meta)
        return jsonify({
            'pk_col': meta.get('pk_col'),
            'contrato_col': meta.get('contrato_col'),
            'blob_col': meta.get('blob_col'),
            'nome_col': meta.get('nome_col'),
            'mime_col': meta.get('mime_col'),
            'required_cols': required,
            'columns': [c.get('COLUMN_NAME') for c in cols],
        })
    finally:
        try:
            cursor.close()
        except Exception:
            pass
        try:
            conn.close()
        except Exception:
            pass


@app.route('/api/pasta-virtual/inserir', methods=['POST'])
def api_pasta_virtual_insert():
    conn = _get_db()
    cursor = conn.cursor()
    try:
        cols = _pv_columns(cursor)
        if not cols:
            return jsonify({'error': 'Tabela `pasta_virtual` não encontrada.'}), 404
        meta = _pv_detect_meta(cols)
        valid_cols = {c['COLUMN_NAME'] for c in cols}
        lower_map = {c['COLUMN_NAME'].lower(): c['COLUMN_NAME'] for c in cols}
        required_cols = set(_pv_required_insert_cols(cols, meta))

        payload = {}
        for k in request.form.keys():
            if k not in valid_cols:
                continue
            if k == meta.get('pk_col'):
                continue
            val = request.form.get(k)
            if val is None:
                continue
            val = str(val).strip()
            if val == '':
                continue
            payload[k] = val

        # Resolve contrato por grupo/cota informado no formulario.
        grupo_in = (request.form.get('grupo') or '').strip()
        cota_in = (request.form.get('cota') or '').strip()
        if meta.get('contrato_col'):
            if not grupo_in or not cota_in:
                return jsonify({'error': 'Informe grupo e cota para localizar o contrato.'}), 400
            cursor.execute(
                "SELECT id FROM contrato WHERE grupo = %s AND cota = %s LIMIT 1",
                (grupo_in, cota_in),
            )
            c_row = cursor.fetchone()
            if not c_row:
                return jsonify({'error': f'Contrato nao encontrado para grupo/cota: {grupo_in}/{cota_in}'}), 404
            payload[meta['contrato_col']] = c_row.get('id')

        # id_funcionario deve vir da sessao do usuario logado.
        id_func_col = lower_map.get('id_funcionario')
        if id_func_col:
            fid = session.get('funcionario_id')
            if not fid:
                return jsonify({'error': 'Sessao invalida. Faca login novamente.'}), 401
            payload[id_func_col] = int(fid)

        if meta.get('blob_col'):
            f = request.files.get('arquivo')
            if f and getattr(f, 'filename', None):
                blob_data = f.read()
                if blob_data:
                    payload[meta['blob_col']] = blob_data
                    if meta.get('nome_col') and not payload.get(meta['nome_col']):
                        payload[meta['nome_col']] = secure_filename(f.filename) or f.filename
                    if meta.get('mime_col') and not payload.get(meta['mime_col']):
                        payload[meta['mime_col']] = f.mimetype or mimetypes.guess_type(f.filename)[0] or 'application/octet-stream'

        missing = [c for c in required_cols if c not in payload]
        if missing:
            return jsonify({
                'error': 'Campos obrigatórios ausentes na pasta_virtual.',
                'missing': missing,
            }), 400

        if not payload:
            return jsonify({'error': 'Nenhum campo válido para inserir.'}), 400

        insert_cols = list(payload.keys())
        sql = (
            f"INSERT INTO {_sql_ident(_PASTA_VIRTUAL_TABLE)} "
            f"({', '.join(_sql_ident(c) for c in insert_cols)}) "
            f"VALUES ({', '.join(['%s'] * len(insert_cols))})"
        )
        cursor.execute(sql, tuple(payload[c] for c in insert_cols))
        conn.commit()
        return jsonify({'ok': True, 'id': cursor.lastrowid})
    except Exception as exc:
        if conn:
            try:
                conn.rollback()
            except Exception:
                pass
        return jsonify({'error': f'Erro ao inserir em pasta_virtual: {exc}'}), 500
    finally:
        try:
            cursor.close()
        except Exception:
            pass
        try:
            conn.close()
        except Exception:
            pass


@app.route('/api/pasta-virtual/<item_id>/download')
def api_pasta_virtual_download(item_id):
    conn = _get_db()
    cursor = conn.cursor()
    try:
        cols = _pv_columns(cursor)
        if not cols:
            return jsonify({'error': 'Tabela `pasta_virtual` não encontrada.'}), 404
        meta = _pv_detect_meta(cols)
        if not meta.get('pk_col') or not meta.get('blob_col'):
            return jsonify({'error': 'Tabela `pasta_virtual` sem coluna de arquivo.'}), 400

        select_expr = [
            f"{_sql_ident(meta['blob_col'])} AS blob_data",
        ]
        if meta['nome_col']:
            select_expr.append(f"{_sql_ident(meta['nome_col'])} AS nome_arquivo")
        if meta['mime_col']:
            select_expr.append(f"{_sql_ident(meta['mime_col'])} AS mime_type")

        cursor.execute(
            "SELECT " + ", ".join(select_expr) +
            f" FROM {_sql_ident(_PASTA_VIRTUAL_TABLE)} "
            f"WHERE {_sql_ident(meta['pk_col'])} = %s LIMIT 1",
            (item_id,),
        )
        row = cursor.fetchone()
        if not row:
            return jsonify({'error': 'Registro não encontrado na pasta_virtual.'}), 404

        blob = row.get('blob_data')
        if isinstance(blob, memoryview):
            blob = blob.tobytes()
        elif blob is not None and not isinstance(blob, (bytes, bytearray)):
            blob = bytes(blob)
        if not blob:
            return jsonify({'error': 'Registro sem arquivo para download.'}), 404

        file_name = row.get('nome_arquivo') if meta['nome_col'] else None
        if not file_name:
            file_name = f"pasta_virtual_{item_id}.bin"
        file_name = secure_filename(str(file_name)) or f"pasta_virtual_{item_id}.bin"

        mime_type = row.get('mime_type') if meta['mime_col'] else None
        if not mime_type:
            guessed, _ = mimetypes.guess_type(file_name)
            mime_type = guessed or _mimetype_for_stored_blob(blob)

        return send_file(
            io.BytesIO(blob),
            as_attachment=True,
            download_name=file_name,
            mimetype=mime_type or 'application/octet-stream',
        )
    finally:
        try:
            cursor.close()
        except Exception:
            pass
        try:
            conn.close()
        except Exception:
            pass


# ---------------------------------------------------------------------------
# API: Upload + Processamento com SSE
# ---------------------------------------------------------------------------

def _safe_txt_dest(temp_dir, upload_filename):
    """Monta caminho seguro dentro de temp_dir para upload (inclui subpastas de importacao por pasta)."""
    if not upload_filename:
        return None
    rel = upload_filename.replace('\\', '/').strip('/')
    parts = []
    for p in rel.split('/'):
        if not p or p in ('.', '..'):
            continue
        s = secure_filename(p)
        if s:
            parts.append(s)
    if not parts or not parts[-1].lower().endswith('.txt'):
        return None
    dest = os.path.join(temp_dir, *parts)
    dest_abs = os.path.abspath(dest)
    temp_abs = os.path.abspath(temp_dir)
    try:
        common = os.path.commonpath([temp_abs, dest_abs])
    except ValueError:
        return None
    if common != temp_abs:
        return None
    return dest_abs


def _validate_incoming_temp_dir(incoming):
    """Aceita apenas pastas criadas por este endpoint (prefixo 'saj_import_' dentro do tempdir)."""
    if not incoming:
        return None
    tmp_root = os.path.abspath(tempfile.gettempdir())
    inc_abs = os.path.abspath(incoming)
    base_name = os.path.basename(inc_abs)
    if not os.path.isdir(inc_abs):
        return None
    if os.path.dirname(inc_abs) != tmp_root:
        return None
    if not base_name.startswith('saj_import_'):
        return None
    return inc_abs


@app.route('/api/upload', methods=['POST'])
def api_upload():
    """Recebe multiplos arquivos TXT e salva em pasta temporaria (estrutura de subpastas preservada).

    Suporta upload em lotes: se o cliente informar o campo 'temp_dir' de um lote
    anterior, os novos arquivos sao acumulados na mesma pasta temporaria.
    """
    files = request.files.getlist('files')
    if not files or all(f.filename == '' for f in files):
        return jsonify({'error': 'Nenhum arquivo enviado'}), 400

    incoming_temp = (request.form.get('temp_dir') or '').strip()
    created_new = False
    if incoming_temp:
        temp_dir = _validate_incoming_temp_dir(incoming_temp)
        if not temp_dir:
            return jsonify({'error': 'temp_dir invalido ou expirado'}), 400
    else:
        temp_dir = tempfile.mkdtemp(prefix='saj_import_')
        created_new = True

    saved = []
    for f in files:
        dest = _safe_txt_dest(temp_dir, f.filename)
        if not dest:
            continue
        os.makedirs(os.path.dirname(dest), exist_ok=True)
        f.save(dest)
        rel_name = os.path.relpath(dest, temp_dir).replace(os.sep, '/')
        saved.append({'name': rel_name, 'size': os.path.getsize(dest)})

    if not saved:
        if created_new:
            shutil.rmtree(temp_dir, ignore_errors=True)
        return jsonify({'error': 'Nenhum arquivo .txt valido enviado'}), 400

    return jsonify({'temp_dir': temp_dir, 'files': saved})


def _classify_log_level(line):
    """Classifica o nivel do log com base no conteudo da linha."""
    upper = line.upper()
    if 'ERRO' in upper or 'FALHA' in upper or 'ERROR' in upper:
        return 'alert'
    if 'SUCESSO' in upper or 'CONCLU' in upper or 'SUCCESS' in upper or 'FINALIZ' in upper:
        return 'success'
    if 'UPDATE' in upper or 'TRACKING' in upper or 'DELTA' in upper:
        return 'update'
    return 'info'


def _sse_event(data_dict):
    """Formata um dict como evento SSE."""
    return f"data: {json.dumps(data_dict, ensure_ascii=False)}\n\n"


_import_pipeline_events_fn = None
_import_jobs = {}
_import_job_lock = threading.Lock()


def _get_iter_importacao_events():
    """Carrega ``Python/importacao_pipeline_events.py`` uma vez (iter_importacao_events)."""
    global _import_pipeline_events_fn
    if _import_pipeline_events_fn is None:
        path = os.path.join(PYTHON_DIR, 'importacao_pipeline_events.py')
        spec = importlib.util.spec_from_file_location('importacao_pipeline_events', path)
        if spec is None or spec.loader is None:
            raise ImportError(f'Nao foi possivel carregar {path}')
        mod = importlib.util.module_from_spec(spec)
        spec.loader.exec_module(mod)
        _import_pipeline_events_fn = mod.iter_importacao_events
    return _import_pipeline_events_fn


def _importacao_background_worker(job_id, temp_dir, funcionario_id):
    """Executa o pipeline num thread daemon e grava eventos em ``_import_jobs``."""
    it_fn = _get_iter_importacao_events()
    job = _import_jobs.get(job_id)
    if not job:
        return
    try:
        for ev in it_fn(
            temp_dir,
            job,
            python_dir=PYTHON_DIR,
            python_exe=PYTHON_EXE,
            subprocess_env=_SUBPROCESS_ENV,
            popen_extra=_POPEN_EXTRA,
            schema_pronto=_schema_pronto_para_importacao,
            classify_log=_classify_log_level,
        ):
            with _import_job_lock:
                job['events'].append(ev)
                if ev.get('type') == 'progress':
                    job['progress'] = int(ev.get('value') or 0)
                elif ev.get('type') == 'status':
                    job['status_text'] = (ev.get('text') or '')[:500]
                if len(job['events']) > 4500:
                    job['events'] = job['events'][-3500:]
                if ev.get('type') in ('done', 'error'):
                    job['result'] = ev
                    break
    except Exception as exc:
        app.logger.exception('importacao background worker')
        with _import_job_lock:
            job['events'].append({'type': 'log', 'level': 'alert', 'text': str(exc)})
            job['events'].append({'type': 'error', 'text': str(exc)})
    finally:
        with _import_job_lock:
            job['running'] = False
            job['active_proc'] = None


@app.route('/api/processar')
def api_processar():
    """Endpoint SSE que executa as fases de importacao, streamando stdout."""
    temp_dir = request.args.get('dir', '')
    if not temp_dir or not os.path.isdir(temp_dir):
        def err():
            yield _sse_event({'type': 'error', 'text': 'Pasta temporaria invalida ou nao encontrada.'})
        return Response(err(), mimetype='text/event-stream')

    def generate():
        it_fn = _get_iter_importacao_events()
        for ev in it_fn(
            temp_dir,
            None,
            python_dir=PYTHON_DIR,
            python_exe=PYTHON_EXE,
            subprocess_env=_SUBPROCESS_ENV,
            popen_extra=_POPEN_EXTRA,
            schema_pronto=_schema_pronto_para_importacao,
            classify_log=_classify_log_level,
        ):
            yield _sse_event(ev)

    return Response(generate(), mimetype='text/event-stream')


@app.route('/api/importacao/background/start', methods=['POST'])
def api_importacao_background_start():
    """Inicia importacao GM em thread no servidor; retorna ``job_id`` para SSE/polling."""
    if not session.get('funcionario_id'):
        return jsonify({'error': 'Nao autenticado.'}), 401
    payload = request.get_json(silent=True) or {}
    raw = (payload.get('temp_dir') or '').strip()
    temp_dir = _validate_incoming_temp_dir(raw)
    if not temp_dir:
        return jsonify({'error': 'Pasta temporaria invalida ou expirada.'}), 400

    fid = session.get('funcionario_id')
    with _import_job_lock:
        existing = None
        for jid, j in _import_jobs.items():
            if j.get('funcionario_id') == fid and j.get('running'):
                existing = jid
                break
        if existing:
            return jsonify(
                {
                    'error': 'Ja existe uma importacao em andamento para sua sessao.',
                    'job_id': existing,
                }
            ), 409
        stale = [k for k, v in list(_import_jobs.items()) if v.get('funcionario_id') == fid and not v.get('running')]
        for k in stale[:8]:
            _import_jobs.pop(k, None)

    job_id = uuid.uuid4().hex
    job = {
        'funcionario_id': fid,
        'temp_dir': temp_dir,
        'running': True,
        'cancel_requested': False,
        'active_proc': None,
        'events': [],
        'progress': 0,
        'status_text': '',
        'result': None,
    }
    with _import_job_lock:
        _import_jobs[job_id] = job

    th = threading.Thread(
        target=_importacao_background_worker,
        args=(job_id, temp_dir, fid),
        daemon=True,
    )
    th.start()
    return jsonify({'job_id': job_id})


@app.route('/api/importacao/background/<job_id>/stream')
def api_importacao_background_stream(job_id):
    """SSE: eventos da importacao em background (parametro ``from`` = indice inicial)."""
    if not session.get('funcionario_id'):
        def err401():
            yield _sse_event({'type': 'error', 'text': 'Nao autenticado.'})
        return Response(err401(), mimetype='text/event-stream')

    with _import_job_lock:
        job = _import_jobs.get(job_id)
    if not job or str(job.get('funcionario_id')) != str(session.get('funcionario_id')):
        def err403():
            yield _sse_event({'type': 'error', 'text': 'Importacao nao encontrada ou sem permissao.'})
        return Response(err403(), mimetype='text/event-stream')

    try:
        start_idx = max(0, int(request.args.get('from') or 0))
    except (TypeError, ValueError):
        start_idx = 0

    def gen():
        last = start_idx
        while True:
            with _import_job_lock:
                evs = list(job['events'])
                running = job['running']
            while last < len(evs):
                ev = evs[last]
                yield _sse_event(ev)
                if ev.get('type') in ('done', 'error'):
                    return
                last += 1
            if not running and last >= len(evs):
                # Cliente ligou com from=len(evs) (ex.: job acabou entre snapshot e EventSource).
                if start_idx >= len(evs) and len(evs) > 0:
                    term = evs[-1]
                    if term.get('type') in ('done', 'error'):
                        yield _sse_event(term)
                return
            time.sleep(0.12)

    return Response(stream_with_context(gen()), mimetype='text/event-stream')


@app.route('/api/importacao/background/<job_id>/snapshot')
def api_importacao_background_snapshot(job_id):
    """JSON: fila de eventos ate ao momento (para reabrir log)."""
    if not session.get('funcionario_id'):
        return jsonify({'error': 'Nao autenticado.'}), 401
    with _import_job_lock:
        job = _import_jobs.get(job_id)
    if not job or str(job.get('funcionario_id')) != str(session.get('funcionario_id')):
        return jsonify({'error': 'Importacao nao encontrada.'}), 404
    return jsonify(
        {
            'events': job['events'],
            'progress': job['progress'],
            'status_text': job['status_text'],
            'running': job['running'],
        }
    )


@app.route('/api/importacao/background/<job_id>/state')
def api_importacao_background_state(job_id):
    """JSON leve para bolha de progresso (polling)."""
    if not session.get('funcionario_id'):
        return jsonify({'error': 'Nao autenticado.'}), 401
    with _import_job_lock:
        job = _import_jobs.get(job_id)
    if not job or str(job.get('funcionario_id')) != str(session.get('funcionario_id')):
        return jsonify({'error': 'Importacao nao encontrada.'}), 404
    return jsonify(
        {
            'progress': job['progress'],
            'status_text': job['status_text'],
            'running': job['running'],
        }
    )


@app.route('/api/importacao/background/<job_id>/cancel', methods=['POST'])
def api_importacao_background_cancel(job_id):
    """Marca cancelamento e envia terminate ao subprocesso activo."""
    if not session.get('funcionario_id'):
        return jsonify({'error': 'Nao autenticado.'}), 401
    with _import_job_lock:
        job = _import_jobs.get(job_id)
        if not job or str(job.get('funcionario_id')) != str(session.get('funcionario_id')):
            return jsonify({'error': 'Importacao nao encontrada.'}), 404
        job['cancel_requested'] = True
        proc = job.get('active_proc')
    if proc:
        try:
            proc.terminate()
        except Exception:
            pass
    return jsonify({'ok': True})


# ---------------------------------------------------------------------------
# API: Distribuicao de funcionarios de cobranca (resumo pos-importacao)
# ---------------------------------------------------------------------------
#
# A tabela `funcionario_cobranca` e puramente relacional:
#
#    funcionario_cobranca(id, id_funcionario FK funcionario, id_contrato FK contrato, ...)
#
# Situacao (critico/atencao/recente), valor_credito e dias_atraso NAO estao
# nessa tabela - sao calculados on-the-fly via JOIN com contrato/parcela.


def _sit_key(sit):
    """Normaliza 'atenção' -> 'atencao' (chave JSON / atributo)."""
    if sit == 'atenção':
        return 'atencao'
    return sit or ''


def _situacao_from_dias(dias):
    try:
        d = int(dias or 0)
    except (TypeError, ValueError):
        d = 0
    if d >= 61:
        return 'critico'
    if d >= 31:
        return 'atenção'
    return 'recente'


def _empty_func_stats():
    return {
        'count': 0, 'value': 0.0,
        'critico_count': 0, 'critico_value': 0.0,
        'atencao_count': 0, 'atencao_value': 0.0,
        'recente_count': 0, 'recente_value': 0.0,
    }


# Filtro SQL reutilizado: apenas perfil operacional de cobranca (nao Gestor/Administrador).
_WHERE_FUNCIONARIO_COBRANCA = (
    "COALESCE(TRIM(nivel_acesso), '') IN ('Cobrança', 'Cobranca')"
)


@app.route('/api/importacao/distribuicao')
def api_distribuicao():
    """Retorna a distribuicao atual gravada em funcionario_cobranca."""
    conn = _get_db()
    cursor = conn.cursor()

    # Lista de funcionarios disponiveis (select / transferencia): somente Cobrança.
    cursor.execute(
        "SELECT id, nome FROM funcionario WHERE " + _WHERE_FUNCIONARIO_COBRANCA + " ORDER BY nome"
    )
    funcionarios_disponiveis = _clean_rows(cursor.fetchall())

    # Contratos atribuidos, com valor/dias calculados via JOIN.
    cursor.execute(
        """
        SELECT fc.id                              AS fc_id,
               fc.id_funcionario,
               fc.id_contrato,
               f.nome                             AS nome_funcionario,
               c.grupo, c.cota, c.numero_contrato,
               c.valor_credito,
               p.nome_completo                    AS nome_devedor,
               p.cpf_cnpj,
               (SELECT DATEDIFF(CURDATE(), MIN(vencimento))
                  FROM parcela
                 WHERE id_contrato = c.id AND status = 'cobranca') AS dias_atraso
        FROM funcionario_cobranca fc
        INNER JOIN funcionario f ON f.id = fc.id_funcionario
        INNER JOIN contrato c    ON c.id = fc.id_contrato
        LEFT JOIN pessoa p       ON p.id = c.id_pessoa
        WHERE c.status = 'cobranca'
        ORDER BY f.nome, c.valor_credito DESC
        """
    )
    rows = _clean_rows(cursor.fetchall())

    # Inicializa o mapa com todos os funcionarios (mesmo sem contratos).
    funcionarios_map = {
        int(f['id']): {
            'id': int(f['id']),
            'nome': f['nome'],
            'stats': _empty_func_stats(),
            'contratos': [],
        }
        for f in funcionarios_disponiveis
    }
    totais = _empty_func_stats()

    for r in rows:
        fid = int(r['id_funcionario'])
        if fid not in funcionarios_map:
            funcionarios_map[fid] = {
                'id': fid,
                'nome': r.get('nome_funcionario') or f'#{fid}',
                'stats': _empty_func_stats(),
                'contratos': [],
            }

        try:
            v = float(r['valor_credito']) if r['valor_credito'] is not None else 0.0
        except (TypeError, ValueError):
            v = 0.0

        situacao = _situacao_from_dias(r.get('dias_atraso'))
        sk = _sit_key(situacao)
        r['situacao'] = situacao  # injeta para o frontend exibir a badge

        bucket = funcionarios_map[fid]
        bucket['contratos'].append(r)
        bucket['stats']['count'] += 1
        bucket['stats']['value'] += v
        if sk in ('critico', 'atencao', 'recente'):
            bucket['stats'][f'{sk}_count'] += 1
            bucket['stats'][f'{sk}_value'] += v

        totais['count'] += 1
        totais['value'] += v
        if sk in ('critico', 'atencao', 'recente'):
            totais[f'{sk}_count'] += 1
            totais[f'{sk}_value'] += v

    cursor.close()
    conn.close()

    return jsonify({
        'funcionarios': list(funcionarios_map.values()),
        'totais': totais,
        'funcionarios_disponiveis': [
            {'id': int(f['id']), 'nome': f['nome']}
            for f in funcionarios_disponiveis
        ],
    })


@app.route('/api/importacao/distribuicao/reassign', methods=['POST'])
def api_distribuicao_reassign():
    """Altera manualmente o funcionario responsavel por um contrato."""
    data = request.get_json(silent=True) or {}
    try:
        fc_id = int(data.get('id'))
        novo_fid = int(data.get('id_funcionario'))
    except (TypeError, ValueError):
        return jsonify({'error': 'Parametros invalidos'}), 400

    conn = _get_db()
    cursor = conn.cursor()

    # Destino deve existir e ser perfil Cobrança (Gestor/Administrador nao recebem contratos aqui).
    cursor.execute(
        "SELECT 1 FROM funcionario WHERE id = %s AND " + _WHERE_FUNCIONARIO_COBRANCA,
        (novo_fid,),
    )
    if not cursor.fetchone():
        cursor.close()
        conn.close()
        return jsonify({'error': 'Funcionario nao encontrado ou sem perfil Cobrança'}), 404

    cursor.execute(
        "UPDATE funcionario_cobranca SET id_funcionario = %s WHERE id = %s",
        (novo_fid, fc_id),
    )
    conn.commit()
    affected = cursor.rowcount
    cursor.close()
    conn.close()

    if affected == 0:
        return jsonify({'error': 'Registro nao encontrado'}), 404
    return jsonify({'ok': True})


@app.route('/api/importacao/distribuicao/transferir', methods=['POST'])
def api_distribuicao_transferir():
    """Transfere todos os contratos atribuidos a um funcionario de origem
    para outro(s) funcionario(s).

    Body JSON:
        id_origem       (int)   - obrigatorio, funcionario que vai ficar sem carga
        modo            (str)   - 'especifico' ou 'igualitaria'
        id_destino      (int)   - obrigatorio quando modo='especifico'

    No modo 'igualitaria' os contratos sao re-balanceados entre os demais
    funcionarios seguindo as mesmas metricas do seed
    (Python/distribuir_funcionarios_cobranca.py): mesma media de valor e
    mesma media de quantidade POR SITUACAO (critico / atencao / recente),
    usando LPT greedy a partir da carga atual dos destinos.
    """
    data = request.get_json(silent=True) or {}
    try:
        id_origem = int(data.get('id_origem'))
    except (TypeError, ValueError):
        return jsonify({'error': 'id_origem invalido'}), 400

    modo = (data.get('modo') or '').strip()
    if modo not in ('especifico', 'igualitaria'):
        return jsonify({'error': "modo deve ser 'especifico' ou 'igualitaria'"}), 400

    id_destino = None
    if modo == 'especifico':
        try:
            id_destino = int(data.get('id_destino'))
        except (TypeError, ValueError):
            return jsonify({'error': 'id_destino invalido'}), 400
        if id_destino == id_origem:
            return jsonify({'error': 'Destino deve ser diferente da origem'}), 400

    conn = _get_db()
    cursor = conn.cursor()
    try:
        cursor.execute(
            "SELECT id, nome FROM funcionario WHERE id = %s", (id_origem,)
        )
        origem = cursor.fetchone()
        if not origem:
            return jsonify({'error': 'Funcionario de origem nao encontrado'}), 404

        if modo == 'especifico':
            cursor.execute(
                "SELECT id, nome FROM funcionario WHERE id = %s AND "
                + _WHERE_FUNCIONARIO_COBRANCA,
                (id_destino,),
            )
            destino = cursor.fetchone()
            if not destino:
                return jsonify({'error': 'Funcionario de destino nao encontrado ou sem perfil Cobrança'}), 404

            cursor.execute(
                "UPDATE funcionario_cobranca SET id_funcionario = %s "
                "WHERE id_funcionario = %s",
                (id_destino, id_origem),
            )
            transferidos = cursor.rowcount
            conn.commit()
            return jsonify({
                'ok': True,
                'modo': modo,
                'transferidos': transferidos,
                'origem': {'id': int(origem['id']), 'nome': origem['nome']},
                'destino': {'id': int(destino['id']), 'nome': destino['nome']},
            })

        # modo == 'igualitaria' ---------------------------------------------
        # 1) Busca os demais funcionarios de cobrança (destinos possiveis).
        cursor.execute(
            "SELECT id, nome FROM funcionario WHERE id <> %s AND "
            + _WHERE_FUNCIONARIO_COBRANCA
            + " AND (ativo IS NULL OR ativo = 1) ORDER BY nome",
            (id_origem,),
        )
        destinos = _clean_rows(cursor.fetchall())
        if not destinos:
            return jsonify({
                'error': 'Nao ha outros funcionarios disponiveis para redistribuir',
            }), 400
        destino_ids = [int(d['id']) for d in destinos]

        # 2) Contratos atualmente atribuidos a origem, com valor e dias de atraso.
        cursor.execute(
            """
            SELECT fc.id                AS fc_id,
                   fc.id_contrato,
                   c.valor_credito,
                   (SELECT DATEDIFF(CURDATE(), MIN(vencimento))
                      FROM parcela
                     WHERE id_contrato = c.id AND status = 'cobranca') AS dias_atraso
            FROM funcionario_cobranca fc
            INNER JOIN contrato c ON c.id = fc.id_contrato
            WHERE fc.id_funcionario = %s AND c.status = 'cobranca'
            """,
            (id_origem,),
        )
        pendentes = _clean_rows(cursor.fetchall())

        # Inclui tambem contratos ja fechados/indenizados que porventura
        # estejam ligados ao origem (para esvaziar totalmente a carga dele).
        cursor.execute(
            "SELECT fc.id AS fc_id FROM funcionario_cobranca fc "
            "WHERE fc.id_funcionario = %s AND fc.id_contrato NOT IN ("
            "SELECT id_contrato FROM funcionario_cobranca fc2 "
            "INNER JOIN contrato c ON c.id = fc2.id_contrato "
            "WHERE fc2.id_funcionario = %s AND c.status = 'cobranca')",
            (id_origem, id_origem),
        )
        fechados = [int(r['fc_id']) for r in cursor.fetchall()]

        if not pendentes and not fechados:
            return jsonify({
                'ok': True, 'modo': modo, 'transferidos': 0,
                'origem': {'id': int(origem['id']), 'nome': origem['nome']},
                'por_destino': [],
                'detalhe': 'Origem nao possuia contratos atribuidos.',
            })

        # 3) Carga atual dos destinos por situacao (ponto de partida do LPT).
        cursor.execute(
            """
            SELECT fc.id_funcionario,
                   c.valor_credito,
                   (SELECT DATEDIFF(CURDATE(), MIN(vencimento))
                      FROM parcela
                     WHERE id_contrato = c.id AND status = 'cobranca') AS dias_atraso
            FROM funcionario_cobranca fc
            INNER JOIN contrato c ON c.id = fc.id_contrato
            WHERE c.status = 'cobranca' AND fc.id_funcionario IN (%s)
            """ % (",".join(["%s"] * len(destino_ids))),
            destino_ids,
        )
        cargas_rows = _clean_rows(cursor.fetchall())

        estados = {
            fid: {s: {'count': 0, 'value': 0.0}
                  for s in ('critico', 'atencao', 'recente')}
            for fid in destino_ids
        }

        def _sit_norm(dias):
            try:
                d = int(dias or 0)
            except (TypeError, ValueError):
                d = 0
            if d >= 61: return 'critico'
            if d >= 31: return 'atencao'
            return 'recente'

        def _to_float(v):
            try:
                return float(v) if v is not None else 0.0
            except (TypeError, ValueError):
                return 0.0

        for cr in cargas_rows:
            fid = int(cr['id_funcionario'])
            if fid not in estados:
                continue
            sit = _sit_norm(cr.get('dias_atraso'))
            estados[fid][sit]['count'] += 1
            estados[fid][sit]['value'] += _to_float(cr.get('valor_credito'))

        # 4) LPT greedy: ordena contratos da origem (abertos) por valor desc
        #    dentro de cada situacao e aloca ao destino de menor carga.
        por_sit = {'critico': [], 'atencao': [], 'recente': []}
        for p in pendentes:
            por_sit[_sit_norm(p.get('dias_atraso'))].append(p)
        for s in por_sit:
            por_sit[s].sort(key=lambda x: _to_float(x.get('valor_credito')),
                            reverse=True)

        por_destino = {fid: {'count': 0, 'value': 0.0, 'fc_ids': []}
                       for fid in destino_ids}
        updates = []  # (novo_fid, fc_id)

        for s in ('critico', 'atencao', 'recente'):
            for p in por_sit[s]:
                valor = _to_float(p.get('valor_credito'))
                def score(fid, _s=s):
                    st = estados[fid][_s]
                    return (st['value'], st['count'])
                chosen = min(destino_ids, key=score)
                estados[chosen][s]['count'] += 1
                estados[chosen][s]['value'] += valor
                por_destino[chosen]['count'] += 1
                por_destino[chosen]['value'] += valor
                por_destino[chosen]['fc_ids'].append(int(p['fc_id']))
                updates.append((chosen, int(p['fc_id'])))

        # Distribui tambem os fechados (round-robin simples).
        for idx, fc_id in enumerate(fechados):
            chosen = destino_ids[idx % len(destino_ids)]
            por_destino[chosen]['fc_ids'].append(fc_id)
            updates.append((chosen, fc_id))

        if updates:
            cursor.executemany(
                "UPDATE funcionario_cobranca SET id_funcionario = %s WHERE id = %s",
                updates,
            )
            conn.commit()

        nome_by_id = {int(d['id']): d['nome'] for d in destinos}
        resumo = [
            {
                'id': fid,
                'nome': nome_by_id.get(fid, f'#{fid}'),
                'count': por_destino[fid]['count'],
                'value': round(por_destino[fid]['value'], 2),
            }
            for fid in destino_ids
        ]

        return jsonify({
            'ok': True,
            'modo': modo,
            'transferidos': len(updates),
            'origem': {'id': int(origem['id']), 'nome': origem['nome']},
            'por_destino': resumo,
        })

    except Exception as exc:
        conn.rollback()
        app.logger.exception('api_distribuicao_transferir: erro inesperado')
        return jsonify({'error': 'Erro na transferencia: ' + str(exc)}), 500
    finally:
        try: cursor.close()
        except Exception: pass
        try: conn.close()
        except Exception: pass


@app.route('/api/importacao/distribuicao/restaurar', methods=['POST'])
def api_distribuicao_restaurar():
    """Reverte a distribuicao atual para a baseline gerada pelo algoritmo
    original (Python/distribuir_funcionarios_cobranca.py).

    Estrategia:
        1. Apaga TODAS as linhas de funcionario_cobranca (remove as
           transferencias manuais / em lote feitas pelo usuario).
        2. Executa o script de distribuicao, que e deterministico: dado o
           mesmo conjunto de contratos abertos e de funcionarios cadastrados,
           produz exatamente a mesma atribuicao da importacao original.

    Resposta:
        { ok: true, total, logs: [..], restauracao_ok: bool }
    """
    conn = _get_db()
    cursor = conn.cursor()
    try:
        cursor.execute("DELETE FROM funcionario_cobranca")
        conn.commit()
    except Exception as exc:
        conn.rollback()
        cursor.close()
        conn.close()
        app.logger.exception('restaurar: falha ao apagar funcionario_cobranca')
        return jsonify({'error': f'Falha ao limpar a distribuicao atual: {exc}'}), 500

    script = os.path.join(PYTHON_DIR, 'distribuir_funcionarios_cobranca.py')
    try:
        proc = subprocess.run(
            [PYTHON_EXE, '-u', script],
            capture_output=True,
            text=True,
            encoding='utf-8',
            errors='replace',
            env=_SUBPROCESS_ENV,
            timeout=300,
        )
    except Exception as exc:
        cursor.close()
        conn.close()
        app.logger.exception('restaurar: erro ao rodar script de distribuicao')
        return jsonify({'error': f'Erro ao executar o script de distribuicao: {exc}'}), 500

    logs = []
    if proc.stdout:
        logs.extend([ln for ln in proc.stdout.splitlines() if ln.strip()])
    if proc.stderr:
        logs.extend([f'[stderr] {ln}' for ln in proc.stderr.splitlines() if ln.strip()])

    cursor.execute("SELECT COUNT(*) AS total FROM funcionario_cobranca")
    row = cursor.fetchone() or {}
    total = int(row.get('total') or 0)
    cursor.close()
    conn.close()

    return jsonify({
        'ok': True,
        'restauracao_ok': (proc.returncode == 0),
        'returncode': proc.returncode,
        'total': total,
        'logs': logs[-80:],  # limita o payload - logs completos ficam no servidor
    })


@app.route('/api/importacao/distribuicao/aprovar', methods=['POST'])
def api_distribuicao_aprovar():
    """No-op de aprovacao. A tabela funcionario_cobranca nao tem coluna de
    status de aprovacao - qualquer linha presente JA e considerada efetiva.
    O endpoint existe apenas para o frontend confirmar que o usuario revisou
    e aceitou a distribuicao proposta."""
    conn = _get_db()
    cursor = conn.cursor()
    cursor.execute("SELECT COUNT(*) AS total FROM funcionario_cobranca")
    row = cursor.fetchone() or {}
    total = int(row.get('total') or 0)
    cursor.close()
    conn.close()
    return jsonify({'ok': True, 'total': total})


def _pode_sms_automatizados_importacao():
    """Gestor ou Administrador (alinhado ao menu Importacao)."""
    return _nivel_normalizado(session.get('funcionario_nivel_acesso')) in (
        'gestor',
        'administrador',
    )


def _format_parcelas_sms_auto(cursor, id_contrato):
    """Parcelas em aberto: prioriza vencidas; senao todas em aberto; senao '-'."""
    cursor.execute(
        """
        SELECT numero_parcela FROM parcela
        WHERE id_contrato = %s AND status = 'cobranca' AND vencimento < CURDATE()
        ORDER BY numero_parcela
        """,
        (int(id_contrato),),
    )
    rows = cursor.fetchall() or []
    if not rows:
        cursor.execute(
            """
            SELECT numero_parcela FROM parcela
            WHERE id_contrato = %s AND status = 'cobranca'
            ORDER BY numero_parcela
            """,
            (int(id_contrato),),
        )
        rows = cursor.fetchall() or []
    if not rows:
        return '-'
    parts = []
    for r in rows:
        np = r.get('numero_parcela')
        if np is None:
            continue
        try:
            parts.append(str(int(np)))
        except (TypeError, ValueError):
            parts.append(str(np))
    return ', '.join(parts) if parts else '-'


def _mensagem_sms_auto_importacao(primeiro_nome, numero_contrato, parcelas_txt, template_id):
    """Monta texto do SMS conforme templates da distribuicao (importacao)."""
    nc = (numero_contrato or '').strip() or '-'
    pn = (primeiro_nome or '').strip() or 'Cliente'
    parcelas_txt = (parcelas_txt or '-').strip() or '-'
    if template_id == 4:
        return (
            f'{pn}, seu contrato com o Consórcio Nacional Chevrolet Ltda encontra-se em atraso. '
            f'Evite medidas judiciais. Ligue 08000012323 e renegocie seu acordo!'
        )
    if template_id == 1:
        return (
            f'{pn}: sua cota foi encaminhada à João Barbosa Assessoria. '
            f'Para regularização, ligue 08000012323.'
        )
    return (
        f'{pn}, não identificamos o pagamento da(s) parcela(s) {parcelas_txt}, referente(s) '
        f'ao contrato {nc}. Entre em contato conosco JOAO BARBOSA ASSESSORIA, 08000012323.'
    )


def _contrato_teve_envio_cobranca_hoje(cursor, id_contrato):
    """True se já existe SMS ou e-mail registrado hoje para o contrato (anti-duplicidade diária)."""
    cid = int(id_contrato)
    cursor.execute(
        """
        SELECT 1 FROM registro_sms
        WHERE id_contrato = %s AND DATE(created_at) = CURDATE()
        LIMIT 1
        """,
        (cid,),
    )
    if cursor.fetchone():
        return True
    cursor.execute(
        """
        SELECT 1 FROM registro_email
        WHERE id_contrato = %s AND DATE(created_at) = CURDATE()
        LIMIT 1
        """,
        (cid,),
    )
    return cursor.fetchone() is not None


def _plain_para_corpo_email_html(msg_plain):
    """Converte texto plano do SMS automático em HTML simples para o MessageCenter."""
    t = (msg_plain or '').strip()
    if not t:
        return '<p></p>'
    esc = html_module.escape(t)
    parts = esc.split('\n')
    return '<p>' + '</p><p>'.join(parts) + '</p>'


def _email_basico_valido(addr):
    s = (addr or '').strip()
    if not s or '@' not in s or len(s) > 254:
        return False
    return True


def _contrato_row_auto_envio(cursor, id_contrato):
    """Uma linha no formato de `_SMS_AUTOM_DISTRIBUICAO_SQL` para um único contrato aberto."""
    cursor.execute(
        """
        SELECT c.id AS id_contrato,
               c.grupo,
               c.cota,
               c.id_pessoa,
               c.numero_contrato,
               p.nome_completo,
               (SELECT DATEDIFF(CURDATE(), MIN(p2.vencimento))
                  FROM parcela p2
                 WHERE p2.id_contrato = c.id AND p2.status = 'cobranca') AS dias_atraso
        FROM contrato c
        LEFT JOIN pessoa p ON p.id = c.id_pessoa
        WHERE c.id = %s AND c.status = 'cobranca'
        """,
        (int(id_contrato),),
    )
    return cursor.fetchone()


def _contratos_rows_auto_envio_batch(cursor, contrato_ids, chunk_size=450):
    """Mesmo formato de `_contrato_row_auto_envio`, em lote (contratos abertos)."""
    if not contrato_ids:
        return {}
    ids = list(contrato_ids)
    merged = {}
    for i in range(0, len(ids), chunk_size):
        chunk = ids[i : i + chunk_size]
        ph = ','.join(['%s'] * len(chunk))
        cursor.execute(
            f"""
            SELECT c.id AS id_contrato,
                   c.grupo,
                   c.cota,
                   c.id_pessoa,
                   c.numero_contrato,
                   p.nome_completo,
                   (SELECT DATEDIFF(CURDATE(), MIN(p2.vencimento))
                      FROM parcela p2
                     WHERE p2.id_contrato = c.id AND p2.status = 'cobranca') AS dias_atraso
            FROM contrato c
            LEFT JOIN pessoa p ON p.id = c.id_pessoa
            WHERE c.id IN ({ph}) AND c.status = 'cobranca'
            """,
            chunk,
        )
        for r in cursor.fetchall() or []:
            merged[int(r['id_contrato'])] = r
    return merged


def _contratos_teve_envio_cobranca_hoje_batch(cursor, contrato_ids, chunk_size=450):
    """Conjunto de ids de contrato que já têm SMS ou e-mail registrados hoje.

    Consultas fatiadas para não estourar limite de placeholders/pacote em bases grandes.
    """
    if not contrato_ids:
        return set()
    ids = list(contrato_ids)
    found = set()
    for i in range(0, len(ids), chunk_size):
        chunk = ids[i : i + chunk_size]
        ph = ','.join(['%s'] * len(chunk))
        cursor.execute(
            f"""
            SELECT DISTINCT id_contrato FROM (
                SELECT id_contrato FROM registro_sms
                WHERE id_contrato IN ({ph}) AND DATE(created_at) = CURDATE()
                UNION ALL
                SELECT id_contrato FROM registro_email
                WHERE id_contrato IN ({ph}) AND DATE(created_at) = CURDATE()
            ) t
            """,
            chunk + chunk,
        )
        for r in cursor.fetchall() or []:
            found.add(int(r['id_contrato']))
    return found


def _telefones_por_pessoas(cursor, pessoa_ids, chunk_size=450):
    """{id_pessoa: [rows telefone]} preservando ordem estável por id."""
    if not pessoa_ids:
        return {}
    ids = list(pessoa_ids)
    out = {}
    for i in range(0, len(ids), chunk_size):
        chunk = ids[i : i + chunk_size]
        ph = ','.join(['%s'] * len(chunk))
        cursor.execute(
            f'SELECT id, id_pessoa, numero FROM telefone WHERE id_pessoa IN ({ph}) ORDER BY id ASC',
            chunk,
        )
        for r in cursor.fetchall() or []:
            pid = int(r['id_pessoa'])
            out.setdefault(pid, []).append(r)
    return out


def _emails_por_pessoas(cursor, pessoa_ids, chunk_size=450):
    """{id_pessoa: [rows email]} preservando ordem estável por id."""
    if not pessoa_ids:
        return {}
    ids = list(pessoa_ids)
    out = {}
    for i in range(0, len(ids), chunk_size):
        chunk = ids[i : i + chunk_size]
        ph = ','.join(['%s'] * len(chunk))
        cursor.execute(
            f'SELECT id, id_pessoa, email FROM email WHERE id_pessoa IN ({ph}) ORDER BY id ASC',
            chunk,
        )
        for r in cursor.fetchall() or []:
            pid = int(r['id_pessoa'])
            out.setdefault(pid, []).append(r)
    return out


# Lista todos os contratos abertos com dias desde MIN(vencimento) das parcelas em aberto.
# Usa a mesma CTE `MenorVencimento` que o Excel (sem subconsulta correlacionada por linha),
# para o mesmo resultado numérico e desempenho aceitável em bases grandes.
_SMS_AUTOM_DISTRIBUICAO_SQL = """
WITH MenorVencimento AS (
    SELECT id_contrato, MIN(vencimento) AS data_vencimento_minima
    FROM parcela
    WHERE status = 'cobranca'
    GROUP BY id_contrato
)
SELECT c.id AS id_contrato,
       c.grupo,
       c.cota,
       c.id_pessoa,
       c.numero_contrato,
       p.nome_completo,
       DATEDIFF(CURRENT_DATE, mv.data_vencimento_minima) AS dias_atraso
FROM contrato c
LEFT JOIN MenorVencimento mv ON c.id = mv.id_contrato
LEFT JOIN pessoa p ON p.id = c.id_pessoa
WHERE c.status = 'cobranca'
ORDER BY c.id
"""

# Export Excel (lista contratos abertos no roteiro de dias); independente da validacao de telefone do preview/POST.
_SMS_AUTOM_EXCEL_SQL = """
WITH MenorVencimento AS (
    SELECT
        id_contrato,
        MIN(vencimento) AS data_vencimento_minima
    FROM parcela
    WHERE status = 'cobranca'
    GROUP BY id_contrato
)
SELECT
    c.id AS contrato_id,
    c.id_pessoa,
    c.numero_contrato,
    c.grupo,
    c.cota,
    mv.data_vencimento_minima,
    DATEDIFF(CURRENT_DATE, mv.data_vencimento_minima) AS dias_atraso
FROM contrato c
JOIN MenorVencimento mv ON c.id = mv.id_contrato
WHERE c.status = 'cobranca'
  AND DATEDIFF(CURRENT_DATE, mv.data_vencimento_minima) IN (0, 16, 31, 61, 85)
ORDER BY dias_atraso DESC
"""


def _sms_autom_excel_maps_por_pessoa(cursor, linhas):
    """Por id_pessoa: coluna Sim/Não, texto telefone(s) e e-mail(s) válidos para disparo (como o preview).

    Retorna (flags_sim_nao, telefones_por_pessoa, emails_por_pessoa).
    Vários números ou e-mails no mesmo devedor são reunidos com '; '.
    """
    pessoa_ids = []
    for row in linhas or []:
        pid = row.get('id_pessoa')
        if pid is None:
            continue
        try:
            pessoa_ids.append(int(pid))
        except (TypeError, ValueError):
            continue
    pessoa_ids = list(dict.fromkeys(pessoa_ids))
    if not pessoa_ids:
        return {}, {}, {}
    tel_map = _telefones_por_pessoas(cursor, pessoa_ids)
    email_map = _emails_por_pessoas(cursor, pessoa_ids)
    flags = {}
    tel_txt = {}
    em_txt = {}
    sep = '; '
    for pid in pessoa_ids:
        nums = []
        for tel in tel_map.get(pid, []):
            digits = _sms_digits_only(tel.get('numero'))
            if not digits:
                continue
            if not _sms_numero_variant_sets_overlap(tel.get('numero'), digits):
                continue
            raw = (tel.get('numero') or '').strip()
            nums.append(raw if raw else digits)
        mails = []
        for em in email_map.get(pid, []):
            if not _email_basico_valido(em.get('email')):
                continue
            mails.append((em.get('email') or '').strip())
        tel_txt[pid] = sep.join(nums)
        em_txt[pid] = sep.join(mails)
        flags[pid] = 'Sim' if (nums and mails) else 'Não'
    return flags, tel_txt, em_txt


def _sms_autom_fill_rota_excel_duas_folhas(cursor, wb, linhas, msg_vazio):
    """Folhas ``SMS`` e ``EMAIL`` separadas: cada aba só contratos com contacto válido naquele canal.

    Coluna ``SMS e e-mail?`` indica se o devedor tem ambos os canais válidos (alinha ao preview).
    """
    flags, tel_por_pessoa, em_por_pessoa = _sms_autom_excel_maps_por_pessoa(cursor, linhas)
    hdr_sms = ('Grupo', 'Cota', 'Grupo/Cota', 'Atraso', 'Telefone(s) disparo', 'SMS e e-mail?')
    hdr_mail = ('Grupo', 'Cota', 'Grupo/Cota', 'Atraso', 'E-mail(s) disparo', 'SMS e e-mail?')
    ws_sms = wb.active
    ws_sms.title = 'SMS'
    ws_mail = wb.create_sheet(title='EMAIL')
    ws_sms.append(hdr_sms)
    ws_mail.append(hdr_mail)

    def _empty_msg_row(msg):
        return (_xlsx_cell_str(msg), '', '', '', '', '', '')

    if not linhas:
        ws_sms.append(_empty_msg_row(msg_vazio))
        ws_mail.append(_empty_msg_row(msg_vazio))
        return

    n_sms = 0
    n_mail = 0
    for row in linhas:
        da = row.get('dias_atraso')
        try:
            da_cell = int(da) if da is not None else ''
        except (TypeError, ValueError):
            da_cell = _xlsx_cell_str(da)
        pid = row.get('id_pessoa')
        col_sim = 'Não'
        tel_cell = ''
        em_cell = ''
        if pid is not None:
            try:
                ip = int(pid)
                col_sim = flags.get(ip, 'Não')
                tel_cell = tel_por_pessoa.get(ip, '') or ''
                em_cell = em_por_pessoa.get(ip, '') or ''
            except (TypeError, ValueError):
                pass
        g = _xlsx_cell_str(row.get('grupo'))
        c = _xlsx_cell_str(row.get('cota'))
        gc = f'{g}/{c}' if (g or c) else ''
        if tel_cell.strip():
            ws_sms.append((g, c, gc, da_cell, _xlsx_cell_str(tel_cell), col_sim))
            n_sms += 1
        if em_cell.strip():
            ws_mail.append((g, c, gc, da_cell, _xlsx_cell_str(em_cell), col_sim))
            n_mail += 1

    msg_sem_sms = (
        'Nenhum contrato neste roteiro com telefone valido para SMS '
        '(mesma validacao do envio / preview).'
    )
    msg_sem_mail = (
        'Nenhum contrato neste roteiro com e-mail valido '
        '(mesma validacao do envio / preview).'
    )
    if n_sms == 0:
        ws_sms.append(_empty_msg_row(msg_sem_sms))
    if n_mail == 0:
        ws_mail.append(_empty_msg_row(msg_sem_mail))


def _sms_autom_excel_linhas_carteira(cursor, contrato_ids):
    """Mesmo critério que `_SMS_AUTOM_EXCEL_SQL`, restrito aos IDs informados (lista visível Cobrança)."""
    if not contrato_ids:
        return []
    placeholders = ','.join(['%s'] * len(contrato_ids))
    sql = _SMS_AUTOM_EXCEL_SQL.strip().replace(
        'ORDER BY dias_atraso DESC',
        'AND c.id IN (' + placeholders + ') ORDER BY dias_atraso DESC',
    )
    cursor.execute(sql, tuple(contrato_ids))
    return cursor.fetchall() or []


# Mesmo filtro que `_SMS_AUTOM_EXCEL_SQL`, com `id_pessoa` para o preview (resumo = mesma lista).
_SMS_AUTOM_PREVIEW_ROTEIRO_SQL = """
WITH MenorVencimento AS (
    SELECT
        id_contrato,
        MIN(vencimento) AS data_vencimento_minima
    FROM parcela
    WHERE status = 'cobranca'
    GROUP BY id_contrato
)
SELECT
    c.id AS id_contrato,
    c.id_pessoa,
    c.grupo,
    c.cota,
    c.numero_contrato,
    DATEDIFF(CURRENT_DATE, mv.data_vencimento_minima) AS dias_atraso
FROM contrato c
JOIN MenorVencimento mv ON c.id = mv.id_contrato
WHERE c.status = 'cobranca'
  AND DATEDIFF(CURRENT_DATE, mv.data_vencimento_minima) IN (0, 16, 31, 61, 85)
ORDER BY dias_atraso DESC
"""


def _sms_automatizados_template_id_por_dias(da_int):
    """Template SMS automaticos importacao: dias desde MIN(vencimento) parcelas abertas.

    DATEDIFF(CURDATE(), vencimento): 0, 16, 31, 61 ou 85 disparam; negativos ou outros
    valores retornam None. None (sem parcela aberta / vencimento) retorna None.
    """
    if da_int is None:
        return None
    try:
        d = int(da_int)
    except (TypeError, ValueError):
        return None
    if d < 0:
        return None
    if d == 0:
        return 1
    if d == 16:
        return 2
    if d == 31:
        return 3
    if d in (61, 85):
        return 4
    return None


def _sms_automatizados_analise(conn):
    """Contagens para o modal de preview (importação): mesma lista que o Excel / Lista SMS-E-mail.

    Usa `_SMS_AUTOM_PREVIEW_ROTEIRO_SQL` (filtro identico a `_SMS_AUTOM_EXCEL_SQL`) e depois
    telefones/e-mails em memoria. Contratos abertos fora dessa lista entram em
    `ignorados_sem_template`. O POST em lote continua a iterar `_SMS_AUTOM_DISTRIBUICAO_SQL`.

    Semântica dos contadores (alinhada ao Excel por contrato):
    - ``sms_previstos`` / ``emails_previstos``: número de **contratos** com pelo menos um envio
      válido naquele canal (igual a ``contratos_com_sms`` / ``contratos_com_email``).
    - ``sms_mensagens_previstas`` / ``email_mensagens_previstas``: total de **mensagens** que o
      POST tentará (um SMS por número válido; um e-mail por endereço válido).
    """
    out = {
        'sms_previstos': 0,
        'emails_previstos': 0,
        'sms_mensagens_previstas': 0,
        'email_mensagens_previstas': 0,
        'contratos_previstos_algum_canal': 0,
        'contratos_com_sms': 0,
        'contratos_com_email': 0,
        'ignorados_sem_entrada': 0,
        'ignorados_sem_template': 0,
        'ignorados_sem_telefone': 0,
        'ignorados_sem_email': 0,
        'ignorados_ja_enviados_hoje': 0,
        'tentativas_bloqueadas_cadastro': 0,
        'contratos_processados': 0,
    }
    with conn.cursor() as cursor:
        cursor.execute(
            "SELECT COUNT(*) AS n FROM contrato WHERE status = %s",
            (_STATUS_BD_COBRANCA,),
        )
        crow = cursor.fetchone() or {}
        total_abertos = int(crow.get('n') or 0)
    out['contratos_processados'] = total_abertos

    with conn.cursor() as cursor:
        cursor.execute(_SMS_AUTOM_PREVIEW_ROTEIRO_SQL)
        roteiro_rows = cursor.fetchall() or []

    n_na_lista = len(roteiro_rows)
    out['ignorados_sem_template'] = max(0, total_abertos - n_na_lista)

    pending = []
    for row in roteiro_rows:
        id_contrato = int(row['id_contrato'])
        id_pessoa = row.get('id_pessoa')
        if id_pessoa is None:
            out['ignorados_sem_template'] += 1
            continue
        pending.append((id_contrato, row, int(id_pessoa)))

    if not pending:
        return out, []

    ids = [p[0] for p in pending]
    with conn.cursor() as cursor:
        envio_hoje = _contratos_teve_envio_cobranca_hoje_batch(cursor, ids)

    pessoa_ids = list({p[2] for p in pending})
    with conn.cursor() as cursor:
        tel_map = _telefones_por_pessoas(cursor, pessoa_ids)
        email_map = _emails_por_pessoas(cursor, pessoa_ids)

    for id_contrato, row, id_pessoa in pending:
        if id_contrato in envio_hoje:
            out['ignorados_ja_enviados_hoje'] += 1
            continue

        telefones = tel_map.get(id_pessoa, [])

        n_ok_sms = 0
        for tel in telefones:
            digits = _sms_digits_only(tel.get('numero'))
            if not digits:
                out['tentativas_bloqueadas_cadastro'] += 1
                continue
            if not _sms_numero_variant_sets_overlap(tel.get('numero'), digits):
                out['tentativas_bloqueadas_cadastro'] += 1
                continue
            out['sms_mensagens_previstas'] += 1
            n_ok_sms += 1
        if n_ok_sms > 0:
            out['contratos_com_sms'] += 1
            out['sms_previstos'] += 1
        elif not telefones:
            out['ignorados_sem_telefone'] += 1

        emails_rows = email_map.get(id_pessoa, [])

        n_ok_mail = 0
        for em in emails_rows:
            if not _email_basico_valido(em.get('email')):
                out['tentativas_bloqueadas_cadastro'] += 1
                continue
            out['email_mensagens_previstas'] += 1
            n_ok_mail += 1
        if n_ok_mail > 0:
            out['contratos_com_email'] += 1
            out['emails_previstos'] += 1
        elif not emails_rows:
            out['ignorados_sem_email'] += 1

        if n_ok_sms > 0 or n_ok_mail > 0:
            out['contratos_previstos_algum_canal'] += 1

    return out, []


def _analise_automacao_carteira(conn, contrato_ids):
    """Igual à análise da Lista SMS/E-mail (distribuição), restrita aos IDs da carteira visível.

    Usado pelo modal de confirmação em Cobrança para mostrar contagens e contratos reais.
    Validação de telefone/e-mail em memória (como o preview de importação); consultas fatiadas.

    Contadores: ``sms_previstos`` / ``emails_previstos`` = contratos por canal;
    ``sms_mensagens_previstas`` / ``email_mensagens_previstas`` = total de envios (vários números ou
    e-mails por contrato). ``contratos_previstos_algum_canal`` = linhas em ``detalhes``.
    """
    uniq = []
    seen = set()
    for x in contrato_ids or []:
        try:
            cid = int(x)
        except (TypeError, ValueError):
            continue
        if cid not in seen:
            seen.add(cid)
            uniq.append(cid)

    out = {
        'carteira_ids': len(uniq),
        'sms_previstos': 0,
        'emails_previstos': 0,
        'sms_mensagens_previstas': 0,
        'email_mensagens_previstas': 0,
        'contratos_previstos_algum_canal': 0,
        'contratos_com_sms': 0,
        'contratos_com_email': 0,
        'ignorados_sem_contrato_aberto': 0,
        'ignorados_sem_pessoa': 0,
        'ignorados_fora_rota': 0,
        'ignorados_ja_enviados_hoje': 0,
        'ignorados_sem_telefone': 0,
        'ignorados_sem_email': 0,
        'tentativas_bloqueadas_cadastro': 0,
        'detalhes': [],
    }

    if not uniq:
        return out

    with conn.cursor() as cursor:
        row_by_id = _contratos_rows_auto_envio_batch(cursor, uniq)
        envio_hoje = _contratos_teve_envio_cobranca_hoje_batch(cursor, uniq)

    pending = []
    for cid in uniq:
        row = row_by_id.get(cid)
        if not row:
            out['ignorados_sem_contrato_aberto'] += 1
            continue

        id_contrato = int(row['id_contrato'])
        id_pessoa = row.get('id_pessoa')
        if id_pessoa is None:
            out['ignorados_sem_pessoa'] += 1
            continue
        id_pessoa = int(id_pessoa)

        dias_atraso = row.get('dias_atraso')
        da_int = None
        if dias_atraso is not None:
            try:
                da_int = int(dias_atraso)
            except (TypeError, ValueError):
                da_int = None

        if _sms_automatizados_template_id_por_dias(da_int) is None:
            out['ignorados_fora_rota'] += 1
            continue

        if id_contrato in envio_hoje:
            out['ignorados_ja_enviados_hoje'] += 1
            continue

        pending.append((id_contrato, row, id_pessoa, da_int))

    if not pending:
        return out

    pessoa_ids = list({p[2] for p in pending})
    with conn.cursor() as cursor:
        tel_map = _telefones_por_pessoas(cursor, pessoa_ids)
        email_map = _emails_por_pessoas(cursor, pessoa_ids)

    for id_contrato, row, id_pessoa, da_int in pending:
        telefones = tel_map.get(id_pessoa, [])

        n_ok_sms = 0
        for tel in telefones:
            digits = _sms_digits_only(tel.get('numero'))
            if not digits:
                out['tentativas_bloqueadas_cadastro'] += 1
                continue
            if not _sms_numero_variant_sets_overlap(tel.get('numero'), digits):
                out['tentativas_bloqueadas_cadastro'] += 1
                continue
            out['sms_mensagens_previstas'] += 1
            n_ok_sms += 1
        if n_ok_sms > 0:
            out['contratos_com_sms'] += 1
            out['sms_previstos'] += 1
        elif not telefones:
            out['ignorados_sem_telefone'] += 1

        emails_rows = email_map.get(id_pessoa, [])

        n_ok_mail = 0
        for em in emails_rows:
            if not _email_basico_valido(em.get('email')):
                out['tentativas_bloqueadas_cadastro'] += 1
                continue
            out['email_mensagens_previstas'] += 1
            n_ok_mail += 1
        if n_ok_mail > 0:
            out['contratos_com_email'] += 1
            out['emails_previstos'] += 1
        elif not emails_rows:
            out['ignorados_sem_email'] += 1

        if n_ok_sms > 0 or n_ok_mail > 0:
            out['detalhes'].append({
                'id_contrato': id_contrato,
                'grupo': row.get('grupo'),
                'cota': row.get('cota'),
                'nome_devedor': (row.get('nome_completo') or '').strip() or '—',
                'dias_atraso': da_int,
                'disparos_sms': n_ok_sms,
                'disparos_email': n_ok_mail,
            })

    out['contratos_previstos_algum_canal'] = len(out['detalhes'])
    out['detalhes'].sort(key=lambda r: (
        (r.get('nome_devedor') or '').lower(),
        int(r.get('id_contrato') or 0),
    ))
    return out


@app.route('/api/importacao/distribuicao/sms-automatizados/preview', methods=['GET'])
def api_distribuicao_sms_automatizados_preview():
    """Preview SMS + e-mail: roteiro 0/16/31/61/85, contagens e ignorados por duplicidade do dia."""
    if not session.get('funcionario_id'):
        return jsonify({'error': 'Nao autenticado. Faca login novamente.'}), 401
    if not _pode_sms_automatizados_importacao():
        return jsonify({'error': 'Acesso restrito a gestores ou administradores.'}), 403

    conn = _get_db()
    try:
        out, _linhas = _sms_automatizados_analise(conn)
    except Exception:
        app.logger.exception('sms-automatizados/preview')
        return jsonify({
            'error': 'Falha ao calcular o preview. Tente novamente; se persistir, verifique os logs do servidor.',
        }), 500
    finally:
        conn.close()

    return jsonify({'ok': True, **out})


def _xlsx_cell_str(val):
    """Texto seguro para celula openpyxl (remove caracteres de controle ilegais)."""
    if val is None:
        return ''
    if isinstance(val, (bytes, bytearray)):
        s = bytes(val).decode('utf-8', errors='replace')
    elif isinstance(val, memoryview):
        s = bytes(val).decode('utf-8', errors='replace')
    elif isinstance(val, (datetime.date, datetime.datetime)):
        s = val.isoformat()
    elif isinstance(val, decimal.Decimal):
        s = format(val, 'f')
    else:
        s = str(val)
    try:
        from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
        s = ILLEGAL_CHARACTERS_RE.sub('', s)
    except Exception:
        pass
    return s


@app.route('/api/importacao/distribuicao/sms-automatizados/excel', methods=['GET'])
def api_distribuicao_sms_automatizados_excel():
    """Excel do roteiro SMS/e-mail: folhas SMS e EMAIL + coluna SMS e e-mail? (`_SMS_AUTOM_EXCEL_SQL`)."""
    if not session.get('funcionario_id'):
        return jsonify({'error': 'Nao autenticado. Faca login novamente.'}), 401
    if not _pode_sms_automatizados_importacao():
        return jsonify({'error': 'Acesso restrito a gestores ou administradores.'}), 403

    try:
        from openpyxl import Workbook
    except ImportError:
        return jsonify({
            'error': 'Biblioteca openpyxl nao instalada. Execute: pip install openpyxl',
        }), 503

    conn = _get_db()
    try:
        wb = Workbook()
        with conn.cursor() as cursor:
            cursor.execute(_SMS_AUTOM_EXCEL_SQL)
            linhas = cursor.fetchall() or []
            _sms_autom_fill_rota_excel_duas_folhas(
                cursor,
                wb,
                linhas,
                'Nenhum contrato no roteiro de dias (0, 16, 31, 61, 85) neste momento.',
            )
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
    except Exception as exc:
        app.logger.exception('sms-automatizados/excel: falha ao consultar ou montar ficheiro')
        return jsonify({'error': f'Falha ao gerar o Excel: {exc}'}), 500
    finally:
        try:
            conn.close()
        except Exception:
            pass

    fname = (
        'sms_email_automatizados_distribuicao_'
        + datetime.datetime.now().strftime('%Y%m%d_%H%M')
        + '.xlsx'
    )
    return send_file(
        buf,
        as_attachment=True,
        download_name=fname,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )


def _negativacao_distribuicao_excel_row(r):
    """Colunas alinhadas à tabela da carteira no módulo Negativação (sem botões)."""
    g = r.get('grupo')
    c = r.get('cota')
    g_str = _xlsx_cell_str(g)
    c_str = _xlsx_cell_str(c)
    gc = f'{g_str}/{c_str}' if (g_str or c_str) else ''
    tipo_ev = (r.get('tipo_evento') or '').strip()
    status_evt = _xlsx_cell_str(tipo_ev if tipo_ev else r.get('status'))
    da = r.get('dias_atraso')
    try:
        da_cell = int(da) if da is not None else ''
    except (TypeError, ValueError):
        da_cell = _xlsx_cell_str(da)
    return (
        g_str,
        c_str,
        gc,
        _xlsx_cell_str(r.get('numero_parcela')),
        da_cell,
        _xlsx_cell_str(status_evt),
        _xlsx_cell_str(r.get('data_negativacao')),
        _xlsx_cell_str(r.get('funcionario_nome')),
    )


_NEG_LISTAGEM_EXCEL_SORT_COLS = frozenset({
    'contrato', 'parcela', 'dias', 'status', 'data', 'operador',
})


def _negativacao_listagem_normalize_excel_sort_col(col, fallback):
    c = (col or '').strip().lower()
    if c in _NEG_LISTAGEM_EXCEL_SORT_COLS:
        return c
    fb = (fallback or 'data').strip().lower()
    return fb if fb in _NEG_LISTAGEM_EXCEL_SORT_COLS else 'data'


def _parse_neg_num_excel_sort(v):
    if v is None:
        return None
    if isinstance(v, bool):
        return None
    if isinstance(v, int):
        return float(v)
    if isinstance(v, float):
        return v
    s = str(v).strip().replace(',', '.')
    if not s:
        return None
    try:
        return float(s)
    except (TypeError, ValueError):
        return None


def _negativacao_listagem_excel_sort_rows(rows, col, dir_):
    """Alinhado a ``sortNegAtivosRowsCopy`` em ``static/negativacao.js``."""
    if not rows:
        return []
    col = _negativacao_listagem_normalize_excel_sort_col(col, 'data')
    dir_l = (dir_ or 'desc').strip().lower()
    mult = -1 if dir_l == 'desc' else 1

    def tie_break(ra, rb):
        ida = _parse_neg_num_excel_sort(ra.get('id'))
        idb = _parse_neg_num_excel_sort(rb.get('id'))
        if ida is not None and idb is not None and ida != idb:
            return -1 if ida < idb else (1 if ida > idb else 0)
        pa = _parse_neg_num_excel_sort(ra.get('id_parcela'))
        pb = _parse_neg_num_excel_sort(rb.get('id_parcela'))
        if pa is not None and pb is not None and pa != pb:
            return -1 if pa < pb else (1 if pa > pb else 0)
        return 0

    def cmp_str(a, b):
        if a < b:
            return -1 * mult
        if a > b:
            return 1 * mult
        return 0

    def cmp_rows(ra, rb):
        c = col
        if c == 'contrato':
            ga = _parse_neg_num_excel_sort(ra.get('grupo'))
            gb = _parse_neg_num_excel_sort(rb.get('grupo'))
            ca = _parse_neg_num_excel_sort(ra.get('cota'))
            cb = _parse_neg_num_excel_sort(rb.get('cota'))
            if ga is not None and gb is not None and ga != gb:
                return -1 if (ga - gb) * mult < 0 else (1 if (ga - gb) * mult > 0 else 0)
            if ca is not None and cb is not None and ca != cb:
                return -1 if (ca - cb) * mult < 0 else (1 if (ca - cb) * mult > 0 else 0)
            sa = f'{ra.get("grupo") or ""}/{ra.get("cota") or ""}'
            sb = f'{rb.get("grupo") or ""}/{rb.get("cota") or ""}'
            sc = cmp_str(sa, sb)
            if sc:
                return sc
            return tie_break(ra, rb)
        if c == 'parcela':
            pa = _parse_neg_num_excel_sort(ra.get('numero_parcela'))
            pb = _parse_neg_num_excel_sort(rb.get('numero_parcela'))
            if pa is None:
                pa = float('-inf') if dir_l == 'desc' else float('inf')
            if pb is None:
                pb = float('-inf') if dir_l == 'desc' else float('inf')
            if pa != pb:
                return -1 if (pa - pb) * mult < 0 else (1 if (pa - pb) * mult > 0 else 0)
            return tie_break(ra, rb)
        if c == 'dias':
            da = _parse_neg_num_excel_sort(ra.get('dias_atraso'))
            db = _parse_neg_num_excel_sort(rb.get('dias_atraso'))
            if da is None:
                da = float('-inf') if dir_l == 'desc' else float('inf')
            if db is None:
                db = float('-inf') if dir_l == 'desc' else float('inf')
            if da != db:
                return -1 if (da - db) * mult < 0 else (1 if (da - db) * mult > 0 else 0)
            return tie_break(ra, rb)
        if c == 'status':
            sta = str(ra.get('tipo_evento') or ra.get('status') or '')
            stb = str(rb.get('tipo_evento') or rb.get('status') or '')
            st = cmp_str(sta, stb)
            if st:
                return st
            return tie_break(ra, rb)
        if c == 'data':
            dta = str(ra.get('data_negativacao') or '').replace(' ', 'T')
            dtb = str(rb.get('data_negativacao') or '').replace(' ', 'T')
            dt = cmp_str(dta, dtb)
            if dt:
                return dt
            return tie_break(ra, rb)
        if c == 'operador':
            oa = str(ra.get('funcionario_nome') or '').lower()
            ob = str(rb.get('funcionario_nome') or '').lower()
            oc = cmp_str(oa, ob)
            if oc:
                return oc
            return tie_break(ra, rb)
        return tie_break(ra, rb)

    return sorted(rows, key=functools.cmp_to_key(lambda a, b: cmp_rows(a, b)))


def _negativacao_listagem_excel_row(r):
    """Colunas para exportacao no modulo Negativacao (Carteira ou Geral)."""
    g = r.get('grupo')
    c = r.get('cota')
    g_str = _xlsx_cell_str(g)
    c_str = _xlsx_cell_str(c)
    gc = f'{g_str}/{c_str}' if (g_str or c_str) else ''
    tipo_ev = (r.get('tipo_evento') or '').strip()
    da = r.get('dias_atraso')
    try:
        da_cell = int(da) if da is not None else ''
    except (TypeError, ValueError):
        da_cell = _xlsx_cell_str(da)
    id_contrato = r.get('id_contrato')
    id_parcela = r.get('id_parcela')
    id_reg = r.get('id')
    try:
        id_contrato_cell = int(id_contrato) if id_contrato is not None else ''
    except (TypeError, ValueError):
        id_contrato_cell = _xlsx_cell_str(id_contrato)
    try:
        id_parcela_cell = int(id_parcela) if id_parcela is not None else ''
    except (TypeError, ValueError):
        id_parcela_cell = _xlsx_cell_str(id_parcela)
    try:
        id_reg_cell = int(id_reg) if id_reg is not None else ''
    except (TypeError, ValueError):
        id_reg_cell = _xlsx_cell_str(id_reg)
    return (
        g_str,
        c_str,
        gc,
        _xlsx_cell_str(r.get('numero_contrato')),
        id_contrato_cell,
        _xlsx_cell_str(r.get('numero_parcela')),
        id_parcela_cell,
        id_reg_cell,
        da_cell,
        _xlsx_cell_str(r.get('status')),
        _xlsx_cell_str(tipo_ev),
        _xlsx_cell_str(r.get('data_negativacao')),
        _xlsx_cell_str(r.get('funcionario_nome')),
        _xlsx_cell_str(r.get('contrato_status')),
        _xlsx_cell_str(r.get('detalhe')),
    )


@app.route(
    '/api/importacao/distribuicao/negativacao-positivacao/excel',
    methods=['GET'],
)
def api_distribuicao_negativacao_positivacao_excel():
    """Excel com duas folhas: lista de negativação e de positivação (carteira, todos operadores)."""
    if not session.get('funcionario_id'):
        return jsonify({'error': 'Nao autenticado. Faca login novamente.'}), 401
    if not _pode_sms_automatizados_importacao():
        return jsonify({'error': 'Acesso restrito a gestores ou administradores.'}), 403

    try:
        from openpyxl import Workbook
    except ImportError:
        return jsonify({
            'error': 'Biblioteca openpyxl nao instalada. Execute: pip install openpyxl',
        }), 503

    hdr = (
        'Grupo',
        'Cota',
        'Grupo/Cota',
        'Parcela',
        'Atraso',
        'Status ou evento',
        'Data',
        'Operador',
    )
    conn = _get_db()
    cursor = conn.cursor()
    try:
        payload = _negativacao_listagem_payload_interno(
            cursor,
            q='',
            tipo_busca='contrato',
            evento='todos',
            status_ativo='',
            data_inicio=None,
            data_fim=None,
            sort_at='data',
            order_at='desc',
            apenas_cobranca=True,
            sem_operador_cobranca=False,
            preview_ativos=False,
            funcionario_id_filtro=None,
        )
    except Exception as exc:
        app.logger.exception('negativacao-positivacao/excel: falha ao montar listagem')
        try:
            cursor.close()
        except Exception:
            pass
        try:
            conn.close()
        except Exception:
            pass
        return jsonify({'error': f'Falha ao consultar dados para o Excel: {exc}'}), 500

    try:
        cursor.close()
    except Exception:
        pass
    try:
        conn.close()
    except Exception:
        pass

    neg_rows = payload.get('ativos_negativados') or []
    pos_rows = payload.get('ativos_positivados') or []

    try:
        wb = Workbook()
        ws_neg = wb.active
        ws_neg.title = 'NEGATIVAÇÃO'
        ws_neg.append(hdr)
        if not neg_rows:
            ws_neg.append(
                ('Nenhum registro na lista de negativação neste momento.', '', '', '', '', '', '', '')
            )
        else:
            for r in neg_rows:
                ws_neg.append(_negativacao_distribuicao_excel_row(r))

        ws_pos = wb.create_sheet(title='POSITIVAÇÃO')
        ws_pos.append(hdr)
        if not pos_rows:
            ws_pos.append(
                ('Nenhum registro na lista de positivação neste momento.', '', '', '', '', '', '', '')
            )
        else:
            for r in pos_rows:
                ws_pos.append(_negativacao_distribuicao_excel_row(r))

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
    except Exception as exc:
        app.logger.exception('negativacao-positivacao/excel: falha ao montar ficheiro')
        return jsonify({'error': f'Falha ao gerar o Excel: {exc}'}), 500

    fname = (
        'negativacao_positivacao_distribuicao_'
        + datetime.datetime.now().strftime('%Y%m%d_%H%M')
        + '.xlsx'
    )
    return send_file(
        buf,
        as_attachment=True,
        download_name=fname,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )


@app.route('/api/importacao/distribuicao/sms-automatizados', methods=['POST'])
def api_distribuicao_sms_automatizados():
    """Envia SMS e e-mail em lote (MessageCenter) para contratos abertos no roteiro automático.

    Regra: DATEDIFF(hoje, MIN vencimento de parcelas abertas) em 0, 16, 31, 61 ou 85 dias;
    mesmo texto nos dois canais. Contratos que já tiveram SMS ou e-mail no dia são ignorados.
    Body JSON opcional: ``{"canais": ["sms"]}``, ``["email"]`` ou ambos (padrao: ambos).

    Apenas Gestor ou Administrador.
    """
    fid = session.get('funcionario_id')
    if not fid:
        return jsonify({'error': 'Nao autenticado. Faca login novamente.'}), 401
    if not _pode_sms_automatizados_importacao():
        return jsonify({'error': 'Acesso restrito a gestores ou administradores.'}), 403

    payload = request.get_json(silent=True) or {}
    canais = {'sms', 'email'}
    raw_canais = payload.get('canais')
    if raw_canais is not None:
        canais = set()
        if isinstance(raw_canais, list):
            for x in raw_canais:
                xl = str(x).strip().lower()
                if xl == 'sms':
                    canais.add('sms')
                elif xl in ('email', 'e-mail', 'mail'):
                    canais.add('email')
        if not canais:
            return jsonify({'error': 'canais deve ser uma lista com "sms" e/ou "email".'}), 400

    remetente_nome = _primeiro_nome(session.get('funcionario_nome'))
    conn = _get_db()
    stats = {
        'envios_sms': 0,
        'envios_email': 0,
        'enviados': 0,
        'falhas': 0,
        'ignorados_sem_entrada': 0,
        'ignorados_sem_template': 0,
        'ignorados_sem_telefone': 0,
        'ignorados_sem_email': 0,
        'ignorados_ja_enviados_hoje': 0,
        'contratos_processados': 0,
    }
    erros_amostra = []

    try:
        with conn.cursor() as cursor:
            cursor.execute(_SMS_AUTOM_DISTRIBUICAO_SQL)
            contratos = cursor.fetchall() or []

        for row in contratos:
            stats['contratos_processados'] += 1
            _auto_envio_contrato_canais(
                conn,
                int(fid),
                row,
                remetente_nome,
                stats,
                erros_amostra,
                canais,
                usar_smtp_google=True,
            )

    finally:
        conn.close()

    stats['enviados'] = stats['envios_sms'] + stats['envios_email']
    return jsonify({
        'ok': True,
        'canais': sorted(canais),
        'enviados': stats['enviados'],
        'envios_sms': stats['envios_sms'],
        'envios_email': stats['envios_email'],
        'falhas': stats['falhas'],
        'ignorados_sem_entrada': stats['ignorados_sem_entrada'],
        'ignorados_sem_template': stats['ignorados_sem_template'],
        'ignorados_sem_telefone': stats['ignorados_sem_telefone'],
        'ignorados_sem_email': stats['ignorados_sem_email'],
        'ignorados_ja_enviados_hoje': stats['ignorados_ja_enviados_hoje'],
        'contratos_processados': stats['contratos_processados'],
        'erros_amostra': erros_amostra,
    })


# ---------------------------------------------------------------------------
# Helpers para API de Busca
# ---------------------------------------------------------------------------

def _get_db():
    """Abre conexão MySQL com PyMySQL usando ``DB_CONFIG`` (variáveis ``DB_*`` / ``MYSQL_*``)."""
    return pymysql.connect(**DB_CONFIG, cursorclass=pymysql.cursors.DictCursor)


def _schema_pronto_para_importacao():
    """Garante que o schema basico foi aplicado (`Banco/criar_banco.py`).

    Retorna (True, None) ou (False, mensagem para o operador).
    """
    tabelas_obrigatorias = (
        "arquivos_gm",
        "empresa",
        "contrato",
        "parcela",
    )
    try:
        conn = _get_db()
        cur = conn.cursor()
        for nome in tabelas_obrigatorias:
            cur.execute("SHOW TABLES LIKE %s", (nome,))
            if not cur.fetchone():
                cur.close()
                conn.close()
                dbn = DB_CONFIG.get("database", "consorcio_gm")
                return (
                    False,
                    (
                        f"Schema ausente: a tabela '{nome}' nao existe no banco '{dbn}'. "
                        "Aplique o script de criacao:  python3 Banco/criar_banco.py  "
                        "(em seguida, se desejar: Banco/seed_funcionarios.py). "
                        "Depois repita a importacao."
                    ),
                )
        cur.close()
        conn.close()
        return True, None
    except Exception as e:
        return (
            False,
            "Nao foi possivel acessar o MariaDB. Verifique se o servico esta no ar, "
            "credenciais (variaveis DB_* / MYSQL_* ou padrao local) e se o banco existe. "
            f"Detalhe: {e}",
        )


def _serialize(obj):
    """Converte tipos nao-serializaveis do MySQL para JSON."""
    if isinstance(obj, (datetime.date, datetime.datetime)):
        return obj.isoformat()
    if isinstance(obj, decimal.Decimal):
        return float(obj)
    if isinstance(obj, bytes):
        try:
            return obj.decode('utf-8')
        except UnicodeDecodeError:
            return obj.decode('latin-1', errors='replace')
    return str(obj)


def _clean_rows(rows):
    return [{k: _serialize(v) if v is not None else None for k, v in r.items()} for r in rows]


def _clean_row(row):
    if not row:
        return None
    return {k: _serialize(v) if v is not None else None for k, v in row.items()}


# ---------------------------------------------------------------------------
# Introspeccao da tabela de bens
# ---------------------------------------------------------------------------
# A tabela de bens pode se chamar `bem` ou `bens` e seu schema pode
# variar (descricao, modelo, marca, codigo, etc). Tambem pode estar
# ligada ao contrato por `id_contrato` OU por (`grupo`, `cota`).
#
# Este helper inspeciona INFORMATION_SCHEMA uma unica vez por processo
# e devolve um dicionario com:
#     {
#         'table':    'bens' | 'bem' | None,
#         'text_cols': [<nomes das colunas textuais>],
#         'join_on':  'id_contrato' | 'grupo_cota' | None,
#     }
# Quando o tracker for ajustado e a tabela for populada, nada precisa
# mudar aqui: a busca passa a encontrar resultados automaticamente.

_BEM_SCHEMA_CACHE = None


def _get_bem_schema():
    """Retorna metadados da tabela de bens. Faz cache por processo."""
    global _BEM_SCHEMA_CACHE
    if _BEM_SCHEMA_CACHE is not None:
        return _BEM_SCHEMA_CACHE

    info = {'table': None, 'text_cols': [], 'join_on': None}
    try:
        conn = _get_db()
        cursor = conn.cursor()
        cursor.execute(
            """
            SELECT TABLE_NAME, COLUMN_NAME, DATA_TYPE
            FROM INFORMATION_SCHEMA.COLUMNS
            WHERE TABLE_SCHEMA = DATABASE()
              AND TABLE_NAME IN ('bem', 'bens')
            """
        )
        rows = cursor.fetchall()
        cursor.close()
        conn.close()

        if not rows:
            _BEM_SCHEMA_CACHE = info
            return info

        # Se ambas existirem, preferimos 'bens' (plural costuma ser o novo).
        tables = {r['TABLE_NAME']: [] for r in rows}
        for r in rows:
            tables[r['TABLE_NAME']].append(r)
        chosen = 'bens' if 'bens' in tables else 'bem'
        cols_rows = tables[chosen]

        text_types = {'varchar', 'char', 'text', 'tinytext', 'mediumtext', 'longtext'}
        all_cols = {r['COLUMN_NAME'].lower() for r in cols_rows}
        text_cols = [r['COLUMN_NAME'] for r in cols_rows
                     if str(r['DATA_TYPE']).lower() in text_types]

        # Prioriza colunas com semantica de descricao.
        preferred = ['descricao', 'descricao_bem', 'desc_bem', 'modelo',
                     'marca', 'categoria', 'codigo', 'codigo_bem', 'nome']
        ordered = [c for c in preferred if c in text_cols]
        ordered += [c for c in text_cols if c not in ordered]

        join_on = None
        if 'id_contrato' in all_cols:
            join_on = 'id_contrato'
        elif 'grupo' in all_cols and 'cota' in all_cols:
            join_on = 'grupo_cota'

        info = {'table': chosen, 'text_cols': ordered, 'join_on': join_on}
    except Exception:
        # Qualquer falha -> nenhum bem disponivel, busca devolve vazio.
        info = {'table': None, 'text_cols': [], 'join_on': None}

    _BEM_SCHEMA_CACHE = info
    return info


def _bem_join_clause(alias_contrato='c', alias_bem='b'):
    """Retorna string do JOIN com a tabela de bens (ou None se indisponivel)."""
    info = _get_bem_schema()
    if not info['table'] or not info['join_on']:
        return None
    if info['join_on'] == 'id_contrato':
        return (f"INNER JOIN {info['table']} {alias_bem} "
                f"ON {alias_bem}.id_contrato = {alias_contrato}.id ")
    # grupo + cota
    return (f"INNER JOIN {info['table']} {alias_bem} "
            f"ON {alias_bem}.grupo = {alias_contrato}.grupo "
            f"AND {alias_bem}.cota = {alias_contrato}.cota ")


def _bem_where_clause(termo, alias_bem='b'):
    """Retorna (sql_fragment, params) para filtrar por termo nas colunas textuais do bem.

    Se nao houver colunas textuais, devolve (None, []).
    """
    info = _get_bem_schema()
    if not info['text_cols']:
        return None, []
    parts = [f"{alias_bem}.`{c}` LIKE %s" for c in info['text_cols']]
    params = [f'%{termo}%'] * len(info['text_cols'])
    return "(" + " OR ".join(parts) + ")", params


def _bem_rotulo_coluna(nome_coluna):
    """Rotulo amigavel a partir do nome da coluna na tabela de bens."""
    if not nome_coluna:
        return ''
    s = str(nome_coluna).replace('_', ' ').strip()
    return s[:1].upper() + s[1:] if s else ''


def _bem_where_clause_from_params(args, alias_bem='b'):
    """WHERE com AND: aceita apenas bem_<nome_coluna> que exista em text_cols (tabela bem/bens)."""
    info = _get_bem_schema()
    text_cols = info.get('text_cols') or []
    if not text_cols:
        return None, []
    lower_to_actual = {}
    for col in text_cols:
        lk = col.lower()
        if lk not in lower_to_actual:
            lower_to_actual[lk] = col
    parts = []
    params = []
    seen_cols = set()
    for key in args:
        if not isinstance(key, str) or not key.startswith('bem_'):
            continue
        suffix = key[4:]
        if not suffix:
            continue
        col = lower_to_actual.get(suffix.lower())
        if not col or col in seen_cols:
            continue
        raw = args.get(key, '') or ''
        val = raw.strip() if isinstance(raw, str) else str(raw).strip()
        if not val:
            continue
        parts.append(f"{alias_bem}.`{col}` LIKE %s")
        params.append(f'%{val}%')
        seen_cols.add(col)
    if not parts:
        return None, []
    return "(" + " AND ".join(parts) + ")", params


def _bem_concat_expr(alias_bem='b'):
    """Expressao SQL que concatena as colunas textuais para mostrar no resultado.

    Exemplo: "CONCAT_WS(' / ', b.descricao, b.modelo)".
    """
    info = _get_bem_schema()
    if not info['text_cols']:
        return "''"
    cols = ", ".join(f"{alias_bem}.`{c}`" for c in info['text_cols'])
    return f"CONCAT_WS(' / ', {cols})"


def _fetch_bens_para_contrato(cursor, contrato):
    """Retorna lista de bens (dicts) ligados ao contrato informado.

    Usa a chave de join detectada automaticamente (id_contrato ou grupo+cota).
    Devolve lista vazia quando a tabela nao existe ou nao possui chave de join.
    """
    info = _get_bem_schema()
    if not info['table'] or not info['join_on']:
        return []

    table = info['table']
    try:
        if info['join_on'] == 'id_contrato':
            cursor.execute(
                f"SELECT * FROM `{table}` WHERE id_contrato = %s",
                (contrato['id'],),
            )
        else:
            cursor.execute(
                f"SELECT * FROM `{table}` WHERE grupo = %s AND cota = %s",
                (contrato.get('grupo'), contrato.get('cota')),
            )
        return _clean_rows(cursor.fetchall())
    except Exception:
        return []


# ---------------------------------------------------------------------------
# API: Busca por Pessoa / Contrato
# ---------------------------------------------------------------------------

@app.route('/api/busca/campos-bem')
def api_busca_campos_bem():
    """Lista colunas textuais da tabela bem/bens para montar filtros na UI (espelha o schema real)."""
    info = _get_bem_schema()
    colunas = []
    for nome in info.get('text_cols') or []:
        colunas.append({'nome': nome, 'rotulo': _bem_rotulo_coluna(nome)})
    return jsonify({
        'ok': True,
        'tabela': info.get('table'),
        'colunas': colunas,
    })


@app.route('/api/busca')
def api_busca():
    tipo = request.args.get('tipo', '').strip()
    termo = request.args.get('termo', '').strip()
    bem_where_cached = None
    bem_params_cached = []

    if tipo == 'bem':
        bem_where_cached, bem_params_cached = _bem_where_clause_from_params(request.args, 'b')
        if not bem_where_cached and termo:
            bem_where_cached, bem_params_cached = _bem_where_clause(termo, 'b')
        if not bem_where_cached:
            return jsonify({'results': [], 'tipo': tipo})
    elif not termo:
        return jsonify({'results': [], 'tipo': tipo})

    conn = _get_db()
    cursor = conn.cursor()

    results = []
    if tipo == 'pessoa':
        status_filtro = _normalize_contrato_status_arg(request.args.get('status', '').strip())
        if status_filtro and status_filtro not in _STATUS_CONTRATO_VALIDOS:
            status_filtro = ''
        if status_filtro:
            # Pessoas (devedor ou avalista) com ao menos um contrato nesse status
            cursor.execute(
                "SELECT DISTINCT p.id, p.cpf_cnpj, p.nome_completo, p.profissao "
                "FROM pessoa p "
                "INNER JOIN contrato c ON (c.id_pessoa = p.id OR c.id_avalista = p.id) "
                "WHERE (p.cpf_cnpj LIKE %s OR p.nome_completo LIKE %s) "
                "AND c.status = %s "
                "LIMIT 50",
                (f'%{termo}%', f'%{termo}%', status_filtro),
            )
        else:
            cursor.execute(
                "SELECT id, cpf_cnpj, nome_completo, profissao "
                "FROM pessoa "
                "WHERE cpf_cnpj LIKE %s OR nome_completo LIKE %s "
                "LIMIT 50",
                (f'%{termo}%', f'%{termo}%'),
            )
        results = _clean_rows(cursor.fetchall())

    elif tipo == 'contrato':
        status_filtro = _normalize_contrato_status_arg(request.args.get('status', '').strip())
        if status_filtro and status_filtro not in _STATUS_CONTRATO_VALIDOS:
            status_filtro = ''
        base_select = (
            "SELECT c.id, c.grupo, c.cota, c.numero_contrato, c.status, "
            "       p.nome_completo AS nome_devedor "
            "FROM contrato c "
            "LEFT JOIN pessoa p ON c.id_pessoa = p.id "
        )
        parts = [p.strip() for p in termo.replace('-', '/').split('/')]
        if len(parts) == 2:
            grupo, cota = parts
            where = "WHERE c.grupo LIKE %s AND c.cota LIKE %s"
            params = [f'%{grupo}%', f'%{cota}%']
        else:
            where = "WHERE (c.grupo LIKE %s OR c.numero_contrato LIKE %s)"
            params = [f'%{termo}%', f'%{termo}%']

        if status_filtro:
            where += " AND c.status = %s"
            params.append(status_filtro)

        cursor.execute(base_select + where + " LIMIT 50", params)
        results = _clean_rows(cursor.fetchall())

    elif tipo == 'bem':
        # Busca contratos por descricao do bem (modelo, marca, etc).
        # Schema da tabela de bens e detectado dinamicamente.
        status_filtro = _normalize_contrato_status_arg(request.args.get('status', '').strip())
        if status_filtro and status_filtro not in _STATUS_CONTRATO_VALIDOS:
            status_filtro = ''
        join_clause = _bem_join_clause('c', 'b')
        bem_where, bem_params = bem_where_cached, list(bem_params_cached)

        if not join_clause or not bem_where:
            # Tabela de bens indisponivel ou sem colunas textuais.
            results = []
        else:
            bem_expr = _bem_concat_expr('b')
            sql = (
                "SELECT c.id, c.grupo, c.cota, c.numero_contrato, c.status, "
                "       p.nome_completo AS nome_devedor, "
                f"       GROUP_CONCAT(DISTINCT {bem_expr} SEPARATOR ' | ') AS bem_descricao "
                "FROM contrato c "
                + join_clause +
                "LEFT JOIN pessoa p ON c.id_pessoa = p.id "
                "WHERE " + bem_where + " "
            )
            params = list(bem_params)
            if status_filtro:
                sql += "AND c.status = %s "
                params.append(status_filtro)
            sql += (
                "GROUP BY c.id, c.grupo, c.cota, c.numero_contrato, c.status, "
                "         p.nome_completo "
                "LIMIT 50"
            )
            cursor.execute(sql, params)
            results = _clean_rows(cursor.fetchall())

    cursor.close()
    conn.close()
    return jsonify({'results': results, 'tipo': tipo})


# Tipos validos para INSERT em `telefone` e `email` (UNIQUE id_pessoa + tipo).
_TELEFONE_TIPOS = (
    'fixo', 'celular', 'comercial', 'comercial_devedor', 'recados', 'outro',
)
_EMAIL_TIPOS = ('principal', 'secundario', 'comercial', 'outro')


def _mysql_table_columns_lower(cursor, table):
    """Nomes de colunas em minúsculas para montar INSERT compatível com o schema."""
    cursor.execute(f'SHOW COLUMNS FROM `{table}`')
    rows = cursor.fetchall() or []
    names = []
    for r in rows:
        if isinstance(r, dict):
            fn = r.get('Field') or r.get('field')
        else:
            fn = r[0] if r else None
        if fn:
            names.append(str(fn).lower())
    return frozenset(names)


def _migrate_contato_fonte_manual_para_terceiro(cursor):
    """Bases que chegaram a ter ENUM com `manual`: converte dados e volta ao schema de 3 valores."""
    for tbl in ('telefone', 'email'):
        try:
            cursor.execute(
                f"UPDATE `{tbl}` SET `fonte` = 'terceiro' WHERE `fonte` = 'manual'"
            )
        except Exception as exc:
            app.logger.debug('_migrate_contato_fonte_manual_para_terceiro update %s: %s', tbl, exc)
        try:
            cursor.execute(
                f"ALTER TABLE `{tbl}` MODIFY COLUMN `fonte` "
                "ENUM('GMAC','enriquecimento','terceiro') DEFAULT 'GMAC'"
            )
        except Exception as exc:
            app.logger.debug('_migrate_contato_fonte_manual_para_terceiro alter %s: %s', tbl, exc)


@app.route('/api/pessoa/<int:pessoa_id>/telefone', methods=['POST'])
def api_pessoa_add_telefone(pessoa_id):
    """Adiciona um registro de telefone (UK id_pessoa+tipo+numero: evita duplicata exata)."""
    if not session.get('funcionario_id'):
        return jsonify({'error': 'Nao autenticado. Faca login novamente.'}), 401
    data = request.get_json(silent=True) or {}
    tipo = (data.get('tipo') or '').strip()
    ramal = (data.get('ramal') or '').strip() or None
    numero = (data.get('numero') or '').strip()
    if not tipo or tipo not in _TELEFONE_TIPOS:
        return jsonify({'error': f'Tipo invalido. Valores: {", ".join(_TELEFONE_TIPOS)}.'}), 400
    if not numero:
        return jsonify({'error': 'Numero e obrigatorio.'}), 400
    conn = _get_db()
    cursor = conn.cursor()
    try:
        cursor.execute("SELECT 1 FROM pessoa WHERE id = %s", (pessoa_id,))
        if not cursor.fetchone():
            return jsonify({'error': 'Pessoa nao encontrada.'}), 404
        cols = _mysql_table_columns_lower(cursor, 'telefone')
        if 'fonte' in cols:
            _migrate_contato_fonte_manual_para_terceiro(cursor)
            cursor.execute(
                "INSERT INTO telefone (id_pessoa, tipo, numero, ramal, fonte) "
                "VALUES (%s, %s, %s, %s, %s)",
                (pessoa_id, tipo, numero, ramal, 'terceiro'),
            )
        else:
            cursor.execute(
                "INSERT INTO telefone (id_pessoa, tipo, numero, ramal) VALUES (%s, %s, %s, %s)",
                (pessoa_id, tipo, numero, ramal),
            )
        new_id = cursor.lastrowid
        conn.commit()
        cursor.execute("SELECT * FROM telefone WHERE id = %s", (new_id,))
        row = _clean_row(cursor.fetchone())
        return jsonify({'ok': True, 'telefone': row})
    except IntegrityError as exc:
        if conn:
            try:
                conn.rollback()
            except Exception:
                pass
        code = exc.args[0] if getattr(exc, 'args', None) else None
        msg = str(exc)
        if code == 1062:
            return jsonify({
                'error': 'Ja existe este numero deste tipo para esta pessoa.',
            }), 409
        if code == 1452:
            return jsonify({'error': 'Pessoa nao encontrada.'}), 404
        return jsonify({'error': f'Falha de integridade ao inserir telefone: {msg}'}), 400
    except Exception as exc:
        if conn:
            try:
                conn.rollback()
            except Exception:
                pass
        return jsonify({'error': f'Erro ao inserir telefone: {exc}'}), 500
    finally:
        try:
            cursor.close()
        except Exception:
            pass
        try:
            conn.close()
        except Exception:
            pass


@app.route('/api/pessoa/<int:pessoa_id>/email', methods=['POST'])
def api_pessoa_add_email(pessoa_id):
    """Adiciona um registro de e-mail (UK id_pessoa+tipo+email: evita duplicata exata)."""
    if not session.get('funcionario_id'):
        return jsonify({'error': 'Nao autenticado. Faca login novamente.'}), 401
    data = request.get_json(silent=True) or {}
    tipo = (data.get('tipo') or '').strip()
    endereco = (data.get('email') or '').strip()
    if not tipo or tipo not in _EMAIL_TIPOS:
        return jsonify({'error': f'Tipo invalido. Valores: {", ".join(_EMAIL_TIPOS)}.'}), 400
    if not endereco:
        return jsonify({'error': 'Email e obrigatorio.'}), 400
    conn = _get_db()
    cursor = conn.cursor()
    try:
        cursor.execute("SELECT 1 FROM pessoa WHERE id = %s", (pessoa_id,))
        if not cursor.fetchone():
            return jsonify({'error': 'Pessoa nao encontrada.'}), 404
        cols = _mysql_table_columns_lower(cursor, 'email')
        if 'fonte' in cols:
            _migrate_contato_fonte_manual_para_terceiro(cursor)
            cursor.execute(
                "INSERT INTO email (id_pessoa, tipo, email, fonte) VALUES (%s, %s, %s, %s)",
                (pessoa_id, tipo, endereco, 'terceiro'),
            )
        else:
            cursor.execute(
                "INSERT INTO email (id_pessoa, tipo, email) VALUES (%s, %s, %s)",
                (pessoa_id, tipo, endereco),
            )
        new_id = cursor.lastrowid
        conn.commit()
        cursor.execute("SELECT * FROM email WHERE id = %s", (new_id,))
        row = _clean_row(cursor.fetchone())
        return jsonify({'ok': True, 'email': row})
    except IntegrityError as exc:
        if conn:
            try:
                conn.rollback()
            except Exception:
                pass
        code = exc.args[0] if getattr(exc, 'args', None) else None
        msg = str(exc)
        if code == 1062:
            return jsonify({
                'error': 'Ja existe este e-mail deste tipo para esta pessoa.',
            }), 409
        if code == 1452:
            return jsonify({'error': 'Pessoa nao encontrada.'}), 404
        return jsonify({'error': f'Falha de integridade ao inserir e-mail: {msg}'}), 400
    except Exception as exc:
        if conn:
            try:
                conn.rollback()
            except Exception:
                pass
        return jsonify({'error': f'Erro ao inserir e-mail: {exc}'}), 500
    finally:
        try:
            cursor.close()
        except Exception:
            pass
        try:
            conn.close()
        except Exception:
            pass


def _discar_numero_funcionario(funcionario_id, raw_numero):
    """Discagem via API B2 (credenciais DISCADOR_* no .env). Retorna dict com ok/error."""
    raw = (raw_numero or '').strip()
    digits = ''.join(ch for ch in raw if ch.isdigit())
    if not digits:
        return {'ok': False, 'error': 'Numero invalido.'}
    if not digits.startswith('55'):
        digits = '55' + digits

    url = (os.environ.get('DISCADOR_URL') or '').strip()
    usuario = (os.environ.get('DISCADOR_USUARIO') or '').strip()
    token = (os.environ.get('DISCADOR_TOKEN') or '').strip()
    if not (url and usuario and token):
        return {
            'ok': False,
            'error': 'Discador nao configurado (env DISCADOR_URL, DISCADOR_USUARIO, DISCADOR_TOKEN).',
        }

    conn = _get_db()
    try:
        with conn.cursor() as cur:
            cur.execute('SELECT ramal FROM funcionario WHERE id = %s', (int(funcionario_id),))
            row = cur.fetchone()
    finally:
        conn.close()

    ramal = (row or {}).get('ramal') if row else None
    if ramal is None or ramal == '':
        return {'ok': False, 'error': 'Seu usuario nao tem ramal cadastrado.'}
    try:
        ramal_int = int(ramal)
    except (TypeError, ValueError):
        return {'ok': False, 'error': 'Ramal invalido no cadastro do funcionario.'}

    body = {
        'dados': {
            'numero_ramal_origem': ramal_int,
            'numero_destino': digits,
            'variaveis': [{}],
        },
    }
    try:
        resp = requests.post(
            url,
            headers={'usuario': usuario, 'token': token, 'Content-Type': 'application/json'},
            json=body,
            timeout=15,
        )
    except requests.RequestException as exc:
        return {'ok': False, 'error': f'Erro de rede ao chamar discador: {exc}'}
    try:
        data = resp.json()
    except Exception:
        data = {'raw': (resp.text or '')[:500]}
    if not resp.ok:
        return {
            'ok': False,
            'error': 'Falha no discador.',
            'status': resp.status_code,
            'detalhe': data,
        }
    return {'ok': True, 'discador': data, 'numero_destino': digits}


@app.route('/api/discar', methods=['POST'])
def api_discar():
    """Proxy seguro: chama a API de discador com ramal do funcionario (credenciais no .env)."""
    fid = session.get('funcionario_id')
    if not fid:
        return jsonify({'error': 'Nao autenticado. Faca login novamente.'}), 401
    payload = request.get_json(silent=True) or {}
    result = _discar_numero_funcionario(fid, payload.get('numero'))
    if not result.get('ok'):
        err = result.get('error') or 'Falha na discagem.'
        code = 400 if 'invalido' in err.lower() or 'ramal' in err.lower() else 502
        if 'nao configurado' in err.lower():
            code = 500
        body = {'error': err}
        if result.get('detalhe'):
            body['detalhe'] = result['detalhe']
        if result.get('status'):
            body['status'] = result['status']
        return jsonify(body), code
    return jsonify({'ok': True, 'discador': result.get('discador')})


WHATSAPP_AB_URL = 'https://joaobarbosa.atenderbem.com/int/enqueueMessageToSend'


@app.route('/api/enviar-whatsapp', methods=['POST'])
def api_enviar_whatsapp():
    """Proxy WhatsApp: usa fila e apikey do usuario logado (`funcionario`) e encaminha para AtenderBem."""
    fid = session.get('funcionario_id')
    if not fid:
        return jsonify({'error': 'Nao autenticado. Faca login novamente.'}), 401
    payload = request.get_json(silent=True) or {}
    raw = (payload.get('numero') or '').strip()
    msg = (payload.get('mensagem') or '').strip()
    if not raw:
        return jsonify({'error': 'Numero invalido.'}), 400
    if not msg:
        return jsonify({'error': 'Mensagem vazia.'}), 400
    if len(msg) > 3000:
        return jsonify({'error': 'Mensagem muito longa (max. 3000 caracteres).'}), 400

    conn = _get_db()
    try:
        with conn.cursor() as cur:
            cur.execute(
                'SELECT fila, apikey FROM funcionario WHERE id = %s',
                (int(fid),),
            )
            row = cur.fetchone()
    finally:
        conn.close()

    fila = (row or {}).get('fila') if row else None
    if fila in (None, ''):
        return jsonify({'error': 'Seu usuario nao tem fila cadastrada para envio de WhatsApp.'}), 400
    try:
        fila_int = int(fila)
    except (TypeError, ValueError):
        return jsonify({'error': 'Fila invalida no cadastro do funcionario.'}), 400

    apikey_raw = (row or {}).get('apikey') if row else None
    if isinstance(apikey_raw, bytes):
        apikey_raw = apikey_raw.decode('utf-8', errors='replace')
    apikey_val = (apikey_raw or '').strip()
    if not apikey_val:
        return jsonify({
            'error': 'Seu usuario nao tem apikey cadastrada para envio de WhatsApp (campo apikey em funcionario).',
        }), 400

    body = {
        'queueId': fila_int,
        'apiKey': apikey_val,
        'number': raw,
        'text': msg,
    }
    try:
        resp = requests.post(
            WHATSAPP_AB_URL,
            json=body,
            headers={
                'Content-Type': 'application/json',
                'Accept': '*/*',
                'User-Agent': 'PostmanRuntime/7.53.0',
            },
            timeout=30,
        )
    except requests.RequestException as exc:
        return jsonify({'error': f'Erro de rede ao enviar WhatsApp: {exc}'}), 502
    try:
        data = resp.json()
    except Exception:
        data = {'raw': (resp.text or '')[:500]}
    if not resp.ok:
        return jsonify({
            'error': 'Falha ao enviar WhatsApp.',
            'status': resp.status_code,
            'detalhe': data,
        }), resp.status_code
    return jsonify({'ok': True, 'whatsapp': data})


# MessageCenter SMS (integracao enviarsms) — credenciais fixas conforme integracao.
SMS_MC_URL = 'https://sistema.messagecenter.com.br/api/Integracao/enviarsms'
SMS_MC_HEADER_NAME = 'apikey'
SMS_MC_HEADER_VALUE = (
    'MC.cC719ae5-22B3-439b-9D63-4D544f79Fffc-788B1CE2-9af2-417F-85Ba-67d3E16a7243'
)

EMAIL_MC_URL = 'https://sistema.messagecenter.com.br/api/Integracao/EnviarEmailHtml'
EMAIL_MC_REMETENTE = os.environ.get(
    'EMAIL_MC_REMETENTE', 'atendimento@joaobarbosa.com.br'
)


def _importacao_envio_email_auto(remetente_nome, destinatario, cliente_primeiro_nome, corpo_html, msg_plain):
    """E-mail do lote automático **só na página Importação** (ver ``usar_smtp_google`` em ``_auto_envio_contrato_canais``).

    Se SMTP Google estiver configurado no ambiente, usa ``send_google_workspace_email``; senão MessageCenter.
    Retorna (ok_bool, erro_str_ou_None, data_dict|None) — mesma forma que ``_messagecenter_post_email_html``.
    """
    if _smtp_google_workspace.google_smtp_config_from_env():
        ok, err = _smtp_google_workspace.send_google_workspace_email(
            [destinatario],
            'Cobranca',
            corpo_html,
            text_body=(msg_plain or '').strip() or None,
        )
        return ok, err, None
    return _messagecenter_post_email_html(
        remetente_nome, destinatario, cliente_primeiro_nome, corpo_html,
    )


def _messagecenter_post_email_html(remetente_nome, destinatario, cliente_primeiro_nome, corpo_html):
    """POST MessageCenter EnviarEmailHtml. Retorna (ok_bool, erro_str_ou_None, data_dict)."""
    destinatario = (destinatario or '').strip()
    if not destinatario:
        return False, 'Destinatario vazio.', None
    query_params = {
        'Destinatario': destinatario,
        'RemetenteNome': remetente_nome or '',
        'RemetenteEmail': EMAIL_MC_REMETENTE,
        'Assunto': 'Cobranca',
        'Identificador': '',
        'ClienteNome': cliente_primeiro_nome or '',
        'ClienteDocumento': '',
        'CentroCusto': '',
        'CamposCustomizados1': '',
        'CamposCustomizados2': '',
        'CamposCustomizados3': '',
        'CamposCustomizados4': '',
        'CamposCustomizados5': '',
    }
    headers = {SMS_MC_HEADER_NAME: SMS_MC_HEADER_VALUE}
    try:
        resp = requests.post(
            EMAIL_MC_URL,
            params=query_params,
            files={'CorpoHtml': (None, corpo_html, 'text/plain; charset=utf-8')},
            headers=headers,
            timeout=60,
        )
    except requests.RequestException as exc:
        return False, f'Erro de rede ao enviar e-mail: {exc}', None
    try:
        data = resp.json()
    except Exception:
        data = {'raw': (resp.text or '')[:500]}
    if not resp.ok:
        return False, f'Falha HTTP {resp.status_code}', data
    return True, None, data


_TELEFONE_NUMERO_DIGITS_SQL = (
    "REPLACE(REPLACE(REPLACE(REPLACE(REPLACE(REPLACE(IFNULL(numero,''),'+',''),'-',''),' ',''),'(',''),')',''),'.','')"
)


def _sms_digits_only(s):
    """Extrai apenas digitos de `telefone.numero` (str, bytes ou memoryview do MySQL)."""
    if s is None:
        return ''
    if isinstance(s, memoryview):
        s = bytes(s)
    if isinstance(s, (bytes, bytearray)):
        s = bytes(s).decode('latin-1', errors='replace').strip()
    elif not isinstance(s, str):
        s = str(s).strip()
    else:
        s = s.strip()
    return ''.join(ch for ch in s if ch.isdigit())


def _sms_phone_variants(digits):
    """Variantes com/sem prefixo 55 para cruzar com `telefone.numero` formatado."""
    d = _sms_digits_only(digits)
    if not d:
        return []
    out = {d}
    if d.startswith('55') and len(d) > 2:
        out.add(d[2:])
    if not d.startswith('55'):
        out.add('55' + d)
    return list(out)


def _sms_numero_variant_sets_overlap(numero_db, digits_destino):
    va = set(_sms_phone_variants(numero_db))
    vb = set(_sms_phone_variants(digits_destino))
    return bool(va & vb)


def _parse_positive_int(val):
    if val is None or val is False:
        return None
    try:
        n = int(val)
    except (TypeError, ValueError):
        return None
    return n if n > 0 else None


def _primeiro_nome(nome_completo):
    """Primeiro token do nome (ex.: remetente / cliente na integracao MessageCenter)."""
    s = (nome_completo or '').strip()
    if not s:
        return ''
    return s.split()[0]


def _validar_contexto_envio_email_html(cursor, id_pessoa_req, id_email_req, email_str, id_contrato_req):
    """Garante que o e-mail pertence a `id_pessoa` e opcionalmente ao contrato informado.

    Retorna (row_dict|None, erro_str|None). row_dict tem email_dest, id_pessoa, nome_completo, cpf_cnpj.
    """
    id_pessoa = _parse_positive_int(id_pessoa_req)
    id_email = _parse_positive_int(id_email_req)
    if not id_pessoa or not id_email:
        return None, 'id_pessoa e id_email sao obrigatorios.'
    email_norm = (email_str or '').strip().lower()
    if not email_norm:
        return None, 'E-mail invalido.'
    cursor.execute(
        'SELECT e.email AS email_dest, e.id_pessoa, p.nome_completo, p.cpf_cnpj '
        'FROM email e INNER JOIN pessoa p ON p.id = e.id_pessoa '
        'WHERE e.id = %s AND e.id_pessoa = %s',
        (id_email, id_pessoa),
    )
    row = cursor.fetchone()
    if not row:
        return None, 'E-mail nao encontrado para esta pessoa.'
    if (row.get('email_dest') or '').strip().lower() != email_norm:
        return None, 'E-mail nao confere com o cadastro.'
    pid = int(row['id_pessoa'])
    id_contrato_p = _parse_positive_int(id_contrato_req)
    if id_contrato_p is not None:
        cursor.execute(
            'SELECT id FROM contrato WHERE id = %s AND (id_pessoa = %s OR id_avalista = %s)',
            (id_contrato_p, pid, pid),
        )
        if not cursor.fetchone():
            return None, 'Contrato informado invalido para esta pessoa.'
    return row, None


def _resolve_ids_registro_email(cursor, id_pessoa, id_email, id_contrato_req):
    """Resolve id_contrato para INSERT em registro_email (contrato obrigatorio no schema).

    Se `id_contrato` vier no payload, ja foi validado em `_validar_contexto_envio_email_html`.
    Caso contrario, usa o primeiro contrato da pessoa (devedor ou avalista), como no SMS.
    """
    pid = int(id_pessoa)
    eid = int(id_email)
    id_contrato_p = _parse_positive_int(id_contrato_req)
    if id_contrato_p is not None:
        return {'id_pessoa': pid, 'id_email': eid, 'id_contrato': id_contrato_p}, None
    cursor.execute(
        'SELECT id FROM contrato WHERE id_pessoa = %s OR id_avalista = %s ORDER BY id ASC LIMIT 1',
        (pid, pid),
    )
    crow = cursor.fetchone()
    if not crow:
        return None, (
            'Esta pessoa nao possui contrato cadastrado; nao e possivel registrar o envio do e-mail.'
        )
    return {'id_pessoa': pid, 'id_email': eid, 'id_contrato': int(crow['id'])}, None


def _resolve_ids_registro_sms(cursor, payload, digits_destino):
    """Resolve id_telefone, id_pessoa e id_contrato para INSERT em registro_sms.

    Retorna (ids_dict|None, erro_str|None). ids_dict tem chaves id_telefone,
    id_pessoa, id_contrato (todos int).
    """
    id_tel_req = _parse_positive_int(payload.get('id_telefone'))
    id_pessoa_req = _parse_positive_int(payload.get('id_pessoa'))
    id_contrato_req = _parse_positive_int(payload.get('id_contrato'))

    if id_tel_req:
        cursor.execute(
            'SELECT id, id_pessoa, numero FROM telefone WHERE id = %s',
            (id_tel_req,),
        )
        row = cursor.fetchone()
        if not row:
            return None, 'Telefone informado nao encontrado.'
        tid = int(row['id'])
        pid = int(row['id_pessoa'])
        if not _sms_numero_variant_sets_overlap(row.get('numero'), digits_destino):
            return None, 'Numero do SMS nao confere com o telefone informado.'
        if id_pessoa_req is not None and id_pessoa_req != pid:
            return None, 'Pessoa informada nao e a dona do telefone.'
    else:
        variants = _sms_phone_variants(digits_destino)
        if not variants:
            return None, 'Numero invalido.'
        ph = ','.join(['%s'] * len(variants))
        sql = (
            f'SELECT id, id_pessoa, numero FROM telefone '
            f'WHERE {_TELEFONE_NUMERO_DIGITS_SQL} IN ({ph})'
        )
        cursor.execute(sql, tuple(variants))
        rows = cursor.fetchall() or []
        if not rows:
            return None, (
                'Nao foi possivel associar o numero a um telefone cadastrado. '
                'Abra o cadastro e use o botao de SMS ao lado do telefone.'
            )
        if id_pessoa_req is not None:
            rows = [r for r in rows if int(r['id_pessoa']) == id_pessoa_req]
        if not rows:
            return None, 'Nenhum telefone cadastrado corresponde a esta pessoa e ao numero.'
        rows.sort(key=lambda r: int(r['id']), reverse=True)
        row = rows[0]
        tid = int(row['id'])
        pid = int(row['id_pessoa'])

    if id_contrato_req is not None:
        cursor.execute(
            'SELECT id FROM contrato WHERE id = %s AND (id_pessoa = %s OR id_avalista = %s)',
            (id_contrato_req, pid, pid),
        )
        crow = cursor.fetchone()
        if not crow:
            return None, 'Contrato informado invalido para esta pessoa.'
        cid = int(crow['id'])
    else:
        cursor.execute(
            'SELECT id FROM contrato WHERE id_pessoa = %s OR id_avalista = %s ORDER BY id ASC LIMIT 1',
            (pid, pid),
        )
        crow = cursor.fetchone()
        if not crow:
            return None, (
                'Esta pessoa nao possui contrato cadastrado; nao e possivel registrar o envio do SMS.'
            )
        cid = int(crow['id'])

    return {'id_telefone': tid, 'id_pessoa': pid, 'id_contrato': cid}, None


def _auto_envio_contrato_canais(
    conn, fid, row, remetente_nome, stats, erros_amostra, canais, usar_smtp_google=False,
):
    """Envia SMS e/ou e-mail automáticos (mesmo texto) para um contrato no roteiro 0/16/31/61/85.

    `canais`: iterable com 'sms' e/ou 'email'. Evita duplicidade: se já houve SMS ou e-mail
    hoje para o contrato, não envia nada nesta chamada.

    ``usar_smtp_google``: apenas ``True`` no POST da **Importação**; noutros fluxos (ex.: Cobrança)
    o e-mail segue sempre MessageCenter ``EnviarEmailHtml``.
    """
    canais = set(canais) if canais else {'sms', 'email'}
    id_contrato = int(row['id_contrato'])
    id_pessoa = row.get('id_pessoa')
    if id_pessoa is None:
        stats['ignorados_sem_template'] += 1
        return
    id_pessoa = int(id_pessoa)

    dias_atraso = row.get('dias_atraso')
    da_int = None
    if dias_atraso is not None:
        try:
            da_int = int(dias_atraso)
        except (TypeError, ValueError):
            da_int = None

    template_id = _sms_automatizados_template_id_por_dias(da_int)
    if template_id is None:
        stats['ignorados_sem_template'] += 1
        return

    with conn.cursor() as cursor:
        if _contrato_teve_envio_cobranca_hoje(cursor, id_contrato):
            stats['ignorados_ja_enviados_hoje'] += 1
            return

    primeiro = _primeiro_nome(row.get('nome_completo'))
    parcelas_txt = '-'
    if template_id in (2, 3):
        with conn.cursor() as cursor:
            parcelas_txt = _format_parcelas_sms_auto(cursor, id_contrato)

    msg = _mensagem_sms_auto_importacao(
        primeiro,
        row.get('numero_contrato'),
        parcelas_txt,
        template_id,
    )
    if len(msg) > 1600:
        msg = msg[:1600]

    corpo_html = _plain_para_corpo_email_html(msg)
    cliente_pn = primeiro or 'Cliente'
    rem_nome = remetente_nome or ''

    if 'sms' in canais:
        with conn.cursor() as cursor:
            cursor.execute(
                'SELECT id, numero FROM telefone WHERE id_pessoa = %s',
                (id_pessoa,),
            )
            telefones = cursor.fetchall() or []

        if not telefones:
            stats['ignorados_sem_telefone'] += 1
        for tel in telefones:
            digits = _sms_digits_only(tel.get('numero'))
            if not digits:
                stats['falhas'] += 1
                if len(erros_amostra) < 20:
                    erros_amostra.append(
                        {'id_contrato': id_contrato, 'erro': 'Numero invalido.'}
                    )
                continue

            payload_resolve = {
                'id_telefone': int(tel['id']),
                'id_pessoa': id_pessoa,
                'id_contrato': id_contrato,
            }
            with conn.cursor() as cursor:
                ids, err = _resolve_ids_registro_sms(cursor, payload_resolve, digits)
                if err:
                    stats['falhas'] += 1
                    if len(erros_amostra) < 20:
                        erros_amostra.append({'id_contrato': id_contrato, 'erro': err})
                    continue

            params = {'Phone': digits, 'Msgtext': msg}
            headers = {SMS_MC_HEADER_NAME: SMS_MC_HEADER_VALUE}
            try:
                resp = requests.get(
                    SMS_MC_URL,
                    params=params,
                    headers=headers,
                    timeout=30,
                )
            except requests.RequestException as exc:
                stats['falhas'] += 1
                if len(erros_amostra) < 20:
                    erros_amostra.append({
                        'id_contrato': id_contrato,
                        'erro': f'Rede: {exc}',
                    })
                continue

            if not resp.ok:
                stats['falhas'] += 1
                if len(erros_amostra) < 20:
                    try:
                        det = resp.json()
                    except Exception:
                        det = (resp.text or '')[:200]
                    erros_amostra.append({
                        'id_contrato': id_contrato,
                        'erro': f'API SMS HTTP {resp.status_code}',
                        'detalhe': det,
                    })
                continue

            try:
                with conn.cursor() as cursor:
                    cursor.execute(
                        'INSERT INTO registro_sms (id_contrato, id_pessoa, id_telefone, id_funcionario, mensagem) '
                        'VALUES (%s, %s, %s, %s, %s)',
                        (
                            ids['id_contrato'],
                            ids['id_pessoa'],
                            ids['id_telefone'],
                            int(fid),
                            msg,
                        ),
                    )
                    conn.commit()
            except Exception as exc:
                try:
                    conn.rollback()
                except Exception:
                    pass
                stats['falhas'] += 1
                app.logger.warning(
                    'registro_sms insert falhou apos SMS ok (lote auto): %s', exc
                )
                if len(erros_amostra) < 20:
                    erros_amostra.append({
                        'id_contrato': id_contrato,
                        'erro': f'Falha ao registrar SMS: {exc}',
                    })
                continue

            stats['envios_sms'] += 1

    if 'email' in canais:
        with conn.cursor() as cursor:
            cursor.execute(
                'SELECT id, email FROM email WHERE id_pessoa = %s',
                (id_pessoa,),
            )
            emails_rows = cursor.fetchall() or []

        if not emails_rows:
            stats['ignorados_sem_email'] += 1
        for em in emails_rows:
            em_addr = (em.get('email') or '').strip()
            if not _email_basico_valido(em_addr):
                stats['falhas'] += 1
                if len(erros_amostra) < 20:
                    erros_amostra.append({
                        'id_contrato': id_contrato,
                        'erro': 'E-mail invalido no cadastro.',
                    })
                continue

            with conn.cursor() as cursor:
                ids_reg, err_reg = _resolve_ids_registro_email(
                    cursor, id_pessoa, int(em['id']), id_contrato,
                )
                if err_reg:
                    stats['falhas'] += 1
                    if len(erros_amostra) < 20:
                        erros_amostra.append({
                            'id_contrato': id_contrato,
                            'erro': err_reg,
                        })
                    continue

            if usar_smtp_google:
                ok_mc, err_mc, det_mc = _importacao_envio_email_auto(
                    rem_nome, em_addr, cliente_pn, corpo_html, msg,
                )
            else:
                ok_mc, err_mc, det_mc = _messagecenter_post_email_html(
                    rem_nome, em_addr, cliente_pn, corpo_html,
                )
            if not ok_mc:
                stats['falhas'] += 1
                if len(erros_amostra) < 20:
                    erros_amostra.append({
                        'id_contrato': id_contrato,
                        'erro': err_mc or 'Falha ao enviar e-mail.',
                        'detalhe': det_mc,
                    })
                continue

            msg_reg = corpo_html[:1600]
            try:
                with conn.cursor() as cursor:
                    cursor.execute(
                        'INSERT INTO registro_email (id_contrato, id_pessoa, id_email, id_funcionario, mensagem) '
                        'VALUES (%s, %s, %s, %s, %s)',
                        (
                            ids_reg['id_contrato'],
                            ids_reg['id_pessoa'],
                            ids_reg['id_email'],
                            int(fid),
                            msg_reg,
                        ),
                    )
                    conn.commit()
            except Exception as exc:
                try:
                    conn.rollback()
                except Exception:
                    pass
                stats['falhas'] += 1
                app.logger.warning(
                    'registro_email insert falhou apos e-mail ok (lote auto): %s', exc
                )
                if len(erros_amostra) < 20:
                    erros_amostra.append({
                        'id_contrato': id_contrato,
                        'erro': f'Falha ao registrar e-mail: {exc}',
                    })
                continue

            stats['envios_email'] += 1


@app.route('/api/enviar-sms', methods=['POST'])
def api_enviar_sms():
    """Proxy: recebe POST JSON; chama MessageCenter com GET (query Phone, Msgtext).

    Apos sucesso, insere em `registro_sms` (id_funcionario da sessao, mensagem,
    id_pessoa, id_telefone, id_contrato). IDs de contexto podem vir no JSON ou
    ser resolvidos pelo numero / telefone cadastrado.
    """
    fid = session.get('funcionario_id')
    if not fid:
        return jsonify({'error': 'Nao autenticado. Faca login novamente.'}), 401
    payload = request.get_json(silent=True) or {}
    raw = (payload.get('numero') or '').strip()
    digits = ''.join(ch for ch in raw if ch.isdigit())
    if not digits:
        return jsonify({'error': 'Numero invalido.'}), 400

    msg = (payload.get('mensagem') or payload.get('Msgtxt') or '').strip()
    if not msg:
        return jsonify({'error': 'Mensagem vazia.'}), 400
    if len(msg) > 1600:
        return jsonify({'error': 'Mensagem muito longa (max. 1600 caracteres).'}), 400

    conn = _get_db()
    try:
        with conn.cursor() as cursor:
            ids, err = _resolve_ids_registro_sms(cursor, payload, digits)
            if err:
                return jsonify({'error': err}), 400
    finally:
        conn.close()

    params = {'Phone': digits, 'Msgtext': msg}
    headers = {SMS_MC_HEADER_NAME: SMS_MC_HEADER_VALUE}
    try:
        resp = requests.get(
            url=SMS_MC_URL,
            params=params,
            headers=headers,
            timeout=30,
        )
    except requests.RequestException as exc:
        return jsonify({'error': f'Erro de rede ao enviar SMS: {exc}'}), 502
    try:
        data = resp.json()
    except Exception:
        data = {'raw': (resp.text or '')[:500]}
    if not resp.ok:
        return jsonify({
            'error': 'Falha ao enviar SMS.',
            'status': resp.status_code,
            'detalhe': data,
        }), 502

    conn = _get_db()
    try:
        with conn.cursor() as cursor:
            cursor.execute(
                'INSERT INTO registro_sms (id_contrato, id_pessoa, id_telefone, id_funcionario, mensagem) '
                'VALUES (%s, %s, %s, %s, %s)',
                (
                    ids['id_contrato'],
                    ids['id_pessoa'],
                    ids['id_telefone'],
                    int(fid),
                    msg,
                ),
            )
            conn.commit()
    except Exception as exc:
        try:
            conn.rollback()
        except Exception:
            pass
        app.logger.warning('registro_sms insert falhou apos SMS ok: %s', exc)
    finally:
        conn.close()

    return jsonify({'ok': True, 'sms': data})


@app.route('/api/enviar-email-html', methods=['POST'])
def api_enviar_email_html():
    """Proxy MessageCenter EnviarEmailHtml: query string + multipart CorpoHtml, header apikey igual ao SMS.

    Apos sucesso, insere em `registro_email` (id_funcionario, mensagem truncada a 1600 chars,
    id_pessoa, id_email, id_contrato), no mesmo espirito de `registro_sms`.
    """
    fid = session.get('funcionario_id')
    if not fid:
        return jsonify({'error': 'Nao autenticado. Faca login novamente.'}), 401
    forbidden_email = _cobranca_json_sem_permissao_envio_email()
    if forbidden_email:
        return forbidden_email
    payload = request.get_json(silent=True) or {}
    corpo = (payload.get('corpo_html') or '').strip()
    if not corpo:
        return jsonify({'error': 'Corpo do e-mail vazio.'}), 400
    if len(corpo) > 200000:
        return jsonify({'error': 'Corpo do e-mail muito longo.'}), 400

    conn = _get_db()
    ids_reg = None
    try:
        with conn.cursor() as cursor:
            row_ctx, err = _validar_contexto_envio_email_html(
                cursor,
                payload.get('id_pessoa'),
                payload.get('id_email'),
                payload.get('email'),
                payload.get('id_contrato'),
            )
            if err:
                return jsonify({'error': err}), 400
            ids_reg, err_reg = _resolve_ids_registro_email(
                cursor,
                row_ctx['id_pessoa'],
                payload.get('id_email'),
                payload.get('id_contrato'),
            )
            if err_reg:
                return jsonify({'error': err_reg}), 400
    finally:
        conn.close()

    remetente_nome = _primeiro_nome(session.get('funcionario_nome'))
    destinatario = (row_ctx.get('email_dest') or '').strip()
    query_params = {
        'Destinatario': destinatario,
        'RemetenteNome': remetente_nome,
        'RemetenteEmail': EMAIL_MC_REMETENTE,
        'Assunto': 'Cobranca',
        'Identificador': '',
        'ClienteNome': _primeiro_nome(row_ctx.get('nome_completo')),
        'ClienteDocumento': '',
        'CentroCusto': '',
        'CamposCustomizados1': '',
        'CamposCustomizados2': '',
        'CamposCustomizados3': '',
        'CamposCustomizados4': '',
        'CamposCustomizados5': '',
    }
    headers = {SMS_MC_HEADER_NAME: SMS_MC_HEADER_VALUE}
    try:
        resp = requests.post(
            EMAIL_MC_URL,
            params=query_params,
            files={'CorpoHtml': (None, corpo, 'text/plain; charset=utf-8')},
            headers=headers,
            timeout=60,
        )
    except requests.RequestException as exc:
        return jsonify({'error': f'Erro de rede ao enviar e-mail: {exc}'}), 502
    try:
        data = resp.json()
    except Exception:
        data = {'raw': (resp.text or '')[:500]}
    if not resp.ok:
        return jsonify({
            'error': 'Falha ao enviar e-mail.',
            'status': resp.status_code,
            'detalhe': data,
        }), 502

    msg_reg = corpo[:1600]
    conn = _get_db()
    try:
        with conn.cursor() as cursor:
            cursor.execute(
                'INSERT INTO registro_email (id_contrato, id_pessoa, id_email, id_funcionario, mensagem) '
                'VALUES (%s, %s, %s, %s, %s)',
                (
                    ids_reg['id_contrato'],
                    ids_reg['id_pessoa'],
                    ids_reg['id_email'],
                    int(fid),
                    msg_reg,
                ),
            )
            conn.commit()
    except Exception as exc:
        try:
            conn.rollback()
        except Exception:
            pass
        app.logger.warning('registro_email insert falhou apos e-mail ok: %s', exc)
    finally:
        conn.close()

    return jsonify({'ok': True, 'email': data})


@app.route('/api/pessoa/<int:pessoa_id>', methods=['GET'])
def api_pessoa_detalhe(pessoa_id):
    conn = _get_db()
    cursor = conn.cursor()

    cursor.execute("SELECT * FROM pessoa WHERE id = %s", (pessoa_id,))
    pessoa = _clean_row(cursor.fetchone())
    if not pessoa:
        cursor.close()
        conn.close()
        return jsonify({'error': 'Pessoa nao encontrada'}), 404

    cursor.execute("SELECT * FROM endereco WHERE id_pessoa = %s", (pessoa_id,))
    enderecos = _clean_rows(cursor.fetchall())

    cursor.execute("SELECT * FROM telefone WHERE id_pessoa = %s", (pessoa_id,))
    telefones = _clean_rows(cursor.fetchall())

    cursor.execute("SELECT * FROM email WHERE id_pessoa = %s", (pessoa_id,))
    emails = _clean_rows(cursor.fetchall())

    cursor.execute(
        "SELECT c.*, p_dev.nome_completo AS nome_devedor, p_aval.nome_completo AS nome_avalista "
        "FROM contrato c "
        "LEFT JOIN pessoa p_dev ON c.id_pessoa = p_dev.id "
        "LEFT JOIN pessoa p_aval ON c.id_avalista = p_aval.id "
        "WHERE c.id_pessoa = %s OR c.id_avalista = %s",
        (pessoa_id, pessoa_id),
    )
    contratos = _clean_rows(cursor.fetchall())

    cursor.close()
    conn.close()
    return jsonify({
        'pessoa': pessoa,
        'enderecos': enderecos,
        'telefones': telefones,
        'emails': emails,
        'contratos': contratos,
    })


def _negativacao_chave_serasa_contrato_nrs16(grupo, cota):
    """Mesma regra que ``montar_linha_detalhe_inclusao`` / ``serasa_conv_txt`` para grupo+cota."""
    g = re.sub(r'\D', '', str(grupo or ''))
    c = re.sub(r'\D', '', str(cota or ''))
    core = (g + c)[:16]
    return core.zfill(16) if core else '0' * 16


def _negativacao_docs_variantes_pefin_15(doc15: str):
    """Variantes de documento a comparar com ``pessoa.cpf_cnpj`` normalizado (15 digitos do TXT)."""
    d = re.sub(r'\D', '', str(doc15 or ''))[-15:].zfill(15)
    out = []
    if d:
        out.append(d)
        s = d.lstrip('0') or '0'
        out.append(s)
        if len(s) >= 11:
            out.append(s[-11:])
        if len(s) >= 14:
            out.append(s[-14:])
    seen = set()
    uniq = []
    for x in out:
        if x not in seen:
            seen.add(x)
            uniq.append(x)
    return uniq


def _negativacao_resolve_contrato_pefin_retorno(cursor, contrato_nrs16: str, doc15: str):
    """Resolve ``contrato.id`` pela chave 16 digitos do TXT e, se necessario, por CPF/CNPJ do devedor."""
    k = re.sub(r'\D', '', str(contrato_nrs16 or ''))[-16:].zfill(16)
    cursor.execute(
        """
        SELECT id, grupo, cota FROM contrato
        WHERE LPAD(SUBSTRING(CONCAT(TRIM(COALESCE(grupo, '')), TRIM(COALESCE(cota, ''))), 1, 16), 16, '0') = %s
        LIMIT 8
        """,
        (k,),
    )
    key_rows = _clean_rows(cursor.fetchall())
    if len(key_rows) == 1:
        return int(key_rows[0]['id'])
    variants = _negativacao_docs_variantes_pefin_15(doc15)
    if not variants:
        return None
    ph = ','.join(['%s'] * len(variants))
    norm_doc = (
        "REPLACE(REPLACE(REPLACE(REPLACE(COALESCE(p.cpf_cnpj, ''), '.', ''), '-', ''), '/', ''), ' ', '')"
    )
    cursor.execute(
        f"""
        SELECT c.id, c.grupo, c.cota FROM contrato c
        INNER JOIN pessoa p ON p.id = c.id_pessoa
        WHERE {norm_doc} IN ({ph})
        LIMIT 40
        """,
        variants,
    )
    doc_rows = _clean_rows(cursor.fetchall())
    narrowed = [
        r for r in doc_rows
        if _negativacao_chave_serasa_contrato_nrs16(r.get('grupo'), r.get('cota')) == k
    ]
    if len(key_rows) > 1:
        allowed = {int(r['id']) for r in key_rows}
        narrowed = [r for r in narrowed if int(r['id']) in allowed]
    if len(narrowed) == 1:
        return int(narrowed[0]['id'])
    return None


def _negativacao_data_evento_de_referencia_pefin(yyyymmdd: str):
    """Data do evento a partir da data de referencia da linha PEFIN (AAAAMMDD)."""
    s = (yyyymmdd or '').strip()
    if len(s) == 8 and s.isdigit():
        try:
            d = datetime.date(int(s[:4]), int(s[4:6]), int(s[6:8]))
            return datetime.datetime.combine(d, datetime.time.min)
        except ValueError:
            pass
    return datetime.datetime.now()


def _negativacao_montar_detalhe_pefin_retorno(
    nome_arquivo: str,
    parsed: dict,
    *,
    pefin_desc,
):
    """Texto para coluna ``detalhe`` (importacao de retorno SERASA)."""
    seq = parsed.get('sequencial_linha_arquivo') or '?'
    cods = parsed.get('codigos_erro') or []
    partes = [
        f'Retorno SERASA PEFIN ficheiro={nome_arquivo!r} seq_linha={seq}',
        f"contrato_nrs_16={parsed.get('contrato_nrs_16')}",
        f"documento_15={parsed.get('documento_15')}",
        f"valor_centavos={parsed.get('valor_centavos')}",
    ]
    if not cods:
        partes.append('Nenhum codigo de erro PEFIN na zona de retorno (531-593).')
    else:
        linhas_erro = []
        for cod in cods:
            desc = pefin_desc(cod) if callable(pefin_desc) else None
            if desc:
                linhas_erro.append(f'{cod} — {desc}')
            else:
                linhas_erro.append(f'{cod} (descricao nao encontrada na tabela local)')
        partes.append('Erros: ' + '; '.join(linhas_erro))
    return ' | '.join(partes)


def _negativacao_parcela_display_key(row):
    """Chave única por parcela dentro do contrato para exibição.

    Inclui `id_contrato` para listagens globais. Prioriza `numero_parcela`.
    """
    try:
        cid = int(row.get('id_contrato') or 0)
    except (TypeError, ValueError):
        cid = 0
    try:
        n = row.get('numero_parcela')
        if n is not None and str(n).strip() != '':
            return (cid, 'num', int(n))
    except (TypeError, ValueError):
        pass
    try:
        if row.get('id_parcela') is not None:
            return (cid, 'id', int(row['id_parcela']))
    except (TypeError, ValueError):
        pass
    return (cid, 'row', int(row.get('id') or 0))


# Status interno: positivação já registrada no histórico; falta o operador disparar o envio ao Serasa.
_NEG_STATUS_AGUARDANDO_POS_SERASA = 'aguardando_positivacao_serasa'
_NEG_STATUS_SERASA_ELEGIVEL_NEGATIVAR = frozenset({'registrado_tracker', 'falhou'})


def _negativacao_row_serasa_flags(row):
    """Marca elegibilidade para envio manual ao Serasa (camadas separadas do registro interno)."""
    st = (row.get('status') or '').strip().lower()
    row['serasa_elegivel_negativar'] = st in _NEG_STATUS_SERASA_ELEGIVEL_NEGATIVAR
    row['serasa_elegivel_positivar'] = st == _NEG_STATUS_AGUARDANDO_POS_SERASA


def _negativacao_serasa_status_por_parcela(cursor, ids):
    """Mapa id_parcela -> status em ``negativacao`` (lower-case)."""
    if not ids:
        return {}
    ph = ','.join(['%s'] * len(ids))
    cursor.execute(
        f'SELECT id_parcela, status FROM negativacao WHERE id_parcela IN ({ph})',
        ids,
    )
    out = {}
    for r in cursor.fetchall() or []:
        try:
            ip = int(r.get('id_parcela'))
        except (TypeError, ValueError):
            continue
        out[ip] = (r.get('status') or '').strip().lower()
    return out


def _negativacao_serasa_erro_elegibilidade_txt(tipo, ids, por_parcela):
    """Mesma regra que o mock ``/api/negativacao/positivar-lote-serasa`` (sem atualizar BD)."""
    if tipo == 'negativar':
        if len(por_parcela) != len(ids):
            return 'Uma ou mais parcelas nao possuem negativacao ativa no cadastro.'
        for ip in ids:
            st = por_parcela.get(ip, '')
            if st not in _NEG_STATUS_SERASA_ELEGIVEL_NEGATIVAR:
                return (
                    'Uma ou mais parcelas nao estao elegiveis '
                    '(negativar: apenas tracker ou falha de envio).'
                )
    else:
        # Positivar: a listagem (Carteira) pode incluir eventos de historico sem linha activa em
        # `negativacao` (ex.: positivado_tracker apaga o registo). A UI filtra com rowCarteiraBulkPositivar.
        pass
    return None


def _negativacao_date_only_sql(val):
    if val is None:
        return None
    if isinstance(val, datetime.datetime):
        return val.date()
    return val


def _negativacao_fetch_serasa_inclusao_payloads(cursor, ids_ordered, *, inner_negativacao=True):
    """Monta lista de dicts para ``montar_linha_detalhe_inclusao`` na ordem de ``ids_ordered``.

    ``inner_negativacao`` (negativar): exige linha activa em ``negativacao``. O valor no TXT (centavos)
    e a soma de ``COALESCE(valor_total, valor_nominal)`` de todas as parcelas do contrato com
    ``status = 'cobranca'`` (uma linha por parcela seleccionada, mesmo montante por contrato).

    Para positivar (TXT exclusao), usa ``LEFT JOIN`` em ``negativacao``, data de referencia a partir
    do historico quando o registo activo ja nao existe, e valor por parcela (parcela alvo).
    """
    if not ids_ordered:
        return []
    ph = ','.join(['%s'] * len(ids_ordered))
    field_ord = ','.join(['%s'] * len(ids_ordered))
    if inner_negativacao:
        neg_join = 'INNER JOIN negativacao n ON n.id_parcela = p.id'
        data_neg_sql = 'n.data_negativacao AS data_negativacao'
        valor_sql = (
            "(SELECT CAST(ROUND(COALESCE(SUM(COALESCE(p2.valor_total, p2.valor_nominal, 0)), 0) * 100) AS SIGNED) "
            "FROM parcela p2 WHERE p2.id_contrato = c.id AND p2.status = 'cobranca')"
        )
    else:
        neg_join = 'LEFT JOIN negativacao n ON n.id_parcela = p.id'
        data_neg_sql = (
            'COALESCE('
            'n.data_negativacao,'
            '(SELECT MAX(nh.data_evento) FROM negativacao_historico nh '
            'WHERE nh.id_parcela = p.id '
            "AND nh.tipo_evento IN ('negativado_manual','negativado_tracker'))"
            ') AS data_negativacao'
        )
        valor_sql = 'CAST(ROUND(COALESCE(p.valor_total, p.valor_nominal, 0) * 100) AS SIGNED)'
    cursor.execute(
        f"""
        SELECT
          p.id AS id_parcela,
          p.vencimento,
          {data_neg_sql},
          dev.cpf_cnpj,
          dev.nome_completo,
          dev.data_nascimento,
          e.logradouro,
          e.bairro,
          e.cidade,
          e.estado AS uf,
          e.cep,
          c.grupo,
          c.cota,
          {valor_sql} AS valor_centavos,
          emp.apelido AS empresa_apelido,
          cred.cpf_cnpj AS credor_documento,
          COALESCE(
            NULLIF(TRIM(cred.nome_completo), ''),
            'GMAC ADMINISTRADORA DE CONSORCIO LTDA'
          ) AS nome_credor
        FROM parcela p
        INNER JOIN contrato c ON c.id = p.id_contrato
        LEFT JOIN pessoa dev ON dev.id = c.id_pessoa
        {neg_join}
        LEFT JOIN endereco e ON e.id = (
          SELECT x.id FROM endereco x
          WHERE x.id_pessoa = dev.id
          ORDER BY CASE x.tipo
            WHEN 'principal' THEN 1
            WHEN 'secundario' THEN 2
            ELSE 9
          END, x.id ASC
          LIMIT 1
        )
        LEFT JOIN empresa emp ON emp.id = c.id_empresa
        LEFT JOIN pessoa cred ON cred.id = emp.id_pessoa
        WHERE p.id IN ({ph})
        ORDER BY FIELD(p.id, {field_ord})
        """,
        list(ids_ordered) + list(ids_ordered),
    )
    rows = _clean_rows(cursor.fetchall())
    payloads = []
    for r in rows:
        ref = _negativacao_date_only_sql(r.get('data_negativacao'))
        if ref is None:
            ref = datetime.date.today()
        venc = _negativacao_date_only_sql(r.get('vencimento'))
        apelido_emp = (r.get('empresa_apelido') or '').strip().upper()
        codigo_credor = '5015' if 'MAPFRE' in apelido_emp else '5016'
        doc_cred = re.sub(r'\D', '', str(r.get('credor_documento') or ''))
        cnpj_credor = doc_cred[:14] if len(doc_cred) >= 14 else None
        try:
            valor_centavos = int(r.get('valor_centavos') or 0)
        except (TypeError, ValueError):
            valor_centavos = 0
        if valor_centavos < 0:
            valor_centavos = 0
        payloads.append({
            'data_ref': ref,
            'data_vencimento': venc or ref,
            'cpf_cnpj': r.get('cpf_cnpj'),
            'nome': r.get('nome_completo'),
            'data_nasc': _negativacao_date_only_sql(r.get('data_nascimento')),
            'logradouro': r.get('logradouro'),
            'bairro': r.get('bairro'),
            'cidade': r.get('cidade'),
            'uf': r.get('uf'),
            'cep': r.get('cep'),
            'grupo': r.get('grupo'),
            'cota': r.get('cota'),
            'nome_credor': r.get('nome_credor'),
            'codigo_credor': codigo_credor,
            'cnpj_credor': cnpj_credor,
            'valor_centavos': valor_centavos,
        })
    return payloads


def _dedupe_negativacao_ativas_exibicao(rows):
    """Uma entrada ativa por parcela lógica (mantém o registro mais recente)."""
    best = {}
    for r in rows:
        k = _negativacao_parcela_display_key(r)
        rid = int(r.get('id') or 0)
        if k not in best or rid > int(best[k].get('id') or 0):
            best[k] = r
    out = list(best.values())
    out.sort(key=lambda x: (-int(x.get('id') or 0),))
    return out


def _negativacao_historico_tipo_padrao_neg_rem(tipo_evento):
    """Tipos sujeitos à deduplicação min/max na timeline do contrato."""
    t = (tipo_evento or '').strip().lower()
    return t.startswith('negativado') or t.startswith('removido') or t == 'positivado_tracker'


def _dedupe_negativacao_historico_exibicao(rows):
    """Por parcela: no máximo uma negativação (primeira data) e uma positivação (última data).

    Observações (`observacao`) são todas preservadas. Outros tipos (ex.: retorno PEFIN,
    ``observacao`` já tratada) são preservados sem deduplicação por parcela.
    """
    obs = [r for r in rows if (r.get('tipo_evento') or '') == 'observacao']
    outros = [r for r in rows if (r.get('tipo_evento') or '') != 'observacao']
    padrao = [r for r in outros if _negativacao_historico_tipo_padrao_neg_rem(r.get('tipo_evento'))]
    especiais = [r for r in outros if not _negativacao_historico_tipo_padrao_neg_rem(r.get('tipo_evento'))]

    grupos = {}
    for r in padrao:
        k = _negativacao_parcela_display_key(r)
        grupos.setdefault(k, []).append(r)

    out = []
    for lst in grupos.values():
        neg = [x for x in lst if str(x.get('tipo_evento') or '').startswith('negativado')]
        rem = [
            x
            for x in lst
            if str(x.get('tipo_evento') or '').startswith('removido')
            or str(x.get('tipo_evento') or '').strip().lower() == 'positivado_tracker'
        ]
        if neg:
            out.append(
                min(neg, key=lambda x: (str(x.get('data_evento') or ''), int(x.get('id') or 0)))
            )
        if rem:
            out.append(
                max(rem, key=lambda x: (str(x.get('data_evento') or ''), int(x.get('id') or 0)))
            )
    out.extend(obs)
    out.extend(especiais)
    out.sort(key=lambda x: (str(x.get('data_evento') or ''), int(x.get('id') or 0)))
    return out


def _ocorrencias_sem_echo_positivacao_negativacao(rows):
    """Remove da timeline genérica ocorrências que só repetem texto de positivação (já em negativacao_historico)."""
    out = []
    for r in rows or []:
        st = (r.get('status') or '').strip().lower()
        desc = (r.get('descricao') or '').lower()
        if st in (_STATUS_BD_COBRANCA, 'aberto') and 'antes negativada' in desc:
            continue
        out.append(r)
    return out


@app.route('/api/contrato/<int:contrato_id>')
def api_contrato_detalhe(contrato_id):
    conn = _get_db()
    cursor = conn.cursor()

    def _safe_all(sql, params=()):
        """Executa SELECT isoladamente; retorna [] em caso de erro (tabela
        ausente, coluna faltando, etc) sem quebrar a resposta inteira."""
        try:
            cursor.execute(sql, params)
            return _clean_rows(cursor.fetchall())
        except Exception as exc:
            app.logger.warning('api_contrato_detalhe: falha em %r (%s)', sql, exc)
            return []

    def _safe_one(sql, params=()):
        try:
            cursor.execute(sql, params)
            return _clean_row(cursor.fetchone())
        except Exception as exc:
            app.logger.warning('api_contrato_detalhe: falha em %r (%s)', sql, exc)
            return None

    try:
        contrato = _safe_one("SELECT * FROM contrato WHERE id = %s", (contrato_id,))
        if not contrato:
            return jsonify({'error': 'Contrato nao encontrado'}), 404

        devedor = None
        devedor_end = devedor_tel = devedor_email = []
        if contrato.get('id_pessoa'):
            pid = contrato['id_pessoa']
            devedor = _safe_one("SELECT * FROM pessoa WHERE id = %s", (pid,))
            devedor_end = _safe_all("SELECT * FROM endereco WHERE id_pessoa = %s", (pid,))
            devedor_tel = _safe_all("SELECT * FROM telefone WHERE id_pessoa = %s", (pid,))
            devedor_email = _safe_all("SELECT * FROM email WHERE id_pessoa = %s", (pid,))

        avalista = None
        avalista_end = avalista_tel = avalista_email = []
        if contrato.get('id_avalista'):
            aid = contrato['id_avalista']
            avalista = _safe_one("SELECT * FROM pessoa WHERE id = %s", (aid,))
            avalista_end = _safe_all("SELECT * FROM endereco WHERE id_pessoa = %s", (aid,))
            avalista_tel = _safe_all("SELECT * FROM telefone WHERE id_pessoa = %s", (aid,))
            avalista_email = _safe_all("SELECT * FROM email WHERE id_pessoa = %s", (aid,))

        parcelas = _safe_all(
            "SELECT * FROM parcela WHERE id_contrato = %s ORDER BY numero_parcela",
            (contrato_id,),
        )
        ocorrencias = _ocorrencias_sem_echo_positivacao_negativacao(
            _safe_all(
                "SELECT * FROM ocorrencia WHERE id_contrato = %s ORDER BY data_arquivo ASC, id ASC",
                (contrato_id,),
            )
        )
        _ensure_tramitacao_fluxo_columns(cursor)
        _tramit_ord = _tramitacao_list_order_sql({n.lower() for n in _tramitacao_column_names(cursor)})
        tramitacoes = _safe_all(
            "SELECT t.*, f.nome AS funcionario_nome "
            "FROM tramitacao t "
            "LEFT JOIN funcionario f ON t.id_funcionario = f.id "
            "WHERE t.id_contrato = %s ORDER BY "
            + _tramit_ord,
            (contrato_id,),
        )

        registros_sms_email = _safe_all(
            "SELECT x.canal, x.id, x.created_at, x.mensagem, x.id_funcionario, f.nome AS funcionario_nome "
            "FROM ( "
            "  SELECT 'sms' AS canal, id, created_at, mensagem, id_funcionario "
            "  FROM registro_sms WHERE id_contrato = %s "
            "  UNION ALL "
            "  SELECT 'email' AS canal, id, created_at, mensagem, id_funcionario "
            "  FROM registro_email WHERE id_contrato = %s "
            ") x "
            "LEFT JOIN funcionario f ON x.id_funcionario = f.id "
            "ORDER BY x.created_at ASC, x.id ASC",
            (contrato_id, contrato_id),
        )

        try:
            bens = _fetch_bens_para_contrato(cursor, contrato)
        except Exception as exc:
            app.logger.warning('api_contrato_detalhe: falha em bens (%s)', exc)
            bens = []

        negativacao_ativas = []
        negativacao_historico = []
        try:
            _ensure_negativacao_table(cursor)
            _ensure_negativacao_historico_table(cursor)
            negativacao_ativas = _dedupe_negativacao_ativas_exibicao(
                _safe_all(
                    "SELECT n.*, f.nome AS funcionario_nome "
                    "FROM negativacao n "
                    "LEFT JOIN funcionario f ON n.id_funcionario = f.id "
                    "WHERE n.id_contrato = %s ORDER BY n.data_negativacao DESC, n.id DESC",
                    (contrato_id,),
                )
            )
            negativacao_historico = _dedupe_negativacao_historico_exibicao(
                _safe_all(
                    "SELECT h.*, f.nome AS funcionario_nome "
                    "FROM negativacao_historico h "
                    "LEFT JOIN funcionario f ON h.id_funcionario = f.id "
                    "WHERE h.id_contrato = %s ORDER BY h.data_evento ASC, h.id ASC",
                    (contrato_id,),
                )
            )
        except Exception as exc:
            app.logger.warning('api_contrato_detalhe: negativacao (%s)', exc)

        return jsonify({
            'contrato': contrato,
            'devedor': devedor,
            'devedor_enderecos': devedor_end,
            'devedor_telefones': devedor_tel,
            'devedor_emails': devedor_email,
            'avalista': avalista,
            'avalista_enderecos': avalista_end,
            'avalista_telefones': avalista_tel,
            'avalista_emails': avalista_email,
            'parcelas': parcelas,
            'ocorrencias': ocorrencias,
            'tramitacoes': tramitacoes,
            'registros_sms_email': registros_sms_email,
            'bens': bens,
            'negativacao_ativas': negativacao_ativas,
            'negativacao_historico': negativacao_historico,
        })
    except Exception as exc:
        app.logger.exception('api_contrato_detalhe: erro inesperado')
        return jsonify({'error': 'Erro ao carregar contrato: ' + str(exc)}), 500
    finally:
        try: cursor.close()
        except Exception: pass
        try: conn.close()
        except Exception: pass


TRAMITACAO_TIPOS = ('ligacao', 'whatsapp', 'email')
TRAMITACAO_CPCS = ('sim', 'nao', 'parente', 'amigo', 'avalista')

FLUXO_MOTIVOS_NA = (
    'caixa_postal',
    'chama_nao_atende',
    'chamada_incompleta',
    'ligacao_caiu',
    'numero_inexistente',
)
FLUXO_CPC_QUEM = ('nao_consorciado_conhece', 'nao_consorciado_nao_conhece')
FLUXO_CPC_QUAL = ('consorciado', 'terceiro', 'avalista')
FLUXO_STATUS_FINAL = (
    'alega_pagamento',
    'agendamento',
    'acordo_firmado',
    'sem_condicoes',
    'sem_interesse',
    'nao_confirma_dados',
    'atende_desliga',
    'ligacao_muda',
)

FLUXO_MOTIVO_LABEL = {
    'caixa_postal': 'Caixa Postal',
    'chama_nao_atende': 'Chama e Não atende',
    'chamada_incompleta': 'Chamada Não Completada',
    'ligacao_caiu': 'Ligação Caiu',
    'numero_inexistente': 'Número não existe',
}
FLUXO_CPC_QUEM_LABEL = {
    'nao_consorciado_conhece': 'Não é o consorciado, o conhece, mas não é o responsável',
    'nao_consorciado_nao_conhece': 'Não é consorciado e não conhece',
}
FLUXO_STATUS_LABEL = {
    'alega_pagamento': 'Alega pagamento',
    'agendamento': 'Agendamento',
    'acordo_firmado': 'Acordo Firmado',
    'sem_condicoes': 'Sem condições financeiras',
    'sem_interesse': 'Sem interesse no pagamento',
    'nao_confirma_dados': 'Não confirma os dados',
    'atende_desliga': 'Atende e desliga',
    'ligacao_muda': 'Ligação ficou muda',
}

# Valores exatos da coluna legada `tramitacao.status` (ENUM), alinhados a Banco/criar_banco.py
FLUXO_STATUS_FINAL_DB = {
    'alega_pagamento': 'alega pagamento',
    'agendamento': 'agendamento',
    'acordo_firmado': 'acordo firmado',
    'sem_condicoes': 'sem condições financeiras',
    'sem_interesse': 'sem interesse no pagamento',
    'nao_confirma_dados': 'não confirma dados',
    'atende_desliga': 'atende e desliga',
    'ligacao_muda': 'ligação ficou muda',
}
FLUXO_MOTIVO_NA_DB = {
    'caixa_postal': 'caixa postal / secretária eletrônica',
    'chama_nao_atende': 'chama e não atende',
    'chamada_incompleta': 'chamada não completada',
    'ligacao_caiu': 'ligação caiu',
    'numero_inexistente': 'numero não existe',
}


def _ensure_tramitacao_fluxo_columns(cursor):
    """Garante colunas esperadas pelo app: `id` (PK), wizard (`fluxo_json`, `status_tramitacao`).

    Idempotente por execução de SHOW COLUMNS (sem cache global), para bases legadas sem `id`.
    """
    try:
        cursor.execute("SHOW COLUMNS FROM tramitacao")
        rows = cursor.fetchall()
        existing = set()
        for r in rows:
            if isinstance(r, dict):
                fn = r.get('Field') or r.get('field')
                if fn:
                    existing.add(str(fn).lower())
            elif r:
                existing.add(str(r[0]).lower())
        if 'id' not in existing:
            cursor.execute(
                "ALTER TABLE tramitacao ADD COLUMN id BIGINT NOT NULL AUTO_INCREMENT PRIMARY KEY FIRST"
            )
            existing.add('id')
        if 'fluxo_json' not in existing:
            cursor.execute(
                "ALTER TABLE tramitacao ADD COLUMN fluxo_json LONGTEXT NULL "
                "COMMENT 'Wizard tramitacao (JSON)'"
            )
        if 'status_tramitacao' not in existing:
            cursor.execute(
                "ALTER TABLE tramitacao ADD COLUMN status_tramitacao VARCHAR(96) NULL "
                "COMMENT 'Rotulo do resultado'"
            )
    except Exception as exc:
        app.logger.warning('_ensure_tramitacao_fluxo_columns: %s', exc)


def _tramitacao_list_order_sql(names_lower):
    """ORDER BY compatível: GM usa `data`; legado costuma ter `created_at` e, após migração, `id`."""
    nl = names_lower if isinstance(names_lower, set) else {str(x).lower() for x in names_lower}
    parts = []
    if 'data' in nl:
        parts.append('t.data DESC')
    if 'created_at' in nl:
        parts.append('t.created_at DESC')
    if 'id' in nl:
        parts.append('t.id DESC')
    if parts:
        return ', '.join(parts)
    return 't.id_contrato DESC'


def _tramitacao_schema(cursor):
    """Nomes na ordem do SHOW COLUMNS, tipos em minúsculas, mapa lower -> nome real."""
    cursor.execute('SHOW COLUMNS FROM tramitacao')
    rows = cursor.fetchall()
    names = []
    types = {}
    tcols = {}
    for r in rows:
        if isinstance(r, dict):
            fn = r.get('Field') or r.get('field')
            tp = (r.get('Type') or r.get('type') or '').lower()
        else:
            fn = r[0] if r else None
            tp = (r[1] or '').lower() if r and len(r) > 1 else ''
        if fn:
            names.append(fn)
            low = fn.lower()
            tcols[low] = fn
            types[low] = tp
    return names, types, tcols


def _tramitacao_column_names(cursor):
    """Retorna lista com nomes das colunas em `tramitacao` (ordem SHOW COLUMNS)."""
    names, _, _ = _tramitacao_schema(cursor)
    return names


def _tramitacao_truncate_descricao(desc, type_low):
    """Encolhe texto para colunas VARCHAR(N) legadas."""
    if desc is None:
        return None
    s = str(desc)
    if 'varchar' not in type_low:
        return s
    try:
        i = type_low.index('varchar(')
        j = type_low.index(')', i)
        n = int(type_low[i + 8 : j])
        return s[:n] if n > 0 else s
    except (ValueError, IndexError):
        return s[:255]


def _fluxo_legacy_status_db(payload):
    """Valor ENUM para coluna legada `status`."""
    if payload.get('atendido') == 'nao':
        mk = payload.get('motivo_nao_atendido')
        return FLUXO_MOTIVO_NA_DB.get(mk, 'chama e não atende')
    if payload.get('modo_indefinido'):
        return 'atende e desliga'
    if payload.get('cpc_correto') == 'nao':
        qm = payload.get('cpc_quem')
        if qm == 'nao_consorciado_conhece':
            return 'não é consorciado, conhece, mas não é responsável'
        return 'não é o consorciado e não conhece'
    sf = payload.get('status_final')
    return FLUXO_STATUS_FINAL_DB.get(sf, 'sem interesse no pagamento')


def _fluxo_legacy_contato(payload):
    """Valor ENUM para coluna legada `contato`."""
    if payload.get('atendido') == 'nao':
        return 'indefinido'
    if payload.get('modo_indefinido'):
        return 'consorciado'
    if payload.get('cpc_correto') == 'nao':
        qm = payload.get('cpc_quem')
        if qm == 'nao_consorciado_conhece':
            return 'consorciado'
        return 'terceiro'
    cq = payload.get('cpc_qual')
    return {'consorciado': 'consorciado', 'terceiro': 'terceiro', 'avalista': 'avalista'}.get(
        cq, 'consorciado'
    )


def _fluxo_legacy_cpc_bit(payload):
    """1 = pessoa certa (fluxo CPC sim ou indefinido tratado como atendido com consorciado)."""
    if payload.get('atendido') == 'nao':
        return 0
    if payload.get('modo_indefinido'):
        return 1
    return 1 if payload.get('cpc_correto') == 'sim' else 0


def _fluxo_legacy_exito_bit(payload):
    if payload.get('atendido') != 'sim':
        return 0
    sf = payload.get('status_final')
    return 1 if sf in ('acordo_firmado', 'agendamento', 'alega_pagamento') else 0


def _fluxo_legacy_classificacao(payload):
    if payload.get('atendido') != 'sim':
        return 'indefinido'
    if payload.get('modo_indefinido'):
        return 'indefinido'
    sf = payload.get('status_final')
    if sf == 'acordo_firmado':
        return 'excelente'
    if sf in ('sem_interesse', 'nao_confirma_dados', 'atende_desliga', 'ligacao_muda'):
        return 'ruim'
    return 'bom'


def _tramitacao_col_boolish(type_low):
    """BIT(1) ou TINYINT usados como booleano na tramitação legada."""
    if not type_low:
        return False
    return 'bit' in type_low or type_low.startswith('tinyint')


def _tramitacao_insert_parts(cursor, *, id_pessoa, id_contrato, canal, cpc, ts, descricao, fid,
                             fluxo_json=None, status_tramitacao=None,
                             carteira_legacy=None, discado_legacy=None,
                             fluxo_payload=None):
    """Monta INSERT alinhado ao schema GM e preenche colunas legadas quando existirem.

    Schemas antigos (BIT/ENUM em criar_banco): `atendido`, `cpc`, `contato`, `exito`, `status`,
    `classificacao`, `tipo` ativa/receptiva, `carteira`, `discado`. O wizard envia `fluxo_payload`
    para derivar esses campos; sem payload, usam-se defaults compatíveis com o CHECK legado (quando
    aplicável).

    Coluna `cpc` pode ser BIT (legado) ou texto (GM): o tipo vem de SHOW COLUMNS.
    """
    _, coltypes, tcols = _tramitacao_schema(cursor)
    cols_sql = []
    vals = []

    def add(lower, val):
        if lower not in tcols:
            return
        cn = str(tcols[lower]).replace('`', '')
        cols_sql.append('`' + cn + '`')
        vals.append(val)

    if 'carteira' in tcols:
        cv = carteira_legacy
        if cv is None:
            cv = 0.0
        try:
            cv = float(cv)
        except (TypeError, ValueError):
            cv = 0.0
        add('carteira', cv)
    if 'discado' in tcols:
        d = ''
        if discado_legacy is not None:
            d = str(discado_legacy).strip()
        if not d:
            d = '—'
        dmax = 20
        dt = coltypes.get('discado', '')
        try:
            if 'varchar' in dt:
                i = dt.index('varchar(')
                j = dt.index(')', i)
                dmax = int(dt[i + 8 : j])
        except (ValueError, IndexError):
            pass
        add('discado', d[:dmax] if dmax > 0 else d)

    if 'atendido' in tcols and _tramitacao_col_boolish(coltypes.get('atendido', '')):
        if fluxo_payload:
            add('atendido', 1 if fluxo_payload.get('atendido') == 'sim' else 0)
        else:
            add('atendido', 1)

    add('id_pessoa', id_pessoa)
    add('id_contrato', id_contrato)
    if 'forma' in tcols:
        add('forma', canal if canal in TRAMITACAO_TIPOS else 'ligacao')

    if 'tipo' in tcols:
        ttyp = coltypes.get('tipo', '')
        if 'ativa' in ttyp or 'receptiva' in ttyp:
            add('tipo', 'ativa')
        else:
            add('tipo', 'ativo')

    if 'cpc' in tcols:
        ctyp = coltypes.get('cpc', '')
        if _tramitacao_col_boolish(ctyp):
            if fluxo_payload:
                add('cpc', _fluxo_legacy_cpc_bit(fluxo_payload))
            else:
                add('cpc', 1 if str(cpc).lower() in ('sim', 'avalista', 'parente', 'amigo') else 0)
        else:
            add('cpc', cpc)

    if 'contato' in tcols:
        if fluxo_payload:
            add('contato', _fluxo_legacy_contato(fluxo_payload))
        else:
            add('contato', 'consorciado')

    if 'exito' in tcols and _tramitacao_col_boolish(coltypes.get('exito', '')):
        if fluxo_payload:
            add('exito', _fluxo_legacy_exito_bit(fluxo_payload))
        else:
            add('exito', 0)

    if 'status' in tcols:
        if fluxo_payload:
            add('status', _fluxo_legacy_status_db(fluxo_payload))
        else:
            add('status', 'sem interesse no pagamento')

    if 'classificacao' in tcols:
        if fluxo_payload:
            add('classificacao', _fluxo_legacy_classificacao(fluxo_payload))
        else:
            add('classificacao', 'indefinido')

    add('data', ts)

    desc_out = descricao
    if 'descricao' in tcols:
        desc_out = _tramitacao_truncate_descricao(descricao, coltypes.get('descricao', ''))
    add('descricao', desc_out)

    add('id_funcionario', fid)
    if fluxo_json is not None:
        add('fluxo_json', fluxo_json)
    if status_tramitacao is not None:
        add('status_tramitacao', status_tramitacao)
    return cols_sql, vals


def _fluxo_map_legacy_cpc(payload):
    """Compatibilidade com coluna cpc da tramitacao legada."""
    if payload.get('atendido') == 'nao':
        return 'nao'
    if payload.get('modo_indefinido'):
        return 'nao'
    if payload.get('cpc_correto') == 'nao':
        return 'nao'
    if payload.get('cpc_correto') != 'sim':
        return 'nao'
    qual = payload.get('cpc_qual')
    if qual == 'consorciado':
        return 'sim'
    if qual == 'avalista':
        return 'avalista'
    if qual == 'terceiro':
        return 'parente'
    return 'nao'


def _fluxo_status_display(payload):
    if payload.get('atendido') == 'nao':
        return 'Ligação Não atendida'
    if payload.get('modo_indefinido'):
        return 'Indefinido'
    sf = payload.get('status_final')
    if sf:
        return FLUXO_STATUS_LABEL.get(sf, sf)
    return '—'


def _fluxo_format_money(val):
    if val is None:
        return '0,00'
    try:
        d = decimal.Decimal(str(val))
        return f'{d:,.2f}'.replace(',', 'X').replace('.', ',').replace('X', '.')
    except (decimal.InvalidOperation, ValueError, TypeError):
        return str(val)


def _fluxo_build_descricao_text(payload, grupo=None, cota=None):
    """Texto legível gravado em tramitacao.descricao."""
    lines = []
    lines.append(f"Carteira (total em aberto): R$ {_fluxo_format_money(payload.get('carteira_devendo'))}")
    lines.append(f"Discado: {payload.get('numero_discado') or '—'}")
    if payload.get('atendido') == 'nao':
        mk = payload.get('motivo_nao_atendido')
        lines.append('Atendido: Não')
        lines.append(f"Motivo: {FLUXO_MOTIVO_LABEL.get(mk, mk or '—')}")
    elif payload.get('modo_indefinido'):
        lines.append('Atendido: Sim — encerrado como indefinido (atendeu e desligou sem detalhes).')
    else:
        lines.append(f"Atendido: {'Sim' if payload.get('atendido') == 'sim' else '—'}")
        if payload.get('cpc_correto') == 'nao':
            lines.append('CPC (pessoa certa): Não')
            qm = payload.get('cpc_quem')
            lines.append(f"Quem atendeu: {FLUXO_CPC_QUEM_LABEL.get(qm, qm or '—')}")
            if payload.get('cpc_etapa_descricao'):
                lines.append(f"Descrição (etapa CPC): {payload['cpc_etapa_descricao']}")
        elif payload.get('cpc_correto') == 'sim':
            lines.append('CPC (pessoa certa): Sim')
            cq = payload.get('cpc_qual')
            cq_lab = {'consorciado': 'Consorciado', 'terceiro': 'Terceiro', 'avalista': 'Avalista'}.get(cq, cq)
            lines.append(f"CPC atendido: {cq_lab or '—'}")
        sf = payload.get('status_final')
        if sf:
            lines.append(f"Status: {FLUXO_STATUS_LABEL.get(sf, sf)}")
        if sf == 'agendamento' and payload.get('agenda_retorno_data'):
            lines.append(f"Agenda — retorno em: {payload.get('agenda_retorno_data')}")
            if payload.get('agenda_retorno_atividade'):
                lines.append(f"Atividade: {payload['agenda_retorno_atividade']}")
        if sf == 'acordo_firmado':
            if payload.get('acordo_data_pagamento'):
                lines.append(f"Acordo — data prevista pagamento: {payload.get('acordo_data_pagamento')}")
            if payload.get('acordo_qtd_parcelas') is not None:
                lines.append(f"Quantidade de parcelas no acordo: {payload.get('acordo_qtd_parcelas')}")
    if payload.get('descricao_final'):
        lines.append('')
        lines.append('Descrição adicional:')
        lines.append(str(payload['descricao_final']))
    lab = _fluxo_status_display(payload)
    return f"[{lab}]\n" + '\n'.join(lines)


def _parse_iso_datetime_agenda(s):
    """Parse datetime-local / ISO para agenda.data."""
    if not s:
        return None
    s = str(s).strip().replace('Z', '')
    s = s.replace('T', ' ')
    for fmt in ('%Y-%m-%d %H:%M:%S', '%Y-%m-%d %H:%M', '%Y-%m-%d'):
        try:
            lim = 19 if fmt != '%Y-%m-%d' else 10
            return datetime.datetime.strptime(s[:lim], fmt)
        except ValueError:
            continue
    return None


def _validate_tramitacao_fluxo_payload(data):
    """Retorna (erro_msg ou None, dict normalizado)."""
    if not isinstance(data, dict):
        return 'JSON inválido.', None
    out = {}
    try:
        cd = data.get('carteira_devendo')
        if cd is not None and cd != '':
            out['carteira_devendo'] = float(cd)
        else:
            out['carteira_devendo'] = None
    except (TypeError, ValueError):
        return 'Valor de carteira inválido.', None

    nd = (data.get('numero_discado') or '').strip()
    if not nd:
        return 'Informe o número discado.', None
    out['numero_discado'] = nd[:64]

    at = (data.get('atendido') or '').strip().lower()
    if at not in ('sim', 'nao'):
        return 'Responda se foi atendido (sim ou não).', None
    out['atendido'] = at

    indef = bool(data.get('modo_indefinido'))
    out['modo_indefinido'] = indef

    if at == 'nao':
        mk = (data.get('motivo_nao_atendido') or '').strip().lower()
        if mk not in FLUXO_MOTIVOS_NA:
            return 'Selecione o motivo (ligação não atendida).', None
        out['motivo_nao_atendido'] = mk
        df = (data.get('descricao_final') or '').strip()
        out['descricao_final'] = df[:20000] if df else ''
        return None, out

    if indef:
        if at != 'sim':
            return 'Encerrar como indefinido só se a ligação foi atendida.', None
        df = (data.get('descricao_final') or '').strip()
        out['descricao_final'] = df[:20000] if df else ''
        out['cpc_correto'] = None
        out['cpc_quem'] = None
        out['cpc_etapa_descricao'] = None
        out['cpc_qual'] = None
        out['status_final'] = None
        return None, out

    cc = (data.get('cpc_correto') or '').strip().lower()
    if cc not in ('sim', 'nao'):
        return 'Responda CPC — contato com a pessoa certa (sim ou não).', None
    out['cpc_correto'] = cc

    if cc == 'nao':
        qm = (data.get('cpc_quem') or '').strip().lower()
        if qm not in FLUXO_CPC_QUEM:
            return 'Selecione quem atendeu.', None
        out['cpc_quem'] = qm
        ed = (data.get('cpc_etapa_descricao') or '').strip()
        out['cpc_etapa_descricao'] = ed[:5000] if ed else ''
        out['cpc_qual'] = None
    else:
        cq = (data.get('cpc_qual') or '').strip().lower()
        if cq not in FLUXO_CPC_QUAL:
            return 'Selecione qual CPC foi atendido.', None
        out['cpc_qual'] = cq
        out['cpc_quem'] = None
        out['cpc_etapa_descricao'] = ''

    sf = (data.get('status_final') or '').strip().lower()
    if sf not in FLUXO_STATUS_FINAL:
        return 'Selecione o status final.', None
    out['status_final'] = sf

    if sf == 'agendamento':
        ag = _parse_iso_datetime_agenda(data.get('agenda_retorno_data'))
        if not ag:
            return 'Informe data e hora do retorno (agenda).', None
        out['agenda_retorno_data'] = data.get('agenda_retorno_data')
        out['agenda_retorno_atividade'] = (data.get('agenda_retorno_atividade') or '').strip()[:255] or (
            'Retorno cobrança — retomar contato'
        )
    else:
        out['agenda_retorno_data'] = None
        out['agenda_retorno_atividade'] = ''

    if sf == 'acordo_firmado':
        dp = _parse_iso_datetime_agenda(data.get('acordo_data_pagamento'))
        if not dp:
            return 'Informe a data prevista do pagamento (acordo).', None
        out['acordo_data_pagamento'] = data.get('acordo_data_pagamento')
        try:
            qp = int(data.get('acordo_qtd_parcelas'))
            if qp < 1:
                raise ValueError
            out['acordo_qtd_parcelas'] = qp
        except (TypeError, ValueError):
            return 'Informe a quantidade de parcelas do acordo (número inteiro ≥ 1).', None
    else:
        out['acordo_data_pagamento'] = None
        out['acordo_qtd_parcelas'] = None

    df = (data.get('descricao_final') or '').strip()
    out['descricao_final'] = df[:20000] if df else ''
    return None, out


def _parse_tramitacao_timestamp(s):
    """Aceita 'YYYY-MM-DDTHH:MM' (datetime-local) ou 'YYYY-MM-DD HH:MM:SS' etc."""
    if s is None:
        return None
    s = str(s).strip()
    if not s:
        return None
    s = s.replace('T', ' ')
    for fmt in ('%Y-%m-%d %H:%M:%S', '%Y-%m-%d %H:%M', '%Y-%m-%d'):
        try:
            lim = 19 if fmt != '%Y-%m-%d' else 10
            return datetime.datetime.strptime(s[:lim], fmt)
        except ValueError:
            continue
    return None


def _tramitacao_row_by_id(cursor, tramitacao_id):
    cursor.execute(
        "SELECT t.*, f.nome AS funcionario_nome "
        "FROM tramitacao t "
        "LEFT JOIN funcionario f ON t.id_funcionario = f.id "
        "WHERE t.id = %s",
        (tramitacao_id,),
    )
    return _clean_row(cursor.fetchone())


def _json_safe(val):
    """Serializa valores para JSON em payload de moderacao."""
    if val is None:
        return None
    if isinstance(val, (datetime.datetime, datetime.date)):
        return val.isoformat()
    if isinstance(val, decimal.Decimal):
        return float(val)
    if isinstance(val, (bytes, bytearray)):
        try:
            return val.decode('utf-8', errors='replace')
        except Exception:
            return str(val)
    return val


def _tramrow_dict_jsonable(row):
    if not row:
        return {}
    out = {}
    for k, v in row.items():
        out[k] = _json_safe(v)
    return out


def _tramitacao_validar_put_body(data):
    """Valida corpo PUT tramitacao legado. Retorna None se OK ou mensagem de erro."""
    tipo = (data.get('tipo') or '').strip().lower()
    cpc = (data.get('cpc') or '').strip().lower()
    descricao = (data.get('descricao') or '')
    if isinstance(descricao, str) and len(descricao) > 20000:
        return 'Descricao muito longa.'
    if tipo not in TRAMITACAO_TIPOS:
        return 'Tipo invalido.'
    if cpc not in TRAMITACAO_CPCS:
        return 'CPC invalido.'
    ts = _parse_tramitacao_timestamp(data.get('data'))
    if not ts:
        return 'Data invalida.'
    return None


def _tramitacao_aplicar_put_payload(cursor, tramitacao_id, data):
    """Aplica atualizacao legado em tramitacao. Retorna (tramitacao_row_or_None, erro_str_or_None)."""
    err_v = _tramitacao_validar_put_body(data)
    if err_v:
        return None, err_v
    tipo = (data.get('tipo') or '').strip().lower()
    cpc = (data.get('cpc') or '').strip().lower()
    descricao = (data.get('descricao') or '')
    ts = _parse_tramitacao_timestamp(data.get('data'))
    canal = tipo
    tcols = {c.lower(): c for c in _tramitacao_column_names(cursor)}
    sets = []
    vals = []
    if 'forma' in tcols:
        sets.append('forma = %s')
        vals.append(canal if canal in TRAMITACAO_TIPOS else 'ligacao')
    if 'tipo' in tcols:
        sets.append('tipo = %s')
        vals.append('ativo')
    sets.extend(['cpc = %s', 'data = %s', 'descricao = %s'])
    vals.extend([cpc, ts, descricao or None, tramitacao_id])
    cursor.execute(
        f"UPDATE tramitacao SET {', '.join(sets)} WHERE id = %s",
        vals,
    )
    return _tramitacao_row_by_id(cursor, tramitacao_id), None


def _moderacao_row_snapshot(cursor, mid):
    cursor.execute(
        """
        SELECT m.*, c.grupo, c.cota,
               fs.nome AS solicitante_nome, fr.nome AS revisor_nome
        FROM solicitacao_moderacao m
        INNER JOIN contrato c ON c.id = m.id_contrato
        LEFT JOIN funcionario fs ON fs.id = m.id_solicitante
        LEFT JOIN funcionario fr ON fr.id = m.id_revisor
        WHERE m.id = %s
        """,
        (mid,),
    )
    return _clean_row(cursor.fetchone())


def _moderacao_contagem_pendente_tramitacao(cursor, ref_id):
    cursor.execute(
        """
        SELECT COUNT(*) AS c FROM solicitacao_moderacao
        WHERE status = 'pendente' AND ref_id = %s
          AND tipo IN ('tramitacao_edit', 'tramitacao_delete')
        """,
        (ref_id,),
    )
    r = cursor.fetchone()
    return int((r or {}).get('c') or 0)


def _moderacao_contagem_pendente_agenda(cursor, ref_id):
    cursor.execute(
        """
        SELECT COUNT(*) AS c FROM solicitacao_moderacao
        WHERE status = 'pendente' AND ref_id = %s
          AND tipo IN ('agenda_edit', 'agenda_delete')
        """,
        (ref_id,),
    )
    r = cursor.fetchone()
    return int((r or {}).get('c') or 0)


@app.route('/api/contrato/<int:contrato_id>/tramitacao', methods=['POST'])
def api_tramitacao_criar(contrato_id):
    fid = session.get('funcionario_id')
    if not fid:
        return jsonify({'error': 'Sessao invalida. Faca login novamente.'}), 401
    data = request.get_json(silent=True) or {}
    tipo = (data.get('tipo') or '').strip().lower()
    cpc = (data.get('cpc') or '').strip().lower()
    descricao = (data.get('descricao') or '')
    if isinstance(descricao, str) and len(descricao) > 20000:
        return jsonify({'error': 'Descricao muito longa.'}), 400
    if tipo not in TRAMITACAO_TIPOS:
        return jsonify({'error': 'Tipo invalido.'}), 400
    if cpc not in TRAMITACAO_CPCS:
        return jsonify({'error': 'CPC invalido.'}), 400
    ts = _parse_tramitacao_timestamp(data.get('data'))
    if not ts:
        return jsonify({'error': 'Data invalida.'}), 400
    conn = _get_db()
    cursor = conn.cursor()
    try:
        _ensure_tramitacao_fluxo_columns(cursor)
        cursor.execute("SELECT id, id_pessoa FROM contrato WHERE id = %s", (contrato_id,))
        c_row = cursor.fetchone()
        if not c_row:
            return jsonify({'error': 'Contrato nao encontrado.'}), 404
        id_pessoa = c_row.get('id_pessoa')
        if not id_pessoa:
            return jsonify({'error': 'Contrato sem pessoa (devedor) vinculada.'}), 400
        cols_sql, insert_vals = _tramitacao_insert_parts(
            cursor,
            id_pessoa=id_pessoa,
            id_contrato=contrato_id,
            canal=tipo,
            cpc=cpc,
            ts=ts,
            descricao=descricao or None,
            fid=int(fid),
            fluxo_json=None,
            status_tramitacao=None,
            discado_legacy=(data.get('numero_discado') or '').strip() or None,
        )
        if not cols_sql:
            return jsonify({'error': 'Tabela tramitacao sem colunas esperadas.'}), 500
        placeholders = ', '.join(['%s'] * len(insert_vals))
        cursor.execute(
            f'INSERT INTO tramitacao ({", ".join(cols_sql)}) VALUES ({placeholders})',
            tuple(insert_vals),
        )
        conn.commit()
        new_id = cursor.lastrowid
        row = _tramitacao_row_by_id(cursor, new_id)
        return jsonify({'tramitacao': row})
    except Exception as exc:
        if conn:
            try:
                conn.rollback()
            except Exception:
                pass
        app.logger.exception('api_tramitacao_criar')
        return jsonify({'error': 'Erro ao salvar: ' + str(exc)}), 500
    finally:
        try:
            cursor.close()
        except Exception:
            pass
        try:
            conn.close()
        except Exception:
            pass


@app.route('/api/tramitacao/<int:tramitacao_id>', methods=['PUT', 'DELETE'])
def api_tramitacao_item(tramitacao_id):
    fid = session.get('funcionario_id')
    if not fid:
        return jsonify({'error': 'Sessao invalida. Faca login novamente.'}), 401
    conn = _get_db()
    cursor = conn.cursor()
    try:
        _ensure_tramitacao_fluxo_columns(cursor)
        _ensure_solicitacao_moderacao_table(cursor)
        cursor.execute(
            "SELECT id, id_contrato FROM tramitacao WHERE id = %s",
            (tramitacao_id,),
        )
        t_row = cursor.fetchone()
        if not t_row:
            return jsonify({'error': 'Tramitacao nao encontrada.'}), 404
        id_contrato = int(t_row['id_contrato'])

        eh_cobranca = _nivel_normalizado(session.get('funcionario_nivel_acesso')) == 'cobranca'

        if request.method == 'DELETE':
            if eh_cobranca:
                if _moderacao_contagem_pendente_tramitacao(cursor, tramitacao_id) > 0:
                    return jsonify({'error': 'Ja existe solicitacao pendente para esta tramitacao.'}), 409
                antes = _tramrow_dict_jsonable(_tramitacao_row_by_id(cursor, tramitacao_id))
                payload = json.dumps({'antes': antes}, ensure_ascii=False)
                cursor.execute(
                    """
                    INSERT INTO solicitacao_moderacao
                    (id_solicitante, tipo, id_contrato, ref_id, payload_json, status)
                    VALUES (%s, 'tramitacao_delete', %s, %s, %s, 'pendente')
                    """,
                    (int(fid), id_contrato, tramitacao_id, payload),
                )
                mid = cursor.lastrowid
                conn.commit()
                return jsonify({
                    'success': True,
                    'pendente_aprovacao': True,
                    'moderacao_id': mid,
                    'mensagem': 'Pedido enviado para aprovacao do gestor ou administrador (pagina Solicitacoes).',
                })
            cursor.execute("DELETE FROM tramitacao WHERE id = %s", (tramitacao_id,))
            conn.commit()
            return jsonify({'success': True})

        data = request.get_json(silent=True) or {}
        if eh_cobranca:
            err_val = _tramitacao_validar_put_body(data)
            if err_val:
                return jsonify({'error': err_val}), 400
            if _moderacao_contagem_pendente_tramitacao(cursor, tramitacao_id) > 0:
                return jsonify({'error': 'Ja existe solicitacao pendente para esta tramitacao.'}), 409
            antes = _tramrow_dict_jsonable(_tramitacao_row_by_id(cursor, tramitacao_id))
            proposta = {
                'tipo': (data.get('tipo') or '').strip().lower(),
                'cpc': (data.get('cpc') or '').strip().lower(),
                'data': data.get('data'),
                'descricao': data.get('descricao'),
            }
            payload = json.dumps({'antes': antes, 'proposta': proposta}, ensure_ascii=False)
            cursor.execute(
                """
                INSERT INTO solicitacao_moderacao
                (id_solicitante, tipo, id_contrato, ref_id, payload_json, status)
                VALUES (%s, 'tramitacao_edit', %s, %s, %s, 'pendente')
                """,
                (int(fid), id_contrato, tramitacao_id, payload),
            )
            mid = cursor.lastrowid
            conn.commit()
            return jsonify({
                'success': True,
                'pendente_aprovacao': True,
                'moderacao_id': mid,
                'mensagem': 'Pedido de alteracao enviado para aprovacao do gestor ou administrador (pagina Solicitacoes).',
            })

        row, err_val = _tramitacao_aplicar_put_payload(cursor, tramitacao_id, data)
        if err_val:
            return jsonify({'error': err_val}), 400
        conn.commit()
        return jsonify({'tramitacao': row})
    except Exception as exc:
        if conn:
            try:
                conn.rollback()
            except Exception:
                pass
        app.logger.exception('api_tramitacao_item')
        return jsonify({'error': 'Erro ao processar: ' + str(exc)}), 500
    finally:
        try:
            cursor.close()
        except Exception:
            pass
        try:
            conn.close()
        except Exception:
            pass


@app.route('/api/contrato/<int:contrato_id>/tramitacao/fluxo', methods=['POST'])
def api_tramitacao_fluxo_criar(contrato_id):
    """Registra tramitação guiada (wizard), opcionalmente criando itens em `agenda`."""
    fid = session.get('funcionario_id')
    if not fid:
        return jsonify({'error': 'Sessao invalida. Faca login novamente.'}), 401
    body = request.get_json(silent=True) or {}
    err, payload = _validate_tramitacao_fluxo_payload(body)
    if err:
        return jsonify({'error': err}), 400

    conn = _get_db()
    cursor = conn.cursor()
    try:
        _ensure_tramitacao_fluxo_columns(cursor)
        cursor.execute(
            "SELECT id, id_pessoa, grupo, cota FROM contrato WHERE id = %s",
            (contrato_id,),
        )
        c_row = cursor.fetchone()
        if not c_row:
            return jsonify({'error': 'Contrato nao encontrado.'}), 404
        id_pessoa = c_row.get('id_pessoa')
        if not id_pessoa:
            return jsonify({'error': 'Contrato sem pessoa (devedor) vinculada.'}), 400
        grupo = c_row.get('grupo')
        cota = c_row.get('cota')

        descricao_txt = _fluxo_build_descricao_text(payload, grupo, cota)
        status_lab = _fluxo_status_display(payload)
        legacy_cpc = _fluxo_map_legacy_cpc(payload)
        ts = datetime.datetime.now()
        fluxo_str = json.dumps(payload, ensure_ascii=False)

        cols_sql, insert_vals = _tramitacao_insert_parts(
            cursor,
            id_pessoa=id_pessoa,
            id_contrato=contrato_id,
            canal='ligacao',
            cpc=legacy_cpc,
            ts=ts,
            descricao=descricao_txt,
            fid=int(fid),
            fluxo_json=fluxo_str,
            status_tramitacao=status_lab[:96],
            carteira_legacy=payload.get('carteira_devendo'),
            discado_legacy=payload.get('numero_discado'),
            fluxo_payload=payload,
        )
        if not cols_sql:
            return jsonify({'error': 'Tabela tramitacao sem colunas esperadas.'}), 500
        placeholders = ', '.join(['%s'] * len(insert_vals))
        cursor.execute(
            f'INSERT INTO tramitacao ({", ".join(cols_sql)}) VALUES ({placeholders})',
            tuple(insert_vals),
        )
        new_id = cursor.lastrowid

        if payload.get('status_final') == 'agendamento' and payload.get('agenda_retorno_data'):
            ag_dt = _parse_iso_datetime_agenda(payload['agenda_retorno_data'])
            if ag_dt:
                cursor.execute(
                    """
                    INSERT INTO agenda (atividade, descricao, data, prioridade, id_contrato, id_funcionario)
                    VALUES (%s, %s, %s, %s, %s, %s)
                    """,
                    (
                        payload.get('agenda_retorno_atividade')
                        or 'Retorno cobrança — retomar contato',
                        (
                            f"Contrato {grupo}/{cota}. Agendado pela tramitação."
                        ),
                        ag_dt,
                        'media',
                        contrato_id,
                        int(fid),
                    ),
                )

        if payload.get('status_final') == 'acordo_firmado':
            ac_dt = _parse_iso_datetime_agenda(payload.get('acordo_data_pagamento'))
            qtp = payload.get('acordo_qtd_parcelas')
            if ac_dt and qtp:
                cursor.execute(
                    """
                    INSERT INTO agenda (atividade, descricao, data, prioridade, id_contrato, id_funcionario)
                    VALUES (%s, %s, %s, %s, %s, %s)
                    """,
                    (
                        f"Verificar acordo — {qtp} parcela(s)",
                        (
                            f"Contrato {grupo}/{cota}. Pagamento previsto em {payload.get('acordo_data_pagamento')}."
                        ),
                        ac_dt,
                        'alta',
                        contrato_id,
                        int(fid),
                    ),
                )

        conn.commit()
        row = _tramitacao_row_by_id(cursor, new_id)
        return jsonify({'tramitacao': row})
    except Exception as exc:
        if conn:
            try:
                conn.rollback()
            except Exception:
                pass
        app.logger.exception('api_tramitacao_fluxo_criar')
        return jsonify({'error': 'Erro ao salvar: ' + str(exc)}), 500
    finally:
        try:
            cursor.close()
        except Exception:
            pass
        try:
            conn.close()
        except Exception:
            pass


# ---------------------------------------------------------------------------
# API: Relatorios
# ---------------------------------------------------------------------------

_RELATORIO_COLUMNS = [
    ('Grupo', 'grupo'),
    ('Cota', 'cota'),
    ('Grupo/Cota', 'grupo_cota'),
    ('CPF / CNPJ', 'cpf_cnpj'),
    ('Nome Devedor', 'nome_devedor'),
    ('Status', 'status'),
    ('Data Arquivo', 'data_arquivo'),
    ('Atraso', 'atraso'),
]


def _build_relatorio_query(tipo, data_inicial, data_final, prioridade=None):
    """Retorna (sql, params) para o relatorio solicitado (exceto 'abertos' — ver _build_relatorio_query_abertos)."""
    params = []

    # Demais tipos: pagos, indenizados, novos
    params = [data_inicial, data_final]

    base = (
        "SELECT DISTINCT c.id, c.grupo, c.cota, "
        "       CONCAT_WS('/', c.grupo, c.cota) AS grupo_cota, "
        "       c.numero_contrato, c.status, "
        "       p.nome_completo AS nome_devedor, p.cpf_cnpj, o.data_arquivo, "
        "       (CASE WHEN parc_ab.min_v_aberto IS NOT NULL "
        "             THEN DATEDIFF(CURRENT_DATE, parc_ab.min_v_aberto) END) AS atraso "
        "FROM ocorrencia o "
        "LEFT JOIN contrato c ON c.id = o.id_contrato "
        "LEFT JOIN pessoa p ON c.id_pessoa = p.id "
        "LEFT JOIN ( "
        "  SELECT id_contrato, MIN(vencimento) AS min_v_aberto "
        "  FROM parcela WHERE status = 'cobranca' GROUP BY id_contrato "
        ") parc_ab ON parc_ab.id_contrato = c.id "
    )

    if tipo == 'novos':
        where = "WHERE o.status = 'cobranca' AND o.descricao LIKE '%%novo%%'"
    elif tipo == 'voltaram':
        where = "WHERE o.status = 'cobranca' AND o.descricao LIKE '%%contrato voltou%%'"
    elif tipo == 'pagos':
        where = "WHERE c.status = 'pago' AND o.status = %s"
        params = [_STATUS_OCORRENCIA_PAGO_TOTAL, data_inicial, data_final]
    elif tipo == 'indenizados':
        where = "WHERE c.status = 'indenizado' AND o.status = 'indenizado'"
    elif tipo == 'pagos_parcialmente':
        # Performado (mesma regra do Performance): parcela quitada com atraso 0–90 d na quitação.
        sql = (
            "SELECT c.id, c.grupo, c.cota, "
            "       CONCAT_WS('/', c.grupo, c.cota) AS grupo_cota, "
            "       c.numero_contrato, c.status, "
            "       pes.nome_completo AS nome_devedor, pes.cpf_cnpj, "
            "       DATE(pq.mx_dt) AS data_arquivo, "
            "       (CASE WHEN parc_ab.min_v_aberto IS NOT NULL "
            "             THEN DATEDIFF(CURRENT_DATE, parc_ab.min_v_aberto) END) AS atraso "
            "FROM ( "
            "  SELECT id_contrato, MAX(data_pagamento) AS mx_dt "
            "  FROM parcela "
            "  WHERE status = 'pago' "
            "    AND data_pagamento >= %s AND data_pagamento <= %s "
            "    AND DATEDIFF(data_pagamento, vencimento) BETWEEN 0 AND 90 "
            "  GROUP BY id_contrato "
            ") pq "
            "JOIN contrato c ON c.id = pq.id_contrato "
            "LEFT JOIN pessoa pes ON c.id_pessoa = pes.id "
            "LEFT JOIN ( "
            "  SELECT id_contrato, MIN(vencimento) AS min_v_aberto "
            "  FROM parcela WHERE status = 'cobranca' GROUP BY id_contrato "
            ") parc_ab ON parc_ab.id_contrato = c.id "
            "ORDER BY pq.mx_dt, c.grupo, c.cota"
        )
        return sql, [data_inicial, data_final]
    else:
        where = "WHERE 1=1"

    where += " AND o.data_arquivo >= %s AND o.data_arquivo <= %s"
    sql = base + where + " ORDER BY o.data_arquivo, c.grupo, c.cota"
    return sql, params


def _fetch_relatorio_rows(tipo, data_inicial, data_final, prioridade=None):
    conn = _get_db()
    cursor = conn.cursor()
    try:
        if tipo == 'abertos':
            _ensure_cobranca_table(cursor)
            data_ref = _parse_data_final_relatorio_abertos(data_final)
            sql, params = _build_relatorio_query_abertos(data_ref, prioridade)
        else:
            sql, params = _build_relatorio_query(tipo, data_inicial, data_final, prioridade)
        cursor.execute(sql, params)
        rows = _clean_rows(cursor.fetchall())
    finally:
        cursor.close()
        conn.close()
    return rows


_RELATORIO_EMAIL_LOTE_MAX_CONTRATOS = 2000
_RELATORIO_EMAIL_LOTE_MAX_MSG_PLAIN = 15000


def _relatorio_parse_contrato_ids_lote(payload):
    """Normaliza lista de ids de contrato do JSON (dedup, ordem preservada, limite)."""
    raw = payload.get('contrato_ids')
    if raw is None:
        return None, 'Lista contrato_ids obrigatoria.'
    if not isinstance(raw, list):
        return None, 'contrato_ids deve ser uma lista.'
    out = []
    seen = set()
    for x in raw:
        try:
            n = int(x)
        except (TypeError, ValueError):
            continue
        if n < 1 or n in seen:
            continue
        seen.add(n)
        out.append(n)
        if len(out) >= _RELATORIO_EMAIL_LOTE_MAX_CONTRATOS:
            break
    if not out:
        return None, 'Informe ao menos um id de contrato valido.'
    return out, None


def _relatorio_email_lote_disparar(fid, contrato_ids, mensagem_plain):
    """Envia e-mail (MessageCenter + registro_email) para cada endereco valido dos devedores dos contratos."""
    msg_plain = (mensagem_plain or '').strip()
    if not msg_plain:
        return None, 'Mensagem vazia.'
    if len(msg_plain) > _RELATORIO_EMAIL_LOTE_MAX_MSG_PLAIN:
        return None, (
            f'Mensagem muito longa (max {_RELATORIO_EMAIL_LOTE_MAX_MSG_PLAIN} caracteres).'
        )

    corpo_html = _plain_para_corpo_email_html(msg_plain)
    if len(corpo_html) > 200000:
        return None, 'Corpo HTML gerado excede o limite.'

    remetente_nome = _primeiro_nome(session.get('funcionario_nome'))
    stats = {
        'envios_email': 0,
        'falhas': 0,
        'ignorados_sem_contrato': 0,
        'ignorados_sem_pessoa': 0,
        'ignorados_sem_email': 0,
        'ignorados_email_invalido': 0,
        'contratos_com_envio': 0,
    }
    erros_amostra = []

    conn = _get_db()
    try:
        placeholders = ','.join(['%s'] * len(contrato_ids))
        with conn.cursor() as cursor:
            cursor.execute(
                f"""
                SELECT c.id AS id_contrato, c.id_pessoa, p.nome_completo
                FROM contrato c
                LEFT JOIN pessoa p ON p.id = c.id_pessoa
                WHERE c.id IN ({placeholders})
                """,
                tuple(contrato_ids),
            )
            por_id = {int(r['id_contrato']): r for r in (cursor.fetchall() or [])}

        pessoa_ids = []
        for cid in contrato_ids:
            row = por_id.get(int(cid))
            if not row:
                stats['ignorados_sem_contrato'] += 1
                continue
            pid = row.get('id_pessoa')
            if pid is None:
                stats['ignorados_sem_pessoa'] += 1
                continue
            pessoa_ids.append(int(pid))
        pessoa_ids = list(dict.fromkeys(pessoa_ids))

        with conn.cursor() as cursor:
            email_map = _emails_por_pessoas(cursor, pessoa_ids)

        for cid in contrato_ids:
            cid_i = int(cid)
            row = por_id.get(cid_i)
            if not row:
                continue
            pid = row.get('id_pessoa')
            if pid is None:
                continue
            id_pessoa = int(pid)
            primeiro = _primeiro_nome(row.get('nome_completo'))
            emails_rows = email_map.get(id_pessoa, [])
            if not emails_rows:
                stats['ignorados_sem_email'] += 1
                continue

            enviou_contrato = False
            for em in emails_rows:
                em_addr = (em.get('email') or '').strip()
                if not _email_basico_valido(em_addr):
                    stats['ignorados_email_invalido'] += 1
                    continue

                ok_mc, err_mc, det_mc = _messagecenter_post_email_html(
                    remetente_nome,
                    em_addr,
                    primeiro or 'Cliente',
                    corpo_html,
                )
                if not ok_mc:
                    stats['falhas'] += 1
                    if len(erros_amostra) < 25:
                        erros_amostra.append({
                            'id_contrato': cid_i,
                            'erro': err_mc or 'Falha MessageCenter',
                            'detalhe': det_mc,
                        })
                    continue

                msg_reg = corpo_html[:1600]
                try:
                    with conn.cursor() as cursor:
                        ids_reg, err_reg = _resolve_ids_registro_email(
                            cursor,
                            id_pessoa,
                            int(em['id']),
                            cid_i,
                        )
                        if err_reg:
                            stats['falhas'] += 1
                            if len(erros_amostra) < 25:
                                erros_amostra.append({
                                    'id_contrato': cid_i,
                                    'erro': err_reg,
                                })
                            continue
                        cursor.execute(
                            'INSERT INTO registro_email '
                            '(id_contrato, id_pessoa, id_email, id_funcionario, mensagem) '
                            'VALUES (%s, %s, %s, %s, %s)',
                            (
                                ids_reg['id_contrato'],
                                ids_reg['id_pessoa'],
                                ids_reg['id_email'],
                                int(fid),
                                msg_reg,
                            ),
                        )
                        conn.commit()
                except Exception as exc:
                    try:
                        conn.rollback()
                    except Exception:
                        pass
                    stats['falhas'] += 1
                    app.logger.warning(
                        'relatorio email lote: insert registro_email falhou: %s', exc
                    )
                    if len(erros_amostra) < 25:
                        erros_amostra.append({
                            'id_contrato': cid_i,
                            'erro': f'Falha ao registrar e-mail: {exc}',
                        })
                    continue

                stats['envios_email'] += 1
                enviou_contrato = True

            if enviou_contrato:
                stats['contratos_com_envio'] += 1

    finally:
        try:
            conn.close()
        except Exception:
            pass

    stats['erros_amostra'] = erros_amostra
    return stats, None


@app.route('/api/relatorios/email-lote', methods=['POST'])
def api_relatorios_email_lote():
    """E-mail em lote para contratos listados em Relatorios (mesma API MessageCenter que /api/enviar-email-html).

    Apenas Gestor ou Administrador. Corpo: texto plano convertido para HTML simples.
    """
    forbidden = _admin_json_forbidden()
    if forbidden:
        return forbidden

    fid = session.get('funcionario_id')
    if not fid:
        return jsonify({'error': 'Nao autenticado. Faca login novamente.'}), 401

    payload = request.get_json(silent=True) or {}
    contrato_ids, err = _relatorio_parse_contrato_ids_lote(payload)
    if err:
        return jsonify({'error': err}), 400

    mensagem = payload.get('mensagem')
    if mensagem is None:
        mensagem = payload.get('mensagem_texto') or payload.get('corpo')

    stats, err = _relatorio_email_lote_disparar(int(fid), contrato_ids, mensagem)
    if err:
        return jsonify({'error': err}), 400

    return jsonify({'ok': True, **stats})


def _validate_relatorio_params():
    tipo = request.args.get('tipo', '').strip()
    data_inicial = request.args.get('data_inicial', '').strip()
    data_final = request.args.get('data_final', '').strip()
    prioridade = request.args.get('prioridade', '').strip()

    if not tipo:
        return None, None, None, None, 'Parametro tipo e obrigatorio.'
    # 'abertos': Data Final = data_arquivo em `cobranca` (obrigatoria; Data Inicial ignorada).
    if tipo == 'abertos' and not data_final:
        return (
            None, None, None, None,
            "Para 'Contratos Abertos', informe a Data Final (data_arquivo na tabela cobranca).",
        )
    if tipo == 'abertos' and data_final:
        try:
            datetime.date.fromisoformat(str(data_final).strip()[:10])
        except ValueError:
            return None, None, None, None, "Data Final invalida. Use o formato AAAA-MM-DD (ex.: 2026-04-22)."
    if tipo != 'abertos' and (not data_inicial or not data_final):
        return None, None, None, None, 'Parametros data_inicial e data_final sao obrigatorios.'
    return tipo, data_inicial or None, data_final or None, prioridade or None, None


def _relatorio_export_periodo_titulo(tipo, dt_ini, dt_fim):
    """Texto de periodo para cabecalhos de export (Excel/PDF)."""
    if tipo != 'abertos':
        return f"{dt_ini or '—'} a {dt_fim or '—'}"
    conn = _get_db()
    cursor = conn.cursor()
    try:
        _ensure_cobranca_table(cursor)
        ref = _relatorio_data_ref_abertos(cursor, dt_fim)
        if ref:
            return f"Snapshot (data GM): {ref.isoformat()}"
        return "Snapshot: —"
    finally:
        cursor.close()
        conn.close()


def _relatorio_export_filename_suffix(tipo, dt_ini, dt_fim):
    if tipo == 'abertos':
        conn = _get_db()
        cursor = conn.cursor()
        try:
            _ensure_cobranca_table(cursor)
            ref = _relatorio_data_ref_abertos(cursor, dt_fim)
            return ref.isoformat() if ref else 'sem_data'
        finally:
            cursor.close()
            conn.close()
    return f"{dt_ini}_{dt_fim}"


_TIPO_LABELS = {
    'abertos': 'Contratos Abertos',
    'pagos': 'Contratos Pagos/Fechados',
    'indenizados': 'Contratos Indenizados',
    'novos': 'Contratos Novos',
    'voltaram': 'Contratos que Voltaram',
    'pagos_parcialmente': 'Contratos pagos parcialmente (performado 0–90 d)',
}


@app.route('/api/relatorios')
def api_relatorios():
    tipo, dt_ini, dt_fim, prioridade, err = _validate_relatorio_params()
    if err:
        return jsonify({'error': err}), 400
    rows = _fetch_relatorio_rows(tipo, dt_ini, dt_fim, prioridade)
    return jsonify({'results': rows, 'tipo': tipo, 'total': len(rows)})


@app.route('/api/relatorios/excel')
def api_relatorios_excel():
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    tipo, dt_ini, dt_fim, prioridade, err = _validate_relatorio_params()
    if err:
        return jsonify({'error': err}), 400

    rows = _fetch_relatorio_rows(tipo, dt_ini, dt_fim, prioridade)
    titulo = _TIPO_LABELS.get(tipo, tipo)
    periodo_titulo = _relatorio_export_periodo_titulo(tipo, dt_ini, dt_fim)
    fname_suffix = _relatorio_export_filename_suffix(tipo, dt_ini, dt_fim)

    wb = Workbook()
    ws = wb.active
    ws.title = 'Relatorio'

    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='3B82F6', end_color='3B82F6', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin', color='D1D5DB'),
        right=Side(style='thin', color='D1D5DB'),
        top=Side(style='thin', color='D1D5DB'),
        bottom=Side(style='thin', color='D1D5DB'),
    )

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(_RELATORIO_COLUMNS))
    title_cell = ws.cell(row=1, column=1, value=f'{titulo}  |  {periodo_titulo}')
    title_cell.font = Font(bold=True, size=13)
    title_cell.alignment = Alignment(horizontal='center')

    for col_idx, (label, _key) in enumerate(_RELATORIO_COLUMNS, start=1):
        cell = ws.cell(row=3, column=col_idx, value=label)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    alt_fill = PatternFill(start_color='F8FAFC', end_color='F8FAFC', fill_type='solid')
    for row_idx, row in enumerate(rows, start=4):
        for col_idx, (_label, key) in enumerate(_RELATORIO_COLUMNS, start=1):
            val = row.get(key, '')
            cell = ws.cell(row=row_idx, column=col_idx, value=val if val is not None else '')
            cell.border = thin_border
            if (row_idx - 4) % 2 == 1:
                cell.fill = alt_fill

    for col_idx, (_label, _key) in enumerate(_RELATORIO_COLUMNS, start=1):
        max_len = len(_label)
        for row_idx in range(4, 4 + len(rows)):
            cell_val = str(ws.cell(row=row_idx, column=col_idx).value or '')
            max_len = max(max_len, len(cell_val))
        ws.column_dimensions[ws.cell(row=3, column=col_idx).column_letter].width = min(max_len + 4, 40)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    fname = f'relatorio_{tipo}_{fname_suffix}.xlsx'
    return send_file(buf, as_attachment=True, download_name=fname,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@app.route('/api/relatorios/pdf')
def api_relatorios_pdf():
    from fpdf import FPDF

    tipo, dt_ini, dt_fim, prioridade, err = _validate_relatorio_params()
    if err:
        return jsonify({'error': err}), 400

    rows = _fetch_relatorio_rows(tipo, dt_ini, dt_fim, prioridade)
    titulo = _TIPO_LABELS.get(tipo, tipo)
    periodo_titulo = _relatorio_export_periodo_titulo(tipo, dt_ini, dt_fim)
    fname_suffix = _relatorio_export_filename_suffix(tipo, dt_ini, dt_fim)

    pdf = FPDF(orientation='L', unit='mm', format='A4')
    pdf.set_auto_page_break(auto=True, margin=15)
    pdf.add_page()

    pdf.set_font('Helvetica', 'B', 14)
    pdf.cell(0, 10, titulo, ln=True, align='C')
    pdf.set_font('Helvetica', '', 10)
    pdf.cell(0, 6, f'Periodo: {periodo_titulo}', ln=True, align='C')
    pdf.ln(6)

    col_widths = [16, 14, 22, 24, 34, 70, 28, 16]
    headers = [label for label, _key in _RELATORIO_COLUMNS]

    pdf.set_font('Helvetica', 'B', 9)
    pdf.set_fill_color(59, 130, 246)
    pdf.set_text_color(255, 255, 255)
    for i, h in enumerate(headers):
        pdf.cell(col_widths[i], 8, h, border=1, align='C', fill=True)
    pdf.ln()

    pdf.set_font('Helvetica', '', 8)
    pdf.set_text_color(30, 41, 59)
    for row_idx, row in enumerate(rows):
        if row_idx % 2 == 1:
            pdf.set_fill_color(248, 250, 252)
            fill = True
        else:
            pdf.set_fill_color(255, 255, 255)
            fill = True

        for i, (_label, key) in enumerate(_RELATORIO_COLUMNS):
            val = str(row.get(key, '') or '')
            pdf.cell(col_widths[i], 7, val[:50], border=1, align='C', fill=fill)
        pdf.ln()

    pdf.ln(4)
    pdf.set_font('Helvetica', 'I', 8)
    pdf.cell(0, 5, f'Total de registros: {len(rows)}', ln=True)

    buf = io.BytesIO()
    pdf.output(buf)
    buf.seek(0)

    fname = f'relatorio_{tipo}_{fname_suffix}.pdf'
    return send_file(buf, as_attachment=True, download_name=fname, mimetype='application/pdf')


# ---------------------------------------------------------------------------
# API: Dashboard
# ---------------------------------------------------------------------------

@app.route('/api/dashboard')
def api_dashboard():
    """Dashboard totalmente reativo.

    Retorna, para os ultimos 12 meses, series mensais com:
      - pagos, indenizados: 1 ocorrencia = 1 contrato distinto no mes
      - novos, retomados: 1 ponto = 1 ocorrencia (status=aberto, mesmo criterio do *Relatorios*),
        para refletir varias ocorrencias do mesmo contrato no mes
      - entradas_safra: 1 contrato distinto (novo|voltou, aberto) alinhado ao eixo do Performance
      - ambos_mes_safra: contratos com ocorrencia novo e ocorrencia voltou no mesmo mes (abertas)

    Todo o recorte (periodo / series) e feito no front-end sobre esse payload,
    sem novas requisicoes.
    """
    conn = _get_db()
    cursor = conn.cursor()
    try:
        _ensure_cobranca_table(cursor)
    except Exception:
        app.logger.exception('api_dashboard: falha ao garantir tabela cobranca')

    now = datetime.date.today()

    # --- Gera lista continua de 12 meses (mais antigo -> mais recente) ---
    start_month = now.replace(day=1)
    for _ in range(11):
        start_month = (start_month - datetime.timedelta(days=1)).replace(day=1)
    window_start = start_month.isoformat()

    all_months = []
    cursor_d = start_month
    while cursor_d <= now:
        all_months.append(cursor_d.strftime('%Y-%m'))
        if cursor_d.month == 12:
            cursor_d = cursor_d.replace(year=cursor_d.year + 1, month=1)
        else:
            cursor_d = cursor_d.replace(month=cursor_d.month + 1)

    def _series(where_clause, params):
        cursor.execute(
            "SELECT DATE_FORMAT(o.data_arquivo, '%%Y-%%m') AS mes, "
            "       COUNT(DISTINCT o.id_contrato) AS total "
            "FROM ocorrencia o "
            + where_clause +
            " AND o.data_arquivo >= %s "
            "GROUP BY mes ORDER BY mes",
            tuple(params) + (window_start,),
        )
        by_month = {r['mes']: int(r['total']) for r in cursor.fetchall()}
        return [by_month.get(m, 0) for m in all_months]

    def _series_ocorrencias_aberto(where_clause, params):
        """1 linha = 1 ocorrencia. Alinha ao relatorio (novo/voltou com status=aberto)."""
        cursor.execute(
            "SELECT DATE_FORMAT(o.data_arquivo, '%%Y-%%m') AS mes, "
            "       COUNT(*) AS total "
            "FROM ocorrencia o "
            + where_clause +
            " AND o.data_arquivo >= %s "
            "GROUP BY mes ORDER BY mes",
            tuple(params) + (window_start,),
        )
        by_month = {r['mes']: int(r['total']) for r in cursor.fetchall()}
        return [by_month.get(m, 0) for m in all_months]

    def _ambos_mes_safra():
        """Qtd. de contratos com ocorrencia 'novo' e ocorrencia 'voltou' no mesmo MES (calendario)."""
        cursor.execute(
            "SELECT DATE_FORMAT(o1.data_arquivo, '%%Y-%%m') AS mes, "
            "       COUNT(DISTINCT o1.id_contrato) AS n "
            "FROM ocorrencia o1 "
            "INNER JOIN ocorrencia o2 "
            "  ON o1.id_contrato = o2.id_contrato AND o1.id <> o2.id "
            " AND YEAR(o1.data_arquivo) = YEAR(o2.data_arquivo) "
            " AND MONTH(o1.data_arquivo) = MONTH(o2.data_arquivo) "
            "WHERE o1.data_arquivo >= %s "
            "  AND o1.status = 'cobranca' AND o1.descricao LIKE '%%novo%%' "
            "  AND o2.status = 'cobranca' AND o2.descricao LIKE '%%contrato voltou%%' "
            "GROUP BY mes",
            (window_start,),
        )
        by_month = {r['mes']: int(r['n'] or 0) for r in cursor.fetchall()}
        return [by_month.get(m, 0) for m in all_months]

    serie_pagos = _series(
        "WHERE o.status IN ('pago total','pago parcial')",
        [],
    )
    serie_indenizados = _series("WHERE o.status = 'indenizado'", [])
    # Ocorrencias (nao 1x por contrato) — alinha ao relatorio: status=aberto + LIKE
    serie_novos = _series_ocorrencias_aberto(
        "WHERE o.status = 'cobranca' AND o.descricao LIKE '%%novo%%'", []
    )
    serie_retomados = _series_ocorrencias_aberto(
        "WHERE o.status = 'cobranca' AND o.descricao LIKE '%%contrato voltou%%'", []
    )
    serie_entradas_safra = _series(
        "WHERE o.status = 'cobranca' AND (o.descricao = 'contrato novo' OR LOWER(o.descricao) LIKE 'contrato voltou%%')",
        [],
    )
    # Mesma regra de "performado" no Performance: parcela quitada com atraso na quitação entre 0 e 90 dias.
    cursor.execute(
        "SELECT DATE_FORMAT(p.data_pagamento, '%%Y-%%m') AS mes, "
        "       COUNT(DISTINCT p.id_contrato) AS total "
        "FROM parcela p "
        "WHERE p.status = 'pago' "
        "AND DATEDIFF(p.data_pagamento, p.vencimento) BETWEEN 0 AND 90 "
        "AND DATE(p.data_pagamento) >= %s "
        "GROUP BY mes ORDER BY mes",
        (window_start,),
    )
    by_pp = {r['mes']: int(r['total']) for r in cursor.fetchall()}
    serie_pagos_parcial = [by_pp.get(m, 0) for m in all_months]

    # Contratos (mes) com as duas naturezas (novo e voltou) no mesmo mes — destaque p/ dupla entrada
    ambos_mes_safra = _ambos_mes_safra()

    # --- KPIs snapshot (tabela `cobranca` = contratos em aberto no ultimo dia GM) ---
    data_dash_ref = _get_data_referencia_arquivos_gm(cursor)
    kpi_em_cobranca = _em_cobranca_count_por_data_ref(cursor, data_dash_ref)

    # --- Grafico doughnut: distribuicao da carteira ---
    cursor.execute("SELECT status, COUNT(*) AS total FROM contrato GROUP BY status")
    pie_raw = {r['status']: int(r['total']) for r in cursor.fetchall()}

    cursor.close()
    conn.close()

    return jsonify({
        'meses': all_months,
        'series': {
            'pagos': serie_pagos,
            'indenizados': serie_indenizados,
            'novos': serie_novos,
            'retomados': serie_retomados,
            'entradas_safra': serie_entradas_safra,
            'pagos_parcial': serie_pagos_parcial,
        },
        'ambos_mes_safra': ambos_mes_safra,
        'snapshot': {
            'em_cobranca': kpi_em_cobranca,
        },
        'pie_chart': pie_raw,
    })


@app.route('/api/dashboard/panel_contratos')
def api_dashboard_panel_contratos():
    """Contratos com ocorrencia no periodo/series ativos (mesma base do export), com busca opcional."""
    ps = str(request.args.get('period_start', '')).strip()
    pe = str(request.args.get('period_end', '')).strip()
    raw = request.args.getlist('series')
    if not raw and request.args.get('series'):
        raw = [x.strip() for x in str(request.args.get('series', '')).split(',') if x.strip()]
    series_keys = [s for s in (raw or []) if s in _PAINEL_DASH_SERIES_KEYS_VALID]
    if not series_keys:
        series_keys = ['pagos', 'indenizados']
    if not ps or not pe or len(ps) < 7 or len(pe) < 7:
        return jsonify({'error': 'Informe period_start e period_end (YYYY-MM).'}), 400
    try:
        y1, m1 = int(ps[:4]), int(ps[5:7])
        y2, m2 = int(pe[:4]), int(pe[5:7])
        d_ini = datetime.date(y1, m1, 1).isoformat()
        d_fim = datetime.date(y2, m2, _last_day_of(y2, m2)).isoformat()
    except (TypeError, ValueError):
        return jsonify({'error': 'Periodo invalido.'}), 400

    union_sql, union_p = _painel_dash_union_contrato_ids_sql(series_keys, d_ini, d_fim)
    if not union_sql:
        return jsonify({'error': 'Nenhuma serie valida.'}), 400

    join_x, and_busca, p_busca = _panel_dash_busca_filtro_sql(
        request.args.get('tipo', 'contrato'),
        request.args.get('termo', ''),
        request.args.get('status', ''),
    )
    params = list(union_p) + p_busca
    count_sql = (
        "SELECT COUNT(DISTINCT c.id) AS n FROM contrato c "
        "LEFT JOIN pessoa p ON c.id_pessoa = p.id "
        + (join_x or '') +
        f"WHERE c.id IN {union_sql} "
        + and_busca
    )
    sql = (
        "SELECT DISTINCT c.id, c.grupo, c.cota, c.numero_contrato, c.status, "
        "       p.nome_completo AS devedor, p.cpf_cnpj AS devedor_cpf_cnpj "
        "FROM contrato c "
        "LEFT JOIN pessoa p ON c.id_pessoa = p.id "
        + (join_x or '') +
        f"WHERE c.id IN {union_sql} "
        + and_busca +
        f" ORDER BY c.grupo, c.cota LIMIT {int(_PAINEL_DASH_LIM)}"
    )

    conn = _get_db()
    cursor = conn.cursor()
    try:
        cursor.execute(count_sql, tuple(params))
        full_total = int((cursor.fetchone() or {}).get('n') or 0)
        cursor.execute(sql, tuple(params))
        rows = _clean_rows(cursor.fetchall())
    finally:
        cursor.close()
        conn.close()
    lim_ok = _PAINEL_DASH_LIM
    return jsonify({
        'results': rows,
        'total': full_total,
        'limited': full_total > lim_ok,
        'limit': lim_ok,
    })


def _panel_perf_contratos_rows(cursor, y, m, safra_index, teto, tipo, termo, st):
    """Cohort Performance (teto) + busca, mesma base do export."""
    enriched = _cohort_enriched_rows_for_export(cursor, y, m, safra_index, teto)
    if not enriched:
        return [], 0, False
    by_id = {e['id_contrato']: e for e in enriched}
    ordered = list(by_id.keys())
    if not ordered:
        return [], 0, False
    join_x, and_busca, p_busca = _panel_dash_busca_filtro_sql(
        (tipo or 'contrato'),
        (termo or '').strip(),
        (st or '').strip(),
    )
    out = []
    chunk = 300
    for i in range(0, len(ordered), chunk):
        part = ordered[i : i + chunk]
        ph = ','.join(['%s'] * len(part))
        sql = (
            "SELECT DISTINCT c.id, c.grupo, c.cota, c.numero_contrato, c.status, "
            "p.nome_completo AS devedor, p.cpf_cnpj AS devedor_cpf_cnpj "
            "FROM contrato c "
            "LEFT JOIN pessoa p ON c.id_pessoa = p.id "
        )
        sql += join_x or ''
        # Sem LIMIT por chunk: o LIMIT antigo cortava o universo (~501) antes de varrer o cohort inteiro.
        sql += f" WHERE c.id IN ({ph})" + (f" {and_busca}" if and_busca else "") + " ORDER BY c.grupo, c.cota"
        params = list(part) + p_busca
        try:
            cursor.execute(sql, tuple(params))
        except Exception:
            app.logger.exception('panel_perf_contratos')
            break
        for row in _clean_rows(cursor.fetchall()):
            cid = int(row['id'])
            meta = by_id.get(cid)
            if not meta:
                continue
            out.append({
                'id': cid,
                'grupo': row.get('grupo'),
                'cota': row.get('cota'),
                'numero_contrato': row.get('numero_contrato'),
                'status': row.get('status'),
                'devedor': row.get('devedor'),
                'devedor_cpf_cnpj': row.get('devedor_cpf_cnpj'),
                'faixa_calendario': meta.get('faixa_calendario'),
                'desempenho': meta.get('desempenho'),
                'prazo_atraso': meta.get('prazo_atraso'),
            })
    out.sort(key=lambda r: (str(r.get('grupo') or ''), str(r.get('cota') or '')))
    total = len(out)
    limited = total > _PAINEL_DASH_LIM
    return out[:_PAINEL_DASH_LIM], total, limited


@app.route('/api/performance/panel_contratos')
def api_performance_panel_contratos():
    mes = str(request.args.get('mes', '')).strip()
    if not mes or len(mes) < 7:
        return jsonify({'error': 'mes obrigatorio (YYYY-MM)'}), 400
    try:
        y, m = int(mes[:4]), int(mes[5:7])
        datetime.date(y, m, 1)
    except ValueError:
        return jsonify({'error': 'mes invalido.'}), 400
    raw_idx = request.args.get('safra_index', 'all')
    if raw_idx in (None, '', 'all'):
        safra_index = None
    else:
        try:
            safra_index = int(raw_idx)
            if safra_index < 0 or safra_index > 3:
                safra_index = None
        except (TypeError, ValueError):
            safra_index = None
    try:
        atraso_teto = int(request.args.get('atraso_teto', 90) or 90)
    except (TypeError, ValueError):
        atraso_teto = 90
    if atraso_teto not in (30, 60, 90):
        atraso_teto = 90

    conn = _get_db()
    cursor = conn.cursor()
    try:
        rows, total, limited = _panel_perf_contratos_rows(
            cursor,
            y,
            m,
            safra_index,
            atraso_teto,
            request.args.get('tipo', 'contrato'),
            request.args.get('termo', ''),
            request.args.get('status', ''),
        )
    finally:
        cursor.close()
        conn.close()
    return jsonify({
        'results': rows,
        'total': total,
        'limited': limited,
        'limit': _PAINEL_DASH_LIM,
    })


# ---------------------------------------------------------------------------
# API: Cobrança — Contratos agrupados por prioridade de parcelas
# ---------------------------------------------------------------------------

def _get_data_referencia_arquivos_gm(cursor):
    """Ultima data de arquivo GM importada. Usada em DATEDIFF de atraso para
    alinhar o painel ao tracker, ao relatorio e ao script de distribuicao
    (nao usar CURDATE() isolado, que desloca a janela 30-90 dias se o
    relogio nao bate com o fim de semana/feriado)."""
    try:
        cursor.execute("SELECT MAX(data_arquivo) AS dr FROM arquivos_gm")
        row = cursor.fetchone()
        if not row or row.get("dr") is None:
            return datetime.date.today()
        dr = row["dr"]
        if isinstance(dr, datetime.datetime):
            return dr.date()
        if isinstance(dr, datetime.date):
            return dr
        s = str(dr)[:10]
        return datetime.date.fromisoformat(s)
    except Exception:
        return datetime.date.today()


def _ensure_cobranca_table(cursor):
    """Cria a tabela `cobranca` (snapshot diario de id_contrato x data do GM) se ainda nao existir."""
    cursor.execute(
        """
        CREATE TABLE IF NOT EXISTS cobranca (
            id BIGINT(20) NOT NULL AUTO_INCREMENT,
            id_contrato BIGINT(20) NOT NULL,
            data_arquivo DATE NOT NULL,
            PRIMARY KEY (id),
            UNIQUE KEY cobranca_id_contrato_IDX (id_contrato, data_arquivo) USING BTREE,
            KEY cobranca_arquivos_gm_fk (data_arquivo),
            CONSTRAINT cobranca_contrato_fk FOREIGN KEY (id_contrato)
                REFERENCES contrato(id) ON DELETE CASCADE
        ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
        """
    )


def _em_cobranca_count_por_data_ref(cursor, data_ref):
    """Conta registros em `cobranca` para a data (alinha a Em Cobranca / snapshot do painel)."""
    if data_ref is None:
        data_ref = _get_data_referencia_arquivos_gm(cursor)
    if isinstance(data_ref, datetime.datetime):
        d = data_ref.date()
    elif isinstance(data_ref, datetime.date):
        d = data_ref
    else:
        s = str(data_ref)[:10]
        try:
            d = datetime.date.fromisoformat(s)
        except ValueError:
            return 0
    cursor.execute(
        "SELECT COUNT(*) AS n FROM cobranca WHERE data_arquivo = %s",
        (d.isoformat(),),
    )
    row = cursor.fetchone()
    return int(row.get('n') or 0) if row else 0


def _parse_data_final_relatorio_abertos(data_final):
    """Converte a Data Final do relatorio (YYYY-MM-DD) em date — obrigatoria para tipo abertos."""
    s = str(data_final).strip()[:10]
    if len(s) != 10:
        raise ValueError('Data Final invalida')
    return datetime.date.fromisoformat(s)


def _relatorio_data_ref_abertos(cursor, data_final):
    """Data de snapshot para textos de export. Preferir Data Final; fallback so ultima GM se vazio."""
    if data_final:
        s = str(data_final).strip()[:10]
        if len(s) == 10:
            try:
                return datetime.date.fromisoformat(s)
            except ValueError:
                pass
    return _get_data_referencia_arquivos_gm(cursor)


def _build_relatorio_query_abertos(data_ref, prioridade=None):
    """Contratos abertos: 1 linha por linha de `cobranca` na data (id_contrato, data_arquivo).

    Nao se usa INNER JOIN em parcela aberto para nao excluir contratos que
    constavam no snapshot daquela `data_arquivo` mas cujas parcelas hoje estao
    fechadas (bater count(*) em cobranca = COUNT no relatorio "Todas").

    Vencimento/dias de atraso: estado *atual* das parcelas abertas (ou NULL se nao houver).
    """
    if isinstance(data_ref, datetime.datetime):
        d = data_ref.date()
    elif isinstance(data_ref, datetime.date):
        d = data_ref
    else:
        d = datetime.date.fromisoformat(str(data_ref)[:10])
    s = d.isoformat()

    sql = (
        "SELECT c.id, c.grupo, c.cota, "
        "       CONCAT_WS('/', c.grupo, c.cota) AS grupo_cota, "
        "       c.numero_contrato, c.status, "
        "       p.nome_completo AS nome_devedor, p.cpf_cnpj, "
        "       snap.data_arquivo AS data_arquivo, "
        "       parc.min_v AS vencimento_mais_antigo, "
        "       (CASE WHEN parc.min_v IS NOT NULL "
        "             THEN DATEDIFF(snap.data_arquivo, parc.min_v) END) AS dias_atraso, "
        "       (CASE WHEN parc.min_v IS NOT NULL "
        "             THEN DATEDIFF(CURRENT_DATE, parc.min_v) END) AS atraso "
        "FROM contrato c "
        "INNER JOIN ( "
        "  SELECT id, id_contrato, data_arquivo "
        "  FROM cobranca "
        "  WHERE data_arquivo = %s "
        ") AS snap ON snap.id_contrato = c.id "
        "LEFT JOIN pessoa p ON c.id_pessoa = p.id "
        "LEFT JOIN ( "
        "  SELECT id_contrato, MIN(vencimento) AS min_v "
        "  FROM parcela "
        "  WHERE status = 'cobranca' "
        "  GROUP BY id_contrato "
        ") parc ON parc.id_contrato = c.id "
        "WHERE 1=1 "
    )
    params = [s]

    if prioridade == 'critico':
        sql += (
            "AND parc.min_v IS NOT NULL "
            "AND DATEDIFF(snap.data_arquivo, parc.min_v) >= 60 "
        )
    elif prioridade == 'atencao':
        sql += (
            "AND parc.min_v IS NOT NULL "
            "AND DATEDIFF(snap.data_arquivo, parc.min_v) >= 30 "
            "AND DATEDIFF(snap.data_arquivo, parc.min_v) < 60 "
        )
    elif prioridade == 'recente':
        sql += (
            "AND parc.min_v IS NOT NULL "
            "AND DATEDIFF(snap.data_arquivo, parc.min_v) >= 1 "
            "AND DATEDIFF(snap.data_arquivo, parc.min_v) < 30 "
        )

    # Contratos sem parcela aberta hoje: NULL em dias; ordena por atraso depois por grupo
    sql += (
        "ORDER BY (parc.min_v IS NULL) ASC, "
        "         DATEDIFF(snap.data_arquivo, parc.min_v) DESC, "
        "         c.grupo, c.cota"
    )
    return sql, params


@app.route('/api/cobranca')
def api_cobranca():
    """Retorna contratos abertos divididos em 3 faixas de prioridade
    baseadas no vencimento mais antigo de parcelas em aberto.

    Universo: snapshot `cobranca` na data do ultimo GM (`data_arquivo`).
    dias_atraso: DATEDIFF(CURRENT_DATE, vencimento mais antigo em aberto).

    O filtro de "operador" agora usa a tabela `funcionario_cobranca`
    (id_funcionario + id_contrato). Aceita tanto `funcionario_id`
    (numerico, preferencial) quanto `operador` (string, retrocompat).
    """
    funcionario_id_raw = request.args.get('funcionario_id', '').strip()
    operador_nome = request.args.get('operador', '').strip()

    funcionario_id = None
    if funcionario_id_raw:
        try:
            funcionario_id = int(funcionario_id_raw)
        except ValueError:
            funcionario_id = None

    # Filtro opcional: contratos com parcela aberta "antiga" e outra paga
    # com número maior (parcelas fora de ordem).
    pdes_raw = request.args.get("parcelas_desordenadas", "").strip().lower()
    parcelas_desordenadas = pdes_raw in ("1", "true", "yes", "sim", "on")

    conn = _get_db()
    cursor = conn.cursor()

    # Garante que a tabela de negativacao existe antes do SELECT que a
    # referencia (evita erro quando ainda e a primeira execucao).
    try:
        _ensure_negativacao_table(cursor)
    except Exception:
        # Se a criacao falhar (permissao, schema bloqueado, etc.), seguimos
        # sem o flag de negativacao em vez de derrubar a tela inteira.
        app.logger.exception('api_cobranca: falha ao garantir tabela negativacao')
    try:
        _ensure_cobranca_table(cursor)
    except Exception:
        app.logger.exception('api_cobranca: falha ao garantir tabela cobranca')

    # Query base: contratos do snapshot `cobranca` (ultimo dia GM) com parcelas
    # em aberto, com join
    # opcional em funcionario_cobranca/funcionario para expor o
    # responsavel de cobranca e (se a tabela existir) com a descricao
    # do bem associado ao contrato.
    # Universo filtrado por cob.data_arquivo = ultimo GM; dias_atraso = dias ate hoje
    # (CURRENT_DATE vs vencimento mais antigo em aberto), alinhado a SMS/automacao.
    data_ref = _get_data_referencia_arquivos_gm(cursor)

    bem_info = _get_bem_schema()
    bem_select = "NULL"
    bem_join = ""
    if bem_info['table'] and bem_info['join_on']:
        bem_concat = _bem_concat_expr('b')
        bem_select = f"GROUP_CONCAT(DISTINCT {bem_concat} SEPARATOR ' | ')"
        if bem_info['join_on'] == 'id_contrato':
            bem_join = f"LEFT JOIN {bem_info['table']} b ON b.id_contrato = c.id "
        else:
            bem_join = (
                f"LEFT JOIN {bem_info['table']} b "
                "ON b.grupo = c.grupo AND b.cota = c.cota "
            )

    # A parcela ALVO e a parcela em aberto mais antiga do contrato. E ela
    # que sera (ou ja foi) negativada. O subquery correlacionado abaixo
    # devolve o ID dessa parcela por contrato (ordena por vencimento e
    # id para ser deterministico).
    base_sql = (
        "SELECT c.id, c.grupo, c.cota, c.numero_contrato, c.status, "
        "       p.nome_completo AS nome_devedor, p.cpf_cnpj, "
        "       c.valor_credito, "
        "       MIN(par.vencimento) AS vencimento_mais_antigo, "
        "       DATEDIFF(CURRENT_DATE, MIN(par.vencimento)) AS dias_atraso, "
        # DISTINCT: o JOIN com `bens` pode duplicar linhas (varios bens no mesmo
        # contrato); o numero de parcelas abertas e por id de parcela, nao por linha.
        "       COUNT(DISTINCT par.id) AS parcelas_abertas, "
        "       fc.id_funcionario AS id_funcionario, "
        "       f.nome AS nome_funcionario, "
        "       ( "
        "         SELECT p2.id FROM parcela p2 "
        "         WHERE p2.id_contrato = c.id AND p2.status = 'cobranca' "
        "         ORDER BY p2.vencimento ASC, p2.id ASC LIMIT 1 "
        "       ) AS id_parcela_alvo, "
        "       ( "
        "         SELECT p3.numero_parcela FROM parcela p3 "
        "         WHERE p3.id_contrato = c.id AND p3.status = 'cobranca' "
        "         ORDER BY p3.vencimento ASC, p3.id ASC LIMIT 1 "
        "       ) AS numero_parcela_alvo, "
        f"       {bem_select} AS bem_descricao "
        "FROM contrato c "
        "INNER JOIN cobranca cob ON cob.id_contrato = c.id AND cob.data_arquivo = %s "
        "INNER JOIN parcela par ON par.id_contrato = c.id AND par.status = 'cobranca' "
        "LEFT JOIN pessoa p ON c.id_pessoa = p.id "
        "LEFT JOIN funcionario_cobranca fc ON fc.id_contrato = c.id "
        "LEFT JOIN funcionario f ON f.id = fc.id_funcionario "
        + bem_join +
        "WHERE c.status = 'cobranca' "
    )
    params = [data_ref]

    if funcionario_id is not None:
        base_sql += " AND fc.id_funcionario = %s "
        params.append(funcionario_id)
    elif operador_nome:
        base_sql += " AND f.nome = %s "
        params.append(operador_nome)

    if parcelas_desordenadas:
        # Mesma regra de universo do painel: último snapshot GM em `cobranca`.
        base_sql += (
            "AND c.id IN ("
            "SELECT DISTINCT p_aberta.id_contrato "
            "FROM parcela p_aberta "
            "INNER JOIN contrato con ON p_aberta.id_contrato = con.id "
            "INNER JOIN parcela p_fechada ON p_aberta.id_contrato = p_fechada.id_contrato "
            "INNER JOIN cobranca cob_d ON cob_d.id_contrato = con.id AND cob_d.data_arquivo = %s "
            "WHERE p_aberta.status = 'cobranca' "
            "  AND p_fechada.status = 'pago' "
            "  AND p_fechada.numero_parcela > p_aberta.numero_parcela "
            "  AND con.status = 'cobranca'"
            ") "
        )
        params.append(data_ref)

    base_sql += (
        "GROUP BY c.id, c.grupo, c.cota, c.numero_contrato, c.status, "
        "         p.nome_completo, p.cpf_cnpj, c.valor_credito, "
        "         fc.id_funcionario, f.nome "
        "ORDER BY dias_atraso DESC"
    )

    cursor.execute(base_sql, params)
    rows = _clean_rows(cursor.fetchall())

    # -------------------------------------------------------------------------
    # Classificacao de negativacao (3 estados):
    #
    #   'negativado'    -> a parcela-alvo ja consta em `negativacao` como
    #                      id_parcela, OU ha registro legado (id_parcela NULL
    #                      de versao antiga do schema) para o id_contrato.
    #   'pendente'      -> 30 < dias_atraso < 90, parcela-alvo ainda nao
    #                      consta em `negativacao` (inclui 1 so parcela em
    #                      aberto, se a janela bater) — "Falta negativar".
    #   'nao_elegivel'  -> fora da janela 31-89, ou ainda sem criterio Serasa
    #                      p/ envio. Nao e o "Ativo" do filtro: ver front.
    #
    # NOTA: a versao anterior consultava `WHERE id_parcela IN (parcelas alvo
    # desta resposta)` e descartava linhas de `negativacao` cujo id_parcela
    # nao estava nessa lista (p.ex. registro legado, ou o IN nao trazia o
    # id exato). Agora buscamos por id_contrato (completo) e comparamos no
    # Python com a parcela-alvo atual.
    #
    # `negativado` (boolean) e mantido apenas para backcompat do front.
    # -------------------------------------------------------------------------
    cids = []
    for r in rows:
        try:
            cids.append(int(r['id']))
        except (TypeError, ValueError, KeyError):
            pass

    try:
        neg_por_contrato = _mapa_parcelas_negativadas_por_contrato(cursor, cids)
    except Exception:
        app.logger.exception('api_cobranca: falha ao consultar negativacao')
        neg_por_contrato = {}

    for r in rows:
        try:
            cid = int(r['id'])
        except (TypeError, ValueError, KeyError):
            cid = None

        try:
            dias_i = int(r.get('dias_atraso') or 0)
        except (TypeError, ValueError):
            dias_i = 0
        try:
            qtd_abertas = int(r.get('parcelas_abertas') or 0)
        except (TypeError, ValueError):
            qtd_abertas = 0

        id_alvo_raw = r.get('id_parcela_alvo')
        id_alvo = None
        if id_alvo_raw is not None:
            try:
                id_alvo = int(id_alvo_raw)
            except (TypeError, ValueError):
                id_alvo = None

        neg_set = neg_por_contrato.get(cid, set()) if cid is not None else set()
        # Historico: negativacao sem id_parcela = contrato ja foi tratado
        # no modelo antigo. Parcela-alvo ja registrada = negativado atual.
        ja_neg = (_NEG_NULL_PARCELA in neg_set) or (
            id_alvo is not None and id_alvo in neg_set
        )
        # Janela Serasa: 31-89 dias na parcela-alvo (1 ou N parcelas abertas).
        elegivel = 30 < dias_i < 90

        if ja_neg:
            status = 'negativado'
        elif elegivel:
            status = 'pendente'
        else:
            status = 'nao_elegivel'

        r['status_negativacao'] = status
        r['elegivel_negativacao'] = elegivel
        r['negativado'] = ja_neg  # backcompat

    # Dividir em 3 blocos de prioridade
    critico = []   # 60-90+ dias
    atencao = []   # 30-60 dias
    recente = []   # 1-30 dias

    for r in rows:
        dias = r.get('dias_atraso') or 0
        if isinstance(dias, str):
            dias = int(float(dias))
        if dias >= 60:
            critico.append(r)
        elif dias >= 30:
            atencao.append(r)
        elif dias >= 1:
            recente.append(r)
        # dias <= 0 => parcela ainda nao venceu, nao entra nos blocos

    # Lista de operadores do filtro: somente perfil Cobrança (mesma regra da importação).
    try:
        cursor.execute(
            "SELECT id, nome FROM funcionario WHERE " + _WHERE_FUNCIONARIO_COBRANCA + " ORDER BY nome"
        )
        funcionarios = [
            {'id': int(r['id']), 'nome': r['nome']}
            for r in cursor.fetchall()
        ]
    except Exception:
        funcionarios = []

    cursor.close()
    conn.close()

    return jsonify({
        'critico': critico,
        'atencao': atencao,
        'recente': recente,
        'total': len(critico) + len(atencao) + len(recente),
        'funcionarios': funcionarios,
        # Mantem a chave antiga por compatibilidade (apenas nomes).
        'operadores': [f['nome'] for f in funcionarios],
    })


# ---------------------------------------------------------------------------
# API: Consorciados e Avalistas
# ---------------------------------------------------------------------------

def _build_pessoa_query(join_field, q, tipo):
    """Helper para construir query de listagem de pessoas vinculadas a contratos."""
    sql = (
        "SELECT p.id, p.nome_completo, p.cpf_cnpj, "
        "       COUNT(DISTINCT c.id) AS total_contratos "
        "FROM pessoa p "
        "INNER JOIN contrato c ON c." + join_field + " = p.id "
    )
    params = []
    where_parts = []

    if q and tipo:
        if tipo == 'nome':
            where_parts.append("p.nome_completo LIKE %s")
            params.append('%' + q + '%')
        elif tipo == 'cpf':
            where_parts.append("p.cpf_cnpj LIKE %s")
            params.append('%' + q + '%')
        elif tipo == 'grupo_cota':
            where_parts.append("(c.grupo LIKE %s OR c.cota LIKE %s OR CONCAT(c.grupo, '/', c.cota) LIKE %s)")
            params.extend(['%' + q + '%', '%' + q + '%', '%' + q + '%'])
        elif tipo == 'bem':
            bem_join = _bem_join_clause('c', 'b')
            bem_where, bem_params = _bem_where_clause(q, 'b')
            if bem_join and bem_where:
                sql += bem_join
                where_parts.append(bem_where)
                params.extend(bem_params)
            else:
                # Tabela de bens indisponivel -> nenhum resultado.
                where_parts.append("1 = 0")
    elif q:
        where_parts.append("(p.nome_completo LIKE %s OR p.cpf_cnpj LIKE %s)")
        params.extend(['%' + q + '%', '%' + q + '%'])

    if where_parts:
        sql += "WHERE " + " AND ".join(where_parts) + " "

    sql += "GROUP BY p.id, p.nome_completo, p.cpf_cnpj "
    sql += "ORDER BY p.nome_completo "

    return sql, params


@app.route('/api/consorciados')
def api_consorciados():
    """Lista pessoas (devedores) vinculadas a contratos via id_pessoa."""
    q = request.args.get('q', '').strip()
    tipo = request.args.get('tipo', '').strip()

    sql, params = _build_pessoa_query('id_pessoa', q, tipo)

    # Sem busca = limite 10 (preview); com busca = todos os resultados
    if not q:
        sql += "LIMIT 10"

    conn = _get_db()
    cursor = conn.cursor()
    cursor.execute(sql, params)
    rows = _clean_rows(cursor.fetchall())
    cursor.close()
    conn.close()

    return jsonify({'results': rows, 'total': len(rows)})


@app.route('/api/avalistas')
def api_avalistas():
    """Lista pessoas (avalistas) vinculadas a contratos via id_avalista."""
    q = request.args.get('q', '').strip()
    tipo = request.args.get('tipo', '').strip()

    sql, params = _build_pessoa_query('id_avalista', q, tipo)

    # Sem busca = limite 10 (preview); com busca = todos os resultados
    if not q:
        sql += "LIMIT 10"

    conn = _get_db()
    cursor = conn.cursor()
    cursor.execute(sql, params)
    rows = _clean_rows(cursor.fetchall())
    cursor.close()
    conn.close()

    return jsonify({'results': rows, 'total': len(rows)})

@app.route('/api/operadores/dashboard')
def api_operadores_dashboard():
    """Retorna estatisticas e lista de contratos agrupados por operador.

    Contratos considerados: snapshot `cobranca` na data do ultimo `arquivos_gm`
    (mesma base do KPI "Em cobranca" do Dashboard e do modulo Cobranca), cruzado
    com `funcionario_cobranca`. dias_atraso = dias ate hoje (CURRENT_DATE vs
    vencimento mais antigo em aberto), igual a /api/cobranca.
    """
    forbidden = _admin_json_forbidden()
    if forbidden:
        return forbidden
    conn = _get_db()
    cursor = conn.cursor()
    try:
        _ensure_cobranca_table(cursor)
    except Exception:
        app.logger.exception('api_operadores_dashboard: falha ao garantir tabela cobranca')
    data_ref = _get_data_referencia_arquivos_gm(cursor)
    if isinstance(data_ref, datetime.datetime):
        dref = data_ref.date()
    elif isinstance(data_ref, datetime.date):
        dref = data_ref
    else:
        dref = datetime.date.fromisoformat(str(data_ref)[:10])
    s_ref = dref.isoformat()

    # 1) Lista todos os funcionarios (para que operadores sem contratos ainda
    #    apareçam no <select> de filtros e no bloco do dashboard).
    cursor.execute(
        """
        SELECT id, nome, COALESCE(ativo, 1) AS ativo
        FROM funcionario
        ORDER BY nome
        """
    )
    funcionarios = _clean_rows(cursor.fetchall())

    def _status_from_ativo(val):
        try:
            return 'ativo' if int(val or 0) == 1 else 'inativo'
        except (TypeError, ValueError):
            return 'ativo'

    operators_map = {}
    for f in funcionarios:
        fid = int(f['id'])
        operators_map[fid] = {
            'id': fid,
            'nome': f['nome'],
            'status_operador': _status_from_ativo(f.get('ativo')),
            'contratos': [],
            'stats': {'total': 0, 'critico': 0, 'atencao': 0, 'recente': 0},
        }

    # 2) Mesma logica de /api/cobranca: snapshot na data do ultimo GM;
    #    dias_atraso ate CURRENT_DATE (nao data_GM).
    cursor.execute(
        """
        SELECT fc.id_funcionario,
               c.id, c.grupo, c.cota, c.numero_contrato, c.valor_credito,
               p.nome_completo AS nome_devedor, p.cpf_cnpj,
               (SELECT DATEDIFF(CURRENT_DATE, MIN(vencimento))
                  FROM parcela
                 WHERE id_contrato = c.id AND status = 'cobranca') AS dias_atraso,
               (SELECT COUNT(*)
                  FROM parcela
                 WHERE id_contrato = c.id AND status = 'cobranca') AS parcelas_abertas,
               (SELECT MIN(vencimento)
                  FROM parcela
                 WHERE id_contrato = c.id AND status = 'cobranca') AS vencimento_mais_antigo
        FROM funcionario_cobranca fc
        INNER JOIN contrato c ON c.id = fc.id_contrato
        INNER JOIN cobranca cob ON cob.id_contrato = c.id AND cob.data_arquivo = %s
        LEFT JOIN pessoa p    ON p.id = c.id_pessoa
        WHERE c.status = 'cobranca'
        ORDER BY fc.id_funcionario, c.valor_credito DESC
        """,
        (s_ref,),
    )
    rows = _clean_rows(cursor.fetchall())

    kpis = {'total': 0, 'critico': 0, 'atencao': 0, 'recente': 0}

    for r in rows:
        fid = int(r['id_funcionario'])
        if fid not in operators_map:
            # Funcionario que nao esta mais na tabela `funcionario` mas ainda
            # tem vinculo na tabela `funcionario_cobranca` (edge case).
            operators_map[fid] = {
                'id': fid,
                'nome': f'#{fid}',
                'status_operador': 'inativo',
                'contratos': [],
                'stats': {'total': 0, 'critico': 0, 'atencao': 0, 'recente': 0},
            }

        situacao = _situacao_from_dias(r.get('dias_atraso'))
        r['situacao'] = situacao  # o frontend usa essa chave para pintar badges
        # Normaliza campos exibidos pelo frontend
        if r.get('dias_atraso') is None:
            r['dias_atraso'] = 0
        if r.get('parcelas_abertas') is None:
            r['parcelas_abertas'] = 0

        bucket = operators_map[fid]
        bucket['contratos'].append(r)
        bucket['stats']['total'] += 1
        kpis['total'] += 1

        sk = _sit_key(situacao)
        if sk in ('critico', 'atencao', 'recente'):
            bucket['stats'][sk] += 1
            kpis[sk] += 1

    cursor.close()
    conn.close()

    # Preserva ordem alfabetica por nome.
    operadores = sorted(operators_map.values(), key=lambda x: (x['nome'] or '').lower())

    return jsonify({
        'operadores': operadores,
        'kpis': kpis,
    })


# ---------------------------------------------------------------------------
# API: Performance (faixas de calendario x desempenho)
# ---------------------------------------------------------------------------

_FAIXA_NOMES = [
    'Faixa 5 (01 - 09)',
    'Faixa 10 (10 - 12)',
    'Faixa 15 (13 - 19)',
    'Faixa 20 (20 - fim)',
]


def _seg_from_delay_days(dd):
    """Recente, Atencao, Critico a partir do atraso (dias) da parcela aberta mais antiga."""
    if dd is None or dd < 1:
        return 'recente'
    if dd <= 30:
        return 'recente'
    if dd <= 60:
        return 'atencao'
    return 'critico'


def _nao_b_from_open_delay(dd):
    """Nao performado: faixas de atraso da parcela aberta (alinhado ao teto 30/60/90 d no grafico)."""
    if dd is None or dd < 1:
        return 'b30'
    d = int(dd)
    if d <= 30:
        return 'b30'
    if d <= 60:
        return 'b60'
    if d <= 90:
        return 'b90'
    return 'bplus'


_SAFRA_ENTRADA_SQL = """
SELECT
    COALESCE(c.id, o.id_contrato) AS id_contrato,
    c.valor_credito AS valor_credito,
    o.data_arquivo AS dt_entrada,
    -- vencimento de referencia: performado -> p_antiga; nao performado -> menor entre aberta/indenizada
    CASE
        WHEN p_antiga.id_contrato IS NOT NULL THEN p_antiga.vencimento
        WHEN p_aberta.id_contrato IS NOT NULL
             AND (p_inden.id_contrato IS NULL OR p_aberta.vencimento <= p_inden.vencimento)
            THEN p_aberta.vencimento
        WHEN p_inden.id_contrato IS NOT NULL THEN p_inden.vencimento
        ELSE NULL
    END AS vencimento,
    p_antiga.data_pagamento AS dt_paga,
    -- VALOR: performado -> parcela paga mais antiga; nao performado -> parcela mais antiga entre aberta e indenizada
    CASE
        WHEN p_antiga.id_contrato IS NOT NULL THEN p_antiga.valor_total
        WHEN p_aberta.id_contrato IS NOT NULL
             AND (p_inden.id_contrato IS NULL OR p_aberta.vencimento <= p_inden.vencimento)
            THEN p_aberta.valor_total
        WHEN p_inden.id_contrato IS NOT NULL THEN p_inden.valor_total
        ELSE NULL
    END AS valor_parcela,
    -- CLASSIFICACAO original: contrato com qualquer parcela paga -> performado
    CASE
        WHEN p_antiga.data_pagamento IS NULL THEN 'nao_performado'
        WHEN DATEDIFF(p_antiga.data_pagamento, p_antiga.vencimento) <= 30 THEN 'd30'
        WHEN DATEDIFF(p_antiga.data_pagamento, p_antiga.vencimento) <= 60 THEN 'd60'
        WHEN DATEDIFF(p_antiga.data_pagamento, p_antiga.vencimento) <= 90 THEN 'd90'
        ELSE 'd90'
    END AS recovery_group,
    CASE WHEN p_antiga.id_contrato IS NOT NULL THEN 1 ELSE 0 END AS is_performado,
    -- delay_open mantido identico ao original (so olha parcelas em aberto).
    -- Contratos com parcela indenizada (sem aberta) ficam com delay_open NULL,
    -- caindo em b30 e entrando na contagem do teto <= 90 d (igual antes).
    CASE
        WHEN p_antiga.id_contrato IS NOT NULL THEN NULL
        ELSE (
            SELECT DATEDIFF(CURDATE(), MIN(p2.vencimento))
            FROM parcela p2
            WHERE p2.id_contrato = c.id AND p2.status = 'cobranca'
        )
    END AS delay_open
FROM (
    SELECT id_contrato, MIN(data_arquivo) AS data_arquivo
    FROM ocorrencia
    WHERE status = 'cobranca'
      AND (descricao = 'contrato novo' OR LOWER(descricao) LIKE 'contrato voltou%%')
      AND data_arquivo >= %s AND data_arquivo <= %s
    GROUP BY id_contrato
) o
LEFT JOIN contrato c ON c.id = o.id_contrato
LEFT JOIN (
    SELECT p1.id_contrato, p1.vencimento, p1.data_pagamento, p1.valor_total
    FROM parcela p1
    INNER JOIN (
        SELECT id_contrato, MIN(vencimento) AS min_vencimento
        FROM parcela
        WHERE status = 'pago'
        GROUP BY id_contrato
    ) p2 ON p1.id_contrato = p2.id_contrato AND p1.vencimento = p2.min_vencimento
    WHERE p1.status = 'pago'
) p_antiga ON p_antiga.id_contrato = c.id
LEFT JOIN (
    SELECT p1.id_contrato, p1.vencimento, p1.valor_total
    FROM parcela p1
    INNER JOIN (
        SELECT id_contrato, MIN(vencimento) AS min_vencimento
        FROM parcela
        WHERE status = 'cobranca'
        GROUP BY id_contrato
    ) pm ON p1.id_contrato = pm.id_contrato AND p1.vencimento = pm.min_vencimento AND p1.status = 'cobranca'
) p_aberta ON p_aberta.id_contrato = c.id
LEFT JOIN (
    SELECT p1.id_contrato, p1.vencimento, p1.valor_total
    FROM parcela p1
    INNER JOIN (
        SELECT id_contrato, MIN(vencimento) AS min_vencimento
        FROM parcela
        WHERE status = 'indenizado'
        GROUP BY id_contrato
    ) pm ON p1.id_contrato = pm.id_contrato AND p1.vencimento = pm.min_vencimento AND p1.status = 'indenizado'
) p_inden ON p_inden.id_contrato = c.id
"""


def _safra_entrada_rows(cursor, d_a, d_b, y, m):
    """Uma linha por (contrato, arquivo) da safra: entrada novo/voltou + exatamente 1 parcela vencida no arquivo."""
    cursor.execute(
        _SAFRA_ENTRADA_SQL,
        (d_a.isoformat(), d_b.isoformat()),
    )
    return cursor.fetchall()


def _valor_metrica_performance_brl(r) -> float:
    """R$ na Performance/Dashboard: valor_total da parcela mais antiga do contrato.

    Esse valor representa exatamente:
      - performado: a parcela paga (que passou de aberto -> fechado);
      - nao performado: a parcela em aberto que originou a classificacao.
    """
    v = r.get('valor_parcela')
    try:
        return float(v) if v is not None and v != '' else 0.0
    except (TypeError, ValueError):
        return 0.0


def _valor_credito_contrato_brl(r) -> float:
    """Soma do valor_credito do contrato na linha da safra (export / agregados)."""
    v = r.get('valor_credito')
    try:
        return float(v) if v is not None and v != '' else 0.0
    except (TypeError, ValueError):
        return 0.0


def _bool_sql(val):
    if isinstance(val, (int, float)):
        return int(val) != 0
    if isinstance(val, (bytes, bytearray)):
        return bool(val) and val not in (b'\x00', b'0')
    return bool(val)


def _decode_sql_str(v):
    if isinstance(v, (bytes, bytearray)):
        return v.decode('utf-8', errors='replace')
    if v is None:
        return ''
    return str(v).strip()


def _row_matches_atraso_teto(r, teto: int) -> bool:
    """Mesma regra do gráfico (sumPagoNao + keysPago/keysNao cumulativos) para 30/60/90 d."""
    is_p = _bool_sql(r.get('is_performado'))
    if is_p:
        rgk = _decode_sql_str(r.get('recovery_group'))
        if not rgk or rgk == 'nao_performado':
            return False
        if teto <= 30:
            return rgk == 'd30'
        if teto <= 60:
            return rgk in ('d30', 'd60')
        return rgk in ('d30', 'd60', 'd90')
    try:
        dd = r.get('delay_open')
        dd_i = int(dd) if dd is not None else None
    except (TypeError, ValueError):
        dd_i = None
    bk = _nao_b_from_open_delay(dd_i)
    if teto <= 30:
        return bk == 'b30'
    if teto <= 60:
        return bk in ('b30', 'b60')
    return bk in ('b30', 'b60', 'b90')


def _export_prazo_atraso_label(r) -> str:
    """Prazo em faixas 30/60/90 alinhado ao gráfico (quitação vs aberto)."""
    is_p = _bool_sql(r.get('is_performado'))
    if is_p:
        rgk = _decode_sql_str(r.get('recovery_group'))
        m = {
            'd30': 'até 30 dias (quitação)',
            'd60': 'até 60 dias (quitação)',
            'd90': 'até 90 dias (quitação)',
            'dplus': 'acima de 90 dias (quitação)',
        }
        return m.get(rgk, rgk or '—')
    try:
        dd = r.get('delay_open')
        dd_i = int(dd) if dd is not None else None
    except (TypeError, ValueError):
        dd_i = None
    bk = _nao_b_from_open_delay(dd_i)
    m = {
        'b30': 'até 30 dias (aberto)',
        'b60': 'até 60 dias (aberto)',
        'b90': 'até 90 dias (aberto)',
        'bplus': 'acima de 90 dias (aberto)',
    }
    return m.get(bk, bk or '—')


def _export_desempenho_label(r) -> str:
    return 'Performado' if _bool_sql(r.get('is_performado')) else 'Não performado'


def _cohort_enriched_rows_for_export(cursor, y, m, safra_index, teto: int):
    """Contratos do cohort (safra) após o mesmo filtro 30/60/90 d do gráfico empilhado."""
    out = []
    seen = set()
    partes = range(4) if safra_index is None else (safra_index,)
    for parte in partes:
        d1, d2 = _safra_bounds(y, m, parte)
        for r in _safra_entrada_rows(cursor, d1, d2, y, m):
            if not _row_matches_atraso_teto(r, teto):
                continue
            cid = r.get('id_contrato')
            if cid is None:
                continue
            try:
                cid = int(cid)
            except (TypeError, ValueError):
                continue
            if cid in seen:
                continue
            seen.add(cid)
            vt = _valor_metrica_performance_brl(r)
            out.append({
                'id_contrato': cid,
                'faixa_calendario': _SAFRA_LABELS_EXPORT[parte],
                'desempenho': _export_desempenho_label(r),
                'prazo_atraso': _export_prazo_atraso_label(r),
                'valor_parcela_entrada_brl': round(vt, 2),
            })
    return out


def _kpi_teto_cumulativo_distinct_global(cursor, y, m):
    """Totais distintos de id_contrato (e valor metrica parcela uma vez por id) com teto 30/60/90 d.

    O grafico soma, por barra, linhas de cohort (_aggregate_performance_faixa) — o mesmo
    contrato pode entrar em mais de uma faixa de calendario no mes, inflando a soma das
    4 faixas. A export e os KPI de visao geral usam 1 contrato = 1 linha (dedup global).
    """
    st = {
        30: {'p': set(), 'n': set(), 'v_p': {}, 'v_n': {}},
        60: {'p': set(), 'n': set(), 'v_p': {}, 'v_n': {}},
        90: {'p': set(), 'n': set(), 'v_p': {}, 'v_n': {}},
    }
    for parte in range(4):
        d1, d2 = _safra_bounds(y, m, parte)
        for r in _safra_entrada_rows(cursor, d1, d2, y, m):
            cid = r.get('id_contrato')
            if cid is None:
                continue
            try:
                cid = int(cid)
            except (TypeError, ValueError):
                continue
            is_p = _bool_sql(r.get('is_performado'))
            v = _valor_metrica_performance_brl(r)
            for teto in (30, 60, 90):
                if not _row_matches_atraso_teto(r, teto):
                    continue
                S = st[teto]
                if is_p:
                    S['p'].add(cid)
                    S['v_p'].setdefault(cid, v)
                else:
                    S['n'].add(cid)
                    S['v_n'].setdefault(cid, v)
    out = {}
    for teto in (30, 60, 90):
        S = st[teto]
        out[str(teto)] = {
            'n_performado': len(S['p']),
            'n_nao': len(S['n']),
            'v_performado': round(sum(S['v_p'].values()), 2),
            'v_nao': round(sum(S['v_n'].values()), 2),
        }
    return out


def _safra_recovery_diff_days(r):
    """Dias entre entrada na safra e quitacao (parcela paga); se so fechado sem dt_paga, usa hoje."""
    dt_ent = r.get('dt_entrada')
    dt_paga = r.get('dt_paga')
    if dt_ent is None:
        return None
    if hasattr(dt_ent, 'date'):
        dt_ent = dt_ent.date()
    if not isinstance(dt_ent, datetime.date):
        return None
    if dt_paga is not None and hasattr(dt_paga, 'date'):
        dt_paga = dt_paga.date()
    if isinstance(dt_paga, datetime.date):
        return (dt_paga - dt_ent).days
    return (datetime.date.today() - dt_ent).days


def _aggregate_performance_faixa(cursor, y, m, parte):
    d_m0a, d_m0b = _safra_bounds(y, m, parte)
    rows = _safra_entrada_rows(cursor, d_m0a, d_m0b, y, m)

    segs = {
        'performado': {'recente': 0, 'atencao': 0, 'critico': 0},
        'nao_performado': {'recente': 0, 'atencao': 0, 'critico': 0},
    }
    val = {
        'performado': {'recente': 0.0, 'atencao': 0.0, 'critico': 0.0},
        'nao_performado': {'recente': 0.0, 'atencao': 0.0, 'critico': 0.0},
    }
    pago = {'d30': 0, 'd60': 0, 'd90': 0, 'dplus': 0}
    pago_v = {k: 0.0 for k in pago}
    pago_cred = {k: 0.0 for k in pago}
    nab = {'b30': 0, 'b60': 0, 'b90': 0, 'bplus': 0}
    nab_v = {k: 0.0 for k in nab}
    nab_cred = {k: 0.0 for k in nab}
    vol_val = 0.0
    vol_cred = 0.0
    for r in rows:
        is_p = _bool_sql(r.get('is_performado'))
        dd = r.get('delay_open')
        seg = _seg_from_delay_days(int(dd) if dd is not None else None)
        keyg = 'performado' if is_p else 'nao_performado'
        v = _valor_metrica_performance_brl(r)
        cred = _valor_credito_contrato_brl(r)
        vol_val += v
        vol_cred += cred
        segs[keyg][seg] += 1
        val[keyg][seg] += v
        if is_p:
            rg = r.get('recovery_group')
            rgk = rg.decode('utf-8') if isinstance(rg, (bytes, bytearray)) else (str(rg) if rg is not None else '')
            if rgk in pago:
                pago[rgk] += 1
                pago_v[rgk] += v
                pago_cred[rgk] += cred
            else:
                pago['dplus'] += 1
                pago_v['dplus'] += v
                pago_cred['dplus'] += cred
        else:
            dd_i = int(dd) if dd is not None else None
            bk = _nao_b_from_open_delay(dd_i)
            nab[bk] += 1
            nab_v[bk] += v
            nab_cred[bk] += cred

    d30, d60, d90, dplus = pago['d30'], pago['d60'], pago['d90'], pago['dplus']
    v30, v60, v90, vplus = pago_v['d30'], pago_v['d60'], pago_v['d90'], pago_v['dplus']
    c30, c60, c90, cplus = pago_cred['d30'], pago_cred['d60'], pago_cred['d90'], pago_cred['dplus']

    return {
        'volume': len(rows),
        'volume_val': vol_val,
        'volume_cred': vol_cred,
        'segmentos': segs,
        'valor_por_segmento_brl': val,
        'performado_por_prazo_quitacao': dict(pago),
        'performado_por_prazo_quitacao_brl': dict(pago_v),
        'performado_por_prazo_quitacao_cred': dict(pago_cred),
        'nao_por_atraso_aberto': dict(nab),
        'nao_por_atraso_aberto_brl': dict(nab_v),
        'nao_por_atraso_aberto_cred': dict(nab_cred),
        'recovery_d30': d30,
        'recovery_v30': v30,
        'recovery_cred30': c30,
        'recovery_d60': d60,
        'recovery_v60': v60,
        'recovery_cred60': c60,
        'recovery_d90': d90,
        'recovery_v90': v90,
        'recovery_cred90': c90,
        'recovery_dplus': dplus,
        'recovery_vplus': vplus,
        'recovery_credplus': cplus,
    }


def _safra_bounds(year, month, parte):
    """parte 0..3: (1-9), (10-12), (13-19), (20-fim). Retorna (date_start, date_end)."""
    last = calendar.monthrange(year, month)[1]
    bounds = [(1, 9), (10, 12), (13, 19), (20, last)]
    d1, d2 = bounds[parte]
    return datetime.date(year, month, d1), datetime.date(year, month, d2)


def _shift_month(year, month, delta):
    idx = year * 12 + (month - 1) + delta
    return idx // 12, idx % 12 + 1


def _count_distinct_ocorrencias(cursor, date_ranges, status_filter=None, desc_novo_only=False):
    """date_ranges: lista de (d_start, d_end) inclusive. status_filter: None (todos), ou lista de status."""
    if not date_ranges:
        return 0
    parts = []
    params = []
    for d1, d2 in date_ranges:
        parts.append('(o.data_arquivo >= %s AND o.data_arquivo <= %s)')
        params.extend([d1.isoformat(), d2.isoformat()])
    where = '(' + ' OR '.join(parts) + ')'
    extra = ''
    if status_filter:
        ph = ','.join(['%s'] * len(status_filter))
        extra += f' AND o.status IN ({ph})'
        params.extend(status_filter)
    if desc_novo_only:
        extra += " AND o.status = 'cobranca' AND o.descricao LIKE '%%novo%%'"
    sql = f'SELECT COUNT(DISTINCT o.id_contrato) AS n FROM ocorrencia o WHERE {where}{extra}'
    cursor.execute(sql, params)
    row = cursor.fetchone()
    return int(row['n'] or 0)


def _daily_series(cursor, d1, d2):
    """Retorna series com UMA barra por dia entre d1 e d2 (inclusive).

    Labels vem no formato 'dd/mm' e as 3 series sao:
      - novos       : ocorrencias com descricao 'contrato novo' (status=aberto)
      - pagos       : ocorrencias com status 'pago total' ou 'pago parcial' (grafico dashboard)
      - indenizados : ocorrencias com status='indenizado'
    """
    if d2 < d1:
        return {'barLabels': [], 'novos': [], 'pagos': [], 'indenizados': []}

    labels = []
    day_keys = []
    cursor_d = d1
    while cursor_d <= d2:
        day_keys.append(cursor_d.isoformat())
        labels.append(cursor_d.strftime('%d/%m'))
        cursor_d = cursor_d + datetime.timedelta(days=1)

    # Para obter valores, precisamos cruzar com a logica da safra (parcela-alvo)
    # Pegamos todas as entradas da safra no periodo d1-d2
    sql_safra = _SAFRA_ENTRADA_SQL
    y_m0, m_m0 = d1.year, d1.month
    cursor.execute(sql_safra, (d1.isoformat(), d2.isoformat()))
    rows_cohort = cursor.fetchall()
    
    novos_map = {}
    novos_val_map = {}
    pagos_map = {}
    pagos_val_map = {}
    
    for r in rows_cohort:
        v = _valor_metrica_performance_brl(r)
        
        # Entrada (Novos)
        dt_e = r.get('dt_entrada')
        if dt_e:
            ke = str(dt_e)
            novos_map[ke] = novos_map.get(ke, 0) + 1
            novos_val_map[ke] = novos_val_map.get(ke, 0.0) + v
            
        # Pagamento (Performado)
        if _bool_sql(r.get('is_performado')):
            dt_p = r.get('dt_paga')
            # Se performado mas sem dt_paga (fechado no banco), 
            # para o grafico diario nao temos o dia exato. 
            # Mas geralmente tem dt_paga se foi via ocorrencia.
            if dt_p:
                kp = str(dt_p)
                pagos_map[kp] = pagos_map.get(kp, 0) + 1
                pagos_val_map[kp] = pagos_val_map.get(kp, 0.0) + v

    # Indenizados: ainda usamos a query de ocorrencia mas sem valor (ou valor estimado)
    # Ja que indenizados nao sao necessariamente da "safra performada" da parcela alvo.
    # Mas para manter consistencia, vamos buscar o valor do contrato/parcela se possivel.
    def _group_by_day_count(sql_where, params):
        cursor.execute(
            "SELECT DATE(o.data_arquivo) AS dia, "
            "       COUNT(DISTINCT o.id_contrato) AS n "
            "FROM ocorrencia o "
            "WHERE o.data_arquivo >= %s AND o.data_arquivo <= %s "
            + sql_where +
            " GROUP BY DATE(o.data_arquivo)",
            (d1.isoformat(), d2.isoformat()) + tuple(params),
        )
        return {str(r['dia']): int(r['n'] or 0) for r in cursor.fetchall()}

    inden_map = _group_by_day_count(" AND o.status = 'indenizado'", [])

    return {
        'barLabels': labels,
        'novos': [novos_map.get(k, 0) for k in day_keys],
        'novos_val': [round(novos_val_map.get(k, 0.0), 2) for k in day_keys],
        'pagos': [pagos_map.get(k, 0) for k in day_keys],
        'pagos_val': [round(pagos_val_map.get(k, 0.0), 2) for k in day_keys],
        'indenizados': [inden_map.get(k, 0) for k in day_keys],
        'indenizados_val': [0 for k in day_keys], # Indenizados valor nao mapeado aqui
    }


@app.route('/api/performance')
def api_performance():
    mes = request.args.get('mes', '').strip()
    if not mes or len(mes) < 7:
        return jsonify({'error': 'Parametro mes obrigatorio (YYYY-MM)'}), 400
    try:
        y, m = int(mes[:4]), int(mes[5:7])
        datetime.date(y, m, 1)
    except ValueError:
        return jsonify({'error': 'mes invalido'}), 400

    conn = _get_db()
    cursor = conn.cursor()

    # KPI: novos contratos no mes (qualquer dia)
    cursor.execute(
        "SELECT COUNT(DISTINCT o.id_contrato) AS n FROM ocorrencia o "
        "WHERE o.status = 'cobranca' AND o.descricao LIKE '%%novo%%' "
        "AND YEAR(o.data_arquivo) = %s AND MONTH(o.data_arquivo) = %s",
        (y, m),
    )
    kpi_novos = int(cursor.fetchone()['n'] or 0)

    # KPI: taxa recuperacao = performados (parcela-alvo paga/fechada) / cohort safra no mes
    last_m = calendar.monthrange(y, m)[1]
    d_m_start = datetime.date(y, m, 1)
    d_m_end = datetime.date(y, m, last_m)
    rows_safra_mes = _safra_entrada_rows(cursor, d_m_start, d_m_end, y, m)
    n_safra_mes = len(rows_safra_mes)
    n_perf_mes = sum(1 for r in rows_safra_mes if _bool_sql(r.get('is_performado')))
    kpi_rec_pct = round(100.0 * n_perf_mes / n_safra_mes, 1) if n_safra_mes else 0.0

    # KPI: safra em destaque (hoje)
    today = datetime.date.today()
    parte_hoje = None
    for p in range(4):
        d1, d2 = _safra_bounds(today.year, today.month, p)
        if d1 <= today <= d2:
            parte_hoje = p
            break
    kpi_faixa_label = _FAIXA_NOMES[parte_hoje] if parte_hoje is not None else '—'

    # KPI: parcelas criticas (aberto, vencimento ha mais de 90 dias)
    cursor.execute(
        "SELECT COUNT(*) AS n FROM parcela par "
        "INNER JOIN contrato c ON c.id = par.id_contrato "
        "WHERE par.status = 'cobranca' AND c.status = 'cobranca' "
        "AND par.vencimento < DATE_SUB(CURDATE(), INTERVAL 90 DAY)"
    )
    kpi_parcelas_crit = int(cursor.fetchone()['n'] or 0)

    safras_out = []
    bar_novos = []
    bar_pagos = []
    bar_indenizados = []
    ch_perf = {
        'performado': {'d30': [], 'd60': [], 'd90': [], 'dplus': []},
        'nao_performado': {'b30': [], 'b60': [], 'b90': [], 'bplus': []},
    }
    ch_val = {
        'performado': {'d30': [], 'd60': [], 'd90': [], 'dplus': []},
        'nao_performado': {'b30': [], 'b60': [], 'b90': [], 'bplus': []},
    }

    for parte in range(4):
        d_m0a, d_m0b = _safra_bounds(y, m, parte)
        ranges_m = [(d_m0a, d_m0b)]

        ag = _aggregate_performance_faixa(cursor, y, m, parte)
        for sk in ('d30', 'd60', 'd90', 'dplus'):
            ch_perf['performado'][sk].append(ag['performado_por_prazo_quitacao'][sk])
            ch_val['performado'][sk].append(round(ag['performado_por_prazo_quitacao_brl'][sk], 2))
        for sk in ('b30', 'b60', 'b90', 'bplus'):
            ch_perf['nao_performado'][sk].append(ag['nao_por_atraso_aberto'][sk])
            ch_val['nao_performado'][sk].append(round(ag['nao_por_atraso_aberto_brl'][sk], 2))

        volume = ag['volume']
        d30 = ag['recovery_d30']
        d60 = ag['recovery_d60']
        d90 = ag['recovery_d90']
        dplus = ag['recovery_dplus']

        detail = _daily_series(cursor, d_m0a, d_m0b)
        # Agora os valores ja vem em buckets (nao cumulativos)
        doughnut_safra = [d30, d60, d90, dplus]

        safras_out.append({
            'index': parte,
            'label': _FAIXA_NOMES[parte],
            'volume': volume,
            'volume_val': ag['volume_val'],
            'd30': d30,
            'v30': ag['recovery_v30'],
            'd60': d60,
            'v60': ag['recovery_v60'],
            'd90': d90,
            'v90': ag['recovery_v90'],
            'dplus': dplus,
            'vplus': ag['recovery_vplus'],
            'desempenho': {
                'contratos_por_segmento': ag['segmentos'],
                'valor_parcela_entrada_por_segmento_brl': ag['valor_por_segmento_brl'],
                'valor_credito_por_segmento_brl': ag['valor_por_segmento_brl'],
                'volume_cohort_safra': ag['volume'],
                'volume_ocorrencias_faixa': ag['volume'],
            },
            'detail': {**detail, 'doughnut': doughnut_safra},
        })

        novos_safra = _count_distinct_ocorrencias(cursor, ranges_m, None, desc_novo_only=True)
        pagos_safra = sum(ag['performado_por_prazo_quitacao'].values())
        inden_safra = _count_distinct_ocorrencias(cursor, ranges_m, ['indenizado'], False)
        bar_novos.append(novos_safra)
        bar_pagos.append(pagos_safra)
        bar_indenizados.append(inden_safra)

    kpi_teto_cumulativo = _kpi_teto_cumulativo_distinct_global(cursor, y, m)
    cursor.close()
    conn.close()

    return jsonify({
        'mes': mes,
        'faixa_labels': _FAIXA_NOMES,
        'kpis': {
            'novos_mes': kpi_novos,
            'faixa_destaque': kpi_faixa_label,
            'safra_destaque': kpi_faixa_label,
            'recuperacao_pct': kpi_rec_pct,
            'safra_cohort_mes': n_safra_mes,
            'safra_performados_mes': n_perf_mes,
            'parcelas_criticas_90d': kpi_parcelas_crit,
        },
        'safras': safras_out,
        'chart_all': {
            'barLabels': _FAIXA_NOMES,
            'count': ch_perf,
            'valor_brl': ch_val,
            'novos': bar_novos,
            'pagos': bar_pagos,
            'indenizados': bar_indenizados,
            'kpi_teto_cumulativo': kpi_teto_cumulativo,
        },
    })


# ---------------------------------------------------------------------------
# API: Performance — exportacao (xlsx / pdf / powerbi)
# ---------------------------------------------------------------------------

_SERIES_LABELS_EXPORT = {
    'novos': 'Contratos Novos',
    'pagos': 'Safra performada (parcela de entrada quitada)',
    'indenizados': 'Contratos Indenizados',
}
_FAIXA_LABELS_EXPORT = {
    'd30': 'Ate 30 dias',
    'd60': '31 a 60 dias',
    'd90': 'Acima de 60 dias',
}
_SAFRA_LABELS_EXPORT = list(_FAIXA_NOMES)

_RESUMO_TIDY_DESC_PERFORMADO = {
    'd30': 'Performado: ate 30d entre pagamento e vencimento (parcela ref.)',
    'd60': 'Performado: 31 a 60d entre pagamento e vencimento',
    'd90': 'Performado: 61 a 90d entre pagamento e vencimento',
    'dplus': 'Performado: acima de 90d entre pagamento e vencimento',
}
_RESUMO_TIDY_DESC_NAO = {
    'b30': 'Nao performado: atraso em aberto ate 30d (menor vencimento em aberto)',
    'b60': 'Nao performado: atraso em aberto 31 a 60d',
    'b90': 'Nao performado: atraso em aberto 61 a 90d',
    'bplus': 'Nao performado: atraso em aberto acima de 90d',
}


def _build_performance_resumo_tidy(mes_ano, safras_summary_rows):
    """Uma linha por dimensao (ideal para tabela dinamica / segmentadores no Excel)."""
    out = []
    for r in safras_summary_rows:
        base = {
            'mes_ano': mes_ano,
            'safra': r['label'],
            'periodo_inicio': r['inicio'],
            'periodo_fim': r['fim'],
        }

        def push(situacao, faixa_codigo, faixa_desc, medida, valor):
            out.append({
                **base,
                'situacao': situacao,
                'faixa_codigo': faixa_codigo,
                'faixa_descricao': faixa_desc,
                'medida': medida,
                'valor': valor,
            })

        push('Cohort', 'total', 'Entradas na safra (contrato novo ou voltou, com parcela vencida no arquivo)',
             'contratos', r['volume'])
        push('Cohort', 'total', 'Entradas na safra — soma valor parcela (metrica Performance)',
             'brl_parcela_metrica', r['volume_parcela_brl'])

        for fk in ('d30', 'd60', 'd90', 'dplus'):
            blk = r['performado'][fk]
            desc = _RESUMO_TIDY_DESC_PERFORMADO[fk]
            push('Performado', fk, desc, 'contratos', blk['qtd'])
            push('Performado', fk, desc, 'brl_parcela_metrica', blk['parcela_brl'])

        for fk in ('b30', 'b60', 'b90', 'bplus'):
            blk = r['nao_performado'][fk]
            desc = _RESUMO_TIDY_DESC_NAO[fk]
            push('Nao_performado', fk, desc, 'contratos', blk['qtd'])
            push('Nao_performado', fk, desc, 'brl_parcela_metrica', blk['parcela_brl'])

    return out


def _resolve_export_payload():
    """Le o JSON do POST de exportacao e valida / normaliza o conteudo."""
    payload = request.get_json(silent=True) or {}
    mes = str(payload.get('mes', '')).strip()
    if not mes or len(mes) < 7:
        return None, 'Parametro mes obrigatorio (YYYY-MM)'
    try:
        y, m = int(mes[:4]), int(mes[5:7])
        datetime.date(y, m, 1)
    except ValueError:
        return None, 'mes invalido'

    raw_idx = payload.get('safra_index')
    if raw_idx in (None, '', 'all'):
        safra_index = None
    else:
        try:
            safra_index = int(raw_idx)
            if safra_index < 0 or safra_index > 3:
                safra_index = None
        except (TypeError, ValueError):
            safra_index = None

    series = [s for s in (payload.get('series') or []) if s in _SERIES_LABELS_EXPORT]
    if not series:
        series = list(_SERIES_LABELS_EXPORT.keys())
    faixas = [f for f in (payload.get('faixas') or []) if f in _FAIXA_LABELS_EXPORT]
    if not faixas:
        faixas = list(_FAIXA_LABELS_EXPORT.keys())

    at_raw = payload.get('atraso_teto', 90)
    try:
        atraso_teto = int(at_raw) if at_raw is not None else 90
    except (TypeError, ValueError):
        atraso_teto = 90
    if atraso_teto not in (30, 60, 90):
        atraso_teto = 90

    return {
        'mes': mes,
        'y': y,
        'm': m,
        'safra_index': safra_index,
        'series': series,
        'faixas': faixas,
        'atraso_teto': atraso_teto,
        'bar_image': payload.get('bar_image') or '',
        'pie_image': payload.get('pie_image') or '',
    }, None


def _fetch_export_dataset(cursor, ctx):
    """Monta os tres blocos de dados que sao usados em qualquer formato."""
    y, m = ctx['y'], ctx['m']
    safra_index = ctx['safra_index']
    series = ctx['series']

    # --- Resumo por safra (sempre as 4, para contexto) ---
    safras_summary = []
    for parte in range(4):
        d_a, d_b = _safra_bounds(y, m, parte)
        ag_sum = _aggregate_performance_faixa(cursor, y, m, parte)
        pp = ag_sum['performado_por_prazo_quitacao']
        pp_v = ag_sum['performado_por_prazo_quitacao_brl']
        pp_c = ag_sum['performado_por_prazo_quitacao_cred']
        npb = ag_sum['nao_por_atraso_aberto']
        npb_v = ag_sum['nao_por_atraso_aberto_brl']
        npb_c = ag_sum['nao_por_atraso_aberto_cred']
        safras_summary.append({
            'label': _SAFRA_LABELS_EXPORT[parte],
            'inicio': d_a.isoformat(),
            'fim': d_b.isoformat(),
            'volume': ag_sum['volume'],
            'volume_parcela_brl': round(ag_sum['volume_val'], 2),
            'volume_credito_brl': round(ag_sum['volume_cred'], 2),
            'd30': ag_sum['recovery_d30'],
            'd60': ag_sum['recovery_d60'],
            'd90': ag_sum['recovery_d90'],
            'dplus': ag_sum['recovery_dplus'],
            'recovery_v30': round(ag_sum['recovery_v30'], 2),
            'recovery_v60': round(ag_sum['recovery_v60'], 2),
            'recovery_v90': round(ag_sum['recovery_v90'], 2),
            'recovery_vplus': round(ag_sum['recovery_vplus'], 2),
            'recovery_cred30': round(ag_sum['recovery_cred30'], 2),
            'recovery_cred60': round(ag_sum['recovery_cred60'], 2),
            'recovery_cred90': round(ag_sum['recovery_cred90'], 2),
            'recovery_credplus': round(ag_sum['recovery_credplus'], 2),
            'performado': {
                fk: {
                    'qtd': int(pp[fk]),
                    'parcela_brl': round(float(pp_v[fk]), 2),
                    'credito_brl': round(float(pp_c[fk]), 2),
                }
                for fk in ('d30', 'd60', 'd90', 'dplus')
            },
            'nao_performado': {
                fk: {
                    'qtd': int(npb[fk]),
                    'parcela_brl': round(float(npb_v[fk]), 2),
                    'credito_brl': round(float(npb_c[fk]), 2),
                }
                for fk in ('b30', 'b60', 'b90', 'bplus')
            },
        })

    resumo_tidy = _build_performance_resumo_tidy(ctx['mes'], safras_summary)

    # --- Serie para o grafico de barras ---
    series_rows = []
    if safra_index is None:
        # Visao geral: uma linha por safra, uma entrada por serie
        for parte in range(4):
            d_a, d_b = _safra_bounds(y, m, parte)
            ranges_m = [(d_a, d_b)]
            rows_faixa = _safra_entrada_rows(cursor, d_a, d_b, y, m)
            values = {
                'novos': _count_distinct_ocorrencias(cursor, ranges_m, None, desc_novo_only=True),
                'pagos': sum(1 for r in rows_faixa if _bool_sql(r.get('is_performado'))),
                'indenizados': _count_distinct_ocorrencias(cursor, ranges_m, ['indenizado'], False),
            }
            for s in series:
                series_rows.append({
                    'safra': _SAFRA_LABELS_EXPORT[parte],
                    'data': d_a.isoformat(),
                    'serie': _SERIES_LABELS_EXPORT[s],
                    'valor': values[s],
                })
    else:
        d_a, d_b = _safra_bounds(y, m, safra_index)
        daily = _daily_series(cursor, d_a, d_b)
        for i, label in enumerate(daily['barLabels']):
            for s in series:
                series_rows.append({
                    'safra': _SAFRA_LABELS_EXPORT[safra_index],
                    'data': label,
                    'serie': _SERIES_LABELS_EXPORT[s],
                    'valor': daily[s][i],
                })

    # --- Distribuicao (doughnut global) ---
    cursor.execute(
        """
        SELECT CASE
            WHEN DATEDIFF(CURDATE(), v.min_v) BETWEEN 1 AND 30 THEN 'd30'
            WHEN DATEDIFF(CURDATE(), v.min_v) BETWEEN 31 AND 60 THEN 'd60'
            WHEN DATEDIFF(CURDATE(), v.min_v) > 60 THEN 'd90'
            ELSE 'out'
        END AS faixa, COUNT(*) AS n
        FROM (
            SELECT c.id, MIN(par.vencimento) AS min_v
            FROM contrato c
            INNER JOIN parcela par ON par.id_contrato = c.id AND par.status = 'cobranca'
            WHERE c.status = 'cobranca'
            GROUP BY c.id
            HAVING min_v < CURDATE()
        ) v
        GROUP BY faixa
        """
    )
    dg = {'d30': 0, 'd60': 0, 'd90': 0}
    for r in cursor.fetchall():
        if r['faixa'] in dg:
            dg[r['faixa']] = int(r['n'] or 0)
    faixas_rows = [
        {'faixa': _FAIXA_LABELS_EXPORT[k], 'total': dg[k]}
        for k in ctx['faixas']
    ]

    # --- Contratos: mesma selecao do grafico (teto 30/60/90 d) + colunas de desempenho/prazo ---
    teto = int(ctx.get('atraso_teto') or 90)
    if teto not in (30, 60, 90):
        teto = 90

    enriched = _cohort_enriched_rows_for_export(cursor, y, m, safra_index, teto)
    if not enriched:
        contratos = []
    else:
        by_id = {e['id_contrato']: e for e in enriched}
        ordered_ids = list(by_id.keys())
        ph = ','.join(['%s'] * len(ordered_ids))
        cursor.execute(
            f"SELECT c.id, c.grupo, c.cota, "
            f"       CONCAT_WS('/', c.grupo, c.cota) AS grupo_cota, "
            f"       (CASE WHEN pv.min_v_aberto IS NOT NULL "
            f"             THEN DATEDIFF(CURRENT_DATE, pv.min_v_aberto) END) AS atraso, "
            f"       c.numero_contrato, c.status, "
            f"c.valor_credito, c.data_adesao, "
            f"p.nome_completo AS devedor, p.cpf_cnpj AS devedor_cpf_cnpj "
            f"FROM contrato c "
            f"LEFT JOIN pessoa p ON p.id = c.id_pessoa "
            f"LEFT JOIN ( "
            f"  SELECT id_contrato, MIN(vencimento) AS min_v_aberto "
            f"  FROM parcela WHERE status = 'cobranca' GROUP BY id_contrato "
            f") pv ON pv.id_contrato = c.id "
            f"WHERE c.id IN ({ph}) "
            f"ORDER BY c.grupo, c.cota",
            tuple(ordered_ids),
        )
        merged = []
        for row in cursor.fetchall():
            cid = int(row['id'])
            meta = by_id.get(cid)
            if not meta:
                continue
            merged.append({
                'id': meta['id_contrato'],
                'faixa_calendario': meta['faixa_calendario'],
                'desempenho': meta['desempenho'],
                'prazo_atraso': meta['prazo_atraso'],
                'valor_parcela_entrada_brl': meta['valor_parcela_entrada_brl'],
                'grupo': row.get('grupo'),
                'cota': row.get('cota'),
                'grupo_cota': row.get('grupo_cota'),
                'atraso': row.get('atraso'),
                'numero_contrato': row.get('numero_contrato'),
                'status': row.get('status'),
                'valor_credito': row.get('valor_credito'),
                'data_adesao': row.get('data_adesao'),
                'devedor': row.get('devedor'),
                'devedor_cpf_cnpj': row.get('devedor_cpf_cnpj'),
            })
        contratos = _clean_rows(merged)

    return {
        'resumo': safras_summary,
        'resumo_tidy': resumo_tidy,
        'series': series_rows,
        'faixas': faixas_rows,
        'contratos': contratos,
    }


def _export_context_labels(ctx):
    safra_lbl = 'Visao geral (todas as safras)' if ctx['safra_index'] is None else _SAFRA_LABELS_EXPORT[ctx['safra_index']]
    series_lbl = ', '.join(_SERIES_LABELS_EXPORT[s] for s in ctx['series'])
    faixas_lbl = ', '.join(_FAIXA_LABELS_EXPORT[f] for f in ctx['faixas'])
    teto = int(ctx.get('atraso_teto') or 90)
    if teto not in (30, 60, 90):
        teto = 90
    teto_lbl = f'Ate {teto} dias (cumulativo: prazo de quitação e atraso em aberto, como no grafico)'
    return safra_lbl, series_lbl, faixas_lbl, teto_lbl


def _format_reais_pt_br_export(value):
    """Valor monetario como texto pt-BR para Excel/LibreOffice (milhar '.' ; decimal ',')."""
    if value is None or value == '':
        return ''
    try:
        n = float(value)
    except (TypeError, ValueError):
        return str(value)
    s = f'{n:,.2f}'
    return s.replace(',', 'X').replace('.', ',').replace('X', '.')


def _write_performance_resumo_safras_sheet(ws, ctx, dataset, header_font, header_fill, header_align, border):
    """Planilha larga: cohort + performado (ate 30 / cum. 31-60 / cum. 61-90) + nao performado (faixas exclusivas)."""
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter

    last_col = 19
    wrap = Alignment(wrap_text=True, vertical='top')
    title_font = Font(bold=True, size=12)
    note_font = Font(size=9)

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_col)
    c = ws.cell(row=1, column=1)
    c.value = f'Resumo por safra — mes {ctx["mes"]} (export alinhado ao painel Performance)'
    c.font = title_font
    c.alignment = wrap

    ws.merge_cells(start_row=2, start_column=1, end_row=3, end_column=last_col)
    c2 = ws.cell(row=2, column=1)
    c2.value = (
        'Cada linha e uma faixa de calendario do mes (entradas na safra). '
        'Colunas "Performado": contratos com parcela quitada; prazo = dias entre pagamento e vencimento da parcela '
        'de referencia. Nesta aba, "ate 30d" e exclusivo; "31-60d" e "61-90d" mostram totais acumulados '
        '(incluem as quitacoes das faixas anteriores). Quitacoes com mais de 90 dias contam na faixa 61-90d '
        '(nao ha coluna de performado >90). '
        'Colunas "Nao performado": contratos em cobranca na parcela de referencia, por atraso da menor parcela '
        'em aberto (b30/b60/b90/b+). R$ parcela = soma da metrica de parcela usada no modulo. '
        'Para tabela dinamica, use a aba "Resumo Safras pivot".'
    )
    c2.font = note_font
    c2.alignment = wrap
    ws.row_dimensions[2].height = 56

    grp_fill_perf = PatternFill(start_color='1D4ED8', end_color='1D4ED8', fill_type='solid')
    grp_fill_nao = PatternFill(start_color='C2410C', end_color='C2410C', fill_type='solid')
    grp_fill_cohort = PatternFill(start_color='047857', end_color='047857', fill_type='solid')

    r0 = 6
    merges = [
        ('A', 'C', 'Safra / periodo'),
        ('D', 'E', 'Cohort (entradas)'),
        ('F', 'G', 'Performado ate 30d'),
        ('H', 'I', 'Performado ate 60d (cumul.)'),
        ('J', 'K', 'Performado ate 90d (cumul.)'),
        ('L', 'M', 'Nao perf. aberto ate 30d'),
        ('N', 'O', 'Nao perf. 31-60d'),
        ('P', 'Q', 'Nao perf. 61-90d'),
        ('R', 'S', 'Nao perf. >90d aberto'),
    ]
    col_pairs = [(1, 3), (4, 5), (6, 7), (8, 9), (10, 11), (12, 13), (14, 15), (16, 17), (18, 19)]
    for i, (_a, _b, title) in enumerate(merges):
        c1, c2 = col_pairs[i]
        ws.merge_cells(start_row=r0, start_column=c1, end_row=r0, end_column=c2)
        cell = ws.cell(row=r0, column=c1, value=title)
        cell.font = Font(bold=True, color='FFFFFF', size=10)
        cell.alignment = header_align
        cell.border = border
        if i == 0:
            cell.fill = header_fill
        elif i == 1:
            cell.fill = grp_fill_cohort
        elif i <= 4:
            cell.fill = grp_fill_perf
        else:
            cell.fill = grp_fill_nao
        for cc in range(c1 + 1, c2 + 1):
            ws.cell(row=r0, column=cc).border = border

    sub = [
        'Safra', 'Inicio', 'Fim',
        'Contratos', 'R$ parcela',
        'N', 'R$ parc.',
        'N', 'R$ parc.',
        'N', 'R$ parc.',
        'N', 'R$ parc.',
        'N', 'R$ parc.',
        'N', 'R$ parc.',
        'N', 'R$ parc.',
    ]
    r1 = r0 + 1
    for col_idx, h in enumerate(sub, start=1):
        cell = ws.cell(row=r1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = border

    brl_fmt = '#,##0.00'
    money_cols = {5, 7, 9, 11, 13, 15, 17, 19}
    data_row = r1 + 1
    first_data_row = data_row
    for r in dataset['resumo']:
        p = r['performado']
        n = r['nao_performado']
        q30, v30 = p['d30']['qtd'], p['d30']['parcela_brl']
        q60, v60 = p['d60']['qtd'], p['d60']['parcela_brl']
        q90, v90 = p['d90']['qtd'], p['d90']['parcela_brl']
        qp, vp = p['dplus']['qtd'], p['dplus']['parcela_brl']
        cum_q_60 = q30 + q60
        cum_v_60 = v30 + v60
        cum_q_90 = cum_q_60 + q90 + qp
        cum_v_90 = cum_v_60 + v90 + vp
        vals = [
            r['label'], r['inicio'], r['fim'],
            r['volume'], r['volume_parcela_brl'],
            q30, v30,
            cum_q_60, cum_v_60,
            cum_q_90, cum_v_90,
            n['b30']['qtd'], n['b30']['parcela_brl'],
            n['b60']['qtd'], n['b60']['parcela_brl'],
            n['b90']['qtd'], n['b90']['parcela_brl'],
            n['bplus']['qtd'], n['bplus']['parcela_brl'],
        ]
        for col_idx, val in enumerate(vals, start=1):
            cell = ws.cell(row=data_row, column=col_idx, value=val)
            cell.border = border
            if col_idx in money_cols and isinstance(val, (int, float)) and not isinstance(val, bool):
                cell.number_format = brl_fmt
        data_row += 1

    last_data_row = data_row - 1
    if last_data_row >= first_data_row:
        total_row = data_row
        tot_font = Font(bold=True)
        ws.cell(row=total_row, column=1, value='TOTAL').font = tot_font
        ws.cell(row=total_row, column=1).border = border
        for cc in range(2, 4):
            ws.cell(row=total_row, column=cc).border = border
        for c in range(4, last_col + 1):
            col_letter = get_column_letter(c)
            cell = ws.cell(
                row=total_row, column=c,
                value=f'=SUM({col_letter}{first_data_row}:{col_letter}{last_data_row})',
            )
            cell.font = tot_font
            cell.border = border
            if c in money_cols:
                cell.number_format = brl_fmt

    widths = [22, 12, 12] + [11] * 16
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _write_performance_resumo_pivot_sheet(wb, dataset, header_font, header_fill, header_align, border):
    """Dados longos + tabela do Excel para o usuario montar dinamicas e segmentadores."""
    from openpyxl.worksheet.table import Table, TableStyleInfo
    from openpyxl.styles import Font, Alignment
    from openpyxl.utils import get_column_letter

    ws = wb.create_sheet('Resumo Safras pivot')
    headers = [
        'Mes_ano', 'Safra', 'Periodo_inicio', 'Periodo_fim',
        'Situacao', 'Faixa_codigo', 'Faixa_descricao', 'Medida', 'Valor',
    ]
    tidy = dataset.get('resumo_tidy') or []
    rows = [
        [
            t['mes_ano'], t['safra'], t['periodo_inicio'], t['periodo_fim'],
            t['situacao'], t['faixa_codigo'], t['faixa_descricao'], t['medida'], t['valor'],
        ]
        for t in tidy
    ]
    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = border
    brl_medidas = {'brl_parcela_metrica'}
    brl_fmt = '#,##0.00'
    for row_idx, r in enumerate(rows, start=2):
        for col_idx, val in enumerate(r, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = border
            if col_idx == 9 and len(r) >= 8 and r[7] in brl_medidas:
                cell.number_format = brl_fmt
    nrows = 1 + len(rows)
    ref = f'A1:{get_column_letter(len(headers))}{max(nrows, 2)}'
    try:
        tab = Table(displayName='tblPerfResumoSafra', ref=ref)
        tab.tableStyleInfo = TableStyleInfo(
            name='TableStyleMedium9', showFirstColumn=False,
            showLastColumn=False, showRowStripes=True, showColumnStripes=False,
        )
        ws.add_table(tab)
    except Exception:
        pass
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 22
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 16
    ws.column_dimensions['F'].width = 10
    ws.column_dimensions['G'].width = 52
    ws.column_dimensions['H'].width = 28
    ws.column_dimensions['I'].width = 14

    ws2 = wb.create_sheet('Resumo pivot notas')
    ws2['A1'] = 'Como usar'
    ws2['A1'].font = Font(bold=True, size=12)
    notes = (
        '1) Na aba "Resumo Safras pivot", selecione qualquer celula da tabela.\n'
        '2) Menu Inserir > Tabela dinamica (ou PivotTable).\n'
        '3) Arraste "Situacao", "Faixa_codigo" ou "Faixa_descricao" para linhas ou filtros.\n'
        '4) Arraste "Medida" para colunas e "Valor" para valores (soma).\n'
        '5) Opcional: Inserir > Segmentacao de dados (Slicers) em Situacao, Safra ou Medida.'
    )
    ws2['A2'] = notes
    ws2['A2'].alignment = Alignment(wrap_text=True, vertical='top')
    ws2.column_dimensions['A'].width = 92
    ws2.row_dimensions[2].height = 88


def _export_to_xlsx(ctx, dataset):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()
    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='3B82F6', end_color='3B82F6', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center')
    thin = Side(style='thin', color='D1D5DB')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def _write_sheet(ws, headers, rows):
        for col_idx, h in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = border
        for row_idx, r in enumerate(rows, start=2):
            for col_idx, val in enumerate(r, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.border = border
        for col_idx, h in enumerate(headers, start=1):
            max_len = len(str(h))
            for row_idx in range(2, 2 + len(rows)):
                v = ws.cell(row=row_idx, column=col_idx).value
                max_len = max(max_len, len(str(v)) if v is not None else 0)
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max_len + 4, 40)

    # Aba 1: Contexto
    ws0 = wb.active
    ws0.title = 'Parametros'
    safra_lbl, series_lbl, faixas_lbl, teto_lbl = _export_context_labels(ctx)
    ws0['A1'] = 'Parametros da exportacao'
    ws0['A1'].font = Font(bold=True, size=13)
    info = [
        ('Mes/Ano', ctx['mes']),
        ('Safra em analise', safra_lbl),
        ('Teto atraso (como no grafico)', teto_lbl),
        ('Series selecionadas', series_lbl),
        ('Faixas de atraso (aba auxiliar)', faixas_lbl),
        ('Resumo Excel', 'Aba "Resumo Safras" (visao larga) + "Resumo Safras pivot" (dados longos; Tabela dinamica) + "Resumo pivot notas".'),
        ('Gerado em', datetime.datetime.now().strftime('%d/%m/%Y %H:%M')),
    ]
    for i, (k, v) in enumerate(info, start=3):
        ws0.cell(row=i, column=1, value=k).font = Font(bold=True)
        ws0.cell(row=i, column=2, value=v)
    ws0.column_dimensions['A'].width = 24
    ws0.column_dimensions['B'].width = 60

    # Aba 2: Resumo das safras (largo + pivot)
    ws1 = wb.create_sheet('Resumo Safras')
    _write_performance_resumo_safras_sheet(ws1, ctx, dataset, header_font, header_fill, header_align, border)
    _write_performance_resumo_pivot_sheet(wb, dataset, header_font, header_fill, header_align, border)

    # Aba: Series (tidy long-format, ideal para PowerBI tambem)
    ws2 = wb.create_sheet('Series Graficas')
    _write_sheet(
        ws2,
        ['Safra', 'Data/Dia', 'Serie', 'Valor'],
        [[r['safra'], r['data'], r['serie'], r['valor']] for r in dataset['series']],
    )

    # Aba 4: Distribuicao
    ws3 = wb.create_sheet('Faixas de Atraso')
    _write_sheet(
        ws3,
        ['Faixa', 'Total de contratos'],
        [[r['faixa'], r['total']] for r in dataset['faixas']],
    )

    # Aba 5: Contratos selecionados
    ws4 = wb.create_sheet('Contratos')
    contratos = dataset['contratos']
    headers = [
        'ID', 'Faixa (calendario)', 'Desempenho', 'Prazo (atraso)', 'Valor parcela (metrica) (R$)',
        'Grupo', 'Cota', 'Grupo/Cota', 'Atraso', 'Nro Contrato', 'Status', 'Valor do Credito',
        'Data de Adesao', 'Devedor', 'CPF/CNPJ',
    ]
    keys = [
        'id', 'faixa_calendario', 'desempenho', 'prazo_atraso', 'valor_parcela_entrada_brl',
        'grupo', 'cota', 'grupo_cota', 'atraso', 'numero_contrato', 'status', 'valor_credito',
        'data_adesao', 'devedor', 'devedor_cpf_cnpj',
    ]
    _monetary_key = {'valor_parcela_entrada_brl', 'valor_credito'}

    def _row_contrato_exp(c):
        row = []
        for k in keys:
            v = c.get(k, '')
            if k in _monetary_key:
                row.append(_format_reais_pt_br_export(v))
            else:
                row.append(v if v is not None else '')
        return row

    _write_sheet(ws4, headers, [_row_contrato_exp(c) for c in contratos])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def _export_to_csv_powerbi(ctx, dataset):
    """Gera um .csv tidy long-format otimizado para 'Get Data -> Text/CSV' no Power BI.

    Cada linha descreve uma observacao unica com as dimensoes necessarias.
    """
    import csv

    buf = io.StringIO()
    writer = csv.writer(buf, delimiter=';')
    # Cabecalho com metadados (PowerBI ignora e o usuario entende)
    safra_lbl, series_lbl, faixas_lbl, teto_lbl = _export_context_labels(ctx)
    writer.writerow(['# Performance JB - export Power BI'])
    writer.writerow(['# Mes/Ano', ctx['mes']])
    writer.writerow(['# Safra', safra_lbl])
    writer.writerow(['# Teto atraso', teto_lbl])
    writer.writerow(['# Series', series_lbl])
    writer.writerow(['# Faixas (aba auxiliar)', faixas_lbl])
    writer.writerow(['# Gerado em', datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')])
    writer.writerow([])

    # Tabela 1 - Series (tidy)
    writer.writerow(['tabela', 'safra', 'data', 'serie', 'valor'])
    for r in dataset['series']:
        writer.writerow(['series', r['safra'], r['data'], r['serie'], r['valor']])

    writer.writerow(['tabela', 'mes_ano', 'safra', 'periodo_inicio', 'periodo_fim',
                     'situacao', 'faixa_codigo', 'faixa_descricao', 'medida', 'valor'])
    for t in dataset.get('resumo_tidy') or []:
        writer.writerow([
            'resumo_safra_tidy',
            t['mes_ano'], t['safra'], t['periodo_inicio'], t['periodo_fim'],
            t['situacao'], t['faixa_codigo'], t['faixa_descricao'], t['medida'], t['valor'],
        ])

    # Tabela 3 - Faixas
    for r in dataset['faixas']:
        writer.writerow(['faixas', '-', '-', r['faixa'], r['total']])

    # Tabela 4 - Contratos
    writer.writerow([])
    writer.writerow([
        'tabela', 'id', 'faixa_calendario', 'desempenho', 'prazo_atraso', 'valor_parcela',
        'grupo', 'cota', 'grupo_cota', 'atraso', 'data_adesao', 'status', 'valor_credito',
        'devedor', 'cpf_cnpj', 'nro_contrato',
    ])
    for c in dataset['contratos']:
        writer.writerow([
            'contratos',
            c.get('id') or '',
            c.get('faixa_calendario') or '',
            c.get('desempenho') or '',
            c.get('prazo_atraso') or '',
            c.get('valor_parcela_entrada_brl') if c.get('valor_parcela_entrada_brl') is not None else 0,
            c.get('grupo') or '',
            c.get('cota') or '',
            c.get('grupo_cota') or (str(c.get('grupo') or '') + '/' + str(c.get('cota') or '')),
            c.get('atraso') if c.get('atraso') is not None else '',
            c.get('data_adesao') or '',
            c.get('status') or '',
            c.get('valor_credito') or 0,
            c.get('devedor') or '',
            c.get('devedor_cpf_cnpj') or '',
            c.get('numero_contrato') or '',
        ])

    data = buf.getvalue().encode('utf-8-sig')  # BOM para Power BI detectar UTF-8
    out = io.BytesIO(data)
    out.seek(0)
    return out


def _decode_data_url_png(data_url):
    """Aceita 'data:image/png;base64,xxxx' -> bytes. Retorna None se invalido."""
    import base64
    if not data_url or ',' not in data_url:
        return None
    try:
        _, b64 = data_url.split(',', 1)
        return base64.b64decode(b64)
    except Exception:
        return None


def _export_to_pdf(ctx, dataset):
    from fpdf import FPDF

    pdf = FPDF(orientation='L', unit='mm', format='A4')
    pdf.set_auto_page_break(auto=True, margin=12)
    pdf.add_page()

    safra_lbl, series_lbl, faixas_lbl, teto_lbl = _export_context_labels(ctx)

    # Cabecalho
    pdf.set_font('Helvetica', 'B', 16)
    pdf.cell(0, 10, 'Performance JB - Relatorio Exportado', ln=True, align='C')
    pdf.set_font('Helvetica', '', 10)
    pdf.cell(0, 6, 'Mes/Ano: ' + ctx['mes'] + '   |   Safra: ' + safra_lbl, ln=True, align='C')
    pdf.cell(0, 6, 'Teto atraso: ' + teto_lbl, ln=True, align='C')
    pdf.cell(0, 6, 'Series: ' + series_lbl, ln=True, align='C')
    pdf.cell(0, 6, 'Faixas (aux): ' + faixas_lbl, ln=True, align='C')
    pdf.cell(0, 6, 'Gerado em ' + datetime.datetime.now().strftime('%d/%m/%Y %H:%M'), ln=True, align='C')
    pdf.ln(4)

    # Graficos (imagens vindas do client)
    tmp_files = []
    try:
        bar_bytes = _decode_data_url_png(ctx['bar_image'])
        pie_bytes = _decode_data_url_png(ctx['pie_image'])

        page_w = pdf.w - 2 * pdf.l_margin
        if bar_bytes and pie_bytes:
            tf_bar = tempfile.NamedTemporaryFile(suffix='.png', delete=False)
            tf_bar.write(bar_bytes); tf_bar.close(); tmp_files.append(tf_bar.name)
            tf_pie = tempfile.NamedTemporaryFile(suffix='.png', delete=False)
            tf_pie.write(pie_bytes); tf_pie.close(); tmp_files.append(tf_pie.name)
            img_h = 75
            col_w = (page_w - 6) / 2
            y0 = pdf.get_y()
            pdf.image(tf_bar.name, x=pdf.l_margin, y=y0, w=col_w, h=img_h)
            pdf.image(tf_pie.name, x=pdf.l_margin + col_w + 6, y=y0, w=col_w, h=img_h)
            pdf.set_y(y0 + img_h + 6)
        elif bar_bytes:
            tf_bar = tempfile.NamedTemporaryFile(suffix='.png', delete=False)
            tf_bar.write(bar_bytes); tf_bar.close(); tmp_files.append(tf_bar.name)
            pdf.image(tf_bar.name, x=pdf.l_margin, w=page_w, h=90)
            pdf.ln(4)

        # Tabela: Resumo das safras (compacto; detalhe no Excel)
        pdf.set_font('Helvetica', 'B', 11)
        pdf.cell(0, 7, 'Resumo das safras (contratos; performado x nao por faixa)', ln=True)
        pdf.set_font('Helvetica', '', 7)
        pdf.multi_cell(0, 4, 'Detalhe com R$ parcela e tabela dinamica: exporte em Excel (abas Resumo Safras).', align='L')
        pdf.ln(1)
        pdf.set_font('Helvetica', 'B', 7)
        pdf.set_fill_color(59, 130, 246)
        pdf.set_text_color(255, 255, 255)
        headers = [
            'Safra', 'Ini', 'Fim', 'N', 'P30', 'P<=60', 'P<=90',
            'N30', 'N60', 'N90', 'N+',
        ]
        col_w = [52, 22, 22, 12, 12, 12, 12, 12, 12, 12, 12]
        for i, h in enumerate(headers):
            pdf.cell(col_w[i], 6, h, border=1, align='C', fill=True)
        pdf.ln()
        pdf.set_font('Helvetica', '', 7)
        pdf.set_text_color(30, 41, 59)
        for idx, r in enumerate(dataset['resumo']):
            fill = idx % 2 == 1
            if fill:
                pdf.set_fill_color(248, 250, 252)
            else:
                pdf.set_fill_color(255, 255, 255)
            p = r['performado']
            n = r['nao_performado']
            pc60 = p['d30']['qtd'] + p['d60']['qtd']
            pc90 = pc60 + p['d90']['qtd'] + p['dplus']['qtd']
            vals = [
                str(r['label'])[:26],
                str(r['inicio'])[5:],
                str(r['fim'])[5:],
                str(r['volume']),
                str(p['d30']['qtd']), str(pc60), str(pc90),
                str(n['b30']['qtd']), str(n['b60']['qtd']), str(n['b90']['qtd']), str(n['bplus']['qtd']),
            ]
            for i, v in enumerate(vals):
                pdf.cell(col_w[i], 6, v, border=1, align='C', fill=True)
            pdf.ln()
        pdf.ln(4)

        # Tabela: Contratos (ate 200 — mais colunas)
        contratos = dataset['contratos'][:200]
        pdf.set_font('Helvetica', 'B', 11)
        pdf.cell(0, 7, 'Contratos selecionados (' + str(len(dataset['contratos'])) + ' total, mostrando ' + str(len(contratos)) + ')', ln=True)
        pdf.set_font('Helvetica', 'B', 7)
        pdf.set_fill_color(59, 130, 246); pdf.set_text_color(255, 255, 255)
        ch = ['Faixa', 'Desemp.', 'Prazo', 'G/C', 'Atraso', 'Nro', 'R$ parc.', 'R$ cr.', 'St.', 'Devedor']
        cw = [30, 16, 22, 18, 12, 18, 18, 18, 8, 54]
        for i, h in enumerate(ch):
            pdf.cell(cw[i], 5, h, border=1, align='C', fill=True)
        pdf.ln()
        pdf.set_font('Helvetica', '', 6)
        pdf.set_text_color(30, 41, 59)
        for idx, c in enumerate(contratos):
            fill = idx % 2 == 1
            if fill: pdf.set_fill_color(248, 250, 252)
            else: pdf.set_fill_color(255, 255, 255)
            vparc = c.get('valor_parcela_entrada_brl')
            try:
                vparcf = 'R$ ' + f"{float(vparc):,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')
            except Exception:
                vparcf = str(vparc or '-')
            valor = c.get('valor_credito')
            try:
                valor_fmt = 'R$ ' + f"{float(valor):,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')
            except Exception:
                valor_fmt = str(valor or '-')
            atr = c.get('atraso')
            atr_s = '' if atr is None else str(int(atr)) if isinstance(atr, (int, float)) else str(atr)[:8]
            row = [
                str(c.get('faixa_calendario') or '')[:18],
                str(c.get('desempenho') or '-')[:8],
                (str(c.get('prazo_atraso') or ''))[:18],
                str(c.get('grupo_cota') or (str(c.get('grupo') or '') + '/' + str(c.get('cota') or '')))[:14],
                atr_s[:8],
                str(c.get('numero_contrato') or '-')[:8],
                vparcf,
                valor_fmt,
                (str(c.get('status') or ''))[:4],
                (str(c.get('devedor') or '-'))[:28],
            ]
            for i, v in enumerate(row):
                pdf.cell(cw[i], 5, v, border=1, align='C', fill=True)
            pdf.ln()

        buf = io.BytesIO()
        pdf.output(buf)
        buf.seek(0)
        return buf
    finally:
        for p in tmp_files:
            try: os.unlink(p)
            except Exception: pass


@app.route('/api/performance/export/<formato>', methods=['POST'])
def api_performance_export(formato):
    formato = (formato or '').lower()
    if formato not in ('xlsx', 'pdf', 'powerbi'):
        return jsonify({'error': 'formato invalido'}), 400

    ctx, err = _resolve_export_payload()
    if err:
        return jsonify({'error': err}), 400

    conn = _get_db()
    cursor = conn.cursor()
    try:
        dataset = _fetch_export_dataset(cursor, ctx)
    finally:
        cursor.close()
        conn.close()

    ts = datetime.datetime.now().strftime('%Y%m%d_%H%M')
    safra_slug = 'geral' if ctx['safra_index'] is None else 'safra' + str(ctx['safra_index'] + 1)
    base = f'performance_{ctx["mes"]}_{safra_slug}_{ts}'

    if formato == 'xlsx':
        buf = _export_to_xlsx(ctx, dataset)
        return send_file(buf, as_attachment=True, download_name=base + '.xlsx',
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    if formato == 'pdf':
        buf = _export_to_pdf(ctx, dataset)
        return send_file(buf, as_attachment=True, download_name=base + '.pdf',
                         mimetype='application/pdf')
    # powerbi -> csv
    buf = _export_to_csv_powerbi(ctx, dataset)
    return send_file(buf, as_attachment=True, download_name=base + '_powerbi.csv',
                     mimetype='text/csv')


# ---------------------------------------------------------------------------
# API: Dashboard — exportacao (xlsx / pdf / powerbi)
# ---------------------------------------------------------------------------

_DASH_SERIES_LABELS = {
    'pagos':       'Contratos Pagos',
    'indenizados': 'Contratos Indenizados',
    'novos':       'Ocorrencias Novo (aberto)',
    'retomados':   'Ocorrencias Voltou (aberto)',
    'pagos_parcial': 'Contratos pagos parcialmente (0–90 d)',
}
_DASH_SERIES_WHERE = {
    'pagos':       ("o.status IN ('pago total','pago parcial')", []),
    'indenizados': ("o.status = 'indenizado'", []),
    'novos':       ("o.status = 'cobranca' AND o.descricao LIKE '%%novo%%'", []),
    'retomados':   ("o.status = 'cobranca' AND o.descricao LIKE '%%contrato voltou%%'", []),
}
# Novos/retomados: soma ocorrencias; demais: contratos distintos no mes
_DASH_SERIES_COUNT_FUNC = {
    'pagos': 'COUNT(DISTINCT o.id_contrato)',
    'indenizados': 'COUNT(DISTINCT o.id_contrato)',
    'novos': 'COUNT(*)',
    'retomados': 'COUNT(*)',
    'pagos_parcial': 'COUNT(DISTINCT p.id_contrato)',
}
_DASH_PIE_LABELS = {
    'cobranca':   'Em Cobranca',
    'pago':       'Pagos',
    'indenizado': 'Indenizados',
}

# Inclui entradas_safra (1 ctt) para OR no painel: mesmo escopo do grafico de "Entradas (safra)"
_PAINEL_DASH_SERIES_WHERE = {**_DASH_SERIES_WHERE}
_PAINEL_DASH_SERIES_WHERE['entradas_safra'] = (
    "o.status = 'cobranca' AND (o.descricao = 'contrato novo' OR LOWER(o.descricao) LIKE 'contrato voltou%%')",
    [],
)
_PAINEL_DASH_LIM = 500

# Chaves aceitas no painel/export (inclui série baseada em parcela, não só ocorrência).
_PAINEL_DASH_SERIES_KEYS_VALID = frozenset(list(_PAINEL_DASH_SERIES_WHERE.keys()) + ['pagos_parcial'])


def _painel_dash_union_contrato_ids_sql(series_keys, d_ini, d_fim):
    """UNION de id_contrato: ocorrências no período + parcelas quitadas com performado 0–90 d."""
    parts = []
    params = []
    for k in series_keys:
        if k == 'pagos_parcial':
            parts.append(
                "SELECT DISTINCT par.id_contrato FROM parcela par "
                "WHERE par.status = 'pago' "
                "AND DATE(par.data_pagamento) >= %s AND DATE(par.data_pagamento) <= %s "
                "AND DATEDIFF(par.data_pagamento, par.vencimento) BETWEEN 0 AND 90"
            )
            params.extend([d_ini, d_fim])
        elif k in _PAINEL_DASH_SERIES_WHERE:
            w, p = _PAINEL_DASH_SERIES_WHERE[k]
            parts.append(
                "SELECT DISTINCT o.id_contrato FROM ocorrencia o "
                f"WHERE ({w}) AND DATE(o.data_arquivo) >= %s AND DATE(o.data_arquivo) <= %s"
            )
            params.extend(list(p) + [d_ini, d_fim])
    if not parts:
        return None, []
    return '(' + ' UNION '.join(parts) + ')', params


def _or_ocorrencia_filtro_series_sql(series_keys, painel_map: dict) -> tuple:
    """(sql_or_subexpr, list_params) para ocorrencia o."""
    parts, params = [], []
    for k in series_keys:
        if not k or k not in painel_map:
            continue
        w, p = painel_map[k]
        parts.append('(' + w + ')')
        params.extend(p)
    if not parts:
        return None, []
    return ' OR '.join(parts), params


def _panel_dash_busca_filtro_sql(tipo: str, termo: str, status_contrato: str) -> tuple:
    """Retorna (join_extra, and_sql, params) para a lista do painel Dashboard (apos WHERE base)."""
    tipo = (tipo or 'contrato').strip()
    termo = (termo or '').strip()
    st = _normalize_contrato_status_arg(status_contrato or '')
    st_ok = st in _STATUS_CONTRATO_VALIDOS
    join_ex = ''
    and_bits = []
    params = []
    if st_ok:
        and_bits.append('c.status = %s')
        params.append(st)
    if not termo:
        if not and_bits:
            return '', '', []
        return join_ex, 'AND ' + ' AND '.join(and_bits), params
    if tipo == 'pessoa':
        t = f'%{termo}%'
        join_ex = 'LEFT JOIN pessoa p_ava ON p_ava.id = c.id_avalista '
        and_bits.append(
            '(p.nome_completo LIKE %s OR p.cpf_cnpj LIKE %s OR p_ava.nome_completo LIKE %s OR p_ava.cpf_cnpj LIKE %s)'
        )
        params.extend([t, t, t, t])
    elif tipo == 'contrato':
        p2 = [p.strip() for p in termo.replace('-', '/').split('/')]
        if len(p2) == 2 and p2[0] and p2[1]:
            and_bits.append('c.grupo LIKE %s AND c.cota LIKE %s')
            params.extend([f'%{p2[0]}%', f'%{p2[1]}%'])
        else:
            t = f'%{termo}%'
            and_bits.append('(c.grupo LIKE %s OR c.cota LIKE %s OR c.numero_contrato LIKE %s)')
            params.extend([t, t, t])
    elif tipo == 'bem':
        jb = _bem_join_clause('c', 'b')
        b_where, b_params = _bem_where_clause(termo, 'b')
        if not jb or not b_where:
            and_bits.append('0=1')
        else:
            join_ex = jb
            and_bits.append('(' + b_where + ')')
            params.extend(b_params)
    else:
        t = f'%{termo}%'
        and_bits.append('(c.grupo LIKE %s OR c.cota LIKE %s)')
        params.extend([t, t])
    return join_ex, 'AND ' + ' AND '.join(and_bits), params


def _month_range_from_yyyymm(ym_start, ym_end):
    """Aceita 'YYYY-MM' start e end (inclusivo) -> lista de meses em ordem."""
    def _parse(ym):
        y, m = int(ym[:4]), int(ym[5:7])
        return y, m
    y1, m1 = _parse(ym_start)
    y2, m2 = _parse(ym_end)
    out = []
    while (y1, m1) <= (y2, m2):
        out.append(f'{y1:04d}-{m1:02d}')
        if m1 == 12:
            y1, m1 = y1 + 1, 1
        else:
            m1 += 1
    return out


def _last_day_of(y, m):
    return calendar.monthrange(y, m)[1]


def _resolve_dash_export_payload():
    payload = request.get_json(silent=True) or {}
    ps = str(payload.get('period_start', '')).strip()
    pe = str(payload.get('period_end', '')).strip()
    try:
        if not ps or not pe or len(ps) < 7 or len(pe) < 7:
            raise ValueError
        y1, m1 = int(ps[:4]), int(ps[5:7])
        y2, m2 = int(pe[:4]), int(pe[5:7])
        datetime.date(y1, m1, 1)
        datetime.date(y2, m2, 1)
        if (y1, m1) > (y2, m2):
            ps, pe = pe, ps
    except (TypeError, ValueError):
        return None, 'Periodo invalido (period_start / period_end no formato YYYY-MM)'

    series = [s for s in (payload.get('series') or []) if s in _DASH_SERIES_LABELS]
    if not series:
        series = ['pagos', 'indenizados']

    pie = [p for p in (payload.get('pie') or []) if p in _DASH_PIE_LABELS]
    if not pie:
        pie = list(_DASH_PIE_LABELS.keys())

    return {
        'period_start': ps,
        'period_end': pe,
        'series': series,
        'pie': pie,
        'line_image': payload.get('line_image') or '',
        'pie_image': payload.get('pie_image') or '',
    }, None


def _fetch_dash_export_dataset(cursor, ctx):
    try:
        _ensure_cobranca_table(cursor)
    except Exception:
        app.logger.exception('_fetch_dash_export: falha ao garantir tabela cobranca')

    meses = _month_range_from_yyyymm(ctx['period_start'], ctx['period_end'])

    # --- Series mensais (long format / tidy) ---
    series_rows = []
    # SQL date window
    y1, m1 = int(ctx['period_start'][:4]), int(ctx['period_start'][5:7])
    y2, m2 = int(ctx['period_end'][:4]), int(ctx['period_end'][5:7])
    d_ini = datetime.date(y1, m1, 1).isoformat()
    d_fim = datetime.date(y2, m2, _last_day_of(y2, m2)).isoformat()

    series_totals = {}  # key -> total no periodo
    series_by_month = {}  # key -> {mes: count}

    for key in ctx['series']:
        if key == 'pagos_parcial':
            cursor.execute(
                "SELECT DATE_FORMAT(p.data_pagamento, '%%Y-%%m') AS mes, "
                "       COUNT(DISTINCT p.id_contrato) AS total "
                "FROM parcela p "
                "WHERE p.status = 'pago' "
                "AND DATEDIFF(p.data_pagamento, p.vencimento) BETWEEN 0 AND 90 "
                "AND DATE(p.data_pagamento) >= %s AND DATE(p.data_pagamento) <= %s "
                "GROUP BY mes ORDER BY mes",
                (d_ini, d_fim),
            )
            by_m = {r['mes']: int(r['total']) for r in cursor.fetchall()}
        else:
            where, params = _DASH_SERIES_WHERE[key]
            count_expr = _DASH_SERIES_COUNT_FUNC.get(key, 'COUNT(DISTINCT o.id_contrato)')
            cursor.execute(
                "SELECT DATE_FORMAT(o.data_arquivo, '%%Y-%%m') AS mes, "
                f"       {count_expr} AS total "
                "FROM ocorrencia o "
                f"WHERE {where} "
                "AND o.data_arquivo >= %s AND o.data_arquivo <= %s "
                "GROUP BY mes ORDER BY mes",
                tuple(params) + (d_ini, d_fim),
            )
            by_m = {r['mes']: int(r['total']) for r in cursor.fetchall()}
        series_by_month[key] = by_m
        series_totals[key] = sum(by_m.values())
        for mes in meses:
            series_rows.append({
                'mes': mes,
                'serie': _DASH_SERIES_LABELS[key],
                'valor': by_m.get(mes, 0),
            })

    # --- Pie chart (distribuicao global atual) ---
    cursor.execute("SELECT status, COUNT(*) AS total FROM contrato GROUP BY status")
    pie_raw = {r['status']: int(r['total']) for r in cursor.fetchall()}
    pie_rows = [
        {'status': _DASH_PIE_LABELS[k], 'total': pie_raw.get(k, 0)}
        for k in ctx['pie']
    ]

    # --- KPIs do periodo: Em Cobranca = snapshot `cobranca` (ultimo dia GM) ---
    data_dash_ref = _get_data_referencia_arquivos_gm(cursor)
    em_cobranca = _em_cobranca_count_por_data_ref(cursor, data_dash_ref)
    kpis = {'em_cobranca': em_cobranca}
    for k in ctx['series']:
        kpis[k] = series_totals.get(k, 0)

    # --- Contratos envolvidos: UNION por série (ocorrência + parcelas performadas 0–90 d) ---
    union_sql, union_p = _painel_dash_union_contrato_ids_sql(ctx['series'], d_ini, d_fim)
    where_series = f"c.id IN {union_sql}" if union_sql else '1=0'

    sql = (
        "SELECT DISTINCT c.id, c.grupo, c.cota, "
        "       CONCAT_WS('/', c.grupo, c.cota) AS grupo_cota, "
        "       (CASE WHEN dash_pvat.min_v_aberto IS NOT NULL "
        "             THEN DATEDIFF(CURRENT_DATE, dash_pvat.min_v_aberto) END) AS atraso, "
        "       c.numero_contrato, c.status, "
        "       c.valor_credito, c.data_adesao, "
        "       CASE "
        "         WHEN dash_pq.id_contrato IS NOT NULL THEN dash_pq.valor_total "
        "         WHEN dash_pa.id_contrato IS NOT NULL "
        "              AND (dash_pi.id_contrato IS NULL OR dash_pa.vencimento <= dash_pi.vencimento) "
        "             THEN dash_pa.valor_total "
        "         WHEN dash_pi.id_contrato IS NOT NULL THEN dash_pi.valor_total "
        "         ELSE NULL "
        "       END AS valor_metrica_parcela, "
        "       p.nome_completo AS devedor, p.cpf_cnpj AS devedor_cpf_cnpj "
        "FROM contrato c "
        "LEFT JOIN pessoa p ON p.id = c.id_pessoa "
        "LEFT JOIN ( "
        "  SELECT id_contrato, MIN(vencimento) AS min_v_aberto "
        "  FROM parcela WHERE status = 'cobranca' GROUP BY id_contrato "
        ") dash_pvat ON dash_pvat.id_contrato = c.id "
        "LEFT JOIN ( "
        "  SELECT p1.id_contrato, p1.vencimento, p1.valor_total "
        "  FROM parcela p1 INNER JOIN ( "
        "    SELECT id_contrato, MIN(vencimento) AS mv "
        "    FROM parcela WHERE status = 'cobranca' GROUP BY id_contrato "
        "  ) t ON p1.id_contrato = t.id_contrato AND p1.vencimento = t.mv AND p1.status = 'cobranca' "
        ") dash_pa ON dash_pa.id_contrato = c.id "
        "LEFT JOIN ( "
        "  SELECT p1.id_contrato, p1.vencimento, p1.valor_total "
        "  FROM parcela p1 INNER JOIN ( "
        "    SELECT id_contrato, MIN(vencimento) AS mv "
        "    FROM parcela WHERE status = 'pago' GROUP BY id_contrato "
        "  ) t ON p1.id_contrato = t.id_contrato AND p1.vencimento = t.mv AND p1.status = 'pago' "
        ") dash_pq ON dash_pq.id_contrato = c.id "
        "LEFT JOIN ( "
        "  SELECT p1.id_contrato, p1.vencimento, p1.valor_total "
        "  FROM parcela p1 INNER JOIN ( "
        "    SELECT id_contrato, MIN(vencimento) AS mv "
        "    FROM parcela WHERE status = 'indenizado' GROUP BY id_contrato "
        "  ) t ON p1.id_contrato = t.id_contrato AND p1.vencimento = t.mv AND p1.status = 'indenizado' "
        ") dash_pi ON dash_pi.id_contrato = c.id "
        f"WHERE {where_series} "
        "ORDER BY c.grupo, c.cota"
    )
    cursor.execute(sql, tuple(union_p))
    contratos = _clean_rows(cursor.fetchall())

    return {
        'meses': meses,
        'series_rows': series_rows,
        'pie_rows': pie_rows,
        'kpis': kpis,
        'contratos': contratos,
    }


def _dash_export_context_labels(ctx):
    series_lbl = ', '.join(_DASH_SERIES_LABELS[s] for s in ctx['series'])
    pie_lbl = ', '.join(_DASH_PIE_LABELS[p] for p in ctx['pie'])
    periodo_lbl = ctx['period_start'] + ' a ' + ctx['period_end']
    return series_lbl, pie_lbl, periodo_lbl


def _dash_export_to_xlsx(ctx, dataset):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()
    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='3B82F6', end_color='3B82F6', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center')
    thin = Side(style='thin', color='D1D5DB')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def _write_sheet(ws, headers, rows):
        for col_idx, h in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = border
        for row_idx, r in enumerate(rows, start=2):
            for col_idx, val in enumerate(r, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.border = border
        for col_idx, h in enumerate(headers, start=1):
            max_len = len(str(h))
            for row_idx in range(2, 2 + len(rows)):
                v = ws.cell(row=row_idx, column=col_idx).value
                max_len = max(max_len, len(str(v)) if v is not None else 0)
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max_len + 4, 40)

    ws0 = wb.active
    ws0.title = 'Parametros'
    series_lbl, pie_lbl, periodo_lbl = _dash_export_context_labels(ctx)
    ws0['A1'] = 'Parametros da exportacao (Dashboard)'
    ws0['A1'].font = Font(bold=True, size=13)
    info = [
        ('Periodo', periodo_lbl),
        ('Series comparadas', series_lbl),
        ('Distribuicao (pizza)', pie_lbl),
        ('Gerado em', datetime.datetime.now().strftime('%d/%m/%Y %H:%M')),
    ]
    for i, (k, v) in enumerate(info, start=3):
        ws0.cell(row=i, column=1, value=k).font = Font(bold=True)
        ws0.cell(row=i, column=2, value=v)
    ws0.column_dimensions['A'].width = 24
    ws0.column_dimensions['B'].width = 60

    ws1 = wb.create_sheet('KPIs')
    k = dataset['kpis']
    kpi_headers = ['KPI', 'Valor']
    kpi_rows = [['Em Cobranca (atual)', k.get('em_cobranca', 0)]]
    for key in ctx['series']:
        kpi_rows.append([_DASH_SERIES_LABELS[key] + ' (periodo)', k.get(key, 0)])
    _write_sheet(ws1, kpi_headers, kpi_rows)

    ws2 = wb.create_sheet('Series Evolucao')
    _write_sheet(ws2, ['Mes', 'Serie', 'Valor'],
                 [[r['mes'], r['serie'], r['valor']] for r in dataset['series_rows']])

    ws3 = wb.create_sheet('Distribuicao Carteira')
    _write_sheet(ws3, ['Status', 'Total'], [[r['status'], r['total']] for r in dataset['pie_rows']])

    ws4 = wb.create_sheet('Contratos')
    headers = ['ID', 'Grupo', 'Cota', 'Grupo/Cota', 'Atraso', 'Nro Contrato', 'Status',
               'Valor parcela (metrica) (R$)', 'Valor credito (ref.) (R$)', 'Data de Adesao',
               'Devedor', 'CPF/CNPJ']
    keys = ['id', 'grupo', 'cota', 'grupo_cota', 'atraso', 'numero_contrato', 'status',
            'valor_metrica_parcela', 'valor_credito', 'data_adesao', 'devedor', 'devedor_cpf_cnpj']
    _dash_m = {'valor_credito', 'valor_metrica_parcela'}

    def _row_dash_contrato(c):
        row = []
        for kk in keys:
            v = c.get(kk, '')
            if kk in _dash_m:
                row.append(_format_reais_pt_br_export(v))
            else:
                row.append(v if v is not None else '')
        return row

    _write_sheet(ws4, headers, [_row_dash_contrato(c) for c in dataset['contratos']])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def _dash_export_to_csv_powerbi(ctx, dataset):
    import csv
    buf = io.StringIO()
    writer = csv.writer(buf, delimiter=';')
    series_lbl, pie_lbl, periodo_lbl = _dash_export_context_labels(ctx)
    writer.writerow(['# Dashboard JB - export Power BI'])
    writer.writerow(['# Periodo', periodo_lbl])
    writer.writerow(['# Series', series_lbl])
    writer.writerow(['# Distribuicao', pie_lbl])
    writer.writerow(['# Gerado em', datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')])
    writer.writerow([])

    # Tabela series (tidy long-format)
    writer.writerow(['tabela', 'mes', 'serie', 'valor'])
    for r in dataset['series_rows']:
        writer.writerow(['series', r['mes'], r['serie'], r['valor']])

    # KPIs
    writer.writerow(['kpi', '-', 'Em Cobranca (atual)', dataset['kpis'].get('em_cobranca', 0)])
    for key in ctx['series']:
        writer.writerow(['kpi', '-', _DASH_SERIES_LABELS[key] + ' (periodo)', dataset['kpis'].get(key, 0)])

    # Pie
    for r in dataset['pie_rows']:
        writer.writerow(['pie', '-', r['status'], r['total']])

    # Contratos
    writer.writerow([
        'tabela', 'id', 'grupo', 'cota', 'grupo_cota', 'atraso', 'numero_contrato', 'status',
        'valor_metrica_parcela', 'valor_credito', 'data_adesao', 'devedor', 'cpf_cnpj',
    ])
    for c in dataset['contratos']:
        writer.writerow([
            'contratos',
            c.get('id') or '',
            c.get('grupo') or '',
            c.get('cota') or '',
            c.get('grupo_cota') or (str(c.get('grupo') or '') + '/' + str(c.get('cota') or '')),
            c.get('atraso') if c.get('atraso') is not None else '',
            c.get('numero_contrato') or '',
            c.get('status') or '',
            c.get('valor_metrica_parcela') if c.get('valor_metrica_parcela') is not None else (
                c.get('valor_credito') or 0),
            c.get('valor_credito') or 0,
            c.get('data_adesao') or '',
            c.get('devedor') or '',
            c.get('devedor_cpf_cnpj') or '',
        ])

    data = buf.getvalue().encode('utf-8-sig')
    out = io.BytesIO(data)
    out.seek(0)
    return out


def _dash_export_to_pdf(ctx, dataset):
    from fpdf import FPDF

    pdf = FPDF(orientation='L', unit='mm', format='A4')
    pdf.set_auto_page_break(auto=True, margin=12)
    pdf.add_page()

    series_lbl, pie_lbl, periodo_lbl = _dash_export_context_labels(ctx)

    pdf.set_font('Helvetica', 'B', 16)
    pdf.cell(0, 10, 'Dashboard JB - Relatorio Exportado', ln=True, align='C')
    pdf.set_font('Helvetica', '', 10)
    pdf.cell(0, 6, 'Periodo: ' + periodo_lbl, ln=True, align='C')
    pdf.cell(0, 6, 'Series: ' + series_lbl, ln=True, align='C')
    pdf.cell(0, 6, 'Distribuicao: ' + pie_lbl, ln=True, align='C')
    pdf.cell(0, 6, 'Gerado em ' + datetime.datetime.now().strftime('%d/%m/%Y %H:%M'), ln=True, align='C')
    pdf.ln(4)

    tmp_files = []
    try:
        line_bytes = _decode_data_url_png(ctx['line_image'])
        pie_bytes = _decode_data_url_png(ctx['pie_image'])

        page_w = pdf.w - 2 * pdf.l_margin
        if line_bytes and pie_bytes:
            tf_l = tempfile.NamedTemporaryFile(suffix='.png', delete=False)
            tf_l.write(line_bytes); tf_l.close(); tmp_files.append(tf_l.name)
            tf_p = tempfile.NamedTemporaryFile(suffix='.png', delete=False)
            tf_p.write(pie_bytes); tf_p.close(); tmp_files.append(tf_p.name)
            img_h = 75
            col_w = (page_w - 6) / 2
            y0 = pdf.get_y()
            pdf.image(tf_l.name, x=pdf.l_margin, y=y0, w=col_w, h=img_h)
            pdf.image(tf_p.name, x=pdf.l_margin + col_w + 6, y=y0, w=col_w, h=img_h)
            pdf.set_y(y0 + img_h + 6)
        elif line_bytes:
            tf_l = tempfile.NamedTemporaryFile(suffix='.png', delete=False)
            tf_l.write(line_bytes); tf_l.close(); tmp_files.append(tf_l.name)
            pdf.image(tf_l.name, x=pdf.l_margin, w=page_w, h=90)
            pdf.ln(4)

        # KPIs
        pdf.set_font('Helvetica', 'B', 11)
        pdf.cell(0, 7, 'KPIs', ln=True)
        pdf.set_font('Helvetica', 'B', 9)
        pdf.set_fill_color(59, 130, 246); pdf.set_text_color(255, 255, 255)
        kpi_headers = ['KPI', 'Valor']
        kcw = [200, 45]
        for i, h in enumerate(kpi_headers):
            pdf.cell(kcw[i], 7, h, border=1, align='C', fill=True)
        pdf.ln()
        pdf.set_font('Helvetica', '', 9)
        pdf.set_text_color(30, 41, 59)
        kpi_items = [('Em Cobranca (atual)', dataset['kpis'].get('em_cobranca', 0))]
        for key in ctx['series']:
            kpi_items.append((_DASH_SERIES_LABELS[key] + ' (periodo)', dataset['kpis'].get(key, 0)))
        for idx, (lab, val) in enumerate(kpi_items):
            fill = idx % 2 == 1
            pdf.set_fill_color(248 if fill else 255, 250 if fill else 255, 252 if fill else 255)
            pdf.cell(kcw[0], 7, str(lab), border=1, align='L', fill=True)
            pdf.cell(kcw[1], 7, str(val), border=1, align='C', fill=True)
            pdf.ln()
        pdf.ln(4)

        # Contratos (ate 300)
        contratos = dataset['contratos'][:300]
        pdf.set_font('Helvetica', 'B', 11)
        pdf.cell(0, 7, 'Contratos envolvidos (' + str(len(dataset['contratos'])) + ' total, mostrando ' + str(len(contratos)) + ')', ln=True)
        pdf.set_font('Helvetica', 'B', 8)
        pdf.set_fill_color(59, 130, 246); pdf.set_text_color(255, 255, 255)
        ch = ['Grupo/Cota', 'Atraso', 'Nro Contrato', 'Status', 'R$ parcela (m.)', 'R$ credito', 'Adesao', 'Devedor', 'CPF/CNPJ']
        cw = [22, 12, 26, 16, 24, 24, 20, 68, 34]
        for i, h in enumerate(ch):
            pdf.cell(cw[i], 6, h, border=1, align='C', fill=True)
        pdf.ln()
        pdf.set_font('Helvetica', '', 7)
        pdf.set_text_color(30, 41, 59)
        for idx, c in enumerate(contratos):
            fill = idx % 2 == 1
            pdf.set_fill_color(248 if fill else 255, 250 if fill else 255, 252 if fill else 255)
            vm = c.get('valor_metrica_parcela')
            vc = c.get('valor_credito')
            try:
                vm_fmt = 'R$ ' + f"{float(vm):,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')
            except Exception:
                vm_fmt = str(vm or '-')
            try:
                vc_fmt = 'R$ ' + f"{float(vc):,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')
            except Exception:
                vc_fmt = str(vc or '-')
            atr = c.get('atraso')
            atr_s = ''
            if atr is not None:
                try:
                    atr_s = str(int(atr))
                except (TypeError, ValueError):
                    atr_s = str(atr)[:8]
            gc_dash = c.get('grupo_cota') or (
                str(c.get('grupo') or '') + '/' + str(c.get('cota') or '')
            )
            row = [
                str(gc_dash)[:22],
                atr_s[:12],
                str(c.get('numero_contrato') or '-'),
                str(c.get('status') or '-'),
                vm_fmt,
                vc_fmt,
                str(c.get('data_adesao') or '-'),
                (str(c.get('devedor') or '-'))[:44],
                str(c.get('devedor_cpf_cnpj') or '-'),
            ]
            for i, v in enumerate(row):
                pdf.cell(cw[i], 6, v, border=1, align='C', fill=True)
            pdf.ln()

        buf = io.BytesIO()
        pdf.output(buf)
        buf.seek(0)
        return buf
    finally:
        for p in tmp_files:
            try: os.unlink(p)
            except Exception: pass


@app.route('/api/dashboard/export/<formato>', methods=['POST'])
def api_dashboard_export(formato):
    formato = (formato or '').lower()
    if formato not in ('xlsx', 'pdf', 'powerbi'):
        return jsonify({'error': 'formato invalido'}), 400

    ctx, err = _resolve_dash_export_payload()
    if err:
        return jsonify({'error': err}), 400

    conn = _get_db()
    cursor = conn.cursor()
    try:
        dataset = _fetch_dash_export_dataset(cursor, ctx)
    finally:
        cursor.close()
        conn.close()

    ts = datetime.datetime.now().strftime('%Y%m%d_%H%M')
    base = f'dashboard_{ctx["period_start"]}_a_{ctx["period_end"]}_{ts}'

    if formato == 'xlsx':
        buf = _dash_export_to_xlsx(ctx, dataset)
        return send_file(buf, as_attachment=True, download_name=base + '.xlsx',
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    if formato == 'pdf':
        buf = _dash_export_to_pdf(ctx, dataset)
        return send_file(buf, as_attachment=True, download_name=base + '.pdf',
                         mimetype='application/pdf')
    buf = _dash_export_to_csv_powerbi(ctx, dataset)
    return send_file(buf, as_attachment=True, download_name=base + '_powerbi.csv',
                     mimetype='text/csv')


# ==========================================
# CRUD funcionário (somente administrador — página Operadores)
# ==========================================

def _coerce_bit_mysql(val, default=1):
    if val is None:
        return default
    if isinstance(val, bool):
        return 1 if val else 0
    s = str(val).strip().lower()
    if s in ('0', 'false', 'off', 'no', 'inativo'):
        return 0
    return 1


def _empty_to_none(val):
    if val is None:
        return None
    if isinstance(val, str) and val.strip() == '':
        return None
    return val


def _parse_optional_int_funcionario_field(val, label):
    """None ou string vazia -> None; caso contrário inteiro ou erro legível."""
    if val is None:
        return None
    if isinstance(val, str) and val.strip() == '':
        return None
    try:
        return int(val)
    except (TypeError, ValueError):
        raise ValueError('%s deve ser um número inteiro ou vazio.' % label)


@app.route('/api/admin/funcionario', methods=['POST'])
def api_admin_funcionario_create():
    """Cadastra um registro em `funcionario`. Exige sessão de administrador."""
    forbidden = _admin_json_forbidden()
    if forbidden:
        return forbidden
    data = request.get_json(force=True, silent=True) or {}
    nome = (data.get('nome') or '').strip()
    cpf_cnpj = (data.get('cpf_cnpj') or '').strip()
    login_v = (data.get('login') or '').strip()
    senha = data.get('senha') or ''
    if not nome or not cpf_cnpj or not login_v:
        return jsonify({'error': 'Nome, CPF e login são obrigatórios.'}), 400
    if len(str(senha)) < 4:
        return jsonify({'error': 'Informe uma senha com pelo menos 4 caracteres.'}), 400

    nivel_acesso = _nivel_acesso_valor_db_amigavel(data.get('nivel_acesso') or 'Cobrança')
    ativo = _coerce_bit_mysql(data.get('ativo'), 1)
    acesso_externo = _coerce_bit_mysql(data.get('acesso_externo'), 0)

    email = _empty_to_none((data.get('email') or '').strip())
    ddd = _empty_to_none((data.get('ddd') or '').strip())
    numero = _empty_to_none((data.get('numero') or '').strip())
    logradouro = _empty_to_none((data.get('logradouro') or '').strip())
    bairro = _empty_to_none((data.get('bairro') or '').strip())
    complemento = _empty_to_none((data.get('complemento') or '').strip())
    cep = _empty_to_none((data.get('cep') or '').strip())
    cidade = _empty_to_none((data.get('cidade') or '').strip())
    estado = _empty_to_none((data.get('estado') or '').strip())
    departamento = _empty_to_none((data.get('departamento') or '').strip())
    sexo = _empty_to_none((data.get('sexo') or '').strip())
    matricula = _empty_to_none((data.get('matricula') or '').strip())
    data_nascimento = _empty_to_none(data.get('data_nascimento'))

    try:
        ramal_v = _parse_optional_int_funcionario_field(data.get('ramal'), 'Ramal')
        fila_v = _parse_optional_int_funcionario_field(data.get('fila'), 'Fila')
    except ValueError as ve:
        return jsonify({'error': str(ve)}), 400
    apikey_v = _empty_to_none((data.get('apikey') or '').strip())
    if apikey_v is not None and len(apikey_v) > 255:
        return jsonify({'error': 'API key deve ter no máximo 255 caracteres.'}), 400

    conn = None
    try:
        conn = _get_db()
        with conn.cursor() as cursor:
            cursor.execute('SELECT id FROM funcionario WHERE cpf_cnpj = %s LIMIT 1', (cpf_cnpj,))
            if cursor.fetchone():
                return jsonify({'error': 'Já existe funcionário com este CPF.'}), 409
            cursor.execute('SELECT id FROM funcionario WHERE login = %s LIMIT 1', (login_v,))
            if cursor.fetchone():
                return jsonify({'error': 'Já existe funcionário com este login.'}), 409
            if ramal_v is not None:
                cursor.execute('SELECT id FROM funcionario WHERE ramal = %s LIMIT 1', (ramal_v,))
                if cursor.fetchone():
                    return jsonify({'error': 'Já existe funcionário com este ramal.'}), 409
            if fila_v is not None:
                cursor.execute('SELECT id FROM funcionario WHERE fila = %s LIMIT 1', (fila_v,))
                if cursor.fetchone():
                    return jsonify({'error': 'Já existe funcionário com esta fila.'}), 409
            if apikey_v is not None:
                cursor.execute('SELECT id FROM funcionario WHERE apikey = %s LIMIT 1', (apikey_v,))
                if cursor.fetchone():
                    return jsonify({'error': 'Já existe funcionário com esta API key.'}), 409

            cursor.execute(
                """
                INSERT INTO funcionario (
                    nome, cpf_cnpj, login, senha, nivel_acesso, ativo, acesso_externo,
                    email, ddd, numero, logradouro, bairro, complemento, cep, cidade, estado,
                    departamento, sexo, matricula, data_nascimento,
                    ramal, fila, apikey
                ) VALUES (
                    %s, %s, %s, %s, %s, %s, %s,
                    %s, %s, %s, %s, %s, %s, %s, %s, %s,
                    %s, %s, %s, %s,
                    %s, %s, %s
                )
                """,
                (
                    nome,
                    cpf_cnpj,
                    login_v,
                    str(senha),
                    nivel_acesso,
                    ativo,
                    acesso_externo,
                    email,
                    ddd,
                    numero,
                    logradouro,
                    bairro,
                    complemento,
                    cep,
                    cidade,
                    estado,
                    departamento,
                    sexo,
                    matricula,
                    data_nascimento,
                    ramal_v,
                    fila_v,
                    apikey_v,
                ),
            )
            new_id = cursor.lastrowid
        conn.commit()
        return jsonify({'ok': True, 'id': new_id})
    except Exception as e:
        if conn is not None:
            conn.rollback()
        app.logger.exception('api_admin_funcionario_create')
        return jsonify({'error': str(e)}), 500
    finally:
        if conn is not None:
            conn.close()


@app.route('/api/admin/funcionario/<int:fid>', methods=['GET', 'PUT'])
def api_admin_funcionario_one(fid):
    """Detalhe (GET) ou atualização (PUT) de um funcionário. Sem devolver senha."""
    forbidden = _admin_json_forbidden()
    if forbidden:
        return forbidden
    conn = None
    try:
        conn = _get_db()
        with conn.cursor() as cursor:
            if request.method == 'GET':
                cursor.execute(
                    """
                    SELECT id, nome, data_nascimento, cpf_cnpj, ativo, login,
                           acesso_externo, email, ddd, numero, logradouro, bairro, complemento,
                           cep, cidade, estado, departamento, nivel_acesso, sexo, matricula,
                           ramal, fila, apikey,
                           created_at, updated_at
                    FROM funcionario WHERE id = %s
                    """,
                    (fid,),
                )
                row = cursor.fetchone()
                if not row:
                    return jsonify({'error': 'Funcionário não encontrado.'}), 404
                return jsonify({'funcionario': _clean_row(row)})

            data = request.get_json(force=True, silent=True) or {}
            cursor.execute('SELECT id FROM funcionario WHERE id = %s', (fid,))
            if not cursor.fetchone():
                return jsonify({'error': 'Funcionário não encontrado.'}), 404

            updates = []
            params = []
            updatable_str = (
                'nome',
                'data_nascimento',
                'cpf_cnpj',
                'email',
                'ddd',
                'numero',
                'logradouro',
                'bairro',
                'complemento',
                'cep',
                'cidade',
                'estado',
                'departamento',
                'nivel_acesso',
                'sexo',
                'matricula',
                'login',
            )
            for key in updatable_str:
                if key not in data:
                    continue
                v = data.get(key)
                if isinstance(v, str) and v.strip() == '':
                    v = None
                if key == 'nivel_acesso' and v is not None:
                    v = _nivel_acesso_valor_db_amigavel(v)
                updates.append(f'`{key}` = %s')
                params.append(v)

            if 'ativo' in data:
                updates.append('`ativo` = %s')
                params.append(_coerce_bit_mysql(data.get('ativo'), 1))
            if 'acesso_externo' in data:
                updates.append('`acesso_externo` = %s')
                params.append(_coerce_bit_mysql(data.get('acesso_externo'), 0))

            pw = data.get('senha')
            if pw is not None and str(pw).strip() != '':
                updates.append('`senha` = %s')
                params.append(str(pw))

            if 'ramal' in data:
                try:
                    ramal_u = _parse_optional_int_funcionario_field(data.get('ramal'), 'Ramal')
                except ValueError as ve:
                    return jsonify({'error': str(ve)}), 400
                updates.append('`ramal` = %s')
                params.append(ramal_u)

            if 'fila' in data:
                try:
                    fila_u = _parse_optional_int_funcionario_field(data.get('fila'), 'Fila')
                except ValueError as ve:
                    return jsonify({'error': str(ve)}), 400
                updates.append('`fila` = %s')
                params.append(fila_u)

            if 'apikey' in data:
                ak = _empty_to_none((data.get('apikey') or '').strip())
                if ak is not None and len(ak) > 255:
                    return jsonify({'error': 'API key deve ter no máximo 255 caracteres.'}), 400
                updates.append('`apikey` = %s')
                params.append(ak)

            if not updates:
                return jsonify({'error': 'Nenhum campo para atualizar.'}), 400

            if 'cpf_cnpj' in data:
                cpf_chk = data.get('cpf_cnpj')
                cursor.execute(
                    'SELECT id FROM funcionario WHERE cpf_cnpj = %s AND id <> %s LIMIT 1',
                    (cpf_chk, fid),
                )
                if cursor.fetchone():
                    return jsonify({'error': 'CPF já cadastrado para outro funcionário.'}), 409
            if 'login' in data:
                login_chk = data.get('login')
                cursor.execute(
                    'SELECT id FROM funcionario WHERE login = %s AND id <> %s LIMIT 1',
                    (login_chk, fid),
                )
                if cursor.fetchone():
                    return jsonify({'error': 'Login já cadastrado para outro funcionário.'}), 409

            if 'ramal' in data:
                try:
                    ramal_dup = _parse_optional_int_funcionario_field(data.get('ramal'), 'Ramal')
                except ValueError as ve:
                    return jsonify({'error': str(ve)}), 400
                if ramal_dup is not None:
                    cursor.execute(
                        'SELECT id FROM funcionario WHERE ramal = %s AND id <> %s LIMIT 1',
                        (ramal_dup, fid),
                    )
                    if cursor.fetchone():
                        return jsonify({'error': 'Ramal já cadastrado para outro funcionário.'}), 409

            if 'fila' in data:
                try:
                    fila_dup = _parse_optional_int_funcionario_field(data.get('fila'), 'Fila')
                except ValueError as ve:
                    return jsonify({'error': str(ve)}), 400
                if fila_dup is not None:
                    cursor.execute(
                        'SELECT id FROM funcionario WHERE fila = %s AND id <> %s LIMIT 1',
                        (fila_dup, fid),
                    )
                    if cursor.fetchone():
                        return jsonify({'error': 'Fila já cadastrada para outro funcionário.'}), 409

            if 'apikey' in data:
                ak_chk = _empty_to_none((data.get('apikey') or '').strip())
                if ak_chk is not None:
                    cursor.execute(
                        'SELECT id FROM funcionario WHERE apikey = %s AND id <> %s LIMIT 1',
                        (ak_chk, fid),
                    )
                    if cursor.fetchone():
                        return jsonify({'error': 'API key já cadastrada para outro funcionário.'}), 409

            params.append(fid)
            sql = 'UPDATE funcionario SET ' + ', '.join(updates) + ' WHERE id = %s'
            cursor.execute(sql, params)
        conn.commit()
        return jsonify({'ok': True})
    except Exception as e:
        if conn is not None:
            conn.rollback()
        app.logger.exception('api_admin_funcionario_one')
        return jsonify({'error': str(e)}), 500
    finally:
        if conn is not None:
            conn.close()


# ==========================================
# AGENDA
# ==========================================
@app.route('/api/funcionarios', methods=['GET'])
def api_funcionarios():
    conn = _get_db()
    cursor = conn.cursor()
    try:
        cursor.execute("SELECT id, nome FROM funcionario WHERE ativo = 1 ORDER BY nome")
        funcis = cursor.fetchall()
        return jsonify(funcis)
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        cursor.close()
        conn.close()


@app.route('/api/funcionario/perfil', methods=['GET'])
def api_funcionario_perfil():
    """Dados do funcionário logado (sem senha nem foto). Inclui funcionario_id para UI (ex.: exclusão do próprio em selects)."""
    fid = session.get('funcionario_id')
    if not fid:
        return jsonify({'error': 'Não autenticado'}), 401
    conn = None
    try:
        conn = _get_db()
        with conn.cursor() as cursor:
            cursor.execute(
                """
                SELECT nome, data_nascimento, cpf_cnpj, ativo, created_at, updated_at,
                       login, acesso_externo, email, ddd, numero, logradouro, bairro,
                       complemento, cep, cidade, estado, departamento, nivel_acesso,
                       sexo, matricula
                FROM funcionario WHERE id = %s
                """,
                (int(fid),),
            )
            row = cursor.fetchone()
    except Exception as e:
        app.logger.exception('api_funcionario_perfil')
        return jsonify({'error': str(e)}), 500
    finally:
        if conn is not None:
            conn.close()
    if not row:
        return jsonify({'error': 'Funcionário não encontrado.'}), 404
    return jsonify({'funcionario': _clean_row(row), 'funcionario_id': int(fid)})


@app.route('/api/agenda', methods=['GET', 'POST'])
def api_agenda():
    conn = _get_db()
    cursor = conn.cursor()
    try:
        if request.method == 'GET':
            month = request.args.get('month')
            year = request.args.get('year')
            
            where = "1=1"
            params = []
            if month and year:
                where = "MONTH(a.data) = %s AND YEAR(a.data) = %s"
                params = [int(month), int(year)]
            
            query = f"""
                SELECT a.*, f.nome as funcionario_nome, c.numero_contrato, c.grupo, c.cota 
                FROM agenda a
                LEFT JOIN funcionario f ON a.id_funcionario = f.id
                LEFT JOIN contrato c ON a.id_contrato = c.id
                WHERE {where}
                ORDER BY a.data ASC
            """
            cursor.execute(query, params)
            tarefas = cursor.fetchall()
            for t in tarefas:
                t['data'] = t['data'].isoformat() if t['data'] else None
            return jsonify(tarefas)

        if request.method == 'POST':
            data = request.json
            atividade = data.get('atividade')
            descricao = data.get('descricao', '')
            data_agenda = data.get('data') # "YYYY-MM-DDTHH:MM" format
            prioridade = data.get('prioridade', 'media')
            id_funcionario = data.get('id_funcionario')
            grupo_cota = data.get('grupo_cota', '').strip()
            
            id_contrato = None
            if grupo_cota:
                parts = tuple(p.strip() for p in grupo_cota.replace('-', '/').split('/', 1))
                if len(parts) == 2:
                    cursor.execute("SELECT id FROM contrato WHERE grupo = %s AND cota = %s", parts)
                    c_row = cursor.fetchone()
                    if c_row:
                        id_contrato = c_row['id']
                    else:
                        return jsonify({'error': 'Contrato com Grupo/Cota informado não foi encontrado.'}), 404
                else:
                    return jsonify({'error': 'Formato de Contrato inválido. Use Grupo/Cota (ex: 50A/0101)'}), 400

            sql = """INSERT INTO agenda 
                     (atividade, descricao, data, prioridade, id_contrato, id_funcionario) 
                     VALUES (%s, %s, %s, %s, %s, %s)"""
            cursor.execute(sql, (atividade, descricao, data_agenda, prioridade, id_contrato, id_funcionario))
            conn.commit()
            return jsonify({'success': True, 'id': cursor.lastrowid})

    except Exception as e:
        app.logger.error("api_agenda error: %s", str(e))
        return jsonify({'error': str(e)}), 500
    finally:
        cursor.close()
        conn.close()

def _agenda_carregar_basico(cursor, id_agenda):
    cursor.execute(
        """
        SELECT a.id, a.id_funcionario, a.id_contrato, a.atividade, a.descricao, a.data,
               a.prioridade, a.status,
               f.nome AS funcionario_nome, c.numero_contrato, c.grupo, c.cota
        FROM agenda a
        LEFT JOIN funcionario f ON a.id_funcionario = f.id
        LEFT JOIN contrato c ON a.id_contrato = c.id
        WHERE a.id = %s
        """,
        (id_agenda,),
    )
    return _clean_row(cursor.fetchone())


def _agenda_sessao_pode_agir(row):
    """Dono do compromisso ou gestor/administrador."""
    if not row:
        return False
    fid_sess = session.get('funcionario_id')
    if fid_sess is None:
        return False
    dono = int(row.get('id_funcionario') or 0)
    return dono == int(fid_sess) or _nivel_normalizado(
        session.get('funcionario_nivel_acesso')
    ) in ('gestor', 'administrador')


@app.route('/api/agenda/<int:id_agenda>', methods=['GET', 'PATCH', 'DELETE'])
def api_agenda_item(id_agenda):
    conn = _get_db()
    cursor = conn.cursor()
    try:
        if request.method == 'GET':
            row = _agenda_carregar_basico(cursor, id_agenda)
            if not row:
                return jsonify({'error': 'Agenda não encontrada.'}), 404
            if not _agenda_sessao_pode_agir(row):
                return jsonify({'error': 'Acesso negado.'}), 403
            return jsonify(row)

        row = _agenda_carregar_basico(cursor, id_agenda)
        if not row:
            return jsonify({'error': 'Agenda não encontrada.'}), 404
        if not _agenda_sessao_pode_agir(row):
            return jsonify({'error': 'Acesso negado.'}), 403

        nivel = _nivel_normalizado(session.get('funcionario_nivel_acesso'))
        eh_cobranca = nivel == 'cobranca'
        fid_sess = int(session.get('funcionario_id'))
        id_ctr = row.get('id_contrato')

        if request.method == 'PATCH':
            data = request.json or {}
            status = data.get('status')
            if status not in ('pendente', 'concluido'):
                return jsonify({'error': 'status invalid'}), 400
            if eh_cobranca and id_ctr is not None:
                _ensure_solicitacao_moderacao_table(cursor)
                if _moderacao_contagem_pendente_agenda(cursor, id_agenda) > 0:
                    return jsonify({'error': 'Ja existe solicitacao pendente para este agendamento.'}), 409
                antes = _tramrow_dict_jsonable(row)
                payload = json.dumps(
                    {'antes': antes, 'proposta': {'status': status}},
                    ensure_ascii=False,
                )
                cursor.execute(
                    """
                    INSERT INTO solicitacao_moderacao
                    (id_solicitante, tipo, id_contrato, ref_id, payload_json, status)
                    VALUES (%s, 'agenda_edit', %s, %s, %s, 'pendente')
                    """,
                    (fid_sess, int(id_ctr), id_agenda, payload),
                )
                mid = cursor.lastrowid
                conn.commit()
                return jsonify({
                    'success': True,
                    'pendente_aprovacao': True,
                    'moderacao_id': mid,
                    'mensagem': 'Pedido enviado para aprovacao do gestor ou administrador (pagina Solicitacoes).',
                })
            cursor.execute("UPDATE agenda SET status = %s WHERE id = %s", (status, id_agenda))
            conn.commit()
            return jsonify({'success': True})

        # DELETE
        if eh_cobranca and id_ctr is not None:
            _ensure_solicitacao_moderacao_table(cursor)
            if _moderacao_contagem_pendente_agenda(cursor, id_agenda) > 0:
                return jsonify({'error': 'Ja existe solicitacao pendente para este agendamento.'}), 409
            antes = _tramrow_dict_jsonable(row)
            payload = json.dumps({'antes': antes}, ensure_ascii=False)
            cursor.execute(
                """
                INSERT INTO solicitacao_moderacao
                (id_solicitante, tipo, id_contrato, ref_id, payload_json, status)
                VALUES (%s, 'agenda_delete', %s, %s, %s, 'pendente')
                """,
                (fid_sess, int(id_ctr), id_agenda, payload),
            )
            mid = cursor.lastrowid
            conn.commit()
            return jsonify({
                'success': True,
                'pendente_aprovacao': True,
                'moderacao_id': mid,
                'mensagem': 'Pedido de exclusao enviado para aprovacao do gestor ou administrador (pagina Solicitacoes).',
            })
        cursor.execute("DELETE FROM agenda WHERE id = %s", (id_agenda,))
        conn.commit()
        return jsonify({'success': True})

    except Exception as e:
        app.logger.exception('api_agenda_item')
        return jsonify({'error': str(e)}), 500
    finally:
        cursor.close()
        conn.close()


# =============================================================================
# Mural de Avisos (tabela `aviso` no MySQL; migracao de data/avisos.json na 1a carga)
# =============================================================================

def _ensure_aviso_table(cursor):
    """Cria a tabela `aviso` alinhada a scripts/criar_banco.py (idempotente)."""
    cursor.execute(
        """
        CREATE TABLE IF NOT EXISTS aviso (
            id BIGINT PRIMARY KEY AUTO_INCREMENT,
            titulo VARCHAR(255) NOT NULL,
            descricao TEXT,
            data_ref DATE NOT NULL,
            created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
            updated_at DATETIME DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP,
            KEY idx_aviso_data_ref (data_ref)
        ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
        """
    )


def _ensure_notificacao_usuario_table(cursor):
    """Tabela de leitura de notificacoes por funcionario (mural + agenda + modulos), idempotente."""
    cursor.execute(
        """
        CREATE TABLE IF NOT EXISTS notificacao_usuario (
            id BIGINT PRIMARY KEY AUTO_INCREMENT,
            id_funcionario INT NOT NULL,
            tipo ENUM(
                'aviso', 'agenda',
                'mensagem', 'solicitacao', 'protocolo'
            ) NOT NULL,
            ref_id BIGINT NOT NULL,
            lida_em TIMESTAMP NOT NULL DEFAULT CURRENT_TIMESTAMP,
            UNIQUE KEY uk_notif_func_tipo_ref (id_funcionario, tipo, ref_id),
            KEY idx_notif_funcionario (id_funcionario),
            CONSTRAINT fk_notificacao_usuario_funcionario
                FOREIGN KEY (id_funcionario) REFERENCES funcionario (id) ON DELETE CASCADE
        ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
        """
    )


def _ensure_solicitacao_moderacao_table(cursor):
    """Pedidos de alteracao/exclusao de tramitacao/agenda (perfil Cobranca -> aprovacao Gestor/Admin)."""
    cursor.execute(
        """
        CREATE TABLE IF NOT EXISTS solicitacao_moderacao (
            id BIGINT PRIMARY KEY AUTO_INCREMENT,
            id_solicitante INT NOT NULL,
            tipo VARCHAR(32) NOT NULL,
            id_contrato BIGINT NOT NULL,
            ref_id BIGINT NOT NULL,
            payload_json TEXT NULL,
            status ENUM('pendente', 'aprovado', 'reprovado') NOT NULL DEFAULT 'pendente',
            id_revisor INT NULL,
            revisado_em DATETIME NULL,
            motivo_reprovacao VARCHAR(512) NULL,
            created_at TIMESTAMP NULL DEFAULT CURRENT_TIMESTAMP,
            KEY idx_sol_mod_status (status),
            KEY idx_sol_mod_solicitante (id_solicitante),
            KEY idx_sol_mod_ref (tipo, ref_id),
            CONSTRAINT fk_sol_mod_solicitante FOREIGN KEY (id_solicitante) REFERENCES funcionario (id),
            CONSTRAINT fk_sol_mod_revisor FOREIGN KEY (id_revisor) REFERENCES funcionario (id),
            CONSTRAINT fk_sol_mod_contrato FOREIGN KEY (id_contrato) REFERENCES contrato (id)
        ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
        """
    )


def _migrate_notificacao_usuario_enum_modulos(cursor, conn):
    """Bancos antigos: amplia ENUM `tipo` para mensagem/solicitacao/protocolo."""
    try:
        cursor.execute(
            """
            SELECT COLUMN_TYPE FROM INFORMATION_SCHEMA.COLUMNS
            WHERE TABLE_SCHEMA = DATABASE()
              AND TABLE_NAME = 'notificacao_usuario'
              AND COLUMN_NAME = 'tipo'
            """
        )
        row = cursor.fetchone()
        if not row:
            return
        ct = row.get('COLUMN_TYPE') if isinstance(row, dict) else row[0]
        ct = (ct or '').lower()
        if 'mensagem' in ct:
            return
        cursor.execute(
            """
            ALTER TABLE notificacao_usuario
            MODIFY COLUMN tipo ENUM(
                'aviso', 'agenda',
                'mensagem', 'solicitacao', 'protocolo'
            ) NOT NULL
            """
        )
        conn.commit()
    except Exception:
        app.logger.exception('_migrate_notificacao_usuario_enum_modulos')


def _aviso_row_to_api_dict(row):
    """Converte linha do banco (dict) no formato esperado pelo front (data_iso)."""
    if not row:
        return None
    dr = row.get("data_ref")
    if dr is not None and hasattr(dr, "isoformat"):
        data_iso = dr.isoformat()[:10]
    else:
        data_iso = (str(dr) if dr is not None else "")[:10]
    out = {
        "id": int(row["id"]),
        "titulo": row.get("titulo") or "",
        "descricao": row.get("descricao") or "",
        "data_iso": data_iso,
    }
    for k in ("created_at", "updated_at"):
        v = row.get(k)
        if v is not None:
            if hasattr(v, "isoformat"):
                out[k] = v.isoformat(timespec="seconds")
            else:
                out[k] = str(v)[:19]
    return out


def _avisos_fix_legacy_pt_br(cursor, conn):
    """Atualiza avisos padrao antigos sem acentuacao (idempotente; nao altera textos personalizados)."""
    fixes = (
        (
            "Nova Politica de Juros",
            "Nova Política de Juros",
            "Atualizacao nas planilhas de calculo exigidas para grupos GM.",
            "Atualização nas planilhas de cálculo exigidas para grupos GM.",
        ),
        (
            "Manutencao no Servidor",
            "Manutenção no Servidor",
            "Agendada uma breve pausa no servidor neste domingo, as 02h.",
            "Agendada uma breve pausa no servidor neste domingo, às 02h.",
        ),
        (
            "Fechamento Mensal",
            "Fechamento Mensal",
            "Lembrete: Os arquivos de repasse devem ser consolidados ate o dia 15.",
            "Lembrete: Os arquivos de repasse devem ser consolidados até o dia 15.",
        ),
    )
    changed = False
    for old_t, new_t, old_d, new_d in fixes:
        cursor.execute(
            "UPDATE aviso SET titulo = %s, descricao = %s WHERE titulo = %s AND descricao = %s",
            (new_t, new_d, old_t, old_d),
        )
        if cursor.rowcount:
            changed = True
    if changed:
        try:
            conn.commit()
        except Exception:
            app.logger.exception("avisos: _avisos_fix_legacy_pt_br commit")


def _avisos_seed_or_migrate(cursor, conn):
    """Tabela vazia: importa data/avisos.json se existir, senao insere avisos padrao; remove o JSON apos import."""
    _avisos_fix_legacy_pt_br(cursor, conn)
    cursor.execute("SELECT COUNT(*) AS c FROM aviso")
    c = int(cursor.fetchone().get("c") or 0)
    if c > 0:
        return
    path_json = os.path.join(os.path.dirname(os.path.abspath(__file__)), "data", "avisos.json")
    if os.path.isfile(path_json):
        try:
            with open(path_json, "r", encoding="utf-8") as f:
                data = json.load(f)
            if isinstance(data, list):
                for a in data:
                    if not isinstance(a, dict):
                        continue
                    tit = (a.get("titulo") or "").strip()[:255]
                    if not tit:
                        continue
                    d_s = (a.get("data_iso") or datetime.date.today().isoformat())[:10]
                    try:
                        datetime.date.fromisoformat(d_s)
                    except ValueError:
                        d_s = datetime.date.today().isoformat()
                    cursor.execute(
                        "INSERT INTO aviso (titulo, descricao, data_ref) VALUES (%s, %s, %s)",
                        (tit, (a.get("descricao") or ""), d_s),
                    )
                conn.commit()
        except Exception:
            app.logger.exception("avisos: migracao a partir de data/avisos.json")
        try:
            os.remove(path_json)
        except OSError:
            pass
        return

    hoje = datetime.date.today()
    seed = [
        (
            "Nova Política de Juros",
            "Atualização nas planilhas de cálculo exigidas para grupos GM.",
            hoje.isoformat(),
        ),
        (
            "Manutenção no Servidor",
            "Agendada uma breve pausa no servidor neste domingo, às 02h.",
            (hoje - datetime.timedelta(days=1)).isoformat(),
        ),
        (
            "Fechamento Mensal",
            "Lembrete: Os arquivos de repasse devem ser consolidados até o dia 15.",
            (hoje - datetime.timedelta(days=10)).isoformat(),
        ),
    ]
    for titulo, descricao, d in seed:
        cursor.execute(
            "INSERT INTO aviso (titulo, descricao, data_ref) VALUES (%s, %s, %s)",
            (titulo, descricao, d),
        )
    conn.commit()


@app.route('/api/avisos', methods=['GET', 'POST'])
def api_avisos_collection():
    conn = _get_db()
    cursor = conn.cursor()
    try:
        _ensure_aviso_table(cursor)
        _avisos_seed_or_migrate(cursor, conn)
    except Exception:
        app.logger.exception("api/avisos: tabela aviso")
        try:
            cursor.close()
        except Exception:
            pass
        try:
            conn.close()
        except Exception:
            pass
        if request.method == "GET":
            return jsonify([]), 200
        return jsonify({"error": "Falha ao preparar avisos"}), 500

    if request.method == "GET":
        try:
            cursor.execute(
                "SELECT id, titulo, descricao, data_ref, created_at, updated_at "
                "FROM aviso ORDER BY data_ref DESC, id DESC"
            )
            rows = _clean_rows(cursor.fetchall())
        finally:
            cursor.close()
            conn.close()
        return jsonify([_aviso_row_to_api_dict(r) for r in rows])

    # POST
    payload = request.get_json(silent=True) or {}
    titulo = (payload.get("titulo") or "").strip()[:255]
    descricao = (payload.get("descricao") or "").strip()
    data_iso = (payload.get("data_iso") or "").strip()

    if not titulo:
        cursor.close()
        conn.close()
        return jsonify({"error": "titulo obrigatorio"}), 400
    if not data_iso:
        data_iso = datetime.date.today().isoformat()
    try:
        datetime.date.fromisoformat(data_iso)
    except ValueError:
        cursor.close()
        conn.close()
        return jsonify({"error": "data_iso invalida (formato YYYY-MM-DD)"}), 400

    try:
        cursor.execute(
            "INSERT INTO aviso (titulo, descricao, data_ref) VALUES (%s, %s, %s)",
            (titulo, descricao, data_iso[:10]),
        )
        conn.commit()
        uid = cursor.lastrowid
        cursor.execute(
            "SELECT id, titulo, descricao, data_ref, created_at, updated_at "
            "FROM aviso WHERE id = %s",
            (uid,),
        )
        r = _clean_row(cursor.fetchone())
    except Exception as e:
        app.logger.exception("api/avisos POST")
        conn.rollback()
        cursor.close()
        conn.close()
        return jsonify({"error": str(e)}), 500
    cursor.close()
    conn.close()
    return jsonify(_aviso_row_to_api_dict(r)), 201


@app.route("/api/avisos/<int:aviso_id>", methods=["PUT", "DELETE"])
def api_avisos_item(aviso_id):
    conn = _get_db()
    cursor = conn.cursor()
    try:
        _ensure_aviso_table(cursor)
        _avisos_seed_or_migrate(cursor, conn)
        cursor.execute(
            "SELECT id, titulo, descricao, data_ref, created_at, updated_at "
            "FROM aviso WHERE id = %s",
            (aviso_id,),
        )
        row = _clean_row(cursor.fetchone())
        if not row:
            return jsonify({"error": "aviso nao encontrado"}), 404

        if request.method == "DELETE":
            antigo = _aviso_row_to_api_dict(row)
            cursor.execute("DELETE FROM aviso WHERE id = %s", (aviso_id,))
            conn.commit()
            return jsonify({"success": True, "removed": antigo})

        payload = request.get_json(silent=True) or {}
        if "titulo" in payload:
            titulo = (payload.get("titulo") or "").strip()[:255]
            if not titulo:
                return jsonify({"error": "titulo obrigatorio"}), 400
        else:
            titulo = row.get("titulo")
        if "descricao" in payload:
            descricao = (payload.get("descricao") or "").strip()
        else:
            descricao = row.get("descricao")
        if "data_iso" in payload:
            data_iso = (payload.get("data_iso") or "").strip()[:10]
            try:
                datetime.date.fromisoformat(data_iso)
            except ValueError:
                return jsonify({"error": "data_iso invalida (formato YYYY-MM-DD)"}), 400
        else:
            dr = row.get("data_ref")
            if hasattr(dr, "isoformat"):
                data_iso = dr.isoformat()[:10]
            else:
                data_iso = str(dr)[:10]

        cursor.execute(
            "UPDATE aviso SET titulo = %s, descricao = %s, data_ref = %s WHERE id = %s",
            (titulo, descricao, data_iso, aviso_id),
        )
        conn.commit()
        cursor.execute(
            "SELECT id, titulo, descricao, data_ref, created_at, updated_at "
            "FROM aviso WHERE id = %s",
            (aviso_id,),
        )
        atual = _clean_row(cursor.fetchone())
        return jsonify(_aviso_row_to_api_dict(atual))
    except Exception as e:
        app.logger.exception("api/avisos item")
        try:
            conn.rollback()
        except Exception:
            pass
        return jsonify({"error": str(e)}), 500
    finally:
        try:
            cursor.close()
        except Exception:
            pass
        try:
            conn.close()
        except Exception:
            pass


# =============================================================================
# Notificacoes (mural de avisos + agenda do operador) — tabela notificacao_usuario
# =============================================================================


def _notif_format_pt_br_dh(value):
    """Formata data/hora MySQL para exibicao no dropdown (naive, fuso do servidor)."""
    if value is None:
        return ""
    if hasattr(value, "strftime"):
        if hasattr(value, "hour"):
            return value.strftime("%d/%m/%Y %H:%M")
        return value.strftime("%d/%m/%Y")
    s = str(value)[:19]
    return s


def _notif_sort_key(value):
    if value is None:
        return datetime.datetime.min
    if isinstance(value, datetime.datetime):
        return value
    if isinstance(value, datetime.date):
        return datetime.datetime.combine(value, datetime.time.min)
    return datetime.datetime.min


@app.route("/api/notificacoes", methods=["GET"])
def api_notificacoes():
    """Lista notificacoes nao lidas: mural, agenda e registros onde o usuario e destinatario (mensagem/solicitacao/protocolo)."""
    raw_fid = session.get("funcionario_id")
    if not raw_fid:
        return jsonify({"error": "Nao autenticado"}), 401
    fid = int(raw_fid)

    conn = _get_db()
    cursor = conn.cursor()
    rows_aviso = []
    rows_agenda = []
    unread_count = 0
    rows_msg = []
    rows_sol = []
    rows_prot = []
    try:
        _ensure_aviso_table(cursor)
        _avisos_fix_legacy_pt_br(cursor, conn)
        _ensure_notificacao_usuario_table(cursor)
        _migrate_notificacao_usuario_enum_modulos(cursor, conn)

        cursor.execute(
            """
            SELECT COUNT(*) AS c
            FROM aviso a
            WHERE NOT EXISTS (
                SELECT 1 FROM notificacao_usuario nu
                WHERE nu.id_funcionario = %s AND nu.tipo = 'aviso' AND nu.ref_id = a.id
            )
            """,
            (fid,),
        )
        count_aviso = int((_clean_row(cursor.fetchone()) or {}).get("c") or 0)

        cursor.execute(
            """
            SELECT COUNT(*) AS c
            FROM agenda a
            WHERE a.id_funcionario = %s
              AND a.status = 'pendente'
              AND a.data <= NOW()
              AND NOT EXISTS (
                SELECT 1 FROM notificacao_usuario nu
                WHERE nu.id_funcionario = %s AND nu.tipo = 'agenda' AND nu.ref_id = a.id
              )
            """,
            (fid, fid),
        )
        count_agenda = int((_clean_row(cursor.fetchone()) or {}).get("c") or 0)

        cursor.execute(
            """
            SELECT COUNT(*) AS c FROM mensagem m
            WHERE m.id_destinatario = %s
              AND NOT EXISTS (
                SELECT 1 FROM notificacao_usuario nu
                WHERE nu.id_funcionario = %s AND nu.tipo = 'mensagem' AND nu.ref_id = m.id
              )
            """,
            (fid, fid),
        )
        count_msg = int((_clean_row(cursor.fetchone()) or {}).get("c") or 0)

        cursor.execute(
            """
            SELECT COUNT(*) AS c FROM solicitacao s
            WHERE s.id_destinatario = %s
              AND NOT EXISTS (
                SELECT 1 FROM notificacao_usuario nu
                WHERE nu.id_funcionario = %s AND nu.tipo = 'solicitacao' AND nu.ref_id = s.id
              )
            """,
            (fid, fid),
        )
        count_sol = int((_clean_row(cursor.fetchone()) or {}).get("c") or 0)

        cursor.execute(
            """
            SELECT COUNT(*) AS c FROM protocolo p
            WHERE p.id_destinatario = %s
              AND NOT EXISTS (
                SELECT 1 FROM notificacao_usuario nu
                WHERE nu.id_funcionario = %s AND nu.tipo = 'protocolo' AND nu.ref_id = p.id
              )
            """,
            (fid, fid),
        )
        count_prot = int((_clean_row(cursor.fetchone()) or {}).get("c") or 0)

        unread_count = count_aviso + count_agenda + count_msg + count_sol + count_prot

        cursor.execute(
            """
            SELECT a.id, a.titulo, a.descricao, a.data_ref, a.created_at
            FROM aviso a
            WHERE NOT EXISTS (
                SELECT 1 FROM notificacao_usuario nu
                WHERE nu.id_funcionario = %s AND nu.tipo = 'aviso' AND nu.ref_id = a.id
            )
            ORDER BY COALESCE(a.created_at, CONCAT(a.data_ref, ' 00:00:00')) DESC, a.id DESC
            LIMIT 40
            """,
            (fid,),
        )
        rows_aviso = _clean_rows(cursor.fetchall())

        cursor.execute(
            """
            SELECT a.id, a.atividade, a.descricao, a.data, a.prioridade, a.status
            FROM agenda a
            WHERE a.id_funcionario = %s
              AND a.status = 'pendente'
              AND a.data <= NOW()
              AND NOT EXISTS (
                SELECT 1 FROM notificacao_usuario nu
                WHERE nu.id_funcionario = %s AND nu.tipo = 'agenda' AND nu.ref_id = a.id
              )
            ORDER BY a.data DESC
            LIMIT 40
            """,
            (fid, fid),
        )
        rows_agenda = _clean_rows(cursor.fetchall())

        cursor.execute(
            """
            SELECT m.id, m.assunto, m.descricao, m.data_envio, fr.nome AS remetente_nome
            FROM mensagem m
            JOIN funcionario fr ON fr.id = m.id_remetente
            WHERE m.id_destinatario = %s
              AND NOT EXISTS (
                SELECT 1 FROM notificacao_usuario nu
                WHERE nu.id_funcionario = %s AND nu.tipo = 'mensagem' AND nu.ref_id = m.id
              )
            ORDER BY m.data_envio DESC, m.id DESC
            LIMIT 40
            """,
            (fid, fid),
        )
        rows_msg = _clean_rows(cursor.fetchall())

        cursor.execute(
            """
            SELECT s.id, s.assunto, s.descricao, s.data_envio, fr.nome AS remetente_nome
            FROM solicitacao s
            JOIN funcionario fr ON fr.id = s.id_remetente
            WHERE s.id_destinatario = %s
              AND NOT EXISTS (
                SELECT 1 FROM notificacao_usuario nu
                WHERE nu.id_funcionario = %s AND nu.tipo = 'solicitacao' AND nu.ref_id = s.id
              )
            ORDER BY s.data_envio DESC, s.id DESC
            LIMIT 40
            """,
            (fid, fid),
        )
        rows_sol = _clean_rows(cursor.fetchall())

        cursor.execute(
            """
            SELECT p.id, p.titulo, p.descricao, p.data_envio, p.created_at, fr.nome AS remetente_nome
            FROM protocolo p
            JOIN funcionario fr ON fr.id = p.id_remetente
            WHERE p.id_destinatario = %s
              AND NOT EXISTS (
                SELECT 1 FROM notificacao_usuario nu
                WHERE nu.id_funcionario = %s AND nu.tipo = 'protocolo' AND nu.ref_id = p.id
              )
            ORDER BY COALESCE(p.created_at, CONCAT(p.data_envio, ' 00:00:00')) DESC, p.id DESC
            LIMIT 40
            """,
            (fid, fid),
        )
        rows_prot = _clean_rows(cursor.fetchall())
    except Exception as e:
        app.logger.exception("api/notificacoes: query")
        return jsonify({"error": str(e)}), 500
    finally:
        try:
            cursor.close()
        except Exception:
            pass
        try:
            conn.close()
        except Exception:
            pass

    merged = []
    for row in rows_aviso:
        created = row.get("created_at")
        data_ref = row.get("data_ref")
        sort_at = created if created is not None else data_ref
        merged.append(
            {
                "tipo": "aviso",
                "ref_id": int(row["id"]),
                "titulo": row.get("titulo") or "",
                "descricao": row.get("descricao") or "",
                "subtitulo": "Mural · %s" % (_notif_format_pt_br_dh(data_ref),),
                "_sk": _notif_sort_key(sort_at),
            }
        )

    for row in rows_agenda:
        dt = row.get("data")
        merged.append(
            {
                "tipo": "agenda",
                "ref_id": int(row["id"]),
                "titulo": row.get("atividade") or "",
                "descricao": row.get("descricao") or "",
                "prioridade": row.get("prioridade") or "",
                "subtitulo": "Agenda · %s" % (_notif_format_pt_br_dh(dt),),
                "_sk": _notif_sort_key(dt),
            }
        )

    for row in rows_msg:
        dt = row.get("data_envio")
        rn = (row.get("remetente_nome") or "").strip()
        merged.append(
            {
                "tipo": "mensagem",
                "ref_id": int(row["id"]),
                "titulo": row.get("assunto") or "",
                "descricao": (row.get("descricao") or "")[:500],
                "subtitulo": "Mensagem · %s%s"
                % (_notif_format_pt_br_dh(dt), (" · " + rn) if rn else ""),
                "_sk": _notif_sort_key(dt),
            }
        )

    for row in rows_sol:
        dt = row.get("data_envio")
        rn = (row.get("remetente_nome") or "").strip()
        merged.append(
            {
                "tipo": "solicitacao",
                "ref_id": int(row["id"]),
                "titulo": row.get("assunto") or "",
                "descricao": (row.get("descricao") or "")[:500],
                "subtitulo": "Solicitação · %s%s"
                % (_notif_format_pt_br_dh(dt), (" · " + rn) if rn else ""),
                "_sk": _notif_sort_key(dt),
            }
        )

    for row in rows_prot:
        dt = row.get("created_at") or row.get("data_envio")
        rn = (row.get("remetente_nome") or "").strip()
        merged.append(
            {
                "tipo": "protocolo",
                "ref_id": int(row["id"]),
                "titulo": row.get("titulo") or "",
                "descricao": (row.get("descricao") or "")[:500],
                "subtitulo": "Protocolo · %s%s"
                % (_notif_format_pt_br_dh(dt), (" · " + rn) if rn else ""),
                "_sk": _notif_sort_key(dt),
            }
        )

    merged.sort(key=lambda x: x["_sk"], reverse=True)
    items = []
    for it in merged[:18]:
        entry = {k: v for k, v in it.items() if k != "_sk"}
        items.append(entry)

    return jsonify({"items": items, "unread_count": unread_count})


@app.route("/api/notificacoes/todas", methods=["GET"])
def api_notificacoes_todas():
    """Mural, agenda e mensagens/solicitações/protocolos recebidos pelo usuario, com flag `lida`."""
    raw_fid = session.get("funcionario_id")
    if not raw_fid:
        return jsonify({"error": "Nao autenticado"}), 401
    fid = int(raw_fid)

    conn = _get_db()
    cursor = conn.cursor()
    rows_aviso = []
    rows_agenda = []
    rows_msg = []
    rows_sol = []
    rows_prot = []
    try:
        _ensure_aviso_table(cursor)
        _avisos_fix_legacy_pt_br(cursor, conn)
        _ensure_notificacao_usuario_table(cursor)
        _migrate_notificacao_usuario_enum_modulos(cursor, conn)

        cursor.execute(
            """
            SELECT a.id, a.titulo, a.descricao, a.data_ref, a.created_at,
                   CASE WHEN nu.id IS NOT NULL THEN 1 ELSE 0 END AS lida
            FROM aviso a
            LEFT JOIN notificacao_usuario nu
              ON nu.id_funcionario = %s AND nu.tipo = 'aviso' AND nu.ref_id = a.id
            ORDER BY COALESCE(a.created_at, CONCAT(a.data_ref, ' 00:00:00')) DESC, a.id DESC
            LIMIT 100
            """,
            (fid,),
        )
        rows_aviso = _clean_rows(cursor.fetchall())

        cursor.execute(
            """
            SELECT a.id, a.atividade, a.descricao, a.data, a.prioridade, a.status,
                   CASE WHEN nu.id IS NOT NULL THEN 1 ELSE 0 END AS lida
            FROM agenda a
            LEFT JOIN notificacao_usuario nu
              ON nu.id_funcionario = %s AND nu.tipo = 'agenda' AND nu.ref_id = a.id
            WHERE a.id_funcionario = %s
              AND a.status = 'pendente'
              AND a.data <= NOW()
            ORDER BY a.data DESC
            LIMIT 100
            """,
            (fid, fid),
        )
        rows_agenda = _clean_rows(cursor.fetchall())

        cursor.execute(
            """
            SELECT m.id, m.assunto, m.descricao, m.data_envio, fr.nome AS remetente_nome,
                   CASE WHEN nu.id IS NOT NULL THEN 1 ELSE 0 END AS lida
            FROM mensagem m
            JOIN funcionario fr ON fr.id = m.id_remetente
            LEFT JOIN notificacao_usuario nu
              ON nu.id_funcionario = %s AND nu.tipo = 'mensagem' AND nu.ref_id = m.id
            WHERE m.id_destinatario = %s
            ORDER BY m.data_envio DESC, m.id DESC
            LIMIT 100
            """,
            (fid, fid),
        )
        rows_msg = _clean_rows(cursor.fetchall())

        cursor.execute(
            """
            SELECT s.id, s.assunto, s.descricao, s.data_envio, fr.nome AS remetente_nome,
                   CASE WHEN nu.id IS NOT NULL THEN 1 ELSE 0 END AS lida
            FROM solicitacao s
            JOIN funcionario fr ON fr.id = s.id_remetente
            LEFT JOIN notificacao_usuario nu
              ON nu.id_funcionario = %s AND nu.tipo = 'solicitacao' AND nu.ref_id = s.id
            WHERE s.id_destinatario = %s
            ORDER BY s.data_envio DESC, s.id DESC
            LIMIT 100
            """,
            (fid, fid),
        )
        rows_sol = _clean_rows(cursor.fetchall())

        cursor.execute(
            """
            SELECT p.id, p.titulo, p.descricao, p.data_envio, p.created_at, fr.nome AS remetente_nome,
                   CASE WHEN nu.id IS NOT NULL THEN 1 ELSE 0 END AS lida
            FROM protocolo p
            JOIN funcionario fr ON fr.id = p.id_remetente
            LEFT JOIN notificacao_usuario nu
              ON nu.id_funcionario = %s AND nu.tipo = 'protocolo' AND nu.ref_id = p.id
            WHERE p.id_destinatario = %s
            ORDER BY COALESCE(p.created_at, CONCAT(p.data_envio, ' 00:00:00')) DESC, p.id DESC
            LIMIT 100
            """,
            (fid, fid),
        )
        rows_prot = _clean_rows(cursor.fetchall())
    except Exception as e:
        app.logger.exception("api/notificacoes/todas")
        return jsonify({"error": str(e)}), 500
    finally:
        try:
            cursor.close()
        except Exception:
            pass
        try:
            conn.close()
        except Exception:
            pass

    merged = []
    for row in rows_aviso:
        created = row.get("created_at")
        data_ref = row.get("data_ref")
        sort_at = created if created is not None else data_ref
        merged.append(
            {
                "tipo": "aviso",
                "ref_id": int(row["id"]),
                "titulo": row.get("titulo") or "",
                "descricao": row.get("descricao") or "",
                "subtitulo": "Mural · %s" % (_notif_format_pt_br_dh(data_ref),),
                "lida": bool(int(row.get("lida") or 0)),
                "_sk": _notif_sort_key(sort_at),
            }
        )

    for row in rows_agenda:
        dt = row.get("data")
        merged.append(
            {
                "tipo": "agenda",
                "ref_id": int(row["id"]),
                "titulo": row.get("atividade") or "",
                "descricao": row.get("descricao") or "",
                "prioridade": row.get("prioridade") or "",
                "subtitulo": "Agenda · %s" % (_notif_format_pt_br_dh(dt),),
                "lida": bool(int(row.get("lida") or 0)),
                "_sk": _notif_sort_key(dt),
            }
        )

    for row in rows_msg:
        dt = row.get("data_envio")
        rn = (row.get("remetente_nome") or "").strip()
        merged.append(
            {
                "tipo": "mensagem",
                "ref_id": int(row["id"]),
                "titulo": row.get("assunto") or "",
                "descricao": (row.get("descricao") or "")[:500],
                "subtitulo": "Mensagem · %s%s"
                % (_notif_format_pt_br_dh(dt), (" · " + rn) if rn else ""),
                "lida": bool(int(row.get("lida") or 0)),
                "_sk": _notif_sort_key(dt),
            }
        )

    for row in rows_sol:
        dt = row.get("data_envio")
        rn = (row.get("remetente_nome") or "").strip()
        merged.append(
            {
                "tipo": "solicitacao",
                "ref_id": int(row["id"]),
                "titulo": row.get("assunto") or "",
                "descricao": (row.get("descricao") or "")[:500],
                "subtitulo": "Solicitação · %s%s"
                % (_notif_format_pt_br_dh(dt), (" · " + rn) if rn else ""),
                "lida": bool(int(row.get("lida") or 0)),
                "_sk": _notif_sort_key(dt),
            }
        )

    for row in rows_prot:
        dt = row.get("created_at") or row.get("data_envio")
        rn = (row.get("remetente_nome") or "").strip()
        merged.append(
            {
                "tipo": "protocolo",
                "ref_id": int(row["id"]),
                "titulo": row.get("titulo") or "",
                "descricao": (row.get("descricao") or "")[:500],
                "subtitulo": "Protocolo · %s%s"
                % (_notif_format_pt_br_dh(dt), (" · " + rn) if rn else ""),
                "lida": bool(int(row.get("lida") or 0)),
                "_sk": _notif_sort_key(dt),
            }
        )

    merged.sort(key=lambda x: x["_sk"], reverse=True)
    items = []
    for it in merged:
        entry = {k: v for k, v in it.items() if k != "_sk"}
        items.append(entry)

    return jsonify({"items": items})


@app.route("/api/notificacoes/lida", methods=["POST"])
def api_notificacoes_marcar_lida():
    """Marca como lida: aviso, agenda ou item recebido (mensagem/solicitacao/protocolo)."""
    raw_fid = session.get("funcionario_id")
    if not raw_fid:
        return jsonify({"error": "Nao autenticado"}), 401
    fid = int(raw_fid)

    payload = request.get_json(silent=True) or {}
    tipo = (payload.get("tipo") or "").strip().lower()
    ref_raw = payload.get("ref_id")
    try:
        ref_id = int(ref_raw)
    except (TypeError, ValueError):
        return jsonify({"error": "ref_id invalido"}), 400

    if tipo not in ("aviso", "agenda", "mensagem", "solicitacao", "protocolo"):
        return jsonify({"error": "tipo invalido"}), 400

    conn = _get_db()
    cursor = conn.cursor()
    try:
        _ensure_notificacao_usuario_table(cursor)
        _migrate_notificacao_usuario_enum_modulos(cursor, conn)
        if tipo == "agenda":
            cursor.execute(
                "SELECT id FROM agenda WHERE id = %s AND id_funcionario = %s",
                (ref_id, fid),
            )
            if not cursor.fetchone():
                return jsonify({"error": "agenda nao encontrada ou nao pertence ao usuario"}), 404
        elif tipo == "aviso":
            cursor.execute("SELECT id FROM aviso WHERE id = %s", (ref_id,))
            if not cursor.fetchone():
                return jsonify({"error": "aviso nao encontrado"}), 404
        elif tipo == "mensagem":
            cursor.execute(
                "SELECT id FROM mensagem WHERE id = %s AND id_destinatario = %s",
                (ref_id, fid),
            )
            if not cursor.fetchone():
                return jsonify({"error": "mensagem nao encontrada ou destinatario invalido"}), 404
        elif tipo == "solicitacao":
            cursor.execute(
                "SELECT id FROM solicitacao WHERE id = %s AND id_destinatario = %s",
                (ref_id, fid),
            )
            if not cursor.fetchone():
                return jsonify({"error": "solicitacao nao encontrada ou destinatario invalido"}), 404
        else:
            cursor.execute(
                "SELECT id FROM protocolo WHERE id = %s AND id_destinatario = %s",
                (ref_id, fid),
            )
            if not cursor.fetchone():
                return jsonify({"error": "protocolo nao encontrado ou destinatario invalido"}), 404

        cursor.execute(
            """
            INSERT INTO notificacao_usuario (id_funcionario, tipo, ref_id)
            VALUES (%s, %s, %s)
            ON DUPLICATE KEY UPDATE lida_em = CURRENT_TIMESTAMP
            """,
            (fid, tipo, ref_id),
        )
        conn.commit()
    except Exception as e:
        app.logger.exception("api/notificacoes/lida")
        try:
            conn.rollback()
        except Exception:
            pass
        return jsonify({"error": str(e)}), 500
    finally:
        try:
            cursor.close()
        except Exception:
            pass
        try:
            conn.close()
        except Exception:
            pass

    return jsonify({"success": True})


# =============================================================================
# Negativação (Serasa) - tabela auto-criada na primeira utilização
# -----------------------------------------------------------------------------
# Regra de negocio:
#   * Um contrato se torna ELEGIVEL a negativacao quando sua parcela em aberto
#     mais antiga tem entre 31 e 89 dias de atraso (inclusive).
#   * Uma vez enviado ao Serasa, NAO e re-enviado automaticamente (o usuario
#     tem que expressamente desejar uma nova negativacao -- futuro).
# =============================================================================

def _ensure_negativacao_table(cursor):
    """Cria a tabela `negativacao` se ainda nao existir.

    Negativacao acontece por PARCELA (nao por contrato). `id_parcela` e a
    chave de unicidade: uma parcela so pode ser gravada aqui uma unica
    vez. Isso resolve o bug anterior em que o simples EXISTS por
    id_contrato fazia o contrato aparecer "negativado" indefinidamente,
    mesmo quando a parcela atual em aberto era outra.
    """
    cursor.execute(
        """
        CREATE TABLE IF NOT EXISTS negativacao (
            id BIGINT PRIMARY KEY AUTO_INCREMENT,
            id_contrato BIGINT NOT NULL,
            id_parcela  BIGINT NOT NULL,
            numero_parcela INT NULL,
            dias_atraso INT NULL,
            data_negativacao DATETIME DEFAULT CURRENT_TIMESTAMP,
            status VARCHAR(32) NOT NULL DEFAULT 'enviado',
            resposta_api TEXT NULL,
            created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
            updated_at DATETIME DEFAULT CURRENT_TIMESTAMP
                                 ON UPDATE CURRENT_TIMESTAMP,
            UNIQUE KEY uk_negativacao_parcela (id_parcela),
            KEY idx_negativacao_contrato (id_contrato),
            CONSTRAINT fk_negativacao_contrato FOREIGN KEY (id_contrato)
                REFERENCES contrato(id) ON DELETE CASCADE,
            CONSTRAINT fk_negativacao_parcela FOREIGN KEY (id_parcela)
                REFERENCES parcela(id) ON DELETE CASCADE
        ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
        """
    )
    _ensure_negativacao_id_funcionario_column(cursor)


def _ensure_negativacao_id_funcionario_column(cursor):
    """Adiciona `id_funcionario` em `negativacao` quando a tabela ja existia sem a coluna."""
    try:
        cursor.execute(
            "SELECT COUNT(*) AS c FROM information_schema.COLUMNS "
            "WHERE TABLE_SCHEMA = DATABASE() AND TABLE_NAME = 'negativacao' "
            "AND COLUMN_NAME = 'id_funcionario'"
        )
        row = cursor.fetchone()
        cnt = int(row['c'] if isinstance(row, dict) else (row[0] if row else 0))
        if cnt == 0:
            cursor.execute(
                "ALTER TABLE negativacao ADD COLUMN id_funcionario INT NULL "
                "DEFAULT NULL AFTER resposta_api"
            )
    except Exception:
        pass


def _ensure_negativacao_historico_table(cursor):
    """Historico append-only de eventos de negativação / positivação por contrato."""
    cursor.execute(
        """
        CREATE TABLE IF NOT EXISTS negativacao_historico (
            id BIGINT PRIMARY KEY AUTO_INCREMENT,
            id_arquivo_gm BIGINT NULL DEFAULT NULL,
            id_contrato BIGINT NOT NULL,
            id_parcela BIGINT NULL,
            numero_parcela INT NULL,
            tipo_evento VARCHAR(40) NOT NULL,
            data_evento DATETIME NOT NULL DEFAULT CURRENT_TIMESTAMP,
            id_funcionario INT NULL,
            dias_atraso INT NULL,
            status_snapshot VARCHAR(64) NULL,
            detalhe TEXT NULL,
            created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
            KEY idx_neg_hist_arquivo_gm (id_arquivo_gm),
            KEY idx_neg_hist_contrato (id_contrato),
            KEY idx_neg_hist_parcela (id_parcela),
            KEY idx_neg_hist_data (data_evento),
            KEY idx_neg_hist_tipo (tipo_evento),
            CONSTRAINT fk_neg_hist_contrato FOREIGN KEY (id_contrato)
                REFERENCES contrato(id) ON DELETE CASCADE,
            CONSTRAINT fk_neg_hist_arquivo_gm FOREIGN KEY (id_arquivo_gm)
                REFERENCES arquivos_gm(id_arquivo_gm) ON DELETE SET NULL
        ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
        """
    )
    _ensure_negativacao_historico_id_arquivo_gm_column(cursor)


def _ensure_negativacao_historico_id_arquivo_gm_column(cursor):
    """Adiciona `id_arquivo_gm` em `negativacao_historico` quando a tabela ja existia sem a coluna."""
    try:
        cursor.execute(
            "SELECT COUNT(*) AS c FROM information_schema.COLUMNS "
            "WHERE TABLE_SCHEMA = DATABASE() AND TABLE_NAME = 'negativacao_historico' "
            "AND COLUMN_NAME = 'id_arquivo_gm'"
        )
        row = cursor.fetchone()
        cnt = int(row['c'] if isinstance(row, dict) else (row[0] if row else 0))
        if cnt == 0:
            cursor.execute(
                "ALTER TABLE negativacao_historico "
                "ADD COLUMN id_arquivo_gm BIGINT NULL DEFAULT NULL AFTER id"
            )
            try:
                cursor.execute(
                    "ALTER TABLE negativacao_historico "
                    "ADD KEY idx_neg_hist_arquivo_gm (id_arquivo_gm)"
                )
            except Exception:
                pass
            try:
                cursor.execute(
                    "ALTER TABLE negativacao_historico "
                    "ADD CONSTRAINT fk_neg_hist_arquivo_gm FOREIGN KEY (id_arquivo_gm) "
                    "REFERENCES arquivos_gm(id_arquivo_gm) ON DELETE SET NULL"
                )
            except Exception:
                pass
    except Exception:
        pass


def _ensure_registro_txt_conteudo_mediumtext(cursor, table_name):
    """Garante coluna `conteudo` MEDIUMTEXT (lotes SERASA podem exceder limite TEXT)."""
    if table_name not in ('registro_txt_negativacao', 'registro_txt_positivacao'):
        return
    try:
        cursor.execute(
            "SELECT DATA_TYPE AS dt FROM information_schema.COLUMNS "
            "WHERE TABLE_SCHEMA = DATABASE() AND TABLE_NAME = %s AND COLUMN_NAME = 'conteudo'",
            (table_name,),
        )
        row = cursor.fetchone()
        if not row:
            return
        dt = (row.get('dt') if isinstance(row, dict) else row[0]) or ''
        if str(dt).lower() != 'mediumtext':
            cursor.execute(
                f"ALTER TABLE `{table_name}` MODIFY COLUMN conteudo MEDIUMTEXT NOT NULL"
            )
    except Exception:
        pass


def _ensure_registro_txt_negativacao_table(cursor):
    cursor.execute(
        """
        CREATE TABLE IF NOT EXISTS registro_txt_negativacao (
            id BIGINT PRIMARY KEY AUTO_INCREMENT,
            nome_arquivo VARCHAR(255) NULL,
            conteudo MEDIUMTEXT NOT NULL,
            id_funcionario INT NOT NULL,
            created_at TIMESTAMP NULL DEFAULT CURRENT_TIMESTAMP,
            updated_at TIMESTAMP NULL DEFAULT CURRENT_TIMESTAMP
                ON UPDATE CURRENT_TIMESTAMP,
            KEY id_funcionario (id_funcionario),
            CONSTRAINT registro_txt_negativacao_ibfk_1 FOREIGN KEY (id_funcionario)
                REFERENCES funcionario(id) ON DELETE CASCADE
        ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
        """
    )
    _ensure_registro_txt_conteudo_mediumtext(cursor, 'registro_txt_negativacao')


def _ensure_registro_txt_positivacao_table(cursor):
    cursor.execute(
        """
        CREATE TABLE IF NOT EXISTS registro_txt_positivacao (
            id BIGINT PRIMARY KEY AUTO_INCREMENT,
            nome_arquivo VARCHAR(255) NULL,
            conteudo MEDIUMTEXT NOT NULL,
            id_funcionario INT NOT NULL,
            created_at TIMESTAMP NULL DEFAULT CURRENT_TIMESTAMP,
            updated_at TIMESTAMP NULL DEFAULT CURRENT_TIMESTAMP
                ON UPDATE CURRENT_TIMESTAMP,
            KEY id_funcionario (id_funcionario),
            CONSTRAINT registro_txt_positivacao_ibfk_1 FOREIGN KEY (id_funcionario)
                REFERENCES funcionario(id) ON DELETE CASCADE
        ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
        """
    )
    _ensure_registro_txt_conteudo_mediumtext(cursor, 'registro_txt_positivacao')


def _registro_txt_serasa_insert(cursor, conn, *, tipo, nome_arquivo, conteudo_bytes, id_funcionario):
    """Grava copia do TXT SERASA gerado em registro_txt_negativacao ou registro_txt_positivacao."""
    if tipo == 'negativar':
        _ensure_registro_txt_negativacao_table(cursor)
        table = 'registro_txt_negativacao'
    else:
        _ensure_registro_txt_positivacao_table(cursor)
        table = 'registro_txt_positivacao'
    conteudo = conteudo_bytes.decode('latin-1', errors='replace')
    cursor.execute(
        f"INSERT INTO {table} (nome_arquivo, conteudo, id_funcionario) VALUES (%s, %s, %s)",
        (nome_arquivo, conteudo, int(id_funcionario)),
    )
    conn.commit()


SERASA_GM_SEQ_INICIAL = 4912
_RE_SERASA_GM_NOME = re.compile(r'^SERASA_GM_(\d{8})(\d{4})', re.IGNORECASE)


def _serasa_gm_parse_nome_arquivo(nome):
    """Extrai data DDMMYYYY e sequencia 4 digitos de ``SERASA_GM_<data><seq>``."""
    if not nome:
        return None, None
    base = str(nome).strip()
    if base.upper().endswith('.TXT'):
        base = base[:-4]
    m = _RE_SERASA_GM_NOME.match(base)
    if not m:
        return None, None
    return m.group(1), int(m.group(2))


def _serasa_gm_montar_nome_arquivo(data_ddmmyyyy, sequencia):
    return f'SERASA_GM_{data_ddmmyyyy}{int(sequencia):04d}.TXT'


def _serasa_gm_ultimo_registro_global(cursor):
    """Registo mais recente entre negativacao e positivacao (por ``created_at``)."""
    _ensure_registro_txt_negativacao_table(cursor)
    _ensure_registro_txt_positivacao_table(cursor)
    candidatos = []
    for table in ('registro_txt_negativacao', 'registro_txt_positivacao'):
        cursor.execute(
            f"SELECT nome_arquivo, created_at FROM `{table}` "
            "ORDER BY created_at DESC, id DESC LIMIT 1"
        )
        row = cursor.fetchone()
        if row:
            candidatos.append(row)
    if not candidatos:
        return None
    return max(
        candidatos,
        key=lambda r: r.get('created_at') if isinstance(r, dict) else r[1],
    )


def _serasa_gm_proxima_sequencia_global(cursor):
    """Proxima sequencia global (+1) e nome com data de hoje (DDMMYYYY)."""
    data_hoje = datetime.datetime.now().strftime('%d%m%Y')
    ultimo = _serasa_gm_ultimo_registro_global(cursor)
    ultimo_nome = None
    if ultimo:
        ultimo_nome = ultimo.get('nome_arquivo') if isinstance(ultimo, dict) else ultimo[0]
    _data_ant, seq_ant = _serasa_gm_parse_nome_arquivo(ultimo_nome)
    if seq_ant is not None:
        sequencia = seq_ant + 1
        if sequencia > 9999:
            raise ValueError('Sequencia SERASA excede 9999; ajuste manualmente.')
    else:
        sequencia = SERASA_GM_SEQ_INICIAL
    nome = _serasa_gm_montar_nome_arquivo(data_hoje, sequencia)
    return {
        'sequencia': sequencia,
        'data_arquivo': data_hoje,
        'nome_arquivo': nome,
        'ultimo_nome_arquivo': ultimo_nome,
    }


def _serasa_gm_parse_sequencia_payload(raw):
    """Valida sequencia 0-9999 a partir do JSON do cliente."""
    if raw is None or (isinstance(raw, str) and not str(raw).strip()):
        return None, 'Informe a sequencia do arquivo SERASA (4 digitos).'
    try:
        if isinstance(raw, str):
            s = raw.strip()
            if not s.isdigit() or len(s) > 4:
                return None, 'Sequencia invalida (use ate 4 digitos numericos).'
            seq = int(s)
        else:
            seq = int(raw)
    except (TypeError, ValueError):
        return None, 'Sequencia invalida.'
    if seq < 0 or seq > 9999:
        return None, 'Sequencia deve estar entre 0 e 9999.'
    return seq, None


def _insert_negativacao_historico(
    cursor,
    id_contrato,
    id_parcela,
    numero_parcela,
    tipo_evento,
    data_evento,
    id_funcionario=None,
    dias_atraso=None,
    status_snapshot=None,
    detalhe=None,
    id_arquivo_gm=None,
):
    """Insere um evento no histórico (sem deduplicação — quem chama decide)."""
    if data_evento is None:
        data_evento = datetime.datetime.now()
    _ensure_negativacao_historico_table(cursor)
    cursor.execute(
        """
        INSERT INTO negativacao_historico (
            id_arquivo_gm, id_contrato, id_parcela, numero_parcela, tipo_evento, data_evento,
            id_funcionario, dias_atraso, status_snapshot, detalhe
        ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
        """,
        (
            id_arquivo_gm,
            id_contrato,
            id_parcela,
            numero_parcela,
            tipo_evento,
            data_evento,
            id_funcionario,
            dias_atraso,
            status_snapshot,
            detalhe,
        ),
    )


# Registros de `negativacao` muito antigos podem nao ter `id_parcela` (modelo
# por contrato). O sentinel abaixo marca isso no mapa de conjuntos.
_NEG_NULL_PARCELA = object()


def _mapa_parcelas_negativadas_por_contrato(cursor, contrato_ids):
    """{id_contrato: set} onde cada set contem id_parcela (int) e, se
    houver negativacao legada sem parcela, o valor _NEG_NULL_PARCELA.

    Necessario em vez de `WHERE id_parcela IN (lista de alvos desta tela)`:
    essa lista nao abrange parcelas de outros contratos nem linhas cujo
    id_parcela nao e o alvo atual, o que deixava o filtro "Ja negativados"
    vazio apesar de existir negativacao no banco.
    """
    if not contrato_ids:
        return {}
    out = {}
    ph = ','.join(['%s'] * len(contrato_ids))
    cursor.execute(
        f"SELECT id_contrato, id_parcela FROM negativacao "
        f"WHERE id_contrato IN ({ph})",
        list(contrato_ids),
    )
    for row in cursor.fetchall():
        cid = int(row['id_contrato'])
        s = out.setdefault(cid, set())
        if row.get('id_parcela') is None:
            s.add(_NEG_NULL_PARCELA)
        else:
            try:
                s.add(int(row['id_parcela']))
            except (TypeError, ValueError):
                s.add(_NEG_NULL_PARCELA)
    return out


# =============================================================================
# Automacoes de cobranca (SMS, e-mail, ligacao)
# -----------------------------------------------------------------------------
# Endpoints proxy entre o frontend e a API que o colega Flavio esta criando.
# Por enquanto retornam sucesso simulado e logam o payload no console; quando a
# API estiver pronta basta substituir o bloco TODO por uma chamada real
# (requests.post / httpx) e propagar a resposta.
# =============================================================================

_AUTOMACAO_TIPOS = {'sms', 'email', 'ligacao', 'sms_email'}


def _automacao_log(tipo, payload, extra=None):
    try:
        print(f"[automacao:{tipo}] payload={json.dumps(payload, ensure_ascii=False)}"
              + (f" extra={extra}" if extra else ''),
              flush=True)
    except Exception:
        pass


def _validar_ids_contratos(ids):
    if not isinstance(ids, list):
        return None, 'contrato_ids deve ser uma lista.'
    out = []
    for v in ids:
        try:
            out.append(int(v))
        except (TypeError, ValueError):
            return None, f'ID invalido na lista: {v!r}'
    if not out:
        return None, 'Nenhum contrato informado.'
    return out, None


_COBRANCA_BLOCO_EXCEL_COLUMNS_BEFORE_PHONES = (
    ('Ordem', 'ordem'),
    ('Grupo', 'grupo'),
    ('Cota', 'cota'),
    ('Nº contrato', 'numero_contrato'),
    ('Nome devedor', 'nome_devedor'),
    ('CPF/CNPJ', 'cpf_cnpj'),
    ('Dias atraso', 'dias_atraso'),
    ('Vencimento mais antigo', 'vencimento_mais_antigo'),
    ('Parcelas abertas', 'parcelas_abertas'),
    ('Nº parcela alvo', 'numero_parcela_alvo'),
    ('Operador cobrança', 'nome_funcionario'),
)

_COBRANCA_BLOCO_EXCEL_COLUMNS_AFTER_PHONES = (
    ('E-mail(s)', 'emails'),
    ('R$ crédito', 'valor_credito'),
    ('Status negativação', 'status_negativacao'),
)


def _cobranca_bloco_excel_column_defs(max_telefones):
    """Colunas fixas + Telefone 1..N (N = máximo de números num contrato da exportação)."""
    cols = list(_COBRANCA_BLOCO_EXCEL_COLUMNS_BEFORE_PHONES)
    for i in range(1, max(0, int(max_telefones or 0)) + 1):
        cols.append((f'Telefone {i}', f'telefone_{i}'))
    cols.extend(_COBRANCA_BLOCO_EXCEL_COLUMNS_AFTER_PHONES)
    return cols


def _contatos_listas_por_pessoa(cursor, pessoa_ids):
    """Telefones (lista por pessoa) e e-mails (texto com '; ')."""
    if not pessoa_ids:
        return {}, {}
    ids = list(dict.fromkeys(pessoa_ids))
    tel_map = _telefones_por_pessoas(cursor, ids)
    email_map = _emails_por_pessoas(cursor, ids)
    sep = '; '
    tel_lists = {}
    em_txt = {}
    for pid in ids:
        nums = []
        for tel in tel_map.get(pid, []):
            n = (tel.get('numero') or '').strip()
            if n:
                nums.append(n)
        mails = []
        for em in email_map.get(pid, []):
            e = (em.get('email') or '').strip()
            if e:
                mails.append(e)
        tel_lists[pid] = nums
        em_txt[pid] = sep.join(mails)
    return tel_lists, em_txt


def _cobranca_bloco_excel_linhas(cursor, contrato_ids):
    """Contratos do bloco (ordem da lista enviada), com contatos para lista de ligação."""
    if not contrato_ids:
        return []
    ids = []
    seen = set()
    for cid in contrato_ids:
        try:
            i = int(cid)
        except (TypeError, ValueError):
            continue
        if i not in seen:
            seen.add(i)
            ids.append(i)
    if not ids:
        return []

    ph = ','.join(['%s'] * len(ids))
    order_ph = ','.join(['%s'] * len(ids))
    sql = (
        'SELECT c.id, c.grupo, c.cota, c.numero_contrato, c.valor_credito, c.id_pessoa, '
        '       p.nome_completo AS nome_devedor, p.cpf_cnpj, '
        '       MIN(par.vencimento) AS vencimento_mais_antigo, '
        '       DATEDIFF(CURRENT_DATE, MIN(par.vencimento)) AS dias_atraso, '
        '       COUNT(DISTINCT par.id) AS parcelas_abertas, '
        '       f.nome AS nome_funcionario, '
        '       ( '
        '         SELECT p2.numero_parcela FROM parcela p2 '
        '         WHERE p2.id_contrato = c.id AND p2.status = %s '
        '         ORDER BY p2.vencimento ASC, p2.id ASC LIMIT 1 '
        '       ) AS numero_parcela_alvo, '
        '       ( '
        '         SELECT p3.id FROM parcela p3 '
        '         WHERE p3.id_contrato = c.id AND p3.status = %s '
        '         ORDER BY p3.vencimento ASC, p3.id ASC LIMIT 1 '
        '       ) AS id_parcela_alvo '
        'FROM contrato c '
        'INNER JOIN parcela par ON par.id_contrato = c.id AND par.status = %s '
        'LEFT JOIN pessoa p ON c.id_pessoa = p.id '
        'LEFT JOIN funcionario_cobranca fc ON fc.id_contrato = c.id '
        'LEFT JOIN funcionario f ON f.id = fc.id_funcionario '
        f'WHERE c.id IN ({ph}) AND c.status = %s '
        'GROUP BY c.id, c.grupo, c.cota, c.numero_contrato, c.valor_credito, c.id_pessoa, '
        '         p.nome_completo, p.cpf_cnpj, f.nome '
        f'ORDER BY FIELD(c.id, {order_ph})'
    )
    params = ['cobranca', 'cobranca', 'cobranca'] + ids + ['cobranca'] + ids
    cursor.execute(sql, params)
    rows = _clean_rows(cursor.fetchall())

    pessoa_ids = []
    cids = []
    for r in rows:
        try:
            cids.append(int(r['id']))
        except (TypeError, ValueError, KeyError):
            pass
        pid = r.get('id_pessoa')
        if pid is not None:
            try:
                pessoa_ids.append(int(pid))
            except (TypeError, ValueError):
                pass

    try:
        neg_por_contrato = _mapa_parcelas_negativadas_por_contrato(cursor, cids)
    except Exception:
        app.logger.exception('cobranca_bloco_excel: falha ao consultar negativacao')
        neg_por_contrato = {}

    tel_lists, em_txt = _contatos_listas_por_pessoa(cursor, pessoa_ids)

    out = []
    for ordem, r in enumerate(rows, start=1):
        try:
            cid = int(r['id'])
        except (TypeError, ValueError, KeyError):
            cid = None
        try:
            dias_i = int(r.get('dias_atraso') or 0)
        except (TypeError, ValueError):
            dias_i = 0
        id_alvo = r.get('id_parcela_alvo')
        try:
            id_alvo = int(id_alvo) if id_alvo is not None else None
        except (TypeError, ValueError):
            id_alvo = None
        neg_set = neg_por_contrato.get(cid, set()) if cid is not None else set()
        ja_neg = (_NEG_NULL_PARCELA in neg_set) or (
            id_alvo is not None and id_alvo in neg_set
        )
        if ja_neg:
            st_neg = 'negativado'
        elif 30 < dias_i < 90:
            st_neg = 'pendente'
        else:
            st_neg = 'nao_elegivel'

        pid = r.get('id_pessoa')
        try:
            pid_i = int(pid) if pid is not None else None
        except (TypeError, ValueError):
            pid_i = None

        venc = r.get('vencimento_mais_antigo')
        if hasattr(venc, 'isoformat'):
            venc = venc.isoformat()

        cred = r.get('valor_credito')
        try:
            cred_f = float(cred) if cred is not None else None
        except (TypeError, ValueError):
            cred_f = None

        out.append({
            'id': cid,
            'ordem': ordem,
            'grupo': r.get('grupo') or '',
            'cota': r.get('cota') or '',
            'numero_contrato': r.get('numero_contrato') or '',
            'nome_devedor': r.get('nome_devedor') or '',
            'cpf_cnpj': r.get('cpf_cnpj') or '',
            'dias_atraso': dias_i,
            'vencimento_mais_antigo': venc or '',
            'parcelas_abertas': int(r.get('parcelas_abertas') or 0),
            'numero_parcela_alvo': r.get('numero_parcela_alvo') or '',
            'nome_funcionario': r.get('nome_funcionario') or '',
            'telefones': list(tel_lists.get(pid_i, [])) if pid_i is not None else [],
            'emails': em_txt.get(pid_i, '') if pid_i is not None else '',
            'valor_credito': cred_f,
            'status_negativacao': st_neg,
        })
    return out


def _write_cobranca_bloco_excel(wb, linhas):
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    ws = wb.active
    ws.title = 'Lista ligacao'
    max_tels = 0
    for row in linhas:
        tels = row.get('telefones') or []
        if len(tels) > max_tels:
            max_tels = len(tels)
    columns = _cobranca_bloco_excel_column_defs(max_tels)
    ncols = len(columns)

    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='047857', end_color='047857', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin', color='D1D5DB'),
        right=Side(style='thin', color='D1D5DB'),
        top=Side(style='thin', color='D1D5DB'),
        bottom=Side(style='thin', color='D1D5DB'),
    )
    alt_fill = PatternFill(start_color='F8FAFC', end_color='F8FAFC', fill_type='solid')
    brl_fmt = '#,##0.00'

    for col_idx, (label, _key) in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    for row_idx, row in enumerate(linhas, start=2):
        tels = row.get('telefones') or []
        for col_idx, (_label, key) in enumerate(columns, start=1):
            if key.startswith('telefone_'):
                try:
                    tel_i = int(key.split('_', 1)[1]) - 1
                except (TypeError, ValueError, IndexError):
                    tel_i = -1
                val = tels[tel_i] if 0 <= tel_i < len(tels) else ''
            else:
                val = row.get(key, '')
            cell = ws.cell(row=row_idx, column=col_idx, value=val if val is not None else '')
            cell.border = thin_border
            if (row_idx - 2) % 2 == 1:
                cell.fill = alt_fill
            if key == 'valor_credito' and isinstance(val, (int, float)) and not isinstance(val, bool):
                cell.number_format = brl_fmt

    widths_before = [8, 10, 8, 14, 32, 16, 11, 16, 12, 12, 22]
    widths_tel = [16] * max_tels
    widths_after = [32, 14, 18]
    widths = widths_before + widths_tel + widths_after
    for col_idx, w in enumerate(widths[:ncols], start=1):
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = w


@app.route('/api/automacao/preview', methods=['POST'])
@app.route('/api/cobranca/sms-email/preview', methods=['POST'])
def api_automacao_preview():
    """Preview de SMS/e-mail para os IDs da carteira (mesma lógica da Lista SMS/E-mail na importação)."""
    if not session.get('funcionario_id'):
        return jsonify({'error': 'Nao autenticado. Faca login novamente.'}), 401
    payload = request.get_json(silent=True) or {}
    ids, err = _validar_ids_contratos(payload.get('contrato_ids'))
    if err:
        return jsonify({'error': err}), 400

    conn = _get_db()
    try:
        data = _analise_automacao_carteira(conn, ids)
    except Exception:
        app.logger.exception('api_automacao_preview: falha ao analisar carteira')
        return jsonify({
            'error': 'Falha ao calcular o preview. Tente novamente; se persistir, verifique os logs do servidor.',
        }), 500
    finally:
        conn.close()

    return jsonify({'ok': True, **data})


@app.route('/api/cobranca/sms-email/excel', methods=['POST'])
def api_cobranca_sms_email_excel():
    """Excel com folhas SMS e EMAIL: mesmo roteiro que a Importação, filtrado aos IDs da carteira."""
    if not session.get('funcionario_id'):
        return jsonify({'error': 'Nao autenticado. Faca login novamente.'}), 401
    payload = request.get_json(silent=True) or {}
    ids, err = _validar_ids_contratos(payload.get('contrato_ids'))
    if err:
        return jsonify({'error': err}), 400

    try:
        from openpyxl import Workbook
    except ImportError:
        return jsonify({
            'error': 'Biblioteca openpyxl nao instalada. Execute: pip install openpyxl',
        }), 503

    conn = _get_db()
    try:
        wb = Workbook()
        with conn.cursor() as cursor:
            linhas = _sms_autom_excel_linhas_carteira(cursor, ids)
            _sms_autom_fill_rota_excel_duas_folhas(
                cursor,
                wb,
                linhas,
                'Nenhum contrato no roteiro de dias (0, 16, 31, 61, 85) nesta lista.',
            )
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
    except Exception as exc:
        app.logger.exception('cobranca/sms-email/excel: falha ao consultar ou montar ficheiro')
        return jsonify({'error': f'Falha ao gerar o Excel: {exc}'}), 500
    finally:
        try:
            conn.close()
        except Exception:
            pass

    fname = (
        'sms_email_carteira_cobranca_'
        + datetime.datetime.now().strftime('%Y%m%d_%H%M')
        + '.xlsx'
    )
    return send_file(
        buf,
        as_attachment=True,
        download_name=fname,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )


@app.route('/api/cobranca/bloco/excel', methods=['POST'])
def api_cobranca_bloco_excel():
    """Excel da lista de contratos de um bloco (ex.: antes das ligações sequenciais)."""
    if not session.get('funcionario_id'):
        return jsonify({'error': 'Nao autenticado. Faca login novamente.'}), 401
    payload = request.get_json(silent=True) or {}
    ids, err = _validar_ids_contratos(payload.get('contrato_ids'))
    if err:
        return jsonify({'error': err}), 400

    nivel = (payload.get('nivel') or 'bloco').strip().lower()

    try:
        from openpyxl import Workbook
    except ImportError:
        return jsonify({
            'error': 'Biblioteca openpyxl nao instalada. Execute: pip install openpyxl',
        }), 503

    conn = _get_db()
    try:
        with conn.cursor() as cursor:
            linhas = _cobranca_bloco_excel_linhas(cursor, ids)
        if not linhas:
            return jsonify({'error': 'Nenhum contrato em cobranca encontrado para os IDs informados.'}), 404

        wb = Workbook()
        _write_cobranca_bloco_excel(wb, linhas)
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
    except Exception as exc:
        app.logger.exception('cobranca/bloco/excel: falha ao gerar ficheiro')
        return jsonify({'error': f'Falha ao gerar o Excel: {exc}'}), 500
    finally:
        try:
            conn.close()
        except Exception:
            pass

    slug = re.sub(r'[^a-z0-9]+', '_', nivel.lower()).strip('_') or 'bloco'
    fname = (
        'lista_ligacao_cobranca_'
        + slug
        + '_'
        + datetime.datetime.now().strftime('%Y%m%d_%H%M')
        + '.xlsx'
    )
    return send_file(
        buf,
        as_attachment=True,
        download_name=fname,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )


_ligacao_bloco_jobs = {}
_ligacao_bloco_lock = threading.Lock()

_LIGACAO_BLOCO_NIVEL_LABEL = {
    'critico': 'Crítico',
    'atencao': 'Atenção',
    'recente': 'Recente',
    'todos': 'Todos os blocos',
}


def _ensure_ligacao_bloco_controle_table(cursor):
    """Controle diario: contrato com sucesso+tramitacao nao volta ao bloco no mesmo dia."""
    cursor.execute(
        """
        CREATE TABLE IF NOT EXISTS ligacao_bloco_controle (
            id BIGINT(20) NOT NULL AUTO_INCREMENT,
            id_funcionario INT(11) NOT NULL,
            id_contrato BIGINT(20) NOT NULL,
            data_ref DATE NOT NULL,
            resultado ENUM('sucesso_tramitado', 'pulado', 'falha') NOT NULL,
            nivel VARCHAR(32) DEFAULT NULL,
            created_at TIMESTAMP NOT NULL DEFAULT CURRENT_TIMESTAMP,
            updated_at TIMESTAMP NOT NULL DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP,
            PRIMARY KEY (id),
            UNIQUE KEY uk_ligacao_bloco_dia (id_funcionario, id_contrato, data_ref),
            KEY idx_ligacao_bloco_contrato (id_contrato),
            CONSTRAINT fk_ligacao_bloco_func FOREIGN KEY (id_funcionario)
                REFERENCES funcionario (id) ON DELETE CASCADE,
            CONSTRAINT fk_ligacao_bloco_contrato FOREIGN KEY (id_contrato)
                REFERENCES contrato (id) ON DELETE CASCADE
        ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
        """
    )


def _ligacao_bloco_data_ref():
    return datetime.date.today()


def _ligacao_bloco_contratos_sucesso_tramitado_hoje(cursor, funcionario_id, contrato_ids):
    """Contratos que ja tiveram ligacao com tramitacao registrada hoje (nao rediscar)."""
    if not contrato_ids:
        return set()
    _ensure_ligacao_bloco_controle_table(cursor)
    ph = ','.join(['%s'] * len(contrato_ids))
    cursor.execute(
        f"""
        SELECT id_contrato FROM ligacao_bloco_controle
        WHERE id_funcionario = %s AND data_ref = %s AND resultado = 'sucesso_tramitado'
          AND id_contrato IN ({ph})
        """,
        [int(funcionario_id), _ligacao_bloco_data_ref()] + list(contrato_ids),
    )
    out = set()
    for row in cursor.fetchall() or []:
        try:
            out.add(int(row['id_contrato']))
        except (TypeError, ValueError, KeyError):
            pass
    return out


def _ligacao_bloco_contratos_com_agenda(cursor, contrato_ids):
    """Contratos com compromisso pendente na agenda (bloqueio de ligacao em cobranca)."""
    if not contrato_ids:
        return {}
    ph = ','.join(['%s'] * len(contrato_ids))
    cursor.execute(
        f"""
        SELECT a.id_contrato, MIN(a.data) AS proxima_data, COUNT(*) AS qtd
        FROM agenda a
        WHERE a.id_contrato IN ({ph}) AND a.status = 'pendente'
        GROUP BY a.id_contrato
        """,
        list(contrato_ids),
    )
    bloqueados = {}
    for row in _clean_rows(cursor.fetchall()):
        try:
            cid = int(row['id_contrato'])
        except (TypeError, ValueError, KeyError):
            continue
        prox = row.get('proxima_data')
        if hasattr(prox, 'strftime'):
            prox_txt = prox.strftime('%d/%m/%Y %H:%M')
        else:
            prox_txt = str(prox or '')
        bloqueados[cid] = {
            'motivo': 'agendamento',
            'motivo_label': 'Agendamento pendente na agenda',
            'proxima_agenda': prox_txt,
            'qtd_agenda': int(row.get('qtd') or 0),
        }
    return bloqueados


def _ligacao_bloco_linha_resumo(row):
    return {
        'contrato_id': row.get('id'),
        'grupo': row.get('grupo') or '',
        'cota': row.get('cota') or '',
        'numero_contrato': row.get('numero_contrato') or '',
        'nome_devedor': row.get('nome_devedor') or '',
    }


def _ligacao_bloco_analisar_preview(cursor, funcionario_id, contrato_ids):
    """Preview antes do disparo: exclusoes e contagem da fila."""
    ids = []
    seen = set()
    for cid in contrato_ids or []:
        try:
            i = int(cid)
        except (TypeError, ValueError):
            continue
        if i not in seen:
            seen.add(i)
            ids.append(i)
    if not ids:
        return {
            'total_solicitados': 0,
            'elegiveis_contratos': 0,
            'ligacoes_previstas': 0,
            'excluidos_ja_tramitados_hoje': [],
            'excluidos_agendamento': [],
            'excluidos_sem_telefone': [],
            'contrato_ids_elegiveis': [],
        }

    linhas = _cobranca_bloco_excel_linhas(cursor, ids)
    por_id = {}
    for row in linhas:
        try:
            por_id[int(row.get('id'))] = row
        except (TypeError, ValueError):
            pass

    ja_tramit = _ligacao_bloco_contratos_sucesso_tramitado_hoje(cursor, funcionario_id, ids)
    com_agenda = _ligacao_bloco_contratos_com_agenda(cursor, ids)

    ex_tramit = []
    ex_agenda = []
    ex_sem_tel = []
    elegiveis_ids = []

    for cid in ids:
        row = por_id.get(cid)
        if cid in ja_tramit:
            base = _ligacao_bloco_linha_resumo(row) if row else {'contrato_id': cid}
            base['motivo'] = 'ja_tramitado_hoje'
            base['motivo_label'] = 'Ligacao com tramitacao ja registrada hoje'
            ex_tramit.append(base)
            continue
        if cid in com_agenda:
            base = _ligacao_bloco_linha_resumo(row) if row else {'contrato_id': cid}
            base.update(com_agenda[cid])
            ex_agenda.append(base)
            continue
        if not row:
            continue
        tels = row.get('telefones') or []
        if isinstance(tels, str):
            tels = [p.strip() for p in tels.split(';') if p.strip()]
        if not tels:
            base = _ligacao_bloco_linha_resumo(row)
            base['motivo'] = 'sem_telefone'
            base['motivo_label'] = 'Sem telefone cadastrado'
            ex_sem_tel.append(base)
            continue
        elegiveis_ids.append(cid)

    queue = _ligacao_bloco_queue_from_linhas([por_id[c] for c in elegiveis_ids if c in por_id])

    return {
        'total_solicitados': len(ids),
        'elegiveis_contratos': len(elegiveis_ids),
        'ligacoes_previstas': len(queue),
        'excluidos_ja_tramitados_hoje': ex_tramit,
        'excluidos_agendamento': ex_agenda,
        'excluidos_sem_telefone': ex_sem_tel,
        'contrato_ids_elegiveis': elegiveis_ids,
    }


def _ligacao_bloco_tem_tramitacao_hoje(cursor, funcionario_id, contrato_id):
    cursor.execute(
        """
        SELECT 1 FROM tramitacao
        WHERE id_contrato = %s AND id_funcionario = %s
          AND DATE(created_at) = CURDATE()
        LIMIT 1
        """,
        (int(contrato_id), int(funcionario_id)),
    )
    return cursor.fetchone() is not None


def _ligacao_bloco_persist_resultado(cursor, funcionario_id, contrato_id, resultado, nivel=None):
    _ensure_ligacao_bloco_controle_table(cursor)
    if resultado == 'sucesso_tramitado':
        cursor.execute(
            """
            INSERT INTO ligacao_bloco_controle
                (id_funcionario, id_contrato, data_ref, resultado, nivel)
            VALUES (%s, %s, %s, %s, %s)
            ON DUPLICATE KEY UPDATE
                resultado = VALUES(resultado),
                nivel = VALUES(nivel),
                updated_at = CURRENT_TIMESTAMP
            """,
            (int(funcionario_id), int(contrato_id), _ligacao_bloco_data_ref(), resultado, nivel),
        )
    elif resultado == 'pulado':
        cursor.execute(
            """
            INSERT INTO ligacao_bloco_controle
                (id_funcionario, id_contrato, data_ref, resultado, nivel)
            VALUES (%s, %s, %s, %s, %s)
            ON DUPLICATE KEY UPDATE
                resultado = IF(resultado = 'sucesso_tramitado', resultado, VALUES(resultado)),
                nivel = VALUES(nivel),
                updated_at = CURRENT_TIMESTAMP
            """,
            (int(funcionario_id), int(contrato_id), _ligacao_bloco_data_ref(), resultado, nivel),
        )


def _ligacao_bloco_contratos_bloqueados_hoje(cursor, funcionario_id):
    _ensure_ligacao_bloco_controle_table(cursor)
    cursor.execute(
        """
        SELECT id_contrato FROM ligacao_bloco_controle
        WHERE id_funcionario = %s AND data_ref = %s AND resultado = 'sucesso_tramitado'
        """,
        (int(funcionario_id), _ligacao_bloco_data_ref()),
    )
    return {int(r['id_contrato']) for r in (cursor.fetchall() or []) if r.get('id_contrato') is not None}


def _ligacao_bloco_skip_indice_ate_elegivel(job, bloqueados):
    queue = job.get('queue') or []
    idx = int(job.get('index') or 0)
    while idx < len(queue):
        try:
            cid = int(queue[idx].get('contrato_id'))
        except (TypeError, ValueError):
            idx += 1
            continue
        if cid not in bloqueados:
            break
        idx += 1
    job['index'] = idx
    if idx >= len(queue):
        job['phase'] = 'completed'
        job['status_text'] = 'Sequencia concluida.'


def _ligacao_bloco_registrar_historico(job, item, resultado_ligacao, tramitacao_registrada=False):
    if not item:
        return
    try:
        cid = int(item.get('contrato_id'))
    except (TypeError, ValueError):
        return
    hist = job.setdefault('historico', [])
    hist.append({
        'contrato_id': cid,
        'grupo': item.get('grupo'),
        'cota': item.get('cota'),
        'numero': item.get('numero'),
        'resultado': resultado_ligacao,
        'tramitacao_registrada': bool(tramitacao_registrada),
    })
    proc = job.setdefault('contratos_processados', set())
    if isinstance(proc, list):
        proc = set(proc)
        job['contratos_processados'] = proc
    proc.add(cid)


def _ligacao_bloco_build_resumo(job):
    """Resumo para modal ao encerrar ou concluir (estilo print do usuario)."""
    queue = job.get('queue') or []
    hist = job.get('historico') or []
    sucessos = int(job.get('successes') or 0)
    pulados = int(job.get('pulados') or 0)
    falhas = int(job.get('failures') or 0)
    idx = int(job.get('index') or 0)

    contratos_total = int(job.get('contratos_total_bloco') or 0)
    if not contratos_total:
        contratos_total = len({h.get('contrato_id') for h in hist if h.get('contrato_id')})

    sucesso_sem_tramit = []
    phase = job.get('phase')
    if phase == 'awaiting_tramitacao':
        cur = job.get('current')
        last = job.get('last_dial') or {}
        if cur and last.get('ok') and not job.get('tramitacao_ok'):
            sucesso_sem_tramit.append({
                'contrato_id': cur.get('contrato_id'),
                'grupo': cur.get('grupo'),
                'cota': cur.get('cota'),
                'nome_devedor': cur.get('nome_devedor'),
                'numero': cur.get('numero'),
            })

    ligacoes_total = len(queue)
    ligacoes_processadas = min(idx, ligacoes_total)

    return {
        'sucessos': sucessos,
        'pulados': pulados,
        'falhas': falhas,
        'ligacoes_processadas': ligacoes_processadas,
        'ligacoes_total': ligacoes_total,
        'contratos_total_bloco': contratos_total,
        'sucesso_sem_tramitacao': sucesso_sem_tramit,
        'cancelado': bool(job.get('cancelled')),
        'nivel': job.get('nivel'),
        'nivel_label': job.get('nivel_label'),
    }


def _ligacao_bloco_queue_from_linhas(linhas):
    """Fila plana: um item por número de telefone, na ordem dos contratos da lista."""
    queue = []
    for row in linhas or []:
        tels = row.get('telefones') or []
        if isinstance(tels, str):
            tels = [p.strip() for p in tels.split(';') if p.strip()]
        if not tels:
            continue
        try:
            cid = int(row.get('id'))
        except (TypeError, ValueError):
            continue
        for tel_i, numero in enumerate(tels, start=1):
            queue.append({
                'contrato_id': cid,
                'grupo': row.get('grupo') or '',
                'cota': row.get('cota') or '',
                'numero_contrato': row.get('numero_contrato') or '',
                'nome_devedor': row.get('nome_devedor') or '',
                'cpf_cnpj': row.get('cpf_cnpj') or '',
                'dias_atraso': row.get('dias_atraso') or 0,
                'vencimento_mais_antigo': row.get('vencimento_mais_antigo') or '',
                'parcelas_abertas': row.get('parcelas_abertas') or 0,
                'numero_parcela_alvo': row.get('numero_parcela_alvo') or '',
                'nome_funcionario': row.get('nome_funcionario') or '',
                'status_negativacao': row.get('status_negativacao') or '',
                'numero': numero,
                'telefone_indice': tel_i,
                'telefones_no_contrato': len(tels),
            })
    return queue


def _ligacao_bloco_job_public(job):
    """Estado serializável do job para polling / painel."""
    queue = job.get('queue') or []
    total = len(queue)
    idx = int(job.get('index') or 0)
    if idx < 0:
        idx = 0
    if idx > total:
        idx = total
    current = job.get('current')
    if not current and 0 <= idx < total:
        current = queue[idx]
    done = idx if job.get('phase') in ('awaiting_continue', 'completed', 'cancelled') else idx
    if job.get('phase') == 'dialing' or job.get('phase') == 'awaiting_tramitacao':
        pos = idx + 1
    elif job.get('phase') == 'awaiting_continue':
        pos = idx + 1
    elif job.get('phase') == 'completed':
        pos = total
    else:
        pos = min(idx + 1, total) if total else 0
    return {
        'job_id': job.get('job_id'),
        'nivel': job.get('nivel'),
        'nivel_label': job.get('nivel_label'),
        'phase': job.get('phase'),
        'running': bool(job.get('running')),
        'cancelled': bool(job.get('cancelled')),
        'index': idx,
        'total': total,
        'posicao': pos,
        'current': current,
        'last_dial': job.get('last_dial'),
        'tramitacao_ok': bool(job.get('tramitacao_ok')),
        'successes': int(job.get('successes') or 0),
        'failures': int(job.get('failures') or 0),
        'pulados': int(job.get('pulados') or 0),
        'skipped_sem_telefone': int(job.get('skipped_sem_telefone') or 0),
        'contratos_total_bloco': int(job.get('contratos_total_bloco') or 0),
        'status_text': job.get('status_text') or '',
        'resumo': _ligacao_bloco_build_resumo(job) if job.get('phase') in (
            'completed', 'cancelled'
        ) else None,
    }


def _ligacao_bloco_dial_worker(job_id):
    """Disca o item atual da fila (thread daemon)."""
    with _ligacao_bloco_lock:
        job = _ligacao_bloco_jobs.get(job_id)
        if not job or job.get('cancelled'):
            return
        queue = job.get('queue') or []
        idx = int(job.get('index') or 0)
        if idx >= len(queue):
            job['phase'] = 'completed'
            job['running'] = False
            job['status_text'] = 'Sequencia concluida.'
            return
        item = queue[idx]
        job['current'] = item
        job['phase'] = 'dialing'
        job['running'] = True
        job['tramitacao_ok'] = False
        job['status_text'] = f"Discando {item.get('numero', '')}..."
        fid = job.get('funcionario_id')

    result = _discar_numero_funcionario(fid, item.get('numero'))

    with _ligacao_bloco_lock:
        job = _ligacao_bloco_jobs.get(job_id)
        if not job or job.get('cancelled'):
            return
        job['running'] = False
        job['last_dial'] = result
        if result.get('ok'):
            job['successes'] = int(job.get('successes') or 0) + 1
            _ligacao_bloco_registrar_historico(job, item, 'sucesso', False)
        else:
            job['failures'] = int(job.get('failures') or 0) + 1
            _ligacao_bloco_registrar_historico(job, item, 'falha', False)
        job['phase'] = 'awaiting_tramitacao'
        job['tramitacao_ok'] = False
        job['status_text'] = (
            'Ligacao concluida. Registre a tramitacao para continuar.'
            if result.get('ok')
            else (result.get('error') or 'Falha na discagem. Registre a tramitacao para continuar.')
        )


def _ligacao_bloco_start_dial(job_id):
    th = threading.Thread(target=_ligacao_bloco_dial_worker, args=(job_id,), daemon=True)
    th.start()


def _ligacao_bloco_get_job(job_id, funcionario_id):
    with _ligacao_bloco_lock:
        job = _ligacao_bloco_jobs.get(job_id)
    if not job or str(job.get('funcionario_id')) != str(funcionario_id):
        return None
    return job


@app.route('/api/cobranca/ligacao-bloco/preview', methods=['POST'])
def api_cobranca_ligacao_bloco_preview():
    """Preview do bloco: exclusoes (agenda, ja tramitado hoje, sem telefone)."""
    fid = session.get('funcionario_id')
    if not fid:
        return jsonify({'error': 'Nao autenticado.'}), 401
    payload = request.get_json(silent=True) or {}
    ids, err = _validar_ids_contratos(payload.get('contrato_ids'))
    if err:
        return jsonify({'error': err}), 400
    conn = _get_db()
    try:
        with conn.cursor() as cursor:
            data = _ligacao_bloco_analisar_preview(cursor, fid, ids)
    finally:
        conn.close()
    return jsonify({'ok': True, **data})


@app.route('/api/cobranca/ligacao-bloco/start', methods=['POST'])
def api_cobranca_ligacao_bloco_start():
    """Inicia sequencia de ligacoes do bloco em background (fila por telefone)."""
    fid = session.get('funcionario_id')
    if not fid:
        return jsonify({'error': 'Nao autenticado.'}), 401
    payload = request.get_json(silent=True) or {}
    ids, err = _validar_ids_contratos(payload.get('contrato_ids'))
    if err:
        return jsonify({'error': err}), 400
    nivel = (payload.get('nivel') or 'bloco').strip().lower()
    nivel_label = _LIGACAO_BLOCO_NIVEL_LABEL.get(nivel, nivel)

    with _ligacao_bloco_lock:
        for jid, j in list(_ligacao_bloco_jobs.items()):
            if str(j.get('funcionario_id')) == str(fid) and j.get('phase') not in ('completed', 'cancelled'):
                payload_out = {
                    'error': 'Ja existe uma sequencia de ligacoes em andamento.',
                    'job_id': jid,
                }
                payload_out.update(_ligacao_bloco_job_public(j))
                return jsonify(payload_out), 409
        stale = [
            k for k, v in list(_ligacao_bloco_jobs.items())
            if str(v.get('funcionario_id')) == str(fid)
            and v.get('phase') in ('completed', 'cancelled')
        ]
        for k in stale[:6]:
            _ligacao_bloco_jobs.pop(k, None)

    conn = _get_db()
    try:
        with conn.cursor() as cursor:
            preview = _ligacao_bloco_analisar_preview(cursor, fid, ids)
            elegiveis = preview.get('contrato_ids_elegiveis') or []
            if not elegiveis:
                return jsonify({
                    'error': 'Nenhum contrato elegivel para ligacao neste bloco.',
                    'preview': preview,
                }), 400
            linhas = _cobranca_bloco_excel_linhas(cursor, elegiveis)
            bloqueados = _ligacao_bloco_contratos_bloqueados_hoje(cursor, fid)
    finally:
        conn.close()

    queue = _ligacao_bloco_queue_from_linhas(linhas)
    queue = [q for q in queue if int(q.get('contrato_id') or 0) not in bloqueados]
    if not queue:
        return jsonify({
            'error': 'Nenhuma ligacao pendente (todos os contratos elegiveis ja foram concluidos hoje).',
            'preview': preview,
        }), 400

    job_id = uuid.uuid4().hex
    job = {
        'job_id': job_id,
        'funcionario_id': fid,
        'nivel': nivel,
        'nivel_label': nivel_label,
        'queue': queue,
        'index': 0,
        'phase': 'starting',
        'running': False,
        'cancelled': False,
        'tramitacao_ok': False,
        'successes': 0,
        'failures': 0,
        'pulados': 0,
        'contratos_total_bloco': preview.get('total_solicitados') or len(ids),
        'skipped_sem_telefone': len(preview.get('excluidos_sem_telefone') or []),
        'preview_start': preview,
        'historico': [],
        'contratos_processados': set(),
        'current': None,
        'last_dial': None,
        'status_text': 'Iniciando sequencia...',
    }
    with _ligacao_bloco_lock:
        _ligacao_bloco_jobs[job_id] = job
        _ligacao_bloco_skip_indice_ate_elegivel(job, bloqueados)

    if job['phase'] == 'completed':
        return jsonify({
            'ok': True,
            'job_id': job_id,
            'completed': True,
            **_ligacao_bloco_job_public(job),
        })

    _ligacao_bloco_start_dial(job_id)
    return jsonify({'ok': True, 'job_id': job_id, **_ligacao_bloco_job_public(job)})


@app.route('/api/cobranca/ligacao-bloco/<job_id>/state')
def api_cobranca_ligacao_bloco_state(job_id):
    """Estado leve para bolha e painel (polling)."""
    fid = session.get('funcionario_id')
    if not fid:
        return jsonify({'error': 'Nao autenticado.'}), 401
    job = _ligacao_bloco_get_job(job_id, fid)
    if not job:
        return jsonify({'error': 'Sequencia nao encontrada.'}), 404
    return jsonify(_ligacao_bloco_job_public(job))


@app.route('/api/cobranca/ligacao-bloco/<job_id>/tramitacao-ok', methods=['POST'])
def api_cobranca_ligacao_bloco_tramitacao_ok(job_id):
    """Valida tramitacao no dia e libera proxima ligacao."""
    fid = session.get('funcionario_id')
    if not fid:
        return jsonify({'error': 'Nao autenticado.'}), 401
    with _ligacao_bloco_lock:
        job = _ligacao_bloco_jobs.get(job_id)
        if not job or str(job.get('funcionario_id')) != str(fid):
            return jsonify({'error': 'Sequencia nao encontrada.'}), 404
        if job.get('cancelled'):
            return jsonify({'error': 'Sequencia cancelada.'}), 400
        if job.get('phase') not in ('awaiting_tramitacao', 'awaiting_continue'):
            return jsonify({'error': 'Nao ha ligacao aguardando tramitacao.'}), 400
        cur = job.get('current')
        if not cur:
            return jsonify({'error': 'Nenhuma ligacao ativa.'}), 400
        try:
            cid = int(cur.get('contrato_id'))
        except (TypeError, ValueError):
            return jsonify({'error': 'Contrato invalido.'}), 400

    conn = _get_db()
    try:
        with conn.cursor() as cursor:
            if not _ligacao_bloco_tem_tramitacao_hoje(cursor, fid, cid):
                return jsonify({
                    'error': 'Registre a tramitacao deste contrato (hoje) antes de continuar.',
                }), 400
            _ligacao_bloco_persist_resultado(
                cursor, fid, cid, 'sucesso_tramitado', job.get('nivel'),
            )
        conn.commit()
    finally:
        conn.close()

    with _ligacao_bloco_lock:
        job = _ligacao_bloco_jobs.get(job_id)
        if not job:
            return jsonify({'error': 'Sequencia nao encontrada.'}), 404
        last = job.get('last_dial') or {}
        res_lig = 'sucesso' if last.get('ok') else 'falha'
        _ligacao_bloco_registrar_historico(job, cur, res_lig, True)
        job['tramitacao_ok'] = True
        job['phase'] = 'awaiting_continue'
        job['status_text'] = 'Tramitacao registrada. Deseja seguir para a proxima ligacao?'
    return jsonify({'ok': True, **_ligacao_bloco_job_public(job)})


@app.route('/api/cobranca/ligacao-bloco/<job_id>/proximo', methods=['POST'])
def api_cobranca_ligacao_bloco_proximo(job_id):
    """Avanca para o proximo numero apos tramitacao obrigatoria."""
    fid = session.get('funcionario_id')
    if not fid:
        return jsonify({'error': 'Nao autenticado.'}), 401
    with _ligacao_bloco_lock:
        job = _ligacao_bloco_jobs.get(job_id)
        if not job or str(job.get('funcionario_id')) != str(fid):
            return jsonify({'error': 'Sequencia nao encontrada.'}), 404
        if job.get('cancelled'):
            return jsonify({'error': 'Sequencia cancelada.'}), 400
        if not job.get('tramitacao_ok'):
            return jsonify({'error': 'Registre a tramitacao antes de prosseguir.'}), 400
        if job.get('phase') != 'awaiting_continue':
            return jsonify({'error': 'Aguarde o fim da ligacao atual.'}), 400
        if job.get('running'):
            return jsonify({'error': 'Discagem em andamento.'}), 409
        queue = job.get('queue') or []
        job['index'] = int(job.get('index') or 0) + 1
        job['tramitacao_ok'] = False
        job['current'] = None
        bloqueados = set()
        conn = _get_db()
        try:
            with conn.cursor() as cursor:
                bloqueados = _ligacao_bloco_contratos_bloqueados_hoje(cursor, fid)
        finally:
            conn.close()
        _ligacao_bloco_skip_indice_ate_elegivel(job, bloqueados)
        if job['index'] >= len(queue):
            job['phase'] = 'completed'
            job['status_text'] = 'Sequencia concluida.'
            return jsonify({
                'ok': True,
                'completed': True,
                'resumo': _ligacao_bloco_build_resumo(job),
                **_ligacao_bloco_job_public(job),
            })
        job['phase'] = 'starting'
        job['status_text'] = 'Preparando proxima ligacao...'

    _ligacao_bloco_start_dial(job_id)
    job = _ligacao_bloco_get_job(job_id, fid)
    return jsonify({'ok': True, 'completed': False, **_ligacao_bloco_job_public(job)})


@app.route('/api/cobranca/ligacao-bloco/<job_id>/pular', methods=['POST'])
def api_cobranca_ligacao_bloco_pular(job_id):
    """Pula a ligacao atual (contrato volta a ser elegivel no bloco no mesmo dia)."""
    fid = session.get('funcionario_id')
    if not fid:
        return jsonify({'error': 'Nao autenticado.'}), 401
    with _ligacao_bloco_lock:
        job = _ligacao_bloco_jobs.get(job_id)
        if not job or str(job.get('funcionario_id')) != str(fid):
            return jsonify({'error': 'Sequencia nao encontrada.'}), 404
        if job.get('cancelled'):
            return jsonify({'error': 'Sequencia cancelada.'}), 400
        if job.get('running'):
            return jsonify({'error': 'Discagem em andamento.'}), 409
        if job.get('phase') not in ('awaiting_tramitacao', 'awaiting_continue'):
            return jsonify({'error': 'Nao ha ligacao para pular.'}), 400
        cur = job.get('current')
        if not cur:
            return jsonify({'error': 'Nenhuma ligacao ativa.'}), 400
        try:
            cid = int(cur.get('contrato_id'))
        except (TypeError, ValueError):
            return jsonify({'error': 'Contrato invalido.'}), 400
        job['pulados'] = int(job.get('pulados') or 0) + 1
        _ligacao_bloco_registrar_historico(job, cur, 'pulado', False)
        job['index'] = int(job.get('index') or 0) + 1
        job['tramitacao_ok'] = False
        job['current'] = None
        queue = job.get('queue') or []

    conn = _get_db()
    try:
        with conn.cursor() as cursor:
            _ligacao_bloco_persist_resultado(cursor, fid, cid, 'pulado', job.get('nivel'))
            bloqueados = _ligacao_bloco_contratos_bloqueados_hoje(cursor, fid)
        conn.commit()
    finally:
        conn.close()

    with _ligacao_bloco_lock:
        job = _ligacao_bloco_jobs.get(job_id)
        _ligacao_bloco_skip_indice_ate_elegivel(job, bloqueados)
        if job['index'] >= len(job.get('queue') or []):
            job['phase'] = 'completed'
            job['status_text'] = 'Sequencia concluida.'
            return jsonify({
                'ok': True,
                'completed': True,
                'resumo': _ligacao_bloco_build_resumo(job),
                **_ligacao_bloco_job_public(job),
            })
        job['phase'] = 'starting'
        job['status_text'] = 'Preparando proxima ligacao...'

    _ligacao_bloco_start_dial(job_id)
    job = _ligacao_bloco_get_job(job_id, fid)
    return jsonify({'ok': True, 'completed': False, **_ligacao_bloco_job_public(job)})


@app.route('/api/cobranca/ligacao-bloco/<job_id>/cancel', methods=['POST'])
def api_cobranca_ligacao_bloco_cancel(job_id):
    """Encerra a sequencia e devolve resumo (sucesso / pulado / falha)."""
    fid = session.get('funcionario_id')
    if not fid:
        return jsonify({'error': 'Nao autenticado.'}), 401
    with _ligacao_bloco_lock:
        job = _ligacao_bloco_jobs.get(job_id)
        if not job or str(job.get('funcionario_id')) != str(fid):
            return jsonify({'error': 'Sequencia nao encontrada.'}), 404
        job['cancelled'] = True
        job['running'] = False
        job['phase'] = 'cancelled'
        job['status_text'] = 'Sequencia encerrada.'
        resumo = _ligacao_bloco_build_resumo(job)
    return jsonify({'ok': True, 'resumo': resumo, **_ligacao_bloco_job_public(job)})


@app.route('/api/automacao/<tipo>', methods=['POST'])
def api_automacao(tipo):
    """Disparo em lote de SMS / e-mail OU disparo unitario de ligacao."""
    if tipo not in _AUTOMACAO_TIPOS:
        return jsonify({'error': f'Tipo desconhecido: {tipo}'}), 400

    payload = request.get_json(silent=True) or {}
    nivel = (payload.get('nivel') or 'na').strip().lower()

    if tipo == 'ligacao':
        try:
            contrato_id = int(payload.get('contrato_id'))
        except (TypeError, ValueError):
            return jsonify({'error': 'contrato_id obrigatorio (inteiro).'}), 400

        _automacao_log('ligacao', {
            'contrato_id': contrato_id,
            'nivel': nivel,
            'nome': payload.get('nome'),
            'telefone': payload.get('telefone'),
        })

        result = _discar_numero_funcionario(session.get('funcionario_id'), payload.get('telefone'))
        if not result.get('ok'):
            return jsonify({
                'error': result.get('error') or 'Falha na discagem.',
                'detalhe': result.get('detalhe'),
            }), 502
        return jsonify({
            'success': True,
            'tipo': 'ligacao',
            'contrato_id': contrato_id,
            'status_chamada': 'iniciada',
            'mensagem': 'Discagem disparada.',
            'discador': result.get('discador'),
        })

    # SMS / email - lote (MessageCenter, mesma regra do roteiro da distribuição;
    # só contratos no roteiro e sem envio SMS/e-mail no dia.)
    ids, err = _validar_ids_contratos(payload.get('contrato_ids'))
    if err:
        return jsonify({'error': err}), 400

    fid = session.get('funcionario_id')
    if not fid:
        return jsonify({'error': 'Nao autenticado. Faca login novamente.'}), 401

    if tipo in ('email', 'sms_email'):
        forbidden_email = _cobranca_json_sem_permissao_envio_email()
        if forbidden_email:
            return forbidden_email

    _automacao_log(tipo, {
        'nivel': nivel,
        'qtd': len(ids),
        'ids': ids[:50],  # log nao precisa imprimir milhares
    })

    remetente_nome = _primeiro_nome(session.get('funcionario_nome'))
    if tipo == 'sms':
        canais = {'sms'}
    elif tipo == 'email':
        canais = {'email'}
    else:
        canais = {'sms', 'email'}
    conn = _get_db()
    stats = {
        'envios_sms': 0,
        'envios_email': 0,
        'falhas': 0,
        'ignorados_sem_template': 0,
        'ignorados_sem_telefone': 0,
        'ignorados_sem_email': 0,
        'ignorados_ja_enviados_hoje': 0,
        'ignorados_fora_rota': 0,
        'ignorados_sem_contrato': 0,
        'contratos_solicitados': len(ids),
    }
    erros_amostra = []

    try:
        for cid in ids:
            with conn.cursor() as cursor:
                row = _contrato_row_auto_envio(cursor, cid)
            if not row:
                stats['ignorados_sem_contrato'] += 1
                continue

            da_int = None
            if row.get('dias_atraso') is not None:
                try:
                    da_int = int(row['dias_atraso'])
                except (TypeError, ValueError):
                    da_int = None

            if _sms_automatizados_template_id_por_dias(da_int) is None:
                stats['ignorados_fora_rota'] += 1
                continue

            _auto_envio_contrato_canais(
                conn,
                int(fid),
                row,
                remetente_nome,
                stats,
                erros_amostra,
                canais,
            )
    finally:
        conn.close()

    if tipo == 'sms':
        enviados = stats['envios_sms']
    elif tipo == 'email':
        enviados = stats['envios_email']
    else:
        enviados = stats['envios_sms'] + stats['envios_email']
    msg = (
        f'SMS: {stats["envios_sms"]} | E-mail: {stats["envios_email"]} | Falhas: {stats["falhas"]} | '
        f'Fora do roteiro (dias): {stats["ignorados_fora_rota"]} | '
        f'Já enviados hoje: {stats["ignorados_ja_enviados_hoje"]} | '
        f'Sem contrato aberto: {stats["ignorados_sem_contrato"]}'
    )
    return jsonify({
        'success': True,
        'tipo': tipo,
        'nivel': nivel,
        'enviados': enviados,
        'envios_sms': stats['envios_sms'],
        'envios_email': stats['envios_email'],
        'falhas': stats['falhas'],
        'ignorados_fora_rota': stats['ignorados_fora_rota'],
        'ignorados_ja_enviados_hoje': stats['ignorados_ja_enviados_hoje'],
        'ignorados_sem_contrato': stats['ignorados_sem_contrato'],
        'ignorados_sem_telefone': stats['ignorados_sem_telefone'],
        'ignorados_sem_email': stats['ignorados_sem_email'],
        'mensagem': msg,
        'erros_amostra': erros_amostra[:20],
        'mock': False,
    })


# =============================================================================
# API: Negativacao (Serasa) - POR PARCELA
# -----------------------------------------------------------------------------
# Regras de negocio (versao 2):
#   - A negativacao e registrada no nivel de PARCELA (nao contrato).
#   - Uma parcela ELEGIVEL: parcela aberto mais antiga com 30 < dias < 90
#     (1 ou N parcelas em aberto) e ainda NAO consta em `negativacao`.
#   - A unicidade e reforcada pelo UNIQUE KEY (id_parcela) + INSERT IGNORE.
#     Isso protege contra:
#       * race condition (dois cliques / dois usuarios simultaneos);
#       * re-envios logicos (o front deveria evitar, mas nunca se confia).
#   - O INSERT ocorre ANTES do disparo para a API externa para que a proxima
#     leitura do painel de cobranca ja apresente o contrato como "negativado"
#     mesmo se a API externa demorar ou falhar. Se a API externa falhar,
#     atualizamos `status` para 'falhou' e guardamos a resposta.
# =============================================================================

def _info_negativacao_por_contrato(cursor, ids, data_referencia):
    """Retorna dict {id_contrato: {...dados da parcela alvo}}.

    ``data_referencia`` = mesma data de referencia de /api/cobranca
    (MAX(data_arquivo) em arquivos_gm) para o calculo de dias de atraso.
    """
    if not ids:
        return {}
    placeholders = ','.join(['%s'] * len(ids))
    q = f"""
        SELECT c.id AS id_contrato,
               c.grupo, c.cota, c.numero_contrato,
               (
                 SELECT COUNT(*) FROM parcela p
                 WHERE p.id_contrato = c.id AND p.status = 'cobranca'
               ) AS parcelas_abertas,
               (
                 SELECT p.id FROM parcela p
                 WHERE p.id_contrato = c.id AND p.status = 'cobranca'
                 ORDER BY p.vencimento ASC, p.id ASC LIMIT 1
               ) AS id_parcela_alvo,
               (
                 SELECT p.numero_parcela FROM parcela p
                 WHERE p.id_contrato = c.id AND p.status = 'cobranca'
                 ORDER BY p.vencimento ASC, p.id ASC LIMIT 1
               ) AS numero_parcela_alvo,
               (
                 SELECT DATEDIFF(%s, MIN(p.vencimento)) FROM parcela p
                 WHERE p.id_contrato = c.id AND p.status = 'cobranca'
               ) AS dias_atraso
        FROM contrato c
        WHERE c.id IN ({placeholders}) AND c.status = 'cobranca'
    """
    cursor.execute(q, (data_referencia, *ids))
    info = _clean_rows(cursor.fetchall())
    return {int(r['id_contrato']): r for r in info}


@app.route('/api/negativacao/enviar', methods=['POST'])
def api_negativacao_enviar():
    payload = request.get_json(silent=True) or {}
    ids, err = _validar_ids_contratos(payload.get('contrato_ids'))
    if err:
        return jsonify({'error': err}), 400

    conn = _get_db()
    cursor = conn.cursor()
    _ensure_negativacao_table(cursor)

    data_ref = _get_data_referencia_arquivos_gm(cursor)

    # --- Fase 1: coleta ------------------------------------------------------
    info_por_id = _info_negativacao_por_contrato(cursor, ids, data_ref)

    # --- Fase 2: CHECK BEFORE CLASSIFICATION --------------------------------
    # Mesma logica de /api/cobranca: mapa por id_contrato (legado NULL + ids).
    neg_map = _mapa_parcelas_negativadas_por_contrato(cursor, ids)

    # --- Fase 3: classificacao (Needs Negative / Already / Not Negative) ----
    elegiveis    = []  # "Needs Negative": serao gravados e enviados agora
    ja_negativ   = []  # "Already Negative": parcela-alvo ja esta registrada
    fora_janela  = []  # atraso fora de (30, 90)
    ausentes     = []  # contrato nao encontrado ou nao-aberto

    for cid in ids:
        r = info_por_id.get(cid)
        if not r:
            ausentes.append({'id_contrato': cid})
            continue

        id_alvo_raw = r.get('id_parcela_alvo')
        try:
            id_alvo = int(id_alvo_raw) if id_alvo_raw is not None else None
        except (TypeError, ValueError):
            id_alvo = None
        try:
            dias_i = int(r.get('dias_atraso') or 0)
        except (TypeError, ValueError):
            dias_i = 0

        neg_set = neg_map.get(cid, set())
        if _NEG_NULL_PARCELA in neg_set or (
            id_alvo is not None and id_alvo in neg_set
        ):
            ja_negativ.append({'id_contrato': cid, 'id_parcela': id_alvo})
            continue
        if not (30 < dias_i < 90):
            fora_janela.append({'id_contrato': cid, 'dias_atraso': dias_i})
            continue
        if id_alvo is None:
            ausentes.append({'id_contrato': cid})
            continue

        elegiveis.append({
            'id_contrato':     cid,
            'id_parcela':      id_alvo,
            'numero_parcela':  r.get('numero_parcela_alvo'),
            'dias_atraso':     dias_i,
            'grupo_cota':      f"{r.get('grupo')}/{r.get('cota')}",
        })

    _automacao_log('negativacao', {
        'qtd_total':         len(ids),
        'qtd_elegiveis':     len(elegiveis),
        'qtd_ja_negativados': len(ja_negativ),
        'qtd_fora_janela':   len(fora_janela),
        'qtd_ausentes':      len(ausentes),
    })

    # --- Fase 4: INSERT primeiro (locking via UNIQUE KEY) -------------------
    # INSERT IGNORE faz o uk_negativacao_parcela virar "seletor": se outra
    # request concorrente ja gravou a mesma parcela, esta tentativa e
    # silenciosamente descartada e rowcount reflete isso.
    gravados = 0
    fid = session.get('funcionario_id')
    try:
        fid = int(fid) if fid is not None else None
    except (TypeError, ValueError):
        fid = None
    _ensure_negativacao_historico_table(cursor)
    if elegiveis:
        resposta_json = '{"mock": true}'
        for e in elegiveis:
            cursor.execute(
                "INSERT IGNORE INTO negativacao "
                "  (id_contrato, id_parcela, numero_parcela, dias_atraso, status, resposta_api, id_funcionario) "
                "VALUES (%s, %s, %s, %s, 'enviado', %s, %s)",
                (
                    e['id_contrato'],
                    e['id_parcela'],
                    e['numero_parcela'],
                    e['dias_atraso'],
                    resposta_json,
                    fid,
                ),
            )
            if cursor.rowcount:
                gravados += 1
                try:
                    cursor.execute(
                        "SELECT data_negativacao FROM negativacao WHERE id_parcela = %s LIMIT 1",
                        (e['id_parcela'],),
                    )
                    row_dn = cursor.fetchone()
                    data_ev = row_dn.get('data_negativacao') if row_dn else None
                    _insert_negativacao_historico(
                        cursor,
                        e['id_contrato'],
                        e['id_parcela'],
                        e.get('numero_parcela'),
                        'negativado_manual',
                        data_ev,
                        fid,
                        e['dias_atraso'],
                        'enviado',
                        'Negativação registrada pelo fluxo manual (envio à API / mock).',
                    )
                except Exception as exc_hist:
                    app.logger.warning('api_negativacao_enviar: historico (%s)', exc_hist)
        conn.commit()

    # --- Fase 5: disparo para a API externa (mock hoje; Flavio depois) ------
    # TODO: substituir pelo POST real. Se falhar, atualizar status p/ 'falhou':
    #   import requests
    #   for e in elegiveis:
    #       try:
    #           r = requests.post('http://api-flavio/serasa/negativacao',
    #                             json={'id_contrato': e['id_contrato'],
    #                                   'id_parcela':  e['id_parcela'],
    #                                   'dias_atraso': e['dias_atraso']},
    #                             timeout=30)
    #           if r.status_code >= 400:
    #               raise RuntimeError(r.text)
    #           cursor.execute(
    #               "UPDATE negativacao SET status='enviado', resposta_api=%s "
    #               "WHERE id_parcela=%s",
    #               (r.text, e['id_parcela']),
    #           )
    #       except Exception as exc:
    #           cursor.execute(
    #               "UPDATE negativacao SET status='falhou', resposta_api=%s "
    #               "WHERE id_parcela=%s",
    #               (str(exc), e['id_parcela']),
    #           )
    #   conn.commit()

    cursor.close()
    conn.close()

    return jsonify({
        'success': True,
        'tipo': 'negativacao',
        'enviados':         len(elegiveis),
        'gravados':         gravados,
        'ja_negativados':   len(ja_negativ),
        'fora_janela':      len(fora_janela),
        'ausentes':         len(ausentes),
        'falhas':           0,
        'mock':             True,
        'mensagem': (
            f'{len(elegiveis)} parcela(s) negativada(s). '
            f'{len(ja_negativ)} parcela(s) ja estavam negativadas. '
            f'{len(fora_janela)} fora da janela (30 < dias < 90).'
        ),
    })


def _negativacao_listagem_data_iso(val):
    """Aceita YYYY-MM-DD; retorna None se invalido."""
    if val is None:
        return None
    s = str(val).strip()[:10]
    if len(s) < 10:
        return None
    try:
        datetime.datetime.strptime(s, '%Y-%m-%d')
        return s
    except ValueError:
        return None


def _negativacao_listagem_order_at(sort_col, order_dir):
    order_dir = 'DESC' if str(order_dir or '').lower() == 'desc' else 'ASC'
    sort_col = (sort_col or 'data').lower()
    clauses = {
        'data': f'n.data_negativacao {order_dir}, n.id {order_dir}',
        'contrato': f'c.grupo {order_dir}, c.cota {order_dir}, n.id {order_dir}',
        'parcela': f'n.numero_parcela {order_dir}, n.id {order_dir}',
        'dias': f'n.dias_atraso {order_dir}, n.id {order_dir}',
        'status': f'n.status {order_dir}, n.id {order_dir}',
        'operador': f'COALESCE(f.nome, "") {order_dir}, n.id {order_dir}',
    }
    return 'ORDER BY ' + clauses.get(sort_col, clauses['data'])


def _negativacao_carteira_where_data_negativacao_efetiva(carteira_dn_ini, carteira_dn_fim):
    """Restringe linhas de `negativacao` ao dia (ou intervalo) da negativação no histórico.

    O modal do contrato usa `negativacao_historico`; filtrar só por `n.data_negativacao` pode incluir
    parcelas cujo evento foi em outro dia. Mantém fallback por `data_negativacao` quando não há
    evento negativado_* no histórico para a parcela (legado).

    Retorna (fragmento_sql, lista_params).
    """
    params_ex = []
    if carteira_dn_ini and carteira_dn_fim:
        inner_date_nh = "DATE(nh.data_evento) BETWEEN %s AND %s"
        params_ex.extend([carteira_dn_ini, carteira_dn_fim])
        fb_date = "DATE(n.data_negativacao) BETWEEN %s AND %s"
        params_fb = [carteira_dn_ini, carteira_dn_fim]
    elif carteira_dn_fim:
        inner_date_nh = "DATE(nh.data_evento) = %s"
        params_ex.append(carteira_dn_fim)
        fb_date = "DATE(n.data_negativacao) = %s"
        params_fb = [carteira_dn_fim]
    elif carteira_dn_ini:
        inner_date_nh = "DATE(nh.data_evento) >= %s"
        params_ex.append(carteira_dn_ini)
        fb_date = "DATE(n.data_negativacao) >= %s"
        params_fb = [carteira_dn_ini]
    else:
        return '', []

    parcela_match_nh = (
        "nh.id_contrato = n.id_contrato "
        "AND ("
        "  (nh.id_parcela IS NOT NULL AND nh.id_parcela = n.id_parcela) "
        "  OR (nh.id_parcela IS NULL AND nh.numero_parcela IS NOT NULL "
        "      AND nh.numero_parcela <=> n.numero_parcela)"
        ")"
    )
    existe_neg_hist = (
        "EXISTS (SELECT 1 FROM negativacao_historico nh "
        "WHERE nh.tipo_evento IN ('negativado_manual','negativado_tracker') "
        f"AND {parcela_match_nh} AND {inner_date_nh})"
    )
    sem_neg_hist_parcela = (
        "NOT EXISTS (SELECT 1 FROM negativacao_historico nh0 "
        "WHERE nh0.tipo_evento IN ('negativado_manual','negativado_tracker') "
        f"AND {parcela_match_nh.replace('nh.', 'nh0.')})"
    )
    sql = f"(({existe_neg_hist}) OR ({sem_neg_hist_parcela} AND ({fb_date})))"
    return sql, params_ex + params_fb


def _negativacao_carteira_sql_date_evento(alias, carteira_dn_ini, carteira_dn_fim):
    """Fragmento DATE(...) sobre data_evento do histórico + params (Carteira)."""
    a = alias
    if carteira_dn_ini and carteira_dn_fim:
        return f"DATE({a}.data_evento) BETWEEN %s AND %s", [carteira_dn_ini, carteira_dn_fim]
    if carteira_dn_fim:
        return f"DATE({a}.data_evento) = %s", [carteira_dn_fim]
    if carteira_dn_ini:
        return f"DATE({a}.data_evento) >= %s", [carteira_dn_ini]
    return '', []


def _negativacao_carteira_deve_incluir_positivas(evento):
    """Mescla positivações do histórico na Carteira (exceto quando o filtro é só negativação)."""
    ev = (evento or '').strip().lower()
    if ev in ('observacao', 'negativacao_retorno', 'positivacao_retorno'):
        return False
    if ev in ('negativado', 'negativado_tracker', 'negativado_manual'):
        return False
    return True


def _negativacao_positivacao_sql_filtro_operador_carteira(fid_op, data_ref_iso):
    """Restringe positivações por operador na visão Carteira (alias de contrato = `c`).

    - Contrato no snapshot GM da data de referência (com parcelas em aberto): deve estar em
      `funcionario_cobranca` com o operador — igual ao painel Cobrança.

    - Contrato **sem** essa correspondência no snapshot (ex.: saiu da carteira mas há positivação
      no histórico): usa-se o **último** `tramitacao.id_funcionario` do contrato (`created_at`,
      depois `id`) como operador “histórico” para o filtro.
    """
    sql = (
        "("
        "EXISTS ("
        "SELECT 1 FROM cobranca cob "
        "INNER JOIN contrato cc ON cc.id = cob.id_contrato AND cc.status = 'cobranca' "
        "INNER JOIN parcela par ON par.id_contrato = cc.id AND par.status = 'cobranca' "
        "INNER JOIN funcionario_cobranca fc ON fc.id_contrato = cc.id "
        "AND fc.id_funcionario = %s "
        "WHERE cob.data_arquivo = %s AND cob.id_contrato = c.id"
        ") OR ("
        "NOT EXISTS ("
        "SELECT 1 FROM cobranca cob "
        "INNER JOIN contrato cc ON cc.id = cob.id_contrato AND cc.status = 'cobranca' "
        "INNER JOIN parcela par ON par.id_contrato = cc.id AND par.status = 'cobranca' "
        "WHERE cob.data_arquivo = %s AND cob.id_contrato = c.id"
        ") AND ("
        "SELECT t.id_funcionario FROM tramitacao t "
        "WHERE t.id_contrato = c.id "
        "ORDER BY t.created_at DESC, t.id DESC LIMIT 1"
        ") = %s"
        ")"
        ")"
    )
    return sql, [fid_op, data_ref_iso, data_ref_iso, fid_op]


def _negativacao_sem_operador_cobranca_contrato_fragment(data_ref_iso):
    """Contratos sem vínculo em `funcionario_cobranca` no snapshot GM e sem `tramitacao` (alias `c`)."""
    d_iso = str(data_ref_iso or '')[:10]
    if len(d_iso) != 10:
        return '1=0', []
    sql = (
        "NOT EXISTS ("
        "SELECT 1 FROM cobranca cob "
        "INNER JOIN contrato cc ON cc.id = cob.id_contrato AND cc.status = 'cobranca' "
        "INNER JOIN parcela par ON par.id_contrato = cc.id AND par.status = 'cobranca' "
        "INNER JOIN funcionario_cobranca fc ON fc.id_contrato = cc.id "
        "WHERE cob.data_arquivo = %s AND cob.id_contrato = c.id"
        ") AND NOT EXISTS (SELECT 1 FROM tramitacao t WHERE t.id_contrato = c.id)"
    )
    return sql, [d_iso]


def _negativacao_listagem_fetch_positivacao_rows(
    cursor,
    dn_ini,
    dn_fim,
    q,
    tipo_busca,
    evento,
    status_ativo,
    clause_cob=None,
    params_cob=None,
    filtro_operador_carteira=None,
    filtro_sem_operador_cobranca=None,
):
    """Eventos removido_* no histórico filtrados por data_evento.

    `clause_cob` + `params_cob`: opcional; na Carteira Cobrança a listagem chama sem esse filtro
    para positivações (mesmo conjunto da visão Geral), pois o contrato pode já ter saído do snapshot.

    `filtro_operador_carteira`: tupla (id_funcionario, data_ref_iso) para filtrar positivações na
    Carteira por operador (snapshot + fc OU fora do snapshot: último operador em tramitacao).

    `filtro_sem_operador_cobranca`: data_ref_iso (YYYY-MM-DD) — positivações cujo contrato não tem
    operador em cobrança no snapshot da data e não possui tramitação (complemento do filtro por operador).
    """
    date_sql, date_params = _negativacao_carteira_sql_date_evento('h', dn_ini, dn_fim)
    if not date_sql:
        return []

    where = [
        "h.tipo_evento IN ('removido_pagamento','removido_manual','positivado_tracker')",
        date_sql,
    ]
    params = list(date_params)
    if clause_cob:
        where.append(clause_cob)
        params.extend(list(params_cob or []))
    if filtro_sem_operador_cobranca:
        frag_s, par_s = _negativacao_sem_operador_cobranca_contrato_fragment(
            filtro_sem_operador_cobranca
        )
        where.append(frag_s)
        params.extend(par_s)
    elif filtro_operador_carteira:
        try:
            fid_op = int(filtro_operador_carteira[0])
            d_iso = str(filtro_operador_carteira[1] or '')[:10]
        except (TypeError, ValueError, IndexError):
            fid_op = None
            d_iso = ''
        if fid_op and len(d_iso) == 10:
            frag_op, par_op = _negativacao_positivacao_sql_filtro_operador_carteira(fid_op, d_iso)
            where.append(frag_op)
            params.extend(par_op)

    if q:
        like = f"%{q}%"
        if tipo_busca == 'contrato':
            clause = (
                "(CAST(c.grupo AS CHAR) LIKE %s OR CAST(c.cota AS CHAR) LIKE %s "
                "OR CONCAT(c.grupo, '/', c.cota) LIKE %s OR CAST(c.numero_contrato AS CHAR) LIKE %s)"
            )
            where.append(clause)
            params.extend([like, like, like, like])
        else:
            where.append("(h.detalhe LIKE %s OR h.tipo_evento LIKE %s)")
            params.extend([like, like])

    ev = (evento or '').strip().lower()
    if status_ativo:
        where.append('h.status_snapshot = %s')
        params.append(status_ativo)
    if ev == 'removido_pagamento':
        where.append("h.tipo_evento IN ('removido_pagamento','positivado_tracker')")
    elif ev == 'positivado_tracker':
        where.append("h.tipo_evento = 'positivado_tracker'")
    elif ev == 'removido_manual':
        where.append("h.tipo_evento = 'removido_manual'")

    wh = 'WHERE ' + ' AND '.join(where)
    cursor.execute(
        f"""
        SELECT h.id AS id_historico, h.id_contrato, h.id_parcela, h.numero_parcela,
               h.tipo_evento, h.data_evento, h.dias_atraso, h.status_snapshot, h.detalhe,
               c.grupo, c.cota, c.numero_contrato, c.status AS contrato_status,
               f.nome AS funcionario_nome
        FROM negativacao_historico h
        JOIN contrato c ON c.id = h.id_contrato
        LEFT JOIN funcionario f ON h.id_funcionario = f.id
        {wh}
        ORDER BY h.data_evento DESC, h.id DESC
        """,
        params,
    )
    out = []
    for raw in _clean_rows(cursor.fetchall()):
        row = dict(raw)
        hid = row.pop('id_historico', None)
        if hid is not None:
            row['id'] = hid
        row['data_negativacao'] = row.pop('data_evento', None)
        row['status'] = ''
        row.pop('status_snapshot', None)
        _negativacao_row_serasa_flags(row)
        out.append(row)
    return out


def _negativacao_apenas_cobranca_exists_clause(cursor, funcionario_id_filtro):
    """EXISTS + params alinhados ao painel `/api/cobranca` (snapshot GM + parcelas abertas).

    Com `funcionario_id_filtro`, restringe aos contratos com vínculo em `funcionario_cobranca`.
    Retorna (sql_fragment, params_list, data_ref_iso).
    """
    _ensure_cobranca_table(cursor)
    data_ref = _get_data_referencia_arquivos_gm(cursor)
    if isinstance(data_ref, datetime.datetime):
        d_iso = data_ref.date().isoformat()
    elif isinstance(data_ref, datetime.date):
        d_iso = data_ref.isoformat()
    else:
        d_iso = str(data_ref)[:10]

    if funcionario_id_filtro is not None:
        sql = (
            "EXISTS (SELECT 1 FROM cobranca cob "
            "INNER JOIN contrato cc ON cc.id = cob.id_contrato AND cc.status = 'cobranca' "
            "INNER JOIN parcela par ON par.id_contrato = cc.id AND par.status = 'cobranca' "
            "INNER JOIN funcionario_cobranca fc ON fc.id_contrato = cc.id "
            "AND fc.id_funcionario = %s "
            "WHERE cob.data_arquivo = %s AND cob.id_contrato = c.id)"
        )
        params = [funcionario_id_filtro, d_iso]
    else:
        sql = (
            "EXISTS (SELECT 1 FROM cobranca cob "
            "INNER JOIN contrato cc ON cc.id = cob.id_contrato AND cc.status = 'cobranca' "
            "INNER JOIN parcela par ON par.id_contrato = cc.id AND par.status = 'cobranca' "
            "WHERE cob.data_arquivo = %s AND cob.id_contrato = c.id)"
        )
        params = [d_iso]
    return sql, params, d_iso


def _split_negativacao_ativos_prioridade(rows):
    """Mesmas faixas do painel Cobrança: crítico >=60, atenção 30–59, recente 1–29 dias."""
    critico, atencao, recente = [], [], []
    for r in rows:
        dias = r.get('dias_atraso')
        try:
            if dias is None:
                d = 0
            elif isinstance(dias, str):
                d = int(float(dias))
            else:
                d = int(dias)
        except (TypeError, ValueError):
            d = 0
        if d >= 60:
            critico.append(r)
        elif d >= 30:
            atencao.append(r)
        elif d >= 1:
            recente.append(r)
    return critico, atencao, recente


def _split_negativacao_ativos_negativados_positivados(rows):
    """Carteira Negativação: negativados (sem evento removido_* / positivado_tracker na linha) vs positivados."""
    negativados, positivados = [], []
    for r in rows:
        t = str((r.get('tipo_evento') or '')).strip().lower()
        if t.startswith('removido') or t == 'positivado_tracker':
            positivados.append(r)
        else:
            negativados.append(r)
    return negativados, positivados




def _negativacao_listagem_payload_interno(
    cursor,
    *,
    q='',
    tipo_busca='contrato',
    evento='todos',
    status_ativo='',
    data_inicio=None,
    data_fim=None,
    sort_at='data',
    order_at='desc',
    apenas_cobranca=False,
    sem_operador_cobranca=False,
    preview_ativos=False,
    funcionario_id_filtro=None,
):
    """Mesmo dict da API GET /api/negativacao/listagem (sem jsonify)."""
    order_at_sql = _negativacao_listagem_order_at(sort_at, order_at)
    _ensure_negativacao_table(cursor)
    _ensure_negativacao_historico_table(cursor)

    where_at = []
    params_at = []

    data_ref_cobranca_resp = None
    clause_cob = None
    params_cob = []
    if apenas_cobranca:
        fid_clause = None if sem_operador_cobranca else funcionario_id_filtro
        clause_cob, params_cob, data_ref_cobranca_resp = (
            _negativacao_apenas_cobranca_exists_clause(cursor, fid_clause)
        )
        if sem_operador_cobranca:
            clause_cob, params_cob = None, []

    carteira_dn_ini = data_inicio
    carteira_dn_fim = data_fim
    carteira_usou_data_padrao_gm = False
    if apenas_cobranca and not carteira_dn_ini and not carteira_dn_fim:
        carteira_dn_fim = data_ref_cobranca_resp
        carteira_usou_data_padrao_gm = bool(data_ref_cobranca_resp)

    if q:
        like = f"%{q}%"
        if tipo_busca == 'contrato':
            clause = (
                "(CAST(c.grupo AS CHAR) LIKE %s OR CAST(c.cota AS CHAR) LIKE %s "
                "OR CONCAT(c.grupo, '/', c.cota) LIKE %s OR CAST(c.numero_contrato AS CHAR) LIKE %s)"
            )
            where_at.append(clause)
            params_at.extend([like, like, like, like])
        else:
            where_at.append(
                "EXISTS (SELECT 1 FROM negativacao_historico h2 WHERE h2.id_contrato = n.id_contrato "
                "AND (h2.detalhe LIKE %s OR h2.tipo_evento LIKE %s))"
            )
            params_at.extend([like, like])

    frag_dn_carteira = ''
    frag_geral_dn = ''
    if apenas_cobranca:
        frag_dn_carteira, par_dn = _negativacao_carteira_where_data_negativacao_efetiva(
            carteira_dn_ini, carteira_dn_fim
        )
        if frag_dn_carteira:
            where_at.append(frag_dn_carteira)
            params_at.extend(par_dn)
    elif data_inicio or data_fim:
        frag_geral_dn, par_g = _negativacao_carteira_where_data_negativacao_efetiva(
            data_inicio, data_fim
        )
        if frag_geral_dn:
            where_at.append(frag_geral_dn)
            params_at.extend(par_g)

    if evento not in ('', 'todos', 'all'):
        if evento in ('negativacao_retorno', 'positivacao_retorno'):
            sub_exist = (
                'EXISTS (SELECT 1 FROM negativacao_historico hx WHERE hx.id_contrato = n.id_contrato '
                'AND hx.tipo_evento = %s'
            )
            params_at.append(evento)
            dn_ini = carteira_dn_ini if apenas_cobranca else data_inicio
            dn_fim = carteira_dn_fim if apenas_cobranca else data_fim
            ds_sql, ds_par = _negativacao_carteira_sql_date_evento('hx', dn_ini, dn_fim)
            if ds_sql:
                sub_exist = sub_exist + ' AND ' + ds_sql
                params_at.extend(ds_par)
            sub_exist += ')'
            where_at.append(sub_exist)
        elif apenas_cobranca:
            if evento in ('positivado', 'observacao', 'removido_pagamento', 'removido_manual', 'positivado_tracker'):
                where_at.append('1=0')
            elif evento == 'negativado_tracker':
                where_at.append("n.status = 'registrado_tracker'")
            elif evento == 'negativado_manual':
                where_at.append("n.status IN ('enviado', 'falhou')")
        else:
            if evento in ('positivado', 'observacao', 'removido_pagamento', 'removido_manual', 'positivado_tracker'):
                where_at.append('1=0')
            elif evento == 'negativado_tracker':
                where_at.append("n.status = 'registrado_tracker'")
            elif evento == 'negativado_manual':
                where_at.append("n.status IN ('enviado', 'falhou')")

    if status_ativo:
        where_at.append('n.status = %s')
        params_at.append(status_ativo)

    if apenas_cobranca:
        if sem_operador_cobranca and data_ref_cobranca_resp:
            frag_sem, par_sem = _negativacao_sem_operador_cobranca_contrato_fragment(
                data_ref_cobranca_resp
            )
            where_at.append(frag_sem)
            params_at.extend(par_sem)
        elif clause_cob:
            where_at.append(clause_cob)
            params_at.extend(params_cob)

    wh_at = ("WHERE " + " AND ".join(where_at)) if where_at else ""

    total_hist = 0
    historico = []

    # Visao geral + filtros padrao + preview: so as 20 parcelas ativas mais recentes.
    # Carteira cobranca ou pesquisa com filtros: todas as linhas que batem no WHERE.
    limite_ativos = None
    if not apenas_cobranca and preview_ativos:
        limite_ativos = 20
    limit_sql_at = ''
    params_limit_at = []
    if limite_ativos is not None:
        limit_sql_at = ' LIMIT %s'
        params_limit_at.append(limite_ativos)

    cursor.execute(
        f"""
        SELECT COUNT(*) AS total
        FROM negativacao n
        JOIN contrato c ON c.id = n.id_contrato
        {wh_at}
        """,
        params_at,
    )
    total_ativos = int((cursor.fetchone() or {}).get('total') or 0)

    cursor.execute(
        f"""
        SELECT n.*, c.grupo, c.cota, c.numero_contrato, c.status AS contrato_status,
               f.nome AS funcionario_nome
        FROM negativacao n
        JOIN contrato c ON c.id = n.id_contrato
        LEFT JOIN funcionario f ON n.id_funcionario = f.id
        {wh_at}
        {order_at_sql}
        {limit_sql_at}
        """,
        params_at + params_limit_at,
    )
    ativos = _dedupe_negativacao_ativas_exibicao(_clean_rows(cursor.fetchall()))
    for r in ativos:
        _negativacao_row_serasa_flags(r)

    frag_dn_para_posit = frag_dn_carteira if apenas_cobranca else frag_geral_dn
    dn_ini_posit = carteira_dn_ini if apenas_cobranca else data_inicio
    dn_fim_posit = carteira_dn_fim if apenas_cobranca else data_fim
    if frag_dn_para_posit and _negativacao_carteira_deve_incluir_positivas(evento):
        if apenas_cobranca:
            posit_kwargs = {}
            if sem_operador_cobranca and data_ref_cobranca_resp:
                posit_kwargs['filtro_sem_operador_cobranca'] = data_ref_cobranca_resp
            elif funcionario_id_filtro is not None and data_ref_cobranca_resp:
                posit_kwargs['filtro_operador_carteira'] = (
                    funcionario_id_filtro,
                    data_ref_cobranca_resp,
                )
            ativos.extend(
                _negativacao_listagem_fetch_positivacao_rows(
                    cursor,
                    dn_ini_posit,
                    dn_fim_posit,
                    q,
                    tipo_busca,
                    evento,
                    status_ativo,
                    **posit_kwargs,
                )
            )
        else:
            ativos.extend(
                _negativacao_listagem_fetch_positivacao_rows(
                    cursor,
                    dn_ini_posit,
                    dn_fim_posit,
                    q,
                    tipo_busca,
                    evento,
                    status_ativo,
                )
            )

    if apenas_cobranca or frag_geral_dn:
        total_ativos = len(ativos)

    payload = {
        'historico': historico,
        'total_historico': total_hist,
        'ativos': ativos,
        'total_ativos': total_ativos,
        'modo_cobranca': bool(apenas_cobranca),
        'ativos_em_preview': bool(not apenas_cobranca and limite_ativos == 20),
    }
    if apenas_cobranca:
        payload['carteira_filtro_data_negativacao'] = {
            'data_inicio': carteira_dn_ini,
            'data_fim': carteira_dn_fim,
            'usou_data_padrao_ultimo_gm': carteira_usou_data_padrao_gm,
        }
    if apenas_cobranca and data_ref_cobranca_resp:
        payload['data_referencia_cobranca'] = data_ref_cobranca_resp
        neg_n, pos_n = _split_negativacao_ativos_negativados_positivados(ativos)
        payload['ativos_negativados'] = neg_n
        payload['ativos_positivados'] = pos_n
        payload['total_ativos_negativados'] = len(neg_n)
        payload['total_ativos_positivados'] = len(pos_n)
        try:
            cursor.execute(
                "SELECT id, nome FROM funcionario WHERE "
                + _WHERE_FUNCIONARIO_COBRANCA
                + " ORDER BY nome"
            )
            payload['funcionarios_cobranca'] = [
                {'id': int(x['id']), 'nome': x['nome']}
                for x in cursor.fetchall()
            ]
        except Exception:
            payload['funcionarios_cobranca'] = []
    return payload


def _negativacao_listagem_kwargs_from_request(req, *, preview_ativos_force=None):
    """Argumentos de ``_negativacao_listagem_payload_interno`` a partir da query string."""
    q = (req.args.get('q') or '').strip()
    tipo_busca = (req.args.get('tipo_busca') or 'contrato').strip().lower()
    if tipo_busca not in ('contrato', 'texto'):
        tipo_busca = 'contrato'

    evento = (req.args.get('evento') or 'todos').strip().lower()
    status_ativo = (req.args.get('status_ativo') or '').strip()
    data_inicio = _negativacao_listagem_data_iso(req.args.get('data_inicio'))
    data_fim = _negativacao_listagem_data_iso(req.args.get('data_fim'))

    sort_at = (req.args.get('sort_ativos') or 'data').strip().lower()
    order_at = (req.args.get('order_ativos') or 'desc').strip().lower()

    _ac = (req.args.get('apenas_cobranca') or '').strip().lower()
    apenas_cobranca = _ac in ('1', 'true', 'yes', 'sim', 'on')

    _so = (req.args.get('sem_operador_cobranca') or '').strip().lower()
    sem_operador_cobranca = bool(apenas_cobranca) and _so in (
        '1',
        'true',
        'yes',
        'sim',
        'on',
    )

    _pv = (req.args.get('preview_ativos') or '').strip().lower()
    preview_ativos = _pv in ('1', 'true', 'yes', 'sim', 'on')
    if preview_ativos_force is not None:
        preview_ativos = bool(preview_ativos_force)

    funcionario_id_filtro = None
    if not sem_operador_cobranca:
        fid_raw = (req.args.get('funcionario_id') or '').strip()
        if fid_raw:
            try:
                funcionario_id_filtro = int(fid_raw)
            except (TypeError, ValueError):
                funcionario_id_filtro = None

    return {
        'q': q,
        'tipo_busca': tipo_busca,
        'evento': evento,
        'status_ativo': status_ativo,
        'data_inicio': data_inicio,
        'data_fim': data_fim,
        'sort_at': sort_at,
        'order_at': order_at,
        'apenas_cobranca': apenas_cobranca,
        'sem_operador_cobranca': sem_operador_cobranca,
        'preview_ativos': preview_ativos,
        'funcionario_id_filtro': funcionario_id_filtro,
    }


@app.route('/api/negativacao/listagem')
def api_negativacao_listagem():
    """Lista parcelas negativadas / positivações no período (painel do módulo).

    Carteira: linhas em `negativacao` no universo snapshot GM (e operador, se filtrado); positivações
    `removido_*` no período ampliam o conjunto como na visão Geral. Com filtro de operador,
    positivações de contratos fora do snapshot usam o último operador em `tramitacao`.
    Com `sem_operador_cobranca=1`, lista apenas contratos sem `funcionario_cobranca` no snapshot da
    data GM e sem tramitação (lacuna entre “todos os operadores” e a soma por operador).

    Geral: com Data início e/ou Data fim, a tabela usa o mesmo critério de evento no histórico
    (negativação + fallback `data_negativacao`) e acrescenta positivações (`removido_*`) no intervalo.
    A resposta inclui `historico` e `total_historico` vazios (compatibilidade; UI unificada na lista).
    """
    kw = _negativacao_listagem_kwargs_from_request(request)

    conn = _get_db()
    cursor = conn.cursor()
    try:
        payload = _negativacao_listagem_payload_interno(cursor, **kw)
        return jsonify(payload)
    finally:
        try:
            cursor.close()
        except Exception:
            pass
        try:
            conn.close()
        except Exception:
            pass


@app.route('/api/negativacao/listagem/excel', methods=['GET'])
def api_negativacao_listagem_excel():
    """Excel com duas folhas (NEGATIVAÇÃO / POSITIVAÇÃO), mesmos filtros que a listagem na UI.

    Ignora ``preview_ativos``: exporta sempre o conjunto completo para os filtros atuais.
    Opcional: ``excel_sort_neg_col``, ``excel_sort_neg_dir``, ``excel_sort_pos_col``,
    ``excel_sort_pos_dir`` (asc/desc), alinhados ao ordenamento das tabelas no navegador.
    """
    if not session.get('funcionario_id'):
        return jsonify({'error': 'Nao autenticado. Faca login novamente.'}), 401

    try:
        from openpyxl import Workbook
    except ImportError:
        return jsonify({
            'error': 'Biblioteca openpyxl nao instalada. Execute: pip install openpyxl',
        }), 503

    kw = _negativacao_listagem_kwargs_from_request(request, preview_ativos_force=False)
    fb_sort = kw['sort_at']
    fb_order = kw['order_at']
    if fb_order not in ('asc', 'desc'):
        fb_order = 'desc'

    neg_sc = _negativacao_listagem_normalize_excel_sort_col(
        request.args.get('excel_sort_neg_col'), fb_sort)
    neg_od = (request.args.get('excel_sort_neg_dir') or fb_order).strip().lower()
    if neg_od not in ('asc', 'desc'):
        neg_od = fb_order

    pos_sc = _negativacao_listagem_normalize_excel_sort_col(
        request.args.get('excel_sort_pos_col'), fb_sort)
    pos_od = (request.args.get('excel_sort_pos_dir') or fb_order).strip().lower()
    if pos_od not in ('asc', 'desc'):
        pos_od = fb_order

    hdr = (
        'Grupo',
        'Cota',
        'Grupo/Cota',
        'Nº contrato',
        'ID contrato',
        'Parcela',
        'ID parcela',
        'ID registro',
        'Atraso',
        'Status parcela',
        'Tipo de evento',
        'Data',
        'Operador',
        'Status contrato',
        'Detalhe',
    )

    conn = _get_db()
    cursor = conn.cursor()
    try:
        payload = _negativacao_listagem_payload_interno(cursor, **kw)
    except Exception as exc:
        app.logger.exception('negativacao/listagem/excel: falha ao montar listagem')
        try:
            cursor.close()
        except Exception:
            pass
        try:
            conn.close()
        except Exception:
            pass
        return jsonify({'error': f'Falha ao consultar dados para o Excel: {exc}'}), 500

    try:
        cursor.close()
    except Exception:
        pass
    try:
        conn.close()
    except Exception:
        pass

    if kw['apenas_cobranca']:
        neg_rows = list(payload.get('ativos_negativados') or [])
        pos_rows = list(payload.get('ativos_positivados') or [])
    else:
        ativos = payload.get('ativos') or []
        neg_rows, pos_rows = _split_negativacao_ativos_negativados_positivados(ativos)

    neg_rows = _negativacao_listagem_excel_sort_rows(neg_rows, neg_sc, neg_od)
    pos_rows = _negativacao_listagem_excel_sort_rows(pos_rows, pos_sc, pos_od)

    try:
        wb = Workbook()
        ws_neg = wb.active
        ws_neg.title = 'NEGATIVAÇÃO'
        ws_neg.append(hdr)
        if not neg_rows:
            ws_neg.append(
                ('Nenhum registro na lista de negativação para os filtros atuais.',) + ('',) * 14
            )
        else:
            for r in neg_rows:
                ws_neg.append(_negativacao_listagem_excel_row(r))

        ws_pos = wb.create_sheet(title='POSITIVAÇÃO')
        ws_pos.append(hdr)
        if not pos_rows:
            ws_pos.append(
                ('Nenhum registro na lista de positivação para os filtros atuais.',) + ('',) * 14
            )
        else:
            for r in pos_rows:
                ws_pos.append(_negativacao_listagem_excel_row(r))

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
    except Exception as exc:
        app.logger.exception('negativacao/listagem/excel: falha ao montar ficheiro')
        return jsonify({'error': f'Falha ao gerar o Excel: {exc}'}), 500

    fname = 'negativacao_listagem_' + datetime.datetime.now().strftime('%Y%m%d_%H%M') + '.xlsx'
    return send_file(
        buf,
        as_attachment=True,
        download_name=fname,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )


@app.route('/api/negativacao/positivar-lote-serasa', methods=['POST'])
def api_negativacao_positivar_lote_serasa_placeholder():
    """Envio em lote ao Serasa (placeholder): positivar ou negativar parcelas do bloco.

    Negativar: todas as parcelas devem ter linha em `negativacao` com status tracker ou falha.

    Positivar: parcelas sem linha em `negativacao` são aceitas (ex.: positivação já refletida pelo GM
    na lista da carteira); parcelas com linha só se status for aguardando_positivacao_serasa.
    Demais estados com linha ativa são rejeitados.

    Campo opcional `faixa`: `negativados` ou `positivados` (apenas rastreio; elegibilidade usa ids).
    """
    payload = request.get_json(silent=True) or {}
    tipo = (payload.get('tipo_operacao') or '').strip().lower()
    if tipo not in ('positivar', 'negativar'):
        return jsonify({'error': 'tipo_operacao deve ser "positivar" ou "negativar".'}), 400
    faixa = (payload.get('faixa') or '').strip().lower()
    if faixa and faixa not in ('negativados', 'positivados'):
        return jsonify({'error': 'faixa inválida (use negativados ou positivados).'}), 400
    raw_ids = payload.get('ids_parcela')
    if not isinstance(raw_ids, list) or len(raw_ids) == 0:
        return jsonify({'error': 'ids_parcela deve ser uma lista não vazia.'}), 400
    ids = []
    for x in raw_ids:
        try:
            ids.append(int(x))
        except (TypeError, ValueError):
            return jsonify({'error': 'ids_parcela contém valor inválido.'}), 400
    ids = list(dict.fromkeys(ids))
    if len(ids) > 50000:
        return jsonify({'error': 'Quantidade de parcelas acima do limite provisório.'}), 400

    conn = _get_db()
    cursor = conn.cursor()
    try:
        _ensure_negativacao_table(cursor)
        ph = ','.join(['%s'] * len(ids))
        cursor.execute(
            f'SELECT id_parcela, status FROM negativacao WHERE id_parcela IN ({ph})',
            ids,
        )
        por_parcela = {}
        for r in cursor.fetchall() or []:
            try:
                ip = int(r.get('id_parcela'))
            except (TypeError, ValueError):
                continue
            por_parcela[ip] = (r.get('status') or '').strip().lower()

        invalidas = []
        if tipo == 'negativar':
            if len(por_parcela) != len(ids):
                return jsonify({
                    'error': 'Uma ou mais parcelas não possuem negativação ativa no cadastro.'
                }), 400
            for ip in ids:
                st = por_parcela.get(ip, '')
                if st not in _NEG_STATUS_SERASA_ELEGIVEL_NEGATIVAR:
                    invalidas.append(ip)
        else:
            for ip in ids:
                st = por_parcela.get(ip)
                if st is None:
                    continue
                if st == _NEG_STATUS_AGUARDANDO_POS_SERASA:
                    continue
                invalidas.append(ip)

        if invalidas:
            msg = (
                'Uma ou mais parcelas não estão elegíveis '
                '(negativar: apenas tracker ou falha de envio; '
                'positivar: sem cadastro ativo ou apenas aguardando envio Serasa).'
            )
            return jsonify({'error': msg}), 400

        resposta_mock = '{"mock":true,"serasa":"placeholder"}'
        if tipo == 'negativar':
            cursor.execute(
                f"""
                UPDATE negativacao
                SET status = 'enviado', resposta_api = %s, updated_at = CURRENT_TIMESTAMP
                WHERE id_parcela IN ({ph})
                  AND LOWER(COALESCE(status, '')) IN ('registrado_tracker', 'falhou')
                """,
                [resposta_mock] + ids,
            )
        else:
            ids_aguardando = [
                ip for ip in ids if por_parcela.get(ip) == _NEG_STATUS_AGUARDANDO_POS_SERASA
            ]
            if ids_aguardando:
                ph2 = ','.join(['%s'] * len(ids_aguardando))
                cursor.execute(
                    f'DELETE FROM negativacao WHERE id_parcela IN ({ph2}) '
                    f"AND LOWER(COALESCE(status, '')) = %s",
                    ids_aguardando + [_NEG_STATUS_AGUARDANDO_POS_SERASA],
                )

        conn.commit()
    except Exception as exc:
        app.logger.exception('api_negativacao_positivar_lote_serasa_placeholder')
        try:
            conn.rollback()
        except Exception:
            pass
        return jsonify({'error': str(exc)}), 500
    finally:
        try:
            cursor.close()
        except Exception:
            pass
        try:
            conn.close()
        except Exception:
            pass

    app.logger.info(
        'negativacao serasa-envio-lote (mock): tipo_operacao=%s faixa=%s qtd=%s',
        tipo,
        faixa or '-',
        len(ids),
    )
    faixa_txt = f' faixa={faixa}' if faixa else ''
    verbo = 'Positivação' if tipo == 'positivar' else 'Negativação'
    return jsonify({
        'success': True,
        'mock': True,
        'tipo_operacao': tipo,
        'faixa': faixa or None,
        'quantidade': len(ids),
        'mensagem': (
            f'{verbo} Serasa (simulação){faixa_txt}: {len(ids)} parcela(s). '
            'Estado interno atualizado; substituir por chamada real à API quando disponível.'
        ),
    })


@app.route('/api/negativacao/serasa-arquivo-txt/sequencia', methods=['GET'])
def api_negativacao_serasa_arquivo_txt_sequencia():
    """Proxima sequencia global SERASA_GM (ultimo registo neg/pos + 1) para o modal de geracao."""
    if not session.get('funcionario_id'):
        return jsonify({'error': 'Nao autenticado. Faca login novamente.'}), 401
    tipo = (request.args.get('tipo_operacao') or '').strip().lower()
    if tipo not in ('positivar', 'negativar'):
        return jsonify({'error': 'tipo_operacao deve ser "positivar" ou "negativar".'}), 400

    conn = _get_db()
    cursor = conn.cursor()
    try:
        info = _serasa_gm_proxima_sequencia_global(cursor)
        return jsonify({'success': True, 'tipo_operacao': tipo, **info})
    except ValueError as exc:
        return jsonify({'error': str(exc)}), 400
    except Exception as exc:
        app.logger.exception('api_negativacao_serasa_arquivo_txt_sequencia')
        return jsonify({'error': str(exc)}), 500
    finally:
        try:
            cursor.close()
        except Exception:
            pass
        try:
            conn.close()
        except Exception:
            pass


@app.route('/api/negativacao/serasa-arquivo-txt', methods=['POST'])
def api_negativacao_serasa_arquivo_txt():
    """Gera ficheiro TXT SERASA-CONVEM para download.

    Negativar: inclusao com linhas de detalhe **1I** (motivo ``00``), layout alinhado ao legado PHP
    (endereco 45 + bairro 20, valor em centavos, bloco ``J10`` + CNPJ/nome credor). Positivar:
    exclusao com linhas **1E** (motivo ``02``), mesma estrutura de ficheiro que
    ``docs/referencias/legacy-php/sistema.geracao.arquivo.negativacao.serasa.php`` (cabecalho + detalhes + rodape), cabecalho
    a partir do modelo ``SERASA_GM_*4910*.TXT``. Os ``ids`` servem para validacao de elegibilidade
    alinhada a UI e ao mock ``positivar-lote-serasa``. Grava copia em
    ``registro_txt_negativacao`` ou ``registro_txt_positivacao`` antes do download.
    """
    if not session.get('funcionario_id'):
        return jsonify({'error': 'Nao autenticado. Faca login novamente.'}), 401

    payload = request.get_json(silent=True) or {}
    tipo = (payload.get('tipo_operacao') or '').strip().lower()
    if tipo not in ('positivar', 'negativar'):
        return jsonify({'error': 'tipo_operacao deve ser "positivar" ou "negativar".'}), 400
    raw_ids = payload.get('ids_parcela')
    if not isinstance(raw_ids, list) or len(raw_ids) == 0:
        return jsonify({'error': 'ids_parcela deve ser uma lista nao vazia.'}), 400
    ids = []
    for x in raw_ids:
        try:
            ids.append(int(x))
        except (TypeError, ValueError):
            return jsonify({'error': 'ids_parcela contem valor invalido.'}), 400
    ids = list(dict.fromkeys(ids))
    if len(ids) > 20000:
        return jsonify({'error': 'Quantidade de parcelas acima do limite para um unico arquivo.'}), 400

    seq, err_seq = _serasa_gm_parse_sequencia_payload(payload.get('sequencia'))
    if err_seq:
        return jsonify({'error': err_seq}), 400

    data_hoje = datetime.datetime.now().strftime('%d%m%Y')
    fname = secure_filename(_serasa_gm_montar_nome_arquivo(data_hoje, seq)) or 'serasa_gm.txt'

    conn = _get_db()
    cursor = conn.cursor()
    body = None
    modo = None
    try:
        _ensure_negativacao_table(cursor)
        if tipo != 'negativar':
            _ensure_negativacao_historico_table(cursor)
        por_parcela = _negativacao_serasa_status_por_parcela(cursor, ids)
        err_elig = _negativacao_serasa_erro_elegibilidade_txt(tipo, ids, por_parcela)
        if err_elig:
            return jsonify({'error': err_elig}), 400

        try:
            mod = _get_serasa_conv_txt()
        except Exception as exc:
            app.logger.exception('serasa_conv_txt: falha ao carregar modulo')
            return jsonify({'error': f'Modulo de layout SERASA indisponivel: {exc}'}), 500

        linhas = _negativacao_fetch_serasa_inclusao_payloads(
            cursor, ids, inner_negativacao=(tipo == 'negativar')
        )
        if len(linhas) != len(ids):
            return jsonify({
                'error': 'Nao foi possivel obter dados de todas as parcelas para o arquivo.',
            }), 400
        modo = 'inclusao' if tipo == 'negativar' else 'exclusao'

        body, _fname_ign = mod.montar_arquivo_txt(modo, linhas)
        _registro_txt_serasa_insert(
            cursor,
            conn,
            tipo=tipo,
            nome_arquivo=fname,
            conteudo_bytes=body,
            id_funcionario=session['funcionario_id'],
        )
    except ValueError as exc:
        return jsonify({'error': str(exc)}), 400
    except FileNotFoundError as exc:
        app.logger.warning('serasa templates ausentes: %s', exc)
        return jsonify({
            'error': (
                'Modelos TXT SERASA nao encontrados. Coloque SERASA_GM_*4912*.TXT e *4910*.TXT '
                'em Python/serasa_templates ou defina SERASA_CONV_TEMPLATE_DIR.'
            ),
        }), 404
    except Exception as exc:
        app.logger.exception('api_negativacao_serasa_arquivo_txt')
        return jsonify({'error': str(exc)}), 500
    finally:
        try:
            cursor.close()
        except Exception:
            pass
        try:
            conn.close()
        except Exception:
            pass

    app.logger.info(
        'negativacao serasa-arquivo-txt: tipo_operacao=%s qtd_ids=%s modo=%s seq=%s fname=%s',
        tipo,
        len(ids),
        modo,
        seq,
        fname,
    )
    return send_file(
        io.BytesIO(body),
        as_attachment=True,
        download_name=fname,
        mimetype='text/plain; charset=iso-8859-1',
    )


@app.route('/api/negativacao/serasa-retorno-txt', methods=['POST'])
def api_negativacao_serasa_retorno_txt():
    """Importa TXT de retorno PEFIN (detalhe ``1I`` / ``1E``) e grava ``negativacao_historico``.

    Campo multipart ``arquivo``: ficheiro de 600 caracteres por linha (latin-1), como os exemplos
    em ``docs/referencias/serasa/amostras-retorno/``. Cabeçalho (``0``) e trailer (``9``) são ignorados.
    """
    if not session.get('funcionario_id'):
        return jsonify({'error': 'Nao autenticado. Faca login novamente.'}), 401
    up = request.files.get('arquivo')
    if not up or not getattr(up, 'filename', None):
        return jsonify({'error': 'Ficheiro em falta (campo ``arquivo``).'}), 400
    nome = secure_filename(up.filename) or 'retorno_pefin.txt'
    raw = up.read()
    if not raw:
        return jsonify({'error': 'Ficheiro vazio.'}), 400
    if len(raw) > 25 * 1024 * 1024:
        return jsonify({'error': 'Ficheiro demasiado grande (limite 25MB).'}), 400

    try:
        pef = _get_serasa_pefin_erros()
    except Exception as exc:
        app.logger.exception('serasa_pefin_erros: falha ao carregar modulo')
        return jsonify({'error': f'Modulo PEFIN indisponivel: {exc}'}), 500

    try:
        fid = int(session.get('funcionario_id'))
    except (TypeError, ValueError):
        fid = None

    texto = raw.decode('latin-1', errors='replace')
    linhas = texto.splitlines()
    lin_len = int(getattr(pef, 'LINHA_DETALHE_PEFIN_LEN', 600))

    inseridos = 0
    ignorados = 0
    nao_resolvidos = []
    err_msgs = []

    conn = _get_db()
    cursor = conn.cursor()
    try:
        _ensure_negativacao_historico_table(cursor)
        for raw_line in linhas:
            line = raw_line.rstrip('\r')
            if not line.strip():
                ignorados += 1
                continue
            if len(line) != lin_len:
                ignorados += 1
                continue
            c0 = line[0]
            if c0 in ('0', '9'):
                ignorados += 1
                continue
            parsed = pef.parse_linha_detalhe_pefin_retorno(line)
            if not parsed:
                ignorados += 1
                continue
            id_contrato = _negativacao_resolve_contrato_pefin_retorno(
                cursor,
                parsed.get('contrato_nrs_16'),
                parsed.get('documento_15'),
            )
            if not id_contrato:
                nao_resolvidos.append({
                    'sequencial': parsed.get('sequencial_linha_arquivo'),
                    'motivo': 'contrato_nao_resolvido',
                })
                continue
            tipo_ev = (
                'negativacao_retorno'
                if (parsed.get('tipo_operacao_ie') or '').upper() == 'I'
                else 'positivacao_retorno'
            )
            detalhe = _negativacao_montar_detalhe_pefin_retorno(
                nome, parsed, pefin_desc=pef.descricao_codigo_erro
            )
            data_ev = _negativacao_data_evento_de_referencia_pefin(
                parsed.get('data_referencia_yyyymmdd')
            )
            _insert_negativacao_historico(
                cursor,
                id_contrato,
                None,
                None,
                tipo_ev,
                data_ev,
                id_funcionario=fid,
                dias_atraso=None,
                status_snapshot=None,
                detalhe=detalhe,
                id_arquivo_gm=None,
            )
            inseridos += 1
        conn.commit()
    except Exception as exc:
        try:
            conn.rollback()
        except Exception:
            pass
        app.logger.exception('api_negativacao_serasa_retorno_txt')
        err_msgs.append(str(exc))
        return jsonify({
            'success': False,
            'error': str(exc),
            'inseridos': inseridos,
            'ignorados': ignorados,
            'nao_resolvidos': nao_resolvidos,
            'erros': err_msgs,
        }), 500
    finally:
        try:
            cursor.close()
        except Exception:
            pass
        try:
            conn.close()
        except Exception:
            pass

    app.logger.info(
        'negativacao serasa-retorno-txt: ficheiro=%s inseridos=%s ignorados=%s nao_resolv=%s',
        nome,
        inseridos,
        ignorados,
        len(nao_resolvidos),
    )
    return jsonify({
        'success': True,
        'inseridos': inseridos,
        'ignorados': ignorados,
        'nao_resolvidos': nao_resolvidos,
        'erros': err_msgs,
    })


@app.route('/api/negativacao/observacao', methods=['POST'])
def api_negativacao_observacao():
    payload = request.get_json(silent=True) or {}
    try:
        id_contrato = int(payload.get('id_contrato'))
    except (TypeError, ValueError):
        return jsonify({'error': 'id_contrato invalido.'}), 400
    detalhe = (payload.get('detalhe') or '').strip()
    if not detalhe:
        return jsonify({'error': 'detalhe e obrigatorio.'}), 400

    fid = session.get('funcionario_id')
    try:
        fid = int(fid) if fid is not None else None
    except (TypeError, ValueError):
        fid = None

    conn = _get_db()
    cursor = conn.cursor()
    try:
        _ensure_negativacao_historico_table(cursor)
        _insert_negativacao_historico(
            cursor,
            id_contrato,
            None,
            None,
            'observacao',
            datetime.datetime.now(),
            fid,
            None,
            None,
            detalhe,
        )
        conn.commit()
        return jsonify({'success': True})
    except Exception as exc:
        app.logger.exception('api_negativacao_observacao')
        try:
            conn.rollback()
        except Exception:
            pass
        return jsonify({'error': str(exc)}), 500
    finally:
        try:
            cursor.close()
        except Exception:
            pass
        try:
            conn.close()
        except Exception:
            pass


@app.route('/api/negativacao/remover-manual', methods=['POST'])
def api_negativacao_remover_manual():
    payload = request.get_json(silent=True) or {}
    try:
        id_parcela = int(payload.get('id_parcela'))
    except (TypeError, ValueError):
        return jsonify({'error': 'id_parcela invalido.'}), 400
    motivo = (payload.get('motivo') or '').strip() or (
        'Remocao manual da negativacao registrada neste painel.'
    )

    fid = session.get('funcionario_id')
    try:
        fid = int(fid) if fid is not None else None
    except (TypeError, ValueError):
        fid = None

    conn = _get_db()
    cursor = conn.cursor()
    try:
        _ensure_negativacao_table(cursor)
        _ensure_negativacao_historico_table(cursor)
        cursor.execute(
            "SELECT * FROM negativacao WHERE id_parcela = %s LIMIT 1",
            (id_parcela,),
        )
        row = cursor.fetchone()
        if not row:
            return jsonify({'error': 'Nenhuma negativacao ativa para esta parcela.'}), 404

        st_atual = (row.get('status') or '').strip().lower()
        if st_atual == _NEG_STATUS_AGUARDANDO_POS_SERASA:
            return jsonify({
                'error': (
                    'Esta parcela já está com positivação interna registrada; '
                    'use o envio ao Serasa (Positivar todos / integração).'
                )
            }), 400

        id_contrato = int(row['id_contrato'])
        data_ev = datetime.datetime.now()
        _insert_negativacao_historico(
            cursor,
            id_contrato,
            id_parcela,
            row.get('numero_parcela'),
            'removido_manual',
            data_ev,
            fid,
            row.get('dias_atraso'),
            row.get('status'),
            motivo,
        )
        cursor.execute(
            """
            UPDATE negativacao
            SET status = %s, resposta_api = %s, updated_at = CURRENT_TIMESTAMP
            WHERE id_parcela = %s
            """,
            (
                _NEG_STATUS_AGUARDANDO_POS_SERASA,
                '{"interno":"positivacao_manual_pendente_serasa"}',
                id_parcela,
            ),
        )
        conn.commit()
        return jsonify({'success': True})
    except Exception as exc:
        app.logger.exception('api_negativacao_remover_manual')
        try:
            conn.rollback()
        except Exception:
            pass
        return jsonify({'error': str(exc)}), 500
    finally:
        try:
            cursor.close()
        except Exception:
            pass
        try:
            conn.close()
        except Exception:
            pass


@app.route('/api/negativacao/registrar-manual-parcela', methods=['POST'])
def api_negativacao_registrar_manual_parcela():
    """Recoloca uma parcela em `negativacao` a partir do histórico (ex.: após positivação).

    Mesma elegibilidade do fluxo de envio em lote: parcela e contrato em aberto,
    atraso entre 31 e 89 dias na data de referência do último arquivo GM.
    """
    payload = request.get_json(silent=True) or {}
    try:
        id_parcela = int(payload.get('id_parcela'))
    except (TypeError, ValueError):
        return jsonify({'error': 'id_parcela invalido.'}), 400

    fid = session.get('funcionario_id')
    try:
        fid = int(fid) if fid is not None else None
    except (TypeError, ValueError):
        fid = None

    conn = _get_db()
    cursor = conn.cursor()
    try:
        _ensure_negativacao_table(cursor)
        _ensure_negativacao_historico_table(cursor)

        cursor.execute(
            "SELECT 1 FROM negativacao WHERE id_parcela = %s LIMIT 1",
            (id_parcela,),
        )
        if cursor.fetchone():
            return jsonify({'error': 'Esta parcela ja possui negativacao ativa no cadastro.'}), 400

        data_ref = _get_data_referencia_arquivos_gm(cursor)
        cursor.execute(
            """
            SELECT p.id, p.id_contrato, p.numero_parcela, p.vencimento,
                   DATEDIFF(%s, p.vencimento) AS dias_atraso
            FROM parcela p
            INNER JOIN contrato c ON c.id = p.id_contrato
            WHERE p.id = %s AND p.status = 'cobranca' AND c.status = 'cobranca'
            LIMIT 1
            """,
            (data_ref, id_parcela),
        )
        row = cursor.fetchone()
        if not row:
            return jsonify({
                'error': 'Parcela nao encontrada em aberto ou contrato nao esta aberto.'
            }), 400

        try:
            dias = int(row.get('dias_atraso') or 0)
        except (TypeError, ValueError):
            dias = 0

        if not (30 < dias < 90):
            return jsonify({
                'error': (
                    f'Elegibilidade: atraso atual {dias} dia(s) na data de referencia '
                    'dos arquivos GM (necessario entre 31 e 89 dias).'
                )
            }), 400

        id_contrato = int(row['id_contrato'])
        num_parc = row.get('numero_parcela')
        resposta_json = '{"origem":"painel_negativacao_registrar_manual_parcela"}'

        cursor.execute(
            "INSERT INTO negativacao "
            "(id_contrato, id_parcela, numero_parcela, dias_atraso, status, resposta_api, id_funcionario) "
            "VALUES (%s, %s, %s, %s, 'enviado', %s, %s)",
            (id_contrato, id_parcela, num_parc, dias, resposta_json, fid),
        )

        cursor.execute(
            "SELECT data_negativacao FROM negativacao WHERE id_parcela = %s LIMIT 1",
            (id_parcela,),
        )
        row_dn = cursor.fetchone()
        data_ev = row_dn.get('data_negativacao') if row_dn else datetime.datetime.now()

        _insert_negativacao_historico(
            cursor,
            id_contrato,
            id_parcela,
            num_parc,
            'negativado_manual',
            data_ev,
            fid,
            dias,
            'enviado',
            'Negativacao registrada manualmente pelo modulo Negativacao (reingresso apos positivacao).',
        )
        conn.commit()
        return jsonify({'success': True})
    except IntegrityError:
        try:
            conn.rollback()
        except Exception:
            pass
        return jsonify({'error': 'Esta parcela ja possui negativacao cadastrada.'}), 400
    except Exception as exc:
        app.logger.exception('api_negativacao_registrar_manual_parcela')
        try:
            conn.rollback()
        except Exception:
            pass
        return jsonify({'error': str(exc)}), 500
    finally:
        try:
            cursor.close()
        except Exception:
            pass
        try:
            conn.close()
        except Exception:
            pass


# ---------------------------------------------------------------------------
# API: Protocolo, Solicitação e Mensagem
# ---------------------------------------------------------------------------

def _optional_positive_int(val):
    if val in (None, '', False):
        return None
    try:
        n = int(val)
    except (TypeError, ValueError):
        return None
    return n if n > 0 else None


def _mensagem_fio_sem_ciclo(cursor, parent_id):
    """Sobe por id_resposta a partir do pai; devolve (ok, erro_ou_none)."""
    seen = set()
    cur = int(parent_id)
    for _ in range(500):
        if cur in seen:
            return False, 'Fio de mensagens inconsistente (ciclo).'
        seen.add(cur)
        cursor.execute('SELECT id_resposta FROM mensagem WHERE id = %s', (cur,))
        row = cursor.fetchone()
        if not row:
            return False, 'Mensagem referenciada nao encontrada.'
        pr = row.get('id_resposta')
        if pr is None:
            return True, None
        cur = int(pr)
    return False, 'Fio de mensagens excede o limite de profundidade.'


def _solicitacao_fio_sem_ciclo(cursor, parent_id):
    """Sobe por id_resposta a partir do pai; devolve (ok, erro_ou_none)."""
    seen = set()
    cur = int(parent_id)
    for _ in range(500):
        if cur in seen:
            return False, 'Fio de solicitacoes inconsistente (ciclo).'
        seen.add(cur)
        cursor.execute('SELECT id_resposta FROM solicitacao WHERE id = %s', (cur,))
        row = cursor.fetchone()
        if not row:
            return False, 'Solicitacao referenciada nao encontrada.'
        pr = row.get('id_resposta')
        if pr is None:
            return True, None
        cur = int(pr)
    return False, 'Fio de solicitacoes excede o limite de profundidade.'


@app.route('/api/mensagem', methods=['POST'])
def api_mensagem_post():
    """Nova mensagem ou resposta: id_remetente = sessao.

    Com ``id_resposta``, o destinatario e o outro participante do pai
    (``id_destinatario`` no JSON e ignorado nesse caso).
    """
    fid = session.get('funcionario_id')
    if not fid:
        return jsonify({'error': 'Nao autenticado.'}), 401
    payload = request.get_json(silent=True) or {}
    fid = int(fid)
    id_resposta = _optional_positive_int(payload.get('id_resposta'))
    assunto = (payload.get('assunto') or '').strip()
    if not assunto:
        return jsonify({'error': 'Assunto obrigatorio.'}), 400
    descricao = (payload.get('descricao') or '').strip() or None

    conn = _get_db()
    cursor = conn.cursor()
    try:
        if id_resposta is None:
            id_ins = None
            try:
                dest_id = int(payload.get('id_destinatario'))
            except (TypeError, ValueError):
                return jsonify({'error': 'Destinatario invalido.'}), 400
            if int(dest_id) == fid:
                return jsonify({'error': 'O destinatario deve ser outro funcionario.'}), 400
            cursor.execute(
                'SELECT id FROM funcionario WHERE id = %s AND ativo = 1',
                (dest_id,),
            )
            if not cursor.fetchone():
                return jsonify({'error': 'Destinatario nao encontrado ou inativo.'}), 400
        else:
            ok_fio, err_fio = _mensagem_fio_sem_ciclo(cursor, id_resposta)
            if not ok_fio:
                return jsonify({'error': err_fio}), 400
            cursor.execute(
                'SELECT id, id_remetente, id_destinatario FROM mensagem WHERE id = %s',
                (id_resposta,),
            )
            pai = cursor.fetchone()
            if not pai:
                return jsonify({'error': 'Mensagem pai nao encontrada.'}), 404
            rem_p = int(pai['id_remetente'])
            des_p = int(pai['id_destinatario'])
            if fid not in (rem_p, des_p):
                return jsonify({'error': 'So participantes do fio podem responder.'}), 403
            dest_id = des_p if fid == rem_p else rem_p
            if int(dest_id) == fid:
                return jsonify({'error': 'Nao e possivel determinar o destinatario da resposta.'}), 400
            cursor.execute(
                'SELECT id FROM funcionario WHERE id = %s AND ativo = 1',
                (dest_id,),
            )
            if not cursor.fetchone():
                return jsonify({'error': 'Destinatario nao encontrado ou inativo.'}), 400
            id_ins = id_resposta

        cursor.execute(
            'INSERT INTO mensagem (id_remetente, id_destinatario, assunto, descricao, id_resposta) '
            'VALUES (%s, %s, %s, %s, %s)',
            (fid, dest_id, assunto, descricao, id_ins),
        )
        conn.commit()
        new_id = cursor.lastrowid
    except Exception as exc:
        conn.rollback()
        app.logger.exception('api_mensagem_post')
        return jsonify({'error': str(exc)}), 500
    finally:
        cursor.close()
        conn.close()
    return jsonify({'ok': True, 'id': new_id})


@app.route('/api/solicitacao', methods=['POST'])
def api_solicitacao_post():
    """Nova solicitacao ou resposta: id_remetente = sessao.

    Com ``id_resposta``, o destinatario e o outro participante do pai;
    ``id_contrato`` omitido herda do pai quando existir.
    """
    fid = session.get('funcionario_id')
    if not fid:
        return jsonify({'error': 'Nao autenticado.'}), 401
    payload = request.get_json(silent=True) or {}
    fid = int(fid)
    id_resposta = _optional_positive_int(payload.get('id_resposta'))
    data_aguardar = (payload.get('data_aguardar') or '').strip()
    if not data_aguardar:
        return jsonify({'error': 'Data a aguardar obrigatoria.'}), 400
    descricao = (payload.get('descricao') or '').strip() or None
    if descricao and len(descricao) > 255:
        return jsonify({'error': 'Descricao: maximo 255 caracteres.'}), 400
    assunto = (payload.get('assunto') or '').strip()
    if not assunto:
        return jsonify({'error': 'Assunto obrigatorio.'}), 400
    if len(assunto) > 255:
        return jsonify({'error': 'Assunto: maximo 255 caracteres.'}), 400

    conn = _get_db()
    cursor = conn.cursor()
    try:
        if id_resposta is None:
            try:
                dest_id = int(payload.get('id_destinatario'))
            except (TypeError, ValueError):
                return jsonify({'error': 'Destinatario invalido.'}), 400
            if int(dest_id) == fid:
                return jsonify({'error': 'O destinatario deve ser outro funcionario.'}), 400
            id_contrato = None
            raw_c = payload.get('id_contrato')
            if raw_c not in (None, ''):
                try:
                    id_contrato = int(raw_c)
                except (TypeError, ValueError):
                    return jsonify({'error': 'ID contrato invalido.'}), 400
            id_ins = None
        else:
            ok_fio, err_fio = _solicitacao_fio_sem_ciclo(cursor, id_resposta)
            if not ok_fio:
                return jsonify({'error': err_fio}), 400
            cursor.execute(
                'SELECT id, id_remetente, id_destinatario, id_contrato FROM solicitacao WHERE id = %s',
                (id_resposta,),
            )
            pai = cursor.fetchone()
            if not pai:
                return jsonify({'error': 'Solicitacao pai nao encontrada.'}), 404
            rem_p = int(pai['id_remetente'])
            des_p = int(pai['id_destinatario'])
            if fid not in (rem_p, des_p):
                return jsonify({'error': 'So participantes do fio podem responder.'}), 403
            dest_id = des_p if fid == rem_p else rem_p
            if int(dest_id) == fid:
                return jsonify({'error': 'Nao e possivel determinar o destinatario da resposta.'}), 400
            id_ins = id_resposta
            id_contrato = None
            raw_c = payload.get('id_contrato')
            if raw_c not in (None, ''):
                try:
                    id_contrato = int(raw_c)
                except (TypeError, ValueError):
                    return jsonify({'error': 'ID contrato invalido.'}), 400
            elif pai.get('id_contrato') is not None:
                id_contrato = int(pai['id_contrato'])

        cursor.execute(
            'SELECT id FROM funcionario WHERE id = %s AND ativo = 1',
            (dest_id,),
        )
        if not cursor.fetchone():
            return jsonify({'error': 'Destinatario nao encontrado ou inativo.'}), 400
        if id_contrato is not None:
            cursor.execute('SELECT id FROM contrato WHERE id = %s', (id_contrato,))
            if not cursor.fetchone():
                return jsonify({'error': 'Contrato nao encontrado.'}), 400

        cursor.execute(
            'INSERT INTO solicitacao (id_remetente, id_destinatario, data_aguardar, descricao, '
            'id_resposta, id_contrato, assunto) VALUES (%s, %s, %s, %s, %s, %s, %s)',
            (fid, dest_id, data_aguardar, descricao, id_ins, id_contrato, assunto),
        )
        conn.commit()
        new_id = cursor.lastrowid
    except Exception as exc:
        conn.rollback()
        app.logger.exception('api_solicitacao_post')
        return jsonify({'error': str(exc)}), 500
    finally:
        cursor.close()
        conn.close()
    return jsonify({'ok': True, 'id': new_id})


@app.route('/api/protocolo', methods=['POST'])
def api_protocolo_post():
    """Registra protocolo interno; o destinatário recebe notificação no sininho."""
    fid = session.get('funcionario_id')
    if not fid:
        return jsonify({'error': 'Nao autenticado.'}), 401
    fid = int(fid)
    payload = request.get_json(silent=True) or {}
    titulo = (payload.get('titulo') or '').strip()
    if not titulo:
        return jsonify({'error': 'Titulo obrigatorio.'}), 400
    titulo = titulo[:50]
    descricao = (payload.get('descricao') or '').strip() or None
    try:
        dest_id = int(payload.get('id_destinatario'))
    except (TypeError, ValueError):
        return jsonify({'error': 'Destinatario invalido.'}), 400
    if int(dest_id) == fid:
        return jsonify({'error': 'O destinatario deve ser outro funcionario.'}), 400
    id_contrato = None
    raw_c = payload.get('id_contrato')
    if raw_c not in (None, ''):
        try:
            id_contrato = int(raw_c)
        except (TypeError, ValueError):
            return jsonify({'error': 'ID contrato invalido.'}), 400

    conn = _get_db()
    cursor = conn.cursor()
    new_id = None
    try:
        cursor.execute(
            'SELECT id FROM funcionario WHERE id = %s AND ativo = 1',
            (dest_id,),
        )
        if not cursor.fetchone():
            return jsonify({'error': 'Destinatario nao encontrado ou inativo.'}), 400
        if id_contrato is not None:
            cursor.execute('SELECT id FROM contrato WHERE id = %s', (id_contrato,))
            if not cursor.fetchone():
                return jsonify({'error': 'Contrato nao encontrado.'}), 400
        cursor.execute(
            'INSERT INTO protocolo (id_remetente, id_destinatario, titulo, descricao, id_contrato) '
            'VALUES (%s, %s, %s, %s, %s)',
            (fid, dest_id, titulo, descricao, id_contrato),
        )
        conn.commit()
        new_id = cursor.lastrowid
    except Exception as exc:
        conn.rollback()
        app.logger.exception('api_protocolo_post')
        return jsonify({'error': str(exc)}), 500
    finally:
        cursor.close()
        conn.close()
    return jsonify({'ok': True, 'id': new_id})


@app.route('/api/solicitacao/moderacao/pendentes', methods=['GET'])
def api_solicitacao_moderacao_pendentes():
    """Lista pedidos pendentes de alteracao/exclusao (tramitacao/agenda) para Gestor/Administrador."""
    forb = _admin_json_forbidden()
    if forb:
        return forb
    conn = _get_db()
    cursor = conn.cursor()
    try:
        _ensure_solicitacao_moderacao_table(cursor)
        cursor.execute(
            """
            SELECT m.*, c.grupo, c.cota, fs.nome AS solicitante_nome
            FROM solicitacao_moderacao m
            INNER JOIN contrato c ON c.id = m.id_contrato
            LEFT JOIN funcionario fs ON fs.id = m.id_solicitante
            WHERE m.status = 'pendente'
            ORDER BY m.created_at DESC
            """
        )
        rows = _clean_rows(cursor.fetchall())
        return jsonify({'results': rows, 'total': len(rows)})
    except Exception as exc:
        app.logger.exception('api_solicitacao_moderacao_pendentes')
        return jsonify({'error': str(exc)}), 500
    finally:
        cursor.close()
        conn.close()


@app.route('/api/solicitacao/moderacao/minhas', methods=['GET'])
def api_solicitacao_moderacao_minhas():
    """Historico do usuario logado (perfil Cobranca): pedidos de moderacao enviados."""
    fid = session.get('funcionario_id')
    if not fid:
        return jsonify({'error': 'Nao autenticado.'}), 401
    conn = _get_db()
    cursor = conn.cursor()
    try:
        _ensure_solicitacao_moderacao_table(cursor)
        cursor.execute(
            """
            SELECT m.*, c.grupo, c.cota
            FROM solicitacao_moderacao m
            INNER JOIN contrato c ON c.id = m.id_contrato
            WHERE m.id_solicitante = %s
            ORDER BY m.created_at DESC
            LIMIT 300
            """,
            (int(fid),),
        )
        rows = _clean_rows(cursor.fetchall())
        return jsonify({'results': rows, 'total': len(rows)})
    except Exception as exc:
        app.logger.exception('api_solicitacao_moderacao_minhas')
        return jsonify({'error': str(exc)}), 500
    finally:
        cursor.close()
        conn.close()


@app.route('/api/solicitacao/moderacao/<int:mid>/decisao', methods=['POST'])
def api_solicitacao_moderacao_decisao(mid):
    """Gestor/Administrador: aprovar ou reprovar pedido de moderacao."""
    forb = _admin_json_forbidden()
    if forb:
        return forb
    body = request.get_json(silent=True) or {}
    acao = (body.get('acao') or '').strip().lower()
    motivo = (body.get('motivo') or '').strip()[:512] or None
    if acao not in ('aprovar', 'reprovar'):
        return jsonify({'error': 'Informe acao: aprovar ou reprovar.'}), 400
    if acao == 'reprovar' and not motivo:
        return jsonify({'error': 'Informe o motivo da reprovacao.'}), 400

    conn = _get_db()
    cursor = conn.cursor()
    try:
        _ensure_solicitacao_moderacao_table(cursor)
        _ensure_tramitacao_fluxo_columns(cursor)
        cursor.execute(
            'SELECT * FROM solicitacao_moderacao WHERE id = %s',
            (mid,),
        )
        mod = _clean_row(cursor.fetchone())
        if not mod:
            return jsonify({'error': 'Solicitacao nao encontrada.'}), 404
        if (mod.get('status') or '').lower() != 'pendente':
            return jsonify({'error': 'Esta solicitacao ja foi tratada.'}), 409

        rev_id = int(session.get('funcionario_id'))
        now = datetime.datetime.now()

        if acao == 'reprovar':
            cursor.execute(
                """
                UPDATE solicitacao_moderacao
                SET status = 'reprovado', id_revisor = %s, revisado_em = %s,
                    motivo_reprovacao = %s
                WHERE id = %s AND status = 'pendente'
                """,
                (rev_id, now, motivo, mid),
            )
            conn.commit()
            if cursor.rowcount == 0:
                return jsonify({'error': 'Nao foi possivel atualizar o registro.'}), 409
            return jsonify({'ok': True})

        tipo = (mod.get('tipo') or '').strip()
        ref_id = int(mod['ref_id'])
        id_ctr = int(mod['id_contrato'])
        payload = {}
        pj = mod.get('payload_json')
        if pj:
            try:
                payload = json.loads(pj)
            except Exception:
                payload = {}

        if tipo == 'tramitacao_delete':
            cursor.execute('SELECT id FROM tramitacao WHERE id = %s', (ref_id,))
            if not cursor.fetchone():
                return jsonify({'error': 'Tramitacao ja foi removida ou nao existe.'}), 409
            cursor.execute('DELETE FROM tramitacao WHERE id = %s', (ref_id,))
        elif tipo == 'tramitacao_edit':
            prop = payload.get('proposta') or {}
            _row, err = _tramitacao_aplicar_put_payload(cursor, ref_id, prop)
            if err:
                return jsonify({'error': 'Falha ao aplicar alteracao: ' + err}), 400
        elif tipo == 'agenda_delete':
            cursor.execute(
                'SELECT id FROM agenda WHERE id = %s AND id_contrato = %s',
                (ref_id, id_ctr),
            )
            if not cursor.fetchone():
                return jsonify({'error': 'Agendamento nao encontrado ou ja removido.'}), 409
            cursor.execute('DELETE FROM agenda WHERE id = %s', (ref_id,))
        elif tipo == 'agenda_edit':
            prop = payload.get('proposta') or {}
            st = prop.get('status')
            if st not in ('pendente', 'concluido'):
                return jsonify({'error': 'Payload da solicitacao invalido.'}), 400
            cursor.execute(
                """
                UPDATE agenda SET status = %s
                WHERE id = %s AND id_contrato = %s
                """,
                (st, ref_id, id_ctr),
            )
            if cursor.rowcount == 0:
                return jsonify({'error': 'Agendamento nao encontrado ou contrato divergente.'}), 409
        else:
            return jsonify({'error': 'Tipo de moderacao desconhecido.'}), 400

        cursor.execute(
            """
            UPDATE solicitacao_moderacao
            SET status = 'aprovado', id_revisor = %s, revisado_em = %s, motivo_reprovacao = NULL
            WHERE id = %s AND status = 'pendente'
            """,
            (rev_id, now, mid),
        )
        conn.commit()
        if cursor.rowcount == 0:
            return jsonify({'error': 'Nao foi possivel concluir a aprovacao.'}), 409
        return jsonify({'ok': True})
    except Exception as exc:
        if conn:
            try:
                conn.rollback()
            except Exception:
                pass
        app.logger.exception('api_solicitacao_moderacao_decisao')
        return jsonify({'error': str(exc)}), 500
    finally:
        cursor.close()
        conn.close()


@app.route('/api/protocolos')
def api_protocolos():
    conn = _get_db()
    cursor = conn.cursor()
    cursor.execute("""
        SELECT p.*, f1.nome AS remetente_nome, f2.nome AS destinatario_nome,
               c.grupo, c.cota
        FROM protocolo p
        JOIN funcionario f1 ON p.id_remetente = f1.id
        JOIN funcionario f2 ON p.id_destinatario = f2.id
        LEFT JOIN contrato c ON p.id_contrato = c.id
        ORDER BY p.data_envio DESC
    """)
    rows = _clean_rows(cursor.fetchall())
    cursor.close()
    conn.close()
    return jsonify({'results': rows, 'total': len(rows)})

@app.route('/api/solicitacoes')
def api_solicitacoes():
    conn = _get_db()
    cursor = conn.cursor()
    cursor.execute("""
        SELECT s.*, f1.nome AS remetente_nome, f2.nome AS destinatario_nome,
               c.grupo, c.cota
        FROM solicitacao s
        JOIN funcionario f1 ON s.id_remetente = f1.id
        JOIN funcionario f2 ON s.id_destinatario = f2.id
        LEFT JOIN contrato c ON s.id_contrato = c.id
        ORDER BY s.data_envio DESC
    """)
    rows = _clean_rows(cursor.fetchall())
    cursor.close()
    conn.close()
    return jsonify({'results': rows, 'total': len(rows)})

@app.route('/api/mensagens')
def api_mensagens():
    conn = _get_db()
    cursor = conn.cursor()
    cursor.execute("""
        SELECT m.*, f1.nome AS remetente_nome, f2.nome AS destinatario_nome
        FROM mensagem m
        JOIN funcionario f1 ON m.id_remetente = f1.id
        JOIN funcionario f2 ON m.id_destinatario = f2.id
        ORDER BY m.data_envio DESC
    """)
    rows = _clean_rows(cursor.fetchall())
    cursor.close()
    conn.close()
    return jsonify({'results': rows, 'total': len(rows)})


if __name__ == '__main__':
    app.run(host='0.0.0.0', debug=True, port=5000)
